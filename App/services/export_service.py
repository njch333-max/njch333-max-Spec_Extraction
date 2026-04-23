from __future__ import annotations

import csv
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from App.services import cleaning_rules, parsing
from App.services.runtime import ensure_job_dirs, safe_filename, utc_now_iso

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", start_color="2F5597")
SUBHEAD_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
SUBHEAD_FILL = PatternFill("solid", start_color="70AD47")
BODY_FONT = Font(name="Arial", size=10)
FLAG_FILL = PatternFill("solid", start_color="FFE699")
SUMMARY_BG = PatternFill("solid", start_color="DDEBF7")
AREA_FONT = Font(name="Arial", bold=True, size=10)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top", horizontal="left")
CENTER = Alignment(horizontal="center", vertical="center")


def build_exports(job_no: str, snapshot: dict[str, Any]) -> dict[str, str]:
    dirs = ensure_job_dirs(job_no)
    stamp = utc_now_iso().replace(":", "").replace("-", "")
    excel_path = dirs["export_dir"] / f"{safe_filename(job_no)}_{stamp}.xlsx"
    csv_path = dirs["export_dir"] / f"{safe_filename(job_no)}_{stamp}.csv"
    _write_excel(excel_path, snapshot)
    _write_csv(csv_path, snapshot)
    return {"excel": str(excel_path), "csv": str(csv_path)}


def build_spec_list_excel(job_no: str, snapshot: dict[str, Any]) -> str:
    dirs = ensure_job_dirs(job_no)
    stamp = utc_now_iso().replace(":", "").replace("-", "")
    excel_path = dirs["export_dir"] / f"{safe_filename(job_no)}_spec_list_{stamp}.xlsx"
    _write_review_excel(excel_path, job_no, snapshot)
    return str(excel_path)


def _write_review_excel(path: Path, job_no: str, snapshot: dict[str, Any]) -> None:
    sections, flagged_items, material_summary_rows, counts = _review_sections_from_snapshot(snapshot)
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _write_review_summary(ws_summary, job_no, snapshot, counts)

    ws_review = wb.create_sheet("By Section")
    _write_by_section(ws_review, sections)
    if flagged_items:
        ws_flagged = wb.create_sheet("Flagged")
        _write_flagged(ws_flagged, flagged_items)
    if material_summary_rows:
        ws_material = wb.create_sheet("Material Summary")
        _write_material_summary(ws_material, material_summary_rows)

    wb.save(path)


def _write_review_summary(ws: Any, job_no: str, snapshot: dict[str, Any], counts: dict[str, Any]) -> None:
    display_job_no = _display_value(snapshot.get("job_no", "")) or job_no
    ws.append([f"Job {display_job_no} - extraction report"])
    ws["A1"].font = Font(name="Arial", bold=True, size=14)
    ws.append([_source_pdf_name(snapshot)])
    ws["A2"].font = Font(name="Arial", italic=True, size=10, color="595959")
    ws.append([])
    metrics = [
        ("Pages", counts.get("page_count", 0)),
        ("Sections", counts.get("section_count", 0)),
        ("Total items", counts.get("item_count", 0)),
        ("Flagged items", counts.get("flagged_count", 0)),
        ("Clean rate", _clean_rate_display(counts)),
        ("", ""),
        ("Bench Tops rows", counts.get("bench_tops", 0)),
        ("Door Colours rows", counts.get("door_colours", 0)),
        ("Handles rows", counts.get("handles", 0)),
    ]
    for key, value in metrics:
        ws.append([key, value])
        row = ws[ws.max_row]
        if key:
            row[0].font = Font(name="Arial", bold=True, size=10)
            row[0].alignment = Alignment(horizontal="left")
            row[0].fill = SUMMARY_BG
            row[1].alignment = Alignment(horizontal="left")
    ws.append([])
    ws.append(["Review instructions"])
    ws[ws.max_row][0].font = Font(name="Arial", bold=True, size=12)
    guide = [
        "1. Open the source PDF alongside this workbook.",
        "2. Go to sheet 'By Section' - it lists every extracted row section by section.",
        "3. For each row, check whether AREA / SPECS / SUPPLIER / NOTES match the PDF cell content.",
        "4. Yellow-highlighted rows were auto-flagged as edge cases - pay extra attention to these.",
        "5. 'Material Summary' sheet filters to the 3 fields you said matter most (Bench Tops / Door Colours / Handles).",
        "6. If you find wrong rows, note the section + row number so Jason can fix the extractor.",
    ]
    for item in guide:
        ws.append([item])
        ws[ws.max_row][0].font = BODY_FONT
    _set_column_widths(ws, [30, 60])


def _write_by_section(ws: Any, sections: list[dict[str, Any]]) -> None:
    headers = ["Section / Area", "Specs / Description", "Supplier", "Notes", "Page", "Flag"]
    ws.append(headers)
    _style_header(ws[1])

    for section in sections:
        items = list(section.get("items", []) or [])
        if not items:
            continue
        title = _section_title_with_metadata(section)
        pages_display = section.get("pages_display", "")
        page_label = f"p{pages_display}" if pages_display else ""
        ws.append([title, "", "", "", page_label, ""])
        _style_subhead(ws[ws.max_row])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=4)
        for item in items:
            flag_reason = _display_value(item.get("flag_reason", ""))
            ws.append(
                [
                    _display_value(item.get("area", "")),
                    _display_value(item.get("specs", "")),
                    _display_value(item.get("supplier", "")),
                    _display_value(item.get("notes", "")),
                    item.get("page", ""),
                    "Y" if flag_reason else "",
                ]
            )
            row = ws[ws.max_row]
            _style_body(row, flagged=bool(flag_reason))
            row[0].font = AREA_FONT

    _set_column_widths(ws, [34, 48, 18, 32, 6, 6])
    ws.freeze_panes = "A2"


def _write_flagged(ws: Any, flagged_items: list[dict[str, Any]]) -> None:
    ws.append(["Section", "Area", "Supplier", "Specs", "Notes", "Flag reason"])
    _style_header(ws[1])
    for item in flagged_items:
        ws.append(
            [
                _display_value(item.get("section_title", "")),
                _display_value(item.get("area", "")),
                _display_value(item.get("supplier", "")),
                _display_value(item.get("specs", "")),
                _display_value(item.get("notes", "")),
                _display_value(item.get("flag_reason", "")),
            ]
        )
        _style_body(ws[ws.max_row], flagged=True)
    _set_column_widths(ws, [30, 26, 16, 42, 30, 30])
    ws.freeze_panes = "A2"


def _write_material_summary(ws: Any, material_rows: list[dict[str, Any]]) -> None:
    ws.append(["This sheet filters to Bench Tops / Door Colours / Handles only - the 3 fields you said matter most for QA."])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws["A1"].font = Font(name="Arial", italic=True, size=10, color="595959")
    ws.append(["Category", "Section", "Area", "Supplier", "Specs / Description", "Notes"])
    _style_header(ws[2])
    for row in material_rows:
        ws.append(
            [
                _display_value(row.get("category", "")),
                _display_value(row.get("section", "")),
                _display_value(row.get("area", "")),
                _display_value(row.get("supplier", "")),
                _display_value(row.get("specs", "")),
                _display_value(row.get("notes", "")),
            ]
        )
        _style_body(ws[ws.max_row])
    _set_column_widths(ws, [14, 32, 28, 16, 42, 30])
    ws.freeze_panes = "A3"


def _review_sections_from_snapshot(
    snapshot: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    sections: list[dict[str, Any]] = []
    flagged_items: list[dict[str, Any]] = []
    material_summary_rows: list[dict[str, Any]] = []
    counts: dict[str, Any] = {
        "item_count": 0,
        "flagged_count": 0,
        "bench_tops": 0,
        "door_colours": 0,
        "handles": 0,
    }

    rooms = _ordered_rooms(snapshot.get("rooms", []))
    imperial_material_mode = parsing._is_imperial_builder(str(snapshot.get("builder_name", "") or "")) and any(
        _mapping_rows(room.get("material_rows", [])) for room in rooms
    )
    if imperial_material_mode:
        for room in rooms:
            material_rows = _ordered_material_rows(room.get("v6_review_rows", [])) or _ordered_material_rows(room.get("material_rows", []))
            if not material_rows:
                continue
            section = {
                "section_title": _imperial_section_title(room, material_rows),
                "metadata": dict(room.get("v6_metadata") or {}),
                "pages": _section_pages(room, material_rows),
                "items": [],
            }
            for material_row in material_rows:
                item = _review_item(
                    section_title=str(section["section_title"]),
                    area=_display_value(material_row.get("area_or_item", "")),
                    specs=_display_value(
                        material_row.get("specs_or_description", "")
                        or material_row.get("specs_description", "")
                        or material_row.get("description", "")
                        or material_row.get("value", "")
                    ),
                    supplier=_display_value(material_row.get("supplier", "")),
                    notes=_display_value(material_row.get("notes", "")),
                    page=_page_display(material_row),
                    flag_reason=_row_flag(material_row),
                )
                item["category"] = _category_name_from_key(_material_category(material_row)) or item.get("category", "")
                section["items"].append(item)
                _track_review_item(counts, flagged_items, material_summary_rows, item)
            sections.append(section)
    else:
        for room in rooms:
            section = {
                "section_title": _room_label(room),
                "metadata": {},
                "pages": _section_pages(room, []),
                "items": [],
            }
            for field_name, field_value, category in _room_field_rows(room):
                item = _review_item(
                    section_title=str(section["section_title"]),
                    area=field_name,
                    specs=field_value,
                    page=_page_display(room),
                    flag_reason=_row_flag(room),
                )
                item["category"] = _category_name_from_key(category) or _classify_review_row(field_name, field_value) or ""
                section["items"].append(item)
                _track_review_item(counts, flagged_items, material_summary_rows, item)
            if section["items"]:
                sections.append(section)

    if imperial_material_mode:
        _finalize_review_counts(sections, counts)
        return sections, flagged_items, _sorted_material_summary(material_summary_rows), counts

    appliance_items: list[dict[str, Any]] = []
    for appliance in _mapping_rows(snapshot.get("appliances", [])):
        if not _include_appliance_row(appliance):
            continue
        specs = _join_parts(
            [
                _display_value(appliance.get("make", "")),
                _display_value(appliance.get("model_no", "")),
                _display_value(appliance.get("overall_size", "")),
                _display_value(appliance.get("product_url", "") or appliance.get("website_url", "")),
            ]
        )
        item = _review_item(
            section_title="APPLIANCES",
            area=_display_value(appliance.get("appliance_type", "")) or "Appliance",
            specs=specs,
            page=_page_display(appliance),
            flag_reason=_row_flag(appliance),
        )
        appliance_items.append(item)
        _track_review_item(counts, flagged_items, material_summary_rows, item)
    if appliance_items:
        sections.append({"section_title": "APPLIANCES", "metadata": {}, "pages": _pages_from_items(appliance_items), "items": appliance_items})

    for section_row in _mapping_rows(snapshot.get("special_sections", [])):
        section_title = _display_value(section_row.get("original_section_label", "")) or _display_value(section_row.get("section_key", "")) or "SPECIAL SECTIONS"
        section = {"section_title": section_title, "metadata": {}, "pages": _section_pages(section_row, []), "items": []}
        fields = section_row.get("fields") or {}
        if isinstance(fields, dict) and fields:
            for field_name, field_value in fields.items():
                item = _review_item(
                    section_title=section_title,
                    area=_display_value(field_name),
                    specs=_display_value(field_value),
                    page=_page_display(section_row),
                    flag_reason=_row_flag(section_row),
                )
                section["items"].append(item)
                _track_review_item(counts, flagged_items, material_summary_rows, item)
        else:
            item = _review_item(
                section_title=section_title,
                area="",
                specs="",
                page=_page_display(section_row),
                flag_reason=_row_flag(section_row),
            )
            section["items"].append(item)
            _track_review_item(counts, flagged_items, material_summary_rows, item)
        if section["items"]:
            sections.append(section)

    others = snapshot.get("others") or {}
    other_items: list[dict[str, Any]] = []
    if isinstance(others, dict):
        for key, value in others.items():
            item = _review_item(
                section_title="OTHERS",
                area=_display_value(key),
                specs=_display_value(value),
            )
            other_items.append(item)
            _track_review_item(counts, flagged_items, material_summary_rows, item)
    elif others:
        item = _review_item(section_title="OTHERS", area="Notes", specs=_display_value(others))
        other_items.append(item)
        _track_review_item(counts, flagged_items, material_summary_rows, item)
    if other_items:
        sections.append({"section_title": "OTHERS", "metadata": {}, "pages": _pages_from_items(other_items), "items": other_items})

    warning_items: list[dict[str, Any]] = []
    for warning in _string_list(snapshot.get("warnings", [])):
        item = _review_item(
            section_title="WARNINGS",
            area="Warning",
            specs=warning,
            flag_reason="Warning",
        )
        warning_items.append(item)
        _track_review_item(counts, flagged_items, material_summary_rows, item)
    if warning_items:
        sections.append({"section_title": "WARNINGS", "metadata": {}, "pages": _pages_from_items(warning_items), "items": warning_items})

    _finalize_review_counts(sections, counts)
    return sections, flagged_items, _sorted_material_summary(material_summary_rows), counts


def _style_header(cells: Any) -> None:
    for cell in cells:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def _style_subhead(cells: Any) -> None:
    for index, cell in enumerate(cells):
        cell.font = SUBHEAD_FONT
        cell.fill = SUBHEAD_FILL
        cell.alignment = Alignment(horizontal="left" if index == 0 else "center", vertical="center", wrap_text=True)
        cell.border = BORDER


def _style_body(cells: Any, flagged: bool = False) -> None:
    for cell in cells:
        cell.font = BODY_FONT
        cell.alignment = WRAP
        cell.border = BORDER
        if flagged:
            cell.fill = FLAG_FILL


def _set_column_widths(ws: Any, widths: list[int]) -> None:
    for column_index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(column_index)].width = width


def _review_item(
    *,
    section_title: str,
    area: Any,
    specs: Any,
    supplier: Any = "",
    notes: Any = "",
    page: Any = "",
    flag_reason: Any = "",
) -> dict[str, Any]:
    display_area = _display_value(area)
    display_specs = _display_value(specs)
    return {
        "section_title": _display_value(section_title),
        "area": display_area,
        "specs": display_specs,
        "supplier": _display_value(supplier),
        "notes": _display_value(notes),
        "page": _page_cell_value(page),
        "flag_reason": _display_value(flag_reason),
        "category": _classify_review_row(display_area, display_specs) or "",
    }


def _track_review_item(
    counts: dict[str, Any],
    flagged_items: list[dict[str, Any]],
    material_summary_rows: list[dict[str, Any]],
    item: dict[str, Any],
) -> None:
    counts["item_count"] = int(counts.get("item_count", 0)) + 1
    category = _category_name_from_key(item.get("category", "")) or _classify_review_row(item.get("area", ""), item.get("specs", "")) or ""
    if category:
        item["category"] = category
    if _display_value(item.get("flag_reason", "")):
        counts["flagged_count"] = int(counts.get("flagged_count", 0)) + 1
        flagged_items.append(dict(item))
    count_key = {
        "Bench Tops": "bench_tops",
        "Door Colours": "door_colours",
        "Handles": "handles",
    }.get(category)
    if not count_key:
        return
    counts[count_key] = int(counts.get(count_key, 0)) + 1
    material_summary_rows.append(
        {
            "_order": len(material_summary_rows),
            "category": category,
            "section": _plain_section_title(item.get("section_title", "")),
            "area": item.get("area", ""),
            "supplier": item.get("supplier", ""),
            "specs": item.get("specs", ""),
            "notes": item.get("notes", ""),
        }
    )


def _finalize_review_counts(sections: list[dict[str, Any]], counts: dict[str, Any]) -> None:
    all_pages: set[int] = set()
    section_count = 0
    for section in sections:
        items = list(section.get("items", []) or [])
        if not items:
            continue
        section_count += 1
        pages: set[int] = set(_extract_page_numbers(section.get("pages", [])))
        for item in items:
            pages.update(_extract_page_numbers(item.get("page", "")))
        ordered_pages = sorted(pages)
        section["pages"] = ordered_pages
        section["pages_display"] = ordered_pages if ordered_pages else ""
        all_pages.update(ordered_pages)
    counts["section_count"] = section_count
    counts["page_count"] = len(all_pages)


def _sorted_material_summary(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    category_order = {"Bench Tops": 0, "Door Colours": 1, "Handles": 2}
    return sorted(
        rows,
        key=lambda row: (
            category_order.get(_display_value(row.get("category", "")), 99),
            _display_value(row.get("section", "")).upper(),
            _display_value(row.get("specs", "")).upper(),
            _safe_int(row.get("_order")) or 0,
        ),
    )


def _section_title_with_metadata(section: dict[str, Any]) -> str:
    title = _display_value(section.get("section_title", "")) or "Section"
    metadata = section.get("metadata") if isinstance(section.get("metadata"), dict) else {}
    parts: list[str] = []
    ceiling_height = _display_value(metadata.get("ceiling_height", ""))
    cabinetry_height = _display_value(metadata.get("cabinetry_height", ""))
    bulkhead = _display_value(metadata.get("bulkhead", ""))
    if ceiling_height:
        parts.append(f"Ceiling {ceiling_height}")
    if cabinetry_height:
        parts.append(f"Cabinetry {cabinetry_height}")
    if bulkhead:
        parts.append(f"Bulkhead: {bulkhead}")
    return f"{title}   ({' | '.join(parts)})" if parts else title


def _source_pdf_name(snapshot: dict[str, Any]) -> str:
    for document in _mapping_rows(snapshot.get("source_documents", [])):
        file_name = _display_value(document.get("file_name", ""))
        if file_name:
            return file_name
    for room in _mapping_rows(snapshot.get("rooms", [])):
        source_file = _display_value(room.get("source_file", ""))
        if source_file:
            return source_file
    return _display_value(snapshot.get("job_no", ""))


def _clean_rate_display(counts: dict[str, Any]) -> str:
    total = int(counts.get("item_count", 0) or 0)
    if total <= 0:
        return "n/a"
    flagged = int(counts.get("flagged_count", 0) or 0)
    return f"{((total - flagged) / total) * 100:.1f}%"


def _imperial_section_title(room: dict[str, Any], material_rows: list[dict[str, Any]]) -> str:
    for row in material_rows:
        provenance = row.get("provenance") if isinstance(row.get("provenance"), dict) else {}
        title = _display_value(provenance.get("section_title", ""))
        if title:
            return title
    title = _display_value(room.get("section_title", "")) or _room_label(room)
    upper_title = title.upper()
    if "SELECTION SHEET" in upper_title or upper_title in {"APPLIANCES", "SINKWARE", "SINKWARE & TAPWARE", "WARNINGS", "OTHERS"}:
        return title
    return f"{title} JOINERY SELECTION SHEET"


def _section_pages(row: dict[str, Any], material_rows: list[dict[str, Any]]) -> list[int]:
    pages: list[int] = []
    for key in ("page_refs", "page_no", "page"):
        pages.extend(_extract_page_numbers(row.get(key, "")))
    for material_row in material_rows:
        for key in ("page_refs", "page_no", "page"):
            pages.extend(_extract_page_numbers(material_row.get(key, "")))
    return sorted(dict.fromkeys(page for page in pages if page > 0))


def _pages_from_items(items: list[dict[str, Any]]) -> list[int]:
    pages: list[int] = []
    for item in items:
        pages.extend(_extract_page_numbers(item.get("page", "")))
    return sorted(dict.fromkeys(page for page in pages if page > 0))


def _category_name_from_key(category: Any) -> str:
    text = _display_value(category)
    normalized = text.strip().lower().replace("-", "_").replace(" ", "_")
    return {
        "bench_tops": "Bench Tops",
        "benchtops": "Bench Tops",
        "bench_top": "Bench Tops",
        "bench_tops_rows": "Bench Tops",
        "door_colours": "Door Colours",
        "door_colors": "Door Colours",
        "door_colour": "Door Colours",
        "door_color": "Door Colours",
        "handles": "Handles",
        "handle": "Handles",
    }.get(normalized, text if text in {"Bench Tops", "Door Colours", "Handles"} else "")


def _classify_review_row(area_text: Any, specs_text: Any) -> str:
    area_upper = _display_value(area_text).upper()
    specs_upper = _display_value(specs_text).upper()
    if "HANDLE" in area_upper or "FINGER PULL" in area_upper:
        return "Handles"
    if any(token in specs_upper for token in ("PUSH TO OPEN", "FINGER PULL", "NO HANDLES", "TOUCH CATCH")):
        return "Handles"
    if ("BENCH" in area_upper and "TOP" in area_upper) or "WFALL" in area_upper or "WATERFALL" in area_upper:
        return "Bench Tops"
    door_markers = (
        "CABINETRY COLOUR",
        "CABINETRY FINISH",
        "DOOR COLOUR",
        "DOOR & PANEL",
        "DOOR/PANEL",
        "TALL DOORS",
        "CABINETRY / FINISHES",
        "BASE COLOUR",
    )
    if any(marker in area_upper for marker in door_markers):
        return "Door Colours"
    return ""


def _page_cell_value(value: Any) -> str | int:
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = _display_value(value)
    if not text:
        return ""
    pages = _extract_page_numbers(text)
    if len(pages) == 1 and text.strip().isdigit():
        return pages[0]
    return text


def _extract_page_numbers(value: Any) -> list[int]:
    if value is None:
        return []
    if isinstance(value, int):
        return [value] if value > 0 else []
    if isinstance(value, float):
        return [int(value)] if value.is_integer() and value > 0 else []
    if isinstance(value, (list, tuple, set)):
        pages: list[int] = []
        for item in value:
            pages.extend(_extract_page_numbers(item))
        return sorted(dict.fromkeys(pages))
    text = _display_value(value)
    if not text:
        return []
    pages: list[int] = []
    for match in re.finditer(r"\d+\s*-\s*\d+|\d+", text):
        token = match.group(0)
        if "-" in token:
            start_text, end_text = re.split(r"\s*-\s*", token, maxsplit=1)
            start = _safe_int(start_text)
            end = _safe_int(end_text)
            if start is None or end is None:
                continue
            if start <= end and end - start <= 100:
                pages.extend(range(start, end + 1))
            else:
                pages.extend([start, end])
        else:
            page = _safe_int(token)
            if page is not None:
                pages.append(page)
    return sorted(dict.fromkeys(page for page in pages if page > 0))


def _plain_section_title(value: Any) -> str:
    title = _display_value(value)
    return title.split("   (", 1)[0]


def _ordered_rooms(value: Any) -> list[dict[str, Any]]:
    rooms = _mapping_rows(value)
    if not any(_safe_int(room.get("room_order")) is not None for room in rooms):
        return rooms
    return [room for _, room in sorted(enumerate(rooms), key=lambda item: (_safe_int(item[1].get("room_order")) is None, _safe_int(item[1].get("room_order")) or 0, item[0]))]


def _ordered_material_rows(value: Any) -> list[dict[str, Any]]:
    rows = _mapping_rows(value)
    if not any(_safe_int(row.get("row_order")) is not None for row in rows):
        return rows
    return [row for _, row in sorted(enumerate(rows), key=lambda item: (_safe_int(item[1].get("row_order")) is None, _safe_int(item[1].get("row_order")) or 0, item[0]))]


def _room_label(room: dict[str, Any]) -> str:
    return _display_value(room.get("original_room_label", "")) or _display_value(room.get("room_key", "")) or "Room"


def _room_field_rows(room: dict[str, Any]) -> list[tuple[str, str, str]]:
    rows: list[tuple[str, str, str]] = []
    benchtop_groups = _split_export_benchtops(room)
    field_specs: list[tuple[str, Any, str]] = [
        ("Bench Tops", room.get("bench_tops", ""), "bench_tops"),
        ("Wall Run Bench Top", benchtop_groups["bench_tops_wall_run"], "bench_tops"),
        ("Island Bench Top", benchtop_groups["bench_tops_island"], "bench_tops"),
        ("Other Bench Top", benchtop_groups["bench_tops_other"], "bench_tops"),
        ("Floating Shelf", room.get("floating_shelf", ""), ""),
        ("Shelf", room.get("shelf", ""), ""),
        ("Door Panel Colours", room.get("door_panel_colours", ""), "door_colours"),
        ("Overhead Door Colours", room.get("door_colours_overheads", ""), "door_colours"),
        ("Base Door Colours", room.get("door_colours_base", ""), "door_colours"),
        ("Tall Door Colours", room.get("door_colours_tall", ""), "door_colours"),
        ("Island Door Colours", room.get("door_colours_island", ""), "door_colours"),
        ("Bar Back Door Colours", room.get("door_colours_bar_back", ""), "door_colours"),
        ("Feature Colour", room.get("feature_colour", ""), "door_colours"),
        ("Toe Kick", room.get("toe_kick", ""), ""),
        ("Bulkheads", room.get("bulkheads", ""), ""),
        ("Handles", room.get("handles", ""), "handles"),
        ("LED", "Yes" if parsing.normalize_space(str(room.get("led", ""))).lower() == "yes" else "", ""),
        ("LED Note", room.get("led_note", ""), ""),
        ("Accessories", room.get("accessories", ""), ""),
        ("Other Items", _display_other_items(room.get("other_items", [])), ""),
        ("Sink", room.get("sink_info", ""), ""),
        ("Basin", room.get("basin_info", ""), ""),
        ("Tap", room.get("tap_info", ""), ""),
        ("Drawers", _normalize_soft_close(room.get("drawers_soft_close", ""), "drawer"), ""),
        ("Hinges", _normalize_soft_close(room.get("hinges_soft_close", ""), "hinge"), ""),
        ("Splashback", room.get("splashback", ""), ""),
        ("Flooring", room.get("flooring", ""), ""),
    ]
    seen: set[tuple[str, str]] = set()
    for label, value, category in field_specs:
        display_value = _display_value(value)
        if not display_value:
            continue
        key = (label, display_value)
        if key in seen:
            continue
        seen.add(key)
        rows.append((label, display_value, category))
    return rows


def _material_category(row: dict[str, Any]) -> str:
    tags = [tag.lower() for tag in parsing._coerce_string_list(row.get("tags", []))]
    text = " ".join(
        [
            _display_value(row.get("area_or_item", "")),
            _display_value(row.get("summary_tag", "")),
            " ".join(tags),
        ]
    ).lower()
    if "handle" in text:
        return "handles"
    if "bench" in text:
        return "bench_tops"
    if any(token in text for token in ("door_colour", "door color", "door colour", "cabinetry colour", "feature cabinetry", "colour")):
        return "door_colours"
    return ""


def _row_flag(row: dict[str, Any]) -> str:
    flags: list[str] = []
    if row.get("needs_review"):
        flags.append("Needs review")
    for key in ("review_hint", "_review_hint", "parser_warning", "row_issue", "issue"):
        value = _display_value(row.get(key, ""))
        if value:
            flags.append(value)
    for key in ("issues", "warnings"):
        for value in _string_list(row.get(key, [])):
            flags.append(value)
    revalidation_status = _display_value(row.get("revalidation_status", "")).lower()
    if revalidation_status and revalidation_status not in {"accepted", "passed", "ok", "clean"}:
        flags.append(f"Revalidation: {revalidation_status}")
    return " | ".join(_dedupe_strings(flags))


def _page_display(row: dict[str, Any]) -> str:
    return _display_value(row.get("page_refs", "")) or _display_value(row.get("page_no", "")) or _display_value(row.get("page", ""))


def _join_parts(parts: list[str]) -> str:
    return " | ".join(part for part in (_display_value(part) for part in parts) if part)


def _dedupe_strings(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        text = _display_value(value).strip()
        key = text.lower()
        if text and key not in seen:
            seen.add(key)
            result.append(text)
    return result


def _safe_int(value: Any) -> int | None:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _write_excel(path: Path, snapshot: dict[str, Any]) -> None:
    wb = Workbook()
    ws_rooms = wb.active
    ws_rooms.title = "Rooms"
    ws_rooms.append(
        [
            "room_key",
            "original_room_label",
            "bench_tops",
            "bench_tops_wall_run",
            "bench_tops_island",
            "bench_tops_other",
            "floating_shelf",
            "shelf",
            "door_panel_colours",
            "door_colours_overheads",
            "door_colours_base",
            "door_colours_tall",
            "door_colours_island",
            "door_colours_bar_back",
            "feature_colour",
            "toe_kick",
            "bulkheads",
            "handles",
            "led",
            "led_note",
            "accessories",
            "other_items",
            "sink_info",
            "basin_info",
            "tap_info",
            "drawers_soft_close",
            "hinges_soft_close",
            "splashback",
            "flooring",
            "source_file",
            "page_refs",
            "evidence_snippet",
            "confidence",
        ]
    )
    for row in _mapping_rows(snapshot.get("rooms", [])):
        benchtop_groups = _split_export_benchtops(row)
        ws_rooms.append(
            [
                _display_value(row.get("room_key", "")),
                _display_value(row.get("original_room_label", "")),
                _display_value(row.get("bench_tops", [])),
                benchtop_groups["bench_tops_wall_run"],
                benchtop_groups["bench_tops_island"],
                benchtop_groups["bench_tops_other"],
                _display_value(row.get("floating_shelf", "")),
                _display_value(row.get("shelf", "")),
                _display_value(row.get("door_panel_colours", [])),
                _display_value(row.get("door_colours_overheads", "")),
                _display_value(row.get("door_colours_base", "")),
                _display_value(row.get("door_colours_tall", "")),
                _display_value(row.get("door_colours_island", "")),
                _display_value(row.get("door_colours_bar_back", "")),
                _display_value(row.get("feature_colour", "")),
                _display_value(row.get("toe_kick", [])),
                _display_value(row.get("bulkheads", [])),
                _display_value(row.get("handles", [])),
                "Yes" if parsing.normalize_space(str(row.get("led", ""))).lower() == "yes" else "No",
                _display_value(row.get("led_note", "")),
                _display_value(row.get("accessories", [])),
                _display_other_items(row.get("other_items", [])),
                _display_value(row.get("sink_info", "")),
                _display_value(row.get("basin_info", "")),
                _display_value(row.get("tap_info", "")),
                _normalize_soft_close(row.get("drawers_soft_close", ""), "drawer"),
                _normalize_soft_close(row.get("hinges_soft_close", ""), "hinge"),
                _display_value(row.get("splashback", "")),
                _display_value(row.get("flooring", "")),
                _display_value(row.get("source_file", "")),
                _display_value(row.get("page_refs", "")),
                _display_value(row.get("evidence_snippet", "")),
                _display_value(row.get("confidence", "")),
            ]
        )

    ws_appliances = wb.create_sheet("Appliances")
    ws_appliances.append(
        [
            "appliance_type",
            "make",
            "model_no",
            "product_url",
            "website_url",
            "overall_size",
            "source_file",
            "page_refs",
            "evidence_snippet",
            "confidence",
        ]
    )
    for row in _mapping_rows(snapshot.get("appliances", [])):
        if not _include_appliance_row(row):
            continue
        product_url = _display_value(row.get("product_url", "") or row.get("website_url", ""))
        ws_appliances.append(
            [
                _display_value(row.get("appliance_type", "")),
                _display_value(row.get("make", "")),
                _display_value(row.get("model_no", "")),
                product_url,
                product_url,
                _display_value(row.get("overall_size", "")),
                _display_value(row.get("source_file", "")),
                _display_value(row.get("page_refs", "")),
                _display_value(row.get("evidence_snippet", "")),
                _display_value(row.get("confidence", "")),
            ]
        )
        current_row = ws_appliances.max_row
        for column in ("D", "E", "F", "I"):
            ws_appliances[f"{column}{current_row}"].alignment = Alignment(wrap_text=True, vertical="top")
        for column in ("D", "E"):
            cell = ws_appliances[f"{column}{current_row}"]
            if cell.value:
                cell.hyperlink = str(cell.value)
                cell.style = "Hyperlink"

    ws_others = wb.create_sheet("Others")
    ws_others.append(["key", "value"])
    others = snapshot.get("others") or {}
    if isinstance(others, dict):
        for key, value in others.items():
            ws_others.append([_display_value(key), _display_value(value)])
    elif others:
        ws_others.append(["notes", _display_value(others)])

    ws_special = wb.create_sheet("Special Sections")
    ws_special.append(
        [
            "section_key",
            "original_section_label",
            "field_key",
            "field_value",
            "source_file",
            "page_refs",
            "evidence_snippet",
            "confidence",
        ]
    )
    for row in _mapping_rows(snapshot.get("special_sections", [])):
        fields = row.get("fields") or {}
        if not isinstance(fields, dict):
            fields = {}
        if fields:
            for field_key, field_value in fields.items():
                ws_special.append(
                    [
                        _display_value(row.get("section_key", "")),
                        _display_value(row.get("original_section_label", "")),
                        _display_value(field_key),
                        _display_value(field_value),
                        _display_value(row.get("source_file", "")),
                        _display_value(row.get("page_refs", "")),
                        _display_value(row.get("evidence_snippet", "")),
                        _display_value(row.get("confidence", "")),
                    ]
                )
        else:
            ws_special.append(
                [
                    _display_value(row.get("section_key", "")),
                    _display_value(row.get("original_section_label", "")),
                    "",
                    "",
                    _display_value(row.get("source_file", "")),
                    _display_value(row.get("page_refs", "")),
                    _display_value(row.get("evidence_snippet", "")),
                    _display_value(row.get("confidence", "")),
                ]
            )

    ws_warnings = wb.create_sheet("Warnings")
    ws_warnings.append(["warning"])
    for warning in _string_list(snapshot.get("warnings", [])):
        ws_warnings.append([warning])

    ws_meta = wb.create_sheet("Meta")
    ws_meta.append(["key", "value"])
    for key in ("job_no", "builder_name", "source_kind", "generated_at"):
        ws_meta.append([key, snapshot.get(key, "")])
    analysis = dict(snapshot.get("analysis") or {})
    for key in ("mode", "parser_strategy", "openai_attempted", "openai_succeeded", "openai_model", "note", "rule_config_updated_at", "worker_pid", "app_build_id"):
        if key in analysis:
            ws_meta.append([f"analysis_{key}", analysis.get(key, "")])
    ws_meta.append(["analysis_rule_flags", json.dumps(cleaning_rules.normalize_rule_flags(analysis.get("rule_flags", {})), ensure_ascii=False)])
    for document in _mapping_rows(snapshot.get("source_documents", [])):
        ws_meta.append(["source_document", f"{_display_value(document.get('role', ''))}: {_display_value(document.get('file_name', ''))}"])

    wb.save(path)


def _write_csv(path: Path, snapshot: dict[str, Any]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["section", "primary_key", "field", "value"])
        for row in _mapping_rows(snapshot.get("rooms", [])):
            room_key = _display_value(row.get("room_key", ""))
            benchtop_groups = _split_export_benchtops(row)
            for field_name in (
                "original_room_label",
                "bench_tops_wall_run",
                "bench_tops_island",
                "bench_tops_other",
                "floating_shelf",
                "shelf",
                "door_colours_overheads",
                "door_colours_base",
                "door_colours_tall",
                "door_colours_island",
                "door_colours_bar_back",
                "feature_colour",
                "led",
                "led_note",
                "sink_info",
                "basin_info",
                "tap_info",
                "splashback",
                "flooring",
                "source_file",
                "page_refs",
                "evidence_snippet",
                "confidence",
            ):
                field_value = benchtop_groups.get(field_name, _display_value(row.get(field_name, "")))
                if field_name == "led":
                    field_value = "Yes" if parsing.normalize_space(str(row.get("led", ""))).lower() == "yes" else "No"
                writer.writerow(["rooms", room_key, field_name, field_value])
            writer.writerow(["rooms", room_key, "drawers_soft_close", _normalize_soft_close(row.get("drawers_soft_close", ""), "drawer")])
            writer.writerow(["rooms", room_key, "hinges_soft_close", _normalize_soft_close(row.get("hinges_soft_close", ""), "hinge")])
            for field_name in ("bench_tops", "door_panel_colours", "toe_kick", "bulkheads", "handles", "accessories"):
                writer.writerow(["rooms", room_key, field_name, _display_value(row.get(field_name, []))])
            writer.writerow(["rooms", room_key, "other_items", _display_other_items(row.get("other_items", []))])
        for row in _mapping_rows(snapshot.get("appliances", [])):
            if not _include_appliance_row(row):
                continue
            row_key = f"{_display_value(row.get('appliance_type', ''))}:{_display_value(row.get('model_no', ''))}"
            product_url = _display_value(row.get("product_url", "") or row.get("website_url", ""))
            for field_name, field_value in (
                ("make", row.get("make", "")),
                ("model_no", row.get("model_no", "")),
                ("product_url", product_url),
                ("website_url", product_url),
                ("overall_size", row.get("overall_size", "")),
                ("source_file", row.get("source_file", "")),
                ("page_refs", row.get("page_refs", "")),
                ("evidence_snippet", row.get("evidence_snippet", "")),
                ("confidence", row.get("confidence", "")),
            ):
                writer.writerow(["appliances", row_key, field_name, _display_value(field_value)])
        others = snapshot.get("others") or {}
        if isinstance(others, dict):
            for key, value in others.items():
                writer.writerow(["others", "others", _display_value(key), _display_value(value)])
        elif others:
            writer.writerow(["others", "others", "notes", _display_value(others)])
        for row in _mapping_rows(snapshot.get("special_sections", [])):
            fields = row.get("fields") or {}
            row_key = _display_value(row.get("section_key", "")) or _display_value(row.get("original_section_label", ""))
            writer.writerow(["special_sections", row_key, "original_section_label", _display_value(row.get("original_section_label", ""))])
            writer.writerow(["special_sections", row_key, "source_file", _display_value(row.get("source_file", ""))])
            writer.writerow(["special_sections", row_key, "page_refs", _display_value(row.get("page_refs", ""))])
            writer.writerow(["special_sections", row_key, "evidence_snippet", _display_value(row.get("evidence_snippet", ""))])
            writer.writerow(["special_sections", row_key, "confidence", _display_value(row.get("confidence", ""))])
            if isinstance(fields, dict):
                for field_name, field_value in fields.items():
                    writer.writerow(["special_sections", row_key, _display_value(field_name), _display_value(field_value)])


def _mapping_rows(value: Any) -> list[dict[str, Any]]:
    if not isinstance(value, (list, tuple)):
        return []
    return [row for row in value if isinstance(row, dict)]


def _include_appliance_row(row: dict[str, Any]) -> bool:
    appliance_type = _display_value(row.get("appliance_type", "")).lower()
    return not any(token in appliance_type for token in ("sink", "basin", "tap", "tub"))


def _string_list(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, (list, tuple, set)):
        return [text for item in value if (text := _display_value(item))]
    text = _display_value(value)
    return [text] if text else []


def _display_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    if isinstance(value, (list, tuple, set)):
        parts = [_display_value(item) for item in value]
        return " | ".join(part for part in parts if part)
    if isinstance(value, dict):
        try:
            return json.dumps(value, ensure_ascii=False)
        except TypeError:
            return str(value)
    return str(value)


def _display_other_items(value: Any) -> str:
    if not isinstance(value, list):
        return ""
    parts: list[str] = []
    for item in value:
        if not isinstance(item, dict):
            continue
        label = _display_value(item.get("label", ""))
        entry_value = _display_value(item.get("value", ""))
        if label and entry_value:
            parts.append(f"{label}: {entry_value}")
    return " | ".join(parts)


def _normalize_soft_close(value: Any, keyword: str) -> str:
    return parsing.normalize_soft_close_value(value, keyword=keyword) or parsing.normalize_soft_close_value(value)


def _split_export_benchtops(row: dict[str, Any]) -> dict[str, str]:
    entries = parsing._coerce_string_list(row.get("bench_tops", []))
    grouped = parsing._split_benchtop_groups(entries)
    return {
        "bench_tops_wall_run": _display_value(row.get("bench_tops_wall_run", "")) or grouped["bench_tops_wall_run"],
        "bench_tops_island": _display_value(row.get("bench_tops_island", "")) or grouped["bench_tops_island"],
        "bench_tops_other": _display_value(row.get("bench_tops_other", "")) or grouped["bench_tops_other"],
    }
