from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from App.services import parsing


SECTION_TITLE_PATTERNS: tuple[tuple[str, str], ...] = (
    ("15", "15 CABINETS"),
    ("16", "16 ELECTRICAL / ALARM SYSTEM / CCTV / SOLAR PV SYSTEM"),
    ("17", "17 APPLIANCES, ACCESSORIES & HOT WATER UNIT"),
    ("18", "18 AIR-CONDITIONING"),
    ("19", "19 PLUMBING & GAS"),
    ("20", "20 PLUMBING FIXTURES & TAPWARE"),
    ("21", "21 MIRRORS"),
    ("22", "22 WINDOW FURNISHINGS"),
    ("23", "23 TILING / HARD FLOORING"),
    ("24", "24 GLASS SPLASHBACK"),
    ("25", "25 CARPET"),
)

ROOM_HEADINGS: dict[str, str] = dict(getattr(parsing, "EVOCA_ROOM_HEADINGS", {}))
TERMINAL_GROUP_VALUES: frozenset[str] = frozenset(
    {
        "not applicable",
        "not included",
        "not required",
        "n/a",
        "#n/a",
        "tbc",
    }
)
LOOKUP_ENTRIES_KEY = "__entries__"
LOOKUP_LINES_KEY = "__raw_text_lines__"
LINE_MATCH_TOLERANCE = 2.0
UNASSIGNED_SOURCE_TEXT_LABEL = "Unassigned Source Text"
APPEND_EXTRA_VALUE_LABEL_KEYS: frozenset[str] = frozenset({"extent"})
KNOWN_GROUP_BOUNDARY_LABELS: frozenset[str] = frozenset(
    {
        "accessories",
        "accessories & toilet suite",
        "alarm system",
        "appliances",
        "basin",
        "basin mixer",
        "bath",
        "bath mixer / spout",
        "benchtops",
        "carpets",
        "cctv",
        "contrasting facings",
        "ducted reverse cycle",
        "fridge water connection",
        "garden taps",
        "gas type",
        "home automation",
        "hot water unit",
        "integrated appliances",
        "main floor tile",
        "mirrors",
        "overhead cupboards",
        "pantry doors",
        "shower",
        "sink",
        "sink mixer",
        "solar pv system",
        "switch plates / gpo's",
        "tub",
        "tub mixer",
        "tv antenna",
        "underbench",
        "underbench including island",
        "vertical / roller blinds",
        "vinyl, hybrid or timber",
        "washing machine taps",
        "wet area",
    }
)
WORKBOOK_HEADERS: tuple[str, ...] = ("Page", "Order", "Room", "Group", "Label", "Value", "Anchor", "Source Text")
WORKBOOK_COLUMN_WIDTHS: tuple[int, ...] = (6, 7, 22, 28, 32, 44, 8, 60)
HEADER_FILL = "222222"
ROOM_BANNER_FILL = "DCE6F1"
ANCHOR_FILL = "FFF2CC"
GROUP_FILL = "F2F2F2"
NOTE_FILL = "FCE4D6"
BORDER_COLOR = "BBBBBB"
TERMINAL_FONT_COLOR = "888888"
ANCHOR_FONT_COLOR = "996600"


def extract_evoca_pdf(pdf_path: str | Path) -> dict[str, Any]:
    path = Path(pdf_path)
    pages = parsing.load_document_pages(path)
    structured = extract_evoca_pages(pages, source_pdf=str(path), document_name=path.name)
    value_lookup = _build_value_lookup(path)
    _rescue_missing_values(structured, value_lookup)
    _update_statistics(structured)
    return structured


def extract_evoca_pages(
    pages: list[dict[str, Any]],
    *,
    source_pdf: str = "",
    document_name: str = "",
) -> dict[str, Any]:
    document: dict[str, Any] = {
        "source_pdf": source_pdf,
        "document_name": document_name or Path(source_pdf).name,
        "builder": "Evoca",
        "schema_version": "evoca_structured_v0",
        "pages": [],
        "sections": [],
        "unstructured_pages": [],
        "diagnostics": {
            "shift_override_groups": 0,
            "shift_overrides_applied": 0,
            "shift_clears_applied": 0,
            "anchor_value_groups": 0,
            "anchor_values_promoted": 0,
            "anchor_value_child_realignments": 0,
            "raw_text_fallback_groups": 0,
            "raw_text_fallback_pairs_filled": 0,
        },
        "statistics": {
            "page_count": len(pages),
            "section_count": 0,
            "room_count": 0,
            "group_count": 0,
            "row_count": 0,
        },
    }
    current_section: dict[str, Any] | None = None
    current_room: dict[str, Any] | None = None
    current_group: dict[str, Any] | None = None

    for page in pages:
        page_no = int(page.get("page_no", 0) or 0)
        tables = _normalize_tables(page.get("table_rows", []))
        page_summary = {
            "page_no": page_no,
            "table_count": len(tables),
            "sections_detected": [],
            "row_count": sum(len(table) for table in tables),
        }
        document["pages"].append(page_summary)
        if not tables:
            raw_text = parsing.normalize_space(str(page.get("raw_text") or page.get("text") or ""))
            if raw_text:
                document["unstructured_pages"].append(
                    {
                        "page_no": page_no,
                        "text_preview": raw_text[:500],
                    }
                )
            continue

        for table_index, table in enumerate(tables):
            row_index = 0
            while row_index < len(table):
                row = table[row_index]
                section = detect_section_title(row)
                if section:
                    if current_section is not None and current_section.get("section_title") == section["section_title"]:
                        _extend_section_page(current_section, page_no)
                    else:
                        current_section = _new_section(section, page_no, len(document["sections"]) + 1)
                        document["sections"].append(current_section)
                    current_room = None
                    current_group = None
                    page_summary["sections_detected"].append(section["section_title"])
                    row_index += 1
                    continue

                if detect_untracked_section_heading(row):
                    current_room = None
                    current_group = None
                    row_index += 1
                    continue

                if current_section is None:
                    row_index += 1
                    continue

                _extend_section_page(current_section, page_no)

                room_label = detect_room_label(row)
                if room_label:
                    current_room = _ensure_room(current_section, room_label, page_no)
                    current_group = None
                    row_index += 1
                    continue

                if detect_unanchored_parent_header(table, row_index):
                    owner = current_room if current_room is not None else current_section
                    group, next_index = _consume_unanchored_parent_group(
                        table,
                        row_index,
                        page_no=page_no,
                        table_index=table_index,
                        owner=owner,
                    )
                    current_group = group
                    row_index = next_index
                    continue

                group_label = detect_group_label(row)
                if group_label:
                    owner = current_room if current_room is not None else current_section
                    group, next_index = _consume_group(
                        table,
                        row_index,
                        page_no=page_no,
                        table_index=table_index,
                        owner=owner,
                    )
                    current_group = group
                    row_index = next_index
                    continue

                if detect_unanchored_group_header(row):
                    owner = current_room if current_room is not None else current_section
                    group, next_index = _consume_unanchored_group(
                        table,
                        row_index,
                        page_no=page_no,
                        table_index=table_index,
                        owner=owner,
                    )
                    current_group = group
                    row_index = next_index
                    continue

                if _row_has_text(row):
                    target = current_group or current_room or current_section
                    _append_unanchored_row(
                        target,
                        row,
                        page_no=page_no,
                        table_index=table_index,
                        row_index=row_index,
                    )
                row_index += 1

    _dedupe_page_sections(document)
    _update_statistics(document)
    return document


def detect_section_title(row: list[str]) -> dict[str, str] | None:
    joined = _row_joined(row).upper()
    for code, title in SECTION_TITLE_PATTERNS:
        title_upper = title.upper()
        if joined == title_upper or joined.startswith(title_upper):
            return {"section_code": code, "section_title": title}
    return None


def detect_untracked_section_heading(row: list[str]) -> bool:
    if detect_section_title(row):
        return False
    joined = _row_joined(row)
    return bool(re.fullmatch(r"(?:[1-9]|[12][0-9]|30)\s+[A-Z][A-Z0-9 /&(),'\\-]+", joined))


def detect_room_label(row: list[str]) -> str:
    if len(row) < 2:
        return ""
    left = _cell(row, 0)
    center = _cell(row, 1)
    value = _value_text(row)
    if left or value:
        return ""
    first_line = _split_lines(center)[0] if _split_lines(center) else ""
    return first_line if first_line in ROOM_HEADINGS else ""


def detect_group_label(row: list[str]) -> str:
    if _cell(row, 0) != "-":
        return ""
    label_lines = _clean_label_lines(_cell(row, 1))
    if not label_lines:
        return ""
    return label_lines[0]


def detect_unanchored_group_header(row: list[str]) -> bool:
    """Detect a visual group header that lacks Evoca's leading '-' marker."""
    if _cell(row, 0):
        return False
    label_lines = _clean_label_lines(_cell(row, 1))
    value_lines = _split_lines(_value_text(row))
    return len(label_lines) >= 2 and len(value_lines) >= 1 and len(label_lines) > len(value_lines)


def detect_unanchored_parent_header(table: list[list[str]], row_index: int) -> bool:
    """Detect a single-line parent header whose properties start on the next '-' row."""
    if row_index + 1 >= len(table):
        return False
    row = table[row_index]
    next_row = table[row_index + 1]
    if _cell(row, 0) or _value_text(row):
        return False
    label_lines = _clean_label_lines(_cell(row, 1))
    if len(label_lines) != 1:
        return False
    if not detect_group_label(next_row) or _value_text(next_row):
        return False
    next_labels = _clean_label_lines(_cell(next_row, 1))
    if next_labels and _is_known_group_boundary_label(next_labels[0]):
        return False
    return len(next_labels) >= 2


def flatten_rows_for_export(structured: dict[str, Any]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for section in structured.get("sections", []):
        section_title = str(section.get("section_title", "") or "")
        for note in section.get("notes", []) or []:
            rows.append(_export_row(section_title, "", "", note))
        for group in section.get("groups", []) or []:
            rows.extend(_export_group(section_title, "", group))
        for room in section.get("rooms", []) or []:
            room_label = str(room.get("room_label", "") or "")
            for note in room.get("notes", []) or []:
                rows.append(_export_row(section_title, room_label, "", note))
            for group in room.get("groups", []) or []:
                rows.extend(_export_group(section_title, room_label, group))
    return rows


def write_structured_workbook(structured: dict[str, Any], xlsx_path: str | Path) -> None:
    from openpyxl import Workbook

    path = Path(xlsx_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    used_sheet_names: set[str] = set()
    styles = _workbook_styles()
    summary_rows: list[dict[str, Any]] = []

    for section in structured.get("sections", []) or []:
        title = str(section.get("section_title", "Section") or "Section")
        sheet_name = _unique_sheet_name(_sheet_name(title), used_sheet_names)
        sheet = workbook.create_sheet(sheet_name)
        summary_rows.append(_write_section_sheet(sheet, section, sheet_name, styles))

    if not workbook.worksheets:
        workbook.create_sheet("Evoca")
    _build_summary_sheet(workbook, structured, summary_rows, styles)
    workbook.save(path)


def write_structured_json(structured: dict[str, Any], json_path: str | Path) -> None:
    path = Path(json_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(structured, indent=2, ensure_ascii=False), encoding="utf-8")


def _normalize_tables(raw_tables: Any) -> list[list[list[str]]]:
    tables: list[list[list[str]]] = []
    for table in raw_tables or []:
        if not isinstance(table, list):
            continue
        normalized_rows: list[list[str]] = []
        for raw_row in table:
            if not isinstance(raw_row, (list, tuple)):
                continue
            row = [parsing.normalize_space(str(cell).replace("\x00", " ")) if cell is not None else "" for cell in raw_row]
            while row and not row[-1]:
                row.pop()
            if any(row):
                normalized_rows.append(row)
        if normalized_rows:
            tables.append(normalized_rows)
    return tables


def _new_section(section: dict[str, str], page_no: int, order: int) -> dict[str, Any]:
    return {
        "section_code": section["section_code"],
        "section_title": section["section_title"],
        "section_order": order,
        "page_start": page_no,
        "page_end": page_no,
        "rooms": [],
        "groups": [],
        "notes": [],
    }


def _extend_section_page(section: dict[str, Any], page_no: int) -> None:
    if not section.get("page_start"):
        section["page_start"] = page_no
    section["page_end"] = max(int(section.get("page_end", page_no) or page_no), page_no)


def _ensure_room(section: dict[str, Any], room_label: str, page_no: int) -> dict[str, Any]:
    for room in section.get("rooms", []):
        if room.get("room_label") == room_label:
            room["page_end"] = max(int(room.get("page_end", page_no) or page_no), page_no)
            return room
    room = {
        "room_label": room_label,
        "room_key": ROOM_HEADINGS.get(room_label, parsing.source_room_key(room_label)),
        "page_start": page_no,
        "page_end": page_no,
        "groups": [],
        "notes": [],
    }
    section.setdefault("rooms", []).append(room)
    return room


def _consume_group(
    table: list[list[str]],
    row_index: int,
    *,
    page_no: int,
    table_index: int,
    owner: dict[str, Any],
) -> tuple[dict[str, Any], int]:
    row = table[row_index]
    label_lines = _clean_label_lines(_cell(row, 1))
    group_label = label_lines[0] if label_lines else "Group"
    child_labels = label_lines[1:]
    values = _split_lines(_value_text(row))
    raw_rows = [_raw_row_record(row, page_no=page_no, table_index=table_index, row_index=row_index)]
    next_index = row_index + 1

    while next_index < len(table):
        next_row = table[next_index]
        if (
            detect_section_title(next_row)
            or detect_untracked_section_heading(next_row)
            or detect_room_label(next_row)
            or detect_group_label(next_row)
            or detect_unanchored_group_header(next_row)
            or detect_unanchored_parent_header(table, next_index)
        ):
            break
        if _row_has_text(next_row):
            raw_rows.append(_raw_row_record(next_row, page_no=page_no, table_index=table_index, row_index=next_index))
            center_lines = _clean_label_lines(_cell(next_row, 1))
            value_lines = _split_lines(_value_text(next_row))
            if center_lines and value_lines:
                child_labels.extend(center_lines)
                values.extend(value_lines)
            elif center_lines:
                if values:
                    values.extend(center_lines)
                else:
                    child_labels.extend(center_lines)
            else:
                values.extend(value_lines)
        next_index += 1

    values = _coalesce_wrapped_value_lines(group_label, child_labels, values)
    group = {
        "group_label": group_label,
        "page_start": page_no,
        "page_end": page_no,
        "raw_rows": raw_rows,
        "value_lines": list(values),
        "_decompose_meta": _decompose_meta(child_labels, values),
        "rows": _build_group_rows(
            group_label,
            child_labels,
            values,
            page_no=page_no,
            table_index=table_index,
            source_row=row_index,
            raw_rows=raw_rows,
        ),
    }
    owner.setdefault("groups", []).append(group)
    return group, next_index


def _consume_unanchored_parent_group(
    table: list[list[str]],
    row_index: int,
    *,
    page_no: int,
    table_index: int,
    owner: dict[str, Any],
) -> tuple[dict[str, Any], int]:
    row = table[row_index]
    property_row = table[row_index + 1]
    group_label_lines = _clean_label_lines(_cell(row, 1))
    group_label = group_label_lines[0] if group_label_lines else "Group"
    labels = _clean_label_lines(_cell(property_row, 1))
    values = _split_lines(_value_text(property_row))
    raw_rows = [
        _raw_row_record(row, page_no=page_no, table_index=table_index, row_index=row_index),
        _raw_row_record(property_row, page_no=page_no, table_index=table_index, row_index=row_index + 1),
    ]
    next_index = row_index + 2

    while next_index < len(table):
        next_row = table[next_index]
        if (
            detect_section_title(next_row)
            or detect_untracked_section_heading(next_row)
            or detect_room_label(next_row)
            or detect_group_label(next_row)
            or detect_unanchored_group_header(next_row)
            or detect_unanchored_parent_header(table, next_index)
        ):
            break
        if _row_has_text(next_row):
            raw_rows.append(_raw_row_record(next_row, page_no=page_no, table_index=table_index, row_index=next_index))
            center_lines = _clean_label_lines(_cell(next_row, 1))
            value_lines = _split_lines(_value_text(next_row))
            if center_lines and value_lines:
                labels.extend(center_lines)
                values.extend(value_lines)
            elif center_lines:
                labels.extend(center_lines)
            else:
                values.extend(value_lines)
        next_index += 1

    values = _coalesce_wrapped_value_lines(group_label, labels, values)
    group = {
        "group_label": group_label,
        "page_start": page_no,
        "page_end": page_no,
        "raw_rows": raw_rows,
        "value_lines": list(values),
        "_decompose_meta": _decompose_meta(labels, values),
        "rows": _build_group_rows(
            group_label,
            labels,
            values,
            page_no=page_no,
            table_index=table_index,
            source_row=row_index,
            raw_rows=raw_rows,
        ),
    }
    owner.setdefault("groups", []).append(group)
    return group, next_index


def _consume_unanchored_group(
    table: list[list[str]],
    row_index: int,
    *,
    page_no: int,
    table_index: int,
    owner: dict[str, Any],
) -> tuple[dict[str, Any], int]:
    row = table[row_index]
    label_lines = _clean_label_lines(_cell(row, 1))
    group_label = label_lines[0] if label_lines else "Group"
    child_labels = label_lines[1:]
    values = _split_lines(_value_text(row))
    raw_rows = [_raw_row_record(row, page_no=page_no, table_index=table_index, row_index=row_index)]
    next_index = row_index + 1

    while next_index < len(table):
        next_row = table[next_index]
        if (
            detect_section_title(next_row)
            or detect_untracked_section_heading(next_row)
            or detect_room_label(next_row)
            or detect_group_label(next_row)
            or detect_unanchored_group_header(next_row)
            or detect_unanchored_parent_header(table, next_index)
        ):
            break
        if not _row_has_text(next_row):
            next_index += 1
            continue
        if _cell(next_row, 0) or _cell(next_row, 1):
            break
        value_lines = _split_lines(_value_text(next_row))
        if not value_lines:
            break
        raw_rows.append(_raw_row_record(next_row, page_no=page_no, table_index=table_index, row_index=next_index))
        values.extend(value_lines)
        next_index += 1

    values = _coalesce_wrapped_value_lines(group_label, child_labels, values)
    group = {
        "group_label": group_label,
        "page_start": page_no,
        "page_end": page_no,
        "raw_rows": raw_rows,
        "value_lines": list(values),
        "_decompose_meta": _decompose_meta(child_labels, values),
        "rows": _build_group_rows(
            group_label,
            child_labels,
            values,
            page_no=page_no,
            table_index=table_index,
            source_row=row_index,
            raw_rows=raw_rows,
        ),
    }
    owner.setdefault("groups", []).append(group)
    return group, next_index


def _build_group_rows(
    group_label: str,
    labels: list[str],
    values: list[str],
    *,
    page_no: int,
    table_index: int,
    source_row: int,
    raw_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    cleaned_labels = [_clean_label(label) for label in labels if _clean_label(label)]
    cleaned_values = [parsing.normalize_space(value) for value in values if parsing.normalize_space(value)]
    all_labels = [group_label, *cleaned_labels]
    if len(all_labels) >= 2 and len(cleaned_values) == 1 and _is_terminal_group_value(cleaned_values[0]):
        for index, label in enumerate(all_labels):
            rows.append(
                _structured_row(
                    label=label,
                    value=cleaned_values[0] if index == 0 else "",
                    page_no=page_no,
                    table_index=table_index,
                    row_index=source_row,
                    raw_rows=raw_rows,
                    is_group_anchor=index == 0,
                )
            )
        return rows
    if not cleaned_labels:
        cleaned_labels = [group_label] if cleaned_values else []
    for index, label in enumerate(cleaned_labels):
        value = cleaned_values[index] if index < len(cleaned_values) else ""
        rows.append(
            _structured_row(
                label=label,
                value=value,
                page_no=page_no,
                table_index=table_index,
                row_index=source_row,
                raw_rows=raw_rows,
                is_group_anchor=label == group_label,
            )
        )
    for extra_index, value in enumerate(cleaned_values[len(cleaned_labels) :], start=len(cleaned_labels)):
        if rows and _should_append_extra_value_to_previous(rows[-1], value):
            rows[-1]["value"] = parsing.normalize_space(f"{rows[-1].get('value', '')} {value}")
            continue
        row = _structured_row(
            label=UNASSIGNED_SOURCE_TEXT_LABEL,
            value=value,
            page_no=page_no,
            table_index=table_index,
            row_index=source_row + extra_index,
            raw_rows=raw_rows,
        )
        row["is_diagnostic"] = True
        rows.append(row)
    if not rows and raw_rows:
        rows.append(
            _structured_row(
                label=group_label,
                value="",
                page_no=page_no,
                table_index=table_index,
                row_index=source_row,
                raw_rows=raw_rows,
            )
        )
    return rows


def _decompose_meta(labels: list[str], values: list[str]) -> dict[str, int]:
    return {
        "labels_count": len([_clean_label(label) for label in labels if _clean_label(label)]),
        "values_count": len([parsing.normalize_space(value) for value in values if parsing.normalize_space(value)]),
    }


def _coalesce_wrapped_value_lines(group_label: str, labels: list[str], values: list[str]) -> list[str]:
    cleaned_values = [parsing.normalize_space(value) for value in values if parsing.normalize_space(value)]
    if len(cleaned_values) <= len(labels) or _norm_label_key(group_label) != "shower":
        return cleaned_values

    label_keys = [_norm_label_key(label) for label in labels]
    index = 0
    while len(cleaned_values) > len(label_keys) and index < min(len(label_keys), len(cleaned_values) - 1):
        if label_keys[index].startswith("shower rail / rose") and _looks_like_wrapped_product_line(
            cleaned_values[index],
            cleaned_values[index + 1],
        ):
            cleaned_values[index : index + 2] = [
                parsing.normalize_space(f"{cleaned_values[index]} {cleaned_values[index + 1]}")
            ]
            continue
        index += 1
    return cleaned_values


def _looks_like_wrapped_product_line(current: str, next_value: str) -> bool:
    current = parsing.normalize_space(current)
    next_value = parsing.normalize_space(next_value)
    if not current or not next_value:
        return False
    lowered_next = next_value.lower()
    if "semi-frameless" in lowered_next or lowered_next in {"gunmetal", "chrome", "black", "white"}:
        return False
    if current.endswith("-") or current.count("(") > current.count(")"):
        return True
    return bool(re.search(r"\b(?:round|rail|rose|head|shower)\b$", current, flags=re.IGNORECASE) and "(" in next_value)


def _should_append_extra_value_to_previous(row: dict[str, Any], value: str) -> bool:
    if not parsing.normalize_space(value):
        return False
    return _norm_label_key(row.get("label", "")) in APPEND_EXTRA_VALUE_LABEL_KEYS


def _append_unanchored_row(
    target: dict[str, Any],
    row: list[str],
    *,
    page_no: int,
    table_index: int,
    row_index: int,
) -> None:
    text = _row_joined(row)
    if not text:
        return
    record = _structured_row(
        label="Note" if target.get("rooms") is None else "Section Note",
        value=text,
        page_no=page_no,
        table_index=table_index,
        row_index=row_index,
        raw_rows=[_raw_row_record(row, page_no=page_no, table_index=table_index, row_index=row_index)],
    )
    if "rows" in target and "group_label" in target:
        target["rows"].append(record)
    else:
        target.setdefault("notes", []).append(record)


def _structured_row(
    *,
    label: str,
    value: str,
    page_no: int,
    table_index: int,
    row_index: int,
    raw_rows: list[dict[str, Any]],
    is_group_anchor: bool = False,
) -> dict[str, Any]:
    row = {
        "label": parsing.normalize_space(label),
        "value": parsing.normalize_space(value),
        "page_no": page_no,
        "row_order": row_index + 1,
        "table_index": table_index,
        "row_index": row_index,
        "raw_cells": raw_rows[0]["raw_cells"] if raw_rows else [],
        "source_rows": raw_rows,
        "source_method": "pdfplumber_table",
    }
    if is_group_anchor:
        row["is_group_anchor"] = True
    return row


def _raw_row_record(row: list[str], *, page_no: int, table_index: int, row_index: int) -> dict[str, Any]:
    return {
        "page_no": page_no,
        "table_index": table_index,
        "row_index": row_index,
        "raw_cells": list(row),
    }


def _export_group(section_title: str, room_label: str, group: dict[str, Any]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    group_label = str(group.get("group_label", "") or "")
    for row in group.get("rows", []) or []:
        rows.append(_export_row(section_title, room_label, group_label, row))
    return rows


def _export_row(section_title: str, room_label: str, group_label: str, row: dict[str, Any]) -> dict[str, str]:
    _ = section_title
    is_anchor = bool(row.get("is_group_anchor"))
    row_type = (
        "anchor"
        if is_anchor
        else "diagnostic"
        if _is_diagnostic_row(row)
        else "note"
        if _is_note_row(row, group_label)
        else "group"
    )
    return {
        "page": str(row.get("page_no", "") or ""),
        "order": str(row.get("row_order", "") or ""),
        "room": room_label,
        "group": group_label,
        "label": str(row.get("label", "") or ""),
        "value": str(row.get("value", "") or ""),
        "anchor": "ANCHOR" if is_anchor else "",
        "source_text": _compact_source_text(row, is_anchor=is_anchor),
        "row_type": row_type,
    }


def _workbook_styles() -> dict[str, Any]:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    thin_side = Side(style="thin", color=BORDER_COLOR)
    return {
        "header_fill": PatternFill("solid", fgColor=HEADER_FILL),
        "room_fill": PatternFill("solid", fgColor=ROOM_BANNER_FILL),
        "anchor_fill": PatternFill("solid", fgColor=ANCHOR_FILL),
        "group_fill": PatternFill("solid", fgColor=GROUP_FILL),
        "note_fill": PatternFill("solid", fgColor=NOTE_FILL),
        "header_font": Font(color="FFFFFF", bold=True),
        "room_font": Font(bold=True),
        "anchor_font": Font(color=ANCHOR_FONT_COLOR, bold=True),
        "note_font": Font(italic=True),
        "terminal_font": Font(color=TERMINAL_FONT_COLOR, italic=True),
        "default_font": Font(color="000000"),
        "border": Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side),
        "wrap_top": Alignment(wrap_text=True, vertical="top"),
    }


def _write_section_sheet(sheet: Any, section: dict[str, Any], sheet_name: str, styles: dict[str, Any]) -> dict[str, Any]:
    sheet.append(list(WORKBOOK_HEADERS))
    _style_header_row(sheet, 1, styles)
    _apply_sheet_layout(sheet)

    row_count = 0
    anchor_count = 0
    note_count = 0

    for note in section.get("notes", []) or []:
        stats = _append_export_row(sheet, _export_row(str(section.get("section_title", "") or ""), "", "", note), styles)
        row_count += stats["rows"]
        anchor_count += stats["anchors"]
        note_count += stats["notes"]
    for group in section.get("groups", []) or []:
        for row in _export_group(str(section.get("section_title", "") or ""), "", group):
            stats = _append_export_row(sheet, row, styles)
            row_count += stats["rows"]
            anchor_count += stats["anchors"]
            note_count += stats["notes"]

    for room in section.get("rooms", []) or []:
        room_label = str(room.get("room_label", "") or "")
        _append_room_banner(sheet, room_label, styles)
        for note in room.get("notes", []) or []:
            stats = _append_export_row(sheet, _export_row(str(section.get("section_title", "") or ""), room_label, "", note), styles)
            row_count += stats["rows"]
            anchor_count += stats["anchors"]
            note_count += stats["notes"]
        for group in room.get("groups", []) or []:
            for row in _export_group(str(section.get("section_title", "") or ""), room_label, group):
                stats = _append_export_row(sheet, row, styles)
                row_count += stats["rows"]
                anchor_count += stats["anchors"]
                note_count += stats["notes"]

    return {
        "section_code": str(section.get("section_code", "") or ""),
        "title": str(section.get("section_title", "") or ""),
        "sheet": sheet_name,
        "page": _page_range(section),
        "rows": row_count,
        "anchors": anchor_count,
        "notes": note_count,
    }


def _build_summary_sheet(
    workbook: Any,
    structured: dict[str, Any],
    summary_rows: list[dict[str, Any]],
    styles: dict[str, Any],
) -> None:
    sheet = workbook.create_sheet("_summary", 0)
    stats = structured.get("statistics", {}) or {}
    diagnostics = structured.get("diagnostics", {}) or {}
    source_pdf = str(structured.get("source_pdf", "") or "")
    source_name = Path(source_pdf).name if source_pdf else str(structured.get("document_name", "") or "")

    header_block = [
        ("Source PDF", source_name),
        ("Pages", stats.get("page_count", "")),
        ("Sections", stats.get("section_count", "")),
        ("Schema version", structured.get("schema_version", "")),
    ]
    for label, value in header_block:
        sheet.append([label, value])
        _style_summary_pair(sheet, sheet.max_row, styles)

    if diagnostics:
        sheet.append(["Diagnostics", ""])
        _style_summary_pair(sheet, sheet.max_row, styles)
        for key in (
            "shift_override_groups",
            "shift_overrides_applied",
            "shift_clears_applied",
            "anchor_value_groups",
            "anchor_values_promoted",
            "anchor_value_child_realignments",
            "raw_text_fallback_groups",
            "raw_text_fallback_pairs_filled",
        ):
            sheet.append([key, diagnostics.get(key, 0)])
            _style_summary_pair(sheet, sheet.max_row, styles)

    sheet.append([])
    table_header_row = sheet.max_row + 1
    sheet.append(["Section #", "Title", "Sheet", "Page", "Rows", "Anchors", "Notes"])
    _style_header_row(sheet, table_header_row, styles)
    for row in summary_rows:
        sheet.append(
            [
                row["section_code"],
                row["title"],
                row["sheet"],
                row["page"],
                row["rows"],
                row["anchors"],
                row["notes"],
            ]
        )
        _style_summary_table_row(sheet, sheet.max_row, styles)
    if summary_rows:
        sheet.append(["", "Total", "", "", sum(row["rows"] for row in summary_rows), sum(row["anchors"] for row in summary_rows), sum(row["notes"] for row in summary_rows)])
        _style_summary_table_row(sheet, sheet.max_row, styles, bold=True)

    for index, width in enumerate((12, 48, 24, 14, 10, 10, 10), start=1):
        sheet.column_dimensions[_column_letter(index)].width = width
    sheet.freeze_panes = f"A{table_header_row + 1}"


def _append_export_row(sheet: Any, row: dict[str, str], styles: dict[str, Any]) -> dict[str, int]:
    sheet.append(
        [
            row["page"],
            row["order"],
            row["room"],
            row["group"],
            row["label"],
            row["value"],
            row["anchor"],
            row["source_text"],
        ]
    )
    row_index = sheet.max_row
    row_type = row.get("row_type", "group")
    fill = styles["anchor_fill"] if row_type == "anchor" else styles["note_fill"] if row_type == "note" else styles["group_fill"]
    for cell in sheet[row_index]:
        cell.fill = fill
        cell.font = styles["note_font"] if row_type == "note" else styles["default_font"]
        cell.border = styles["border"]
        cell.alignment = styles["wrap_top"]
    if row.get("anchor"):
        sheet.cell(row=row_index, column=7).font = styles["anchor_font"]
    if _is_terminal_group_value(row.get("value", "")):
        sheet.cell(row=row_index, column=6).font = styles["terminal_font"]
    return {
        "rows": 1,
        "anchors": 1 if row.get("anchor") else 0,
        "notes": 1 if row_type == "note" else 0,
    }


def _append_room_banner(sheet: Any, room_label: str, styles: dict[str, Any]) -> None:
    sheet.append([f"room: {room_label}", "", "", "", "", "", "", ""])
    row_index = sheet.max_row
    sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=len(WORKBOOK_HEADERS))
    for cell in sheet[row_index]:
        cell.fill = styles["room_fill"]
        cell.font = styles["room_font"]
        cell.border = styles["border"]
        cell.alignment = styles["wrap_top"]


def _style_header_row(sheet: Any, row_index: int, styles: dict[str, Any]) -> None:
    for cell in sheet[row_index]:
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.border = styles["border"]
        cell.alignment = styles["wrap_top"]


def _style_summary_pair(sheet: Any, row_index: int, styles: dict[str, Any]) -> None:
    for cell in sheet[row_index]:
        cell.border = styles["border"]
        cell.alignment = styles["wrap_top"]
    sheet.cell(row=row_index, column=1).fill = styles["header_fill"]
    sheet.cell(row=row_index, column=1).font = styles["header_font"]


def _style_summary_table_row(sheet: Any, row_index: int, styles: dict[str, Any], *, bold: bool = False) -> None:
    for cell in sheet[row_index]:
        cell.fill = styles["group_fill"]
        cell.border = styles["border"]
        cell.alignment = styles["wrap_top"]
        if bold:
            cell.font = styles["room_font"]


def _apply_sheet_layout(sheet: Any) -> None:
    for index, width in enumerate(WORKBOOK_COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[_column_letter(index)].width = width
    sheet.freeze_panes = "A2"


def _column_letter(index: int) -> str:
    from openpyxl.utils import get_column_letter

    return get_column_letter(index)


def _page_range(section: dict[str, Any]) -> str:
    start = str(section.get("page_start", "") or "")
    end = str(section.get("page_end", "") or "")
    if not start:
        return ""
    return start if not end or end == start else f"{start}-{end}"


def _compact_source_text(row: dict[str, Any], *, is_anchor: bool) -> str:
    label = parsing.normalize_space(str(row.get("label", "") or ""))
    value = parsing.normalize_space(str(row.get("value", "") or ""))
    if is_anchor:
        return f"- {label}" if label else ""
    if _is_note_row(row, ""):
        return value or label
    if label and value:
        return f"{label}: {value}"
    return label or value


def _is_note_row(row: dict[str, Any], group_label: str) -> bool:
    _ = group_label
    return parsing.normalize_space(str(row.get("label", "") or "")).lower() in {"note", "section note"}


def _is_diagnostic_row(row: dict[str, Any]) -> bool:
    return bool(row.get("is_diagnostic")) or (
        _norm_label_key(row.get("label", "")) == _norm_label_key(UNASSIGNED_SOURCE_TEXT_LABEL)
    )


def _build_value_lookup(pdf_path: Path) -> dict[int, dict[str, Any]]:
    import pdfplumber

    lookup: dict[int, dict[str, Any]] = {}
    with pdfplumber.open(str(pdf_path)) as pdf:
        for page_index, page in enumerate(pdf.pages, start=1):
            page_lookup: dict[str, Any] = {}
            entries: list[dict[str, Any]] = []
            text_tables = page.extract_tables(
                table_settings={
                    "vertical_strategy": "lines",
                    "horizontal_strategy": "text",
                }
            ) or []
            for table_index, table in enumerate(text_tables):
                for row_index, raw_row in enumerate(table or []):
                    entry = _lookup_entry_from_row(raw_row)
                    if entry is None:
                        continue
                    label, value = entry
                    key = _norm_label_key(label)
                    if not key:
                        continue
                    entries.append(
                        {
                            "key": key,
                            "label": label,
                            "value": value,
                            "table_index": table_index,
                            "row_index": row_index,
                        }
                    )
                    if value:
                        page_lookup.setdefault(key, []).append(value)
            if entries:
                page_lookup[LOOKUP_ENTRIES_KEY] = entries
            raw_lines = _extract_word_lines(page)
            if raw_lines:
                page_lookup[LOOKUP_LINES_KEY] = raw_lines
            if page_lookup:
                lookup[page_index] = page_lookup
    return lookup


def _extract_word_lines(page: Any) -> list[dict[str, Any]]:
    words = page.extract_words(x_tolerance=1, y_tolerance=3, keep_blank_chars=False) or []
    normalized_words: list[dict[str, Any]] = []
    for word in words:
        text = parsing.normalize_space(str(word.get("text", "") or ""))
        if not text:
            continue
        top = float(word.get("top", 0.0) or 0.0)
        bottom = float(word.get("bottom", top) or top)
        normalized_words.append(
            {
                "text": text,
                "x0": float(word.get("x0", 0.0) or 0.0),
                "x1": float(word.get("x1", 0.0) or 0.0),
                "top": top,
                "bottom": bottom,
                "center": (top + bottom) / 2.0,
            }
        )
    normalized_words.sort(key=lambda item: (item["center"], item["x0"]))

    lines: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None
    for word in normalized_words:
        if current is None or abs(float(word["center"]) - float(current["center"])) > LINE_MATCH_TOLERANCE:
            current = {
                "top": float(word["top"]),
                "bottom": float(word["bottom"]),
                "center": float(word["center"]),
                "words": [],
            }
            lines.append(current)
        current["words"].append(word)
        current["top"] = min(float(current["top"]), float(word["top"]))
        current["bottom"] = max(float(current["bottom"]), float(word["bottom"]))
        current["center"] = (float(current["top"]) + float(current["bottom"])) / 2.0

    for line in lines:
        line["words"].sort(key=lambda item: item["x0"])
        line["text"] = parsing.normalize_space(" ".join(str(word["text"]) for word in line["words"]))
    return lines


def _lookup_pair_from_row(raw_row: Any) -> tuple[str, str] | None:
    entry = _lookup_entry_from_row(raw_row)
    if entry is None:
        return None
    label, value = entry
    if not value:
        return None
    return label, value


def _lookup_entry_from_row(raw_row: Any) -> tuple[str, str] | None:
    if not isinstance(raw_row, (list, tuple)):
        return None
    cells = [parsing.normalize_space(str(cell).replace("\x00", " ")) if cell is not None else "" for cell in raw_row]
    non_empty = [cell for cell in cells if cell]
    if len(non_empty) == 1:
        label = _clean_label(non_empty[0])
        value = ""
    elif len(non_empty) == 2:
        label = _clean_label(non_empty[0])
        value = parsing.normalize_space(non_empty[1])
    else:
        return None
    if not label:
        return None
    if label in {"-", "–", "—"}:
        return None
    if detect_section_title([label, value]) or detect_section_title([label]):
        return None
    return label, value


def _rescue_missing_values(structured: dict[str, Any], lookup: dict[int, dict[str, Any]]) -> None:
    consumable: dict[int, dict[str, Any]] = {}
    for page, page_values in lookup.items():
        page_copy: dict[str, Any] = {}
        for key, values in page_values.items():
            if key in {LOOKUP_ENTRIES_KEY, LOOKUP_LINES_KEY}:
                page_copy[key] = [dict(entry) for entry in values]
            else:
                page_copy[key] = list(values)
        consumable[page] = page_copy
    diagnostics = structured.setdefault("diagnostics", {})
    for key in (
        "shift_override_groups",
        "shift_overrides_applied",
        "shift_clears_applied",
        "anchor_value_groups",
        "anchor_values_promoted",
        "anchor_value_child_realignments",
        "raw_text_fallback_groups",
        "raw_text_fallback_pairs_filled",
    ):
        diagnostics.setdefault(key, 0)
    group_context = _build_rescue_group_context(structured, consumable)
    for section in structured.get("sections", []) or []:
        for group in section.get("groups", []) or []:
            _rescue_group(group, consumable, diagnostics, group_context)
        for room in section.get("rooms", []) or []:
            for group in room.get("groups", []) or []:
                _rescue_group(group, consumable, diagnostics, group_context)


def _build_rescue_group_context(
    structured: dict[str, Any],
    consumable: dict[int, dict[str, Any]],
) -> dict[str, Any]:
    boundaries = _collect_rescue_boundaries(structured)
    blocks_by_page: dict[int, list[dict[str, Any]]] = {}
    raw_blocks_by_page: dict[int, list[dict[str, Any]]] = {}
    for page, page_lookup in consumable.items():
        entries = page_lookup.get(LOOKUP_ENTRIES_KEY) or []
        raw_lines = page_lookup.get(LOOKUP_LINES_KEY) or []
        page_boundaries = boundaries.get(page) or {}
        group_keys = page_boundaries.get("groups", set())
        if not group_keys:
            continue
        room_keys = page_boundaries.get("rooms", set())
        if entries:
            blocks = _build_page_group_blocks(entries, group_keys, room_keys)
            if blocks:
                blocks_by_page[page] = blocks
        if raw_lines:
            raw_blocks = _build_page_raw_text_blocks(raw_lines, group_keys, room_keys)
            if raw_blocks:
                raw_blocks_by_page[page] = raw_blocks
    return {"blocks_by_page": blocks_by_page, "raw_blocks_by_page": raw_blocks_by_page, "cursors": {}, "raw_cursors": {}}


def _collect_rescue_boundaries(structured: dict[str, Any]) -> dict[int, dict[str, set[str]]]:
    boundaries: dict[int, dict[str, set[str]]] = {}
    for section in structured.get("sections", []) or []:
        for group in section.get("groups", []) or []:
            _record_group_boundary(boundaries, group)
        for room in section.get("rooms", []) or []:
            room_key = _norm_label_key(room.get("room_label", ""))
            room_pages = _page_numbers_from_range(room.get("page_start"), room.get("page_end"))
            for page in room_pages:
                boundaries.setdefault(page, {"groups": set(), "rooms": set()})["rooms"].add(room_key)
            for group in room.get("groups", []) or []:
                _record_group_boundary(boundaries, group)
    return boundaries


def _record_group_boundary(boundaries: dict[int, dict[str, set[str]]], group: dict[str, Any]) -> None:
    group_key = _norm_label_key(group.get("group_label", ""))
    if not group_key or not _is_known_group_boundary_label(group_key):
        return
    pages = {
        int(raw.get("page_no", 0) or 0)
        for raw in group.get("raw_rows", []) or []
        if int(raw.get("page_no", 0) or 0)
    }
    if not pages:
        pages = _page_numbers_from_range(group.get("page_start"), group.get("page_end"))
    for page in pages:
        boundaries.setdefault(page, {"groups": set(), "rooms": set()})["groups"].add(group_key)


def _page_numbers_from_range(start: Any, end: Any) -> set[int]:
    start_page = int(start or 0)
    end_page = int(end or start_page or 0)
    if not start_page:
        return set()
    return set(range(start_page, max(start_page, end_page) + 1))


def _build_page_group_blocks(
    entries: list[dict[str, Any]],
    group_keys: set[str],
    room_keys: set[str],
) -> list[dict[str, Any]]:
    blocks: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None
    for entry in entries:
        key = str(entry.get("key", "") or "")
        if not key:
            continue
        value = parsing.normalize_space(str(entry.get("value", "") or ""))
        if key in room_keys:
            current = None
            continue
        if key in group_keys:
            current = {"group_key": key, "values": {}}
            blocks.append(current)
            if value:
                current["values"].setdefault(key, []).append(value)
            continue
        if current is not None and value:
            current["values"].setdefault(key, []).append(value)
    return blocks


def _build_page_raw_text_blocks(
    raw_lines: list[dict[str, Any]],
    group_keys: set[str],
    room_keys: set[str],
) -> list[dict[str, Any]]:
    boundary_lines: list[dict[str, Any]] = []
    for index, line in enumerate(raw_lines):
        words = line.get("words", []) or []
        group_key = _match_group_boundary_key(words, group_keys)
        room_key = _match_room_boundary_key(line, room_keys)
        is_section = detect_section_title([str(line.get("text", "") or "")]) is not None
        if group_key or room_key or is_section:
            boundary_lines.append(
                {
                    "index": index,
                    "top": float(line.get("top", 0.0) or 0.0),
                    "group_key": group_key,
                    "is_boundary": True,
                }
            )

    blocks: list[dict[str, Any]] = []
    for boundary_index, boundary in enumerate(boundary_lines):
        group_key = str(boundary.get("group_key", "") or "")
        if not group_key:
            continue
        start_index = int(boundary["index"])
        end_index = len(raw_lines)
        bottom = float("inf")
        for next_boundary in boundary_lines[boundary_index + 1 :]:
            end_index = int(next_boundary["index"])
            bottom = float(next_boundary["top"])
            break
        block_lines = [
            line
            for line in raw_lines[start_index:end_index]
            if float(line.get("center", line.get("top", 0.0)) or 0.0) >= float(boundary["top"]) - LINE_MATCH_TOLERANCE
            and float(line.get("center", line.get("top", 0.0)) or 0.0) < bottom - LINE_MATCH_TOLERANCE
        ]
        blocks.append(
            {
                "group_key": group_key,
                "top": float(boundary["top"]),
                "bottom": bottom,
                "lines": block_lines,
            }
        )
    return blocks


def _next_group_rescue_lookup(group: dict[str, Any], group_context: dict[str, Any]) -> dict[str, list[str]] | None:
    rows = group.get("rows", []) or []
    first_row = rows[0] if rows else {}
    page = int(first_row.get("page_no", 0) or group.get("page_start", 0) or 0)
    group_key = _norm_label_key(group.get("group_label", ""))
    if not page or not group_key:
        return None
    blocks = group_context.get("blocks_by_page", {}).get(page) or []
    if not blocks:
        return None
    cursors = group_context.setdefault("cursors", {}).setdefault(page, {})
    start = int(cursors.get(group_key, 0) or 0)
    for index in range(start, len(blocks)):
        block = blocks[index]
        if block.get("group_key") != group_key:
            continue
        cursors[group_key] = index + 1
        return _copy_lookup_values(block.get("values", {}) or {})
    cursors[group_key] = len(blocks)
    return None


def _next_group_raw_text_lookup(group: dict[str, Any], group_context: dict[str, Any]) -> dict[str, list[str]] | None:
    rows = group.get("rows", []) or []
    first_row = rows[0] if rows else {}
    page = int(first_row.get("page_no", 0) or group.get("page_start", 0) or 0)
    group_key = _norm_label_key(group.get("group_label", ""))
    if not page or not group_key:
        return None
    blocks = group_context.get("raw_blocks_by_page", {}).get(page) or []
    if not blocks:
        return None
    cursors = group_context.setdefault("raw_cursors", {}).setdefault(page, {})
    start = int(cursors.get(group_key, 0) or 0)
    for index in range(start, len(blocks)):
        block = blocks[index]
        if block.get("group_key") != group_key:
            continue
        cursors[group_key] = index + 1
        return _parse_raw_text_pairs_for_group(group, block)
    cursors[group_key] = len(blocks)
    return None


def _copy_lookup_values(values_by_key: dict[str, list[str]]) -> dict[str, list[str]]:
    return {key: list(values) for key, values in values_by_key.items()}


def _parse_raw_text_pairs_for_group(group: dict[str, Any], block: dict[str, Any]) -> dict[str, list[str]]:
    labels = [
        parsing.normalize_space(str(row.get("label", "") or ""))
        for row in group.get("rows", []) or []
        if not row.get("is_group_anchor")
        and not _is_note_row(row, "")
        and not _is_diagnostic_row(row)
        and _norm_label_key(row.get("label", "")) != "continuation"
    ]
    label_specs = _label_token_specs(labels)
    if not label_specs:
        return {}

    lookup: dict[str, list[str]] = {}
    lines = block.get("lines", []) or []
    consumed_line_indexes: set[int] = set()
    for line_index, line in enumerate(lines):
        if line_index in consumed_line_indexes:
            continue
        words = line.get("words", []) or []
        tokens = [_norm_word_token(str(word.get("text", "") or "")) for word in words]
        index = 0
        if tokens and tokens[0] == "":
            index = 1
        while index < len(tokens):
            match = _match_label_tokens(tokens, index, label_specs)
            if match is None:
                break
            label_key, token_count = match
            value_start = index + token_count
            next_label_index = len(tokens)
            scan = value_start
            while scan < len(tokens):
                if _match_label_tokens(tokens, scan, label_specs) is not None:
                    next_label_index = scan
                    break
                scan += 1
            value = parsing.normalize_space(
                " ".join(str(words[word_index].get("text", "") or "") for word_index in range(value_start, next_label_index))
            )
            if not value and next_label_index >= len(tokens):
                value = _next_raw_text_continuation_value(lines, line_index, label_specs, consumed_line_indexes)
            if value and not _is_raw_text_noise_value(value):
                lookup.setdefault(label_key, []).append(value)
            if next_label_index >= len(tokens):
                break
            index = next_label_index
    return lookup


def _label_token_specs(labels: list[str]) -> list[tuple[str, tuple[str, ...]]]:
    specs: list[tuple[str, tuple[str, ...]]] = []
    seen: set[str] = set()
    for label in labels:
        key = _norm_label_key(label)
        if not key or key in seen:
            continue
        tokens = tuple(_norm_word_token(token) for token in key.split() if _norm_word_token(token))
        if not tokens:
            continue
        specs.append((key, tokens))
        seen.add(key)
    return sorted(specs, key=lambda item: len(item[1]), reverse=True)


def _match_label_tokens(
    tokens: list[str],
    index: int,
    label_specs: list[tuple[str, tuple[str, ...]]],
) -> tuple[str, int] | None:
    for key, label_tokens in label_specs:
        end = index + len(label_tokens)
        if end > len(tokens):
            continue
        if tuple(tokens[index:end]) == label_tokens:
            return key, len(label_tokens)
    return None


def _next_raw_text_continuation_value(
    lines: list[dict[str, Any]],
    line_index: int,
    label_specs: list[tuple[str, tuple[str, ...]]],
    consumed_line_indexes: set[int],
) -> str:
    next_index = line_index + 1
    if next_index >= len(lines) or next_index in consumed_line_indexes:
        return ""
    next_words = lines[next_index].get("words", []) or []
    next_tokens = [_norm_word_token(str(word.get("text", "") or "")) for word in next_words]
    if not next_tokens or _match_label_tokens(next_tokens, 0, label_specs) is not None:
        return ""
    value = parsing.normalize_space(" ".join(str(word.get("text", "") or "") for word in next_words))
    if _is_raw_text_noise_value(value):
        return ""
    if value:
        consumed_line_indexes.add(next_index)
    return value


def _is_raw_text_noise_value(value: Any) -> bool:
    normalized = parsing.normalize_space(str(value or ""))
    if not normalized:
        return True
    lowered = normalized.lower()
    return bool(
        re.search(r"\bpage\s+\d+\s+of\s+\d+\b", lowered)
        or "client initials" in lowered
        or "initials:" in lowered
    )


def _apply_raw_text_fallback(
    rows: list[dict[str, Any]],
    raw_text_lookup: dict[str, list[str]] | None,
    diagnostics: dict[str, int],
) -> None:
    if not raw_text_lookup:
        return
    filled = 0
    for row in rows:
        if row.get("is_group_anchor") or _is_note_row(row, "") or _is_diagnostic_row(row):
            continue
        key = _norm_label_key(row.get("label", ""))
        if not key or key == "continuation":
            continue
        if parsing.normalize_space(str(row.get("value") or "")):
            continue
        candidates = raw_text_lookup.get(key)
        if not candidates:
            continue
        value = parsing.normalize_space(candidates.pop(0))
        if not value:
            continue
        row["value"] = value
        row["source_method"] = "pdfplumber_raw_text_fallback"
        filled += 1
    if filled:
        diagnostics["raw_text_fallback_groups"] += 1
        diagnostics["raw_text_fallback_pairs_filled"] += filled


def _row_rescue_lookup(
    row: dict[str, Any],
    consumable: dict[int, dict[str, Any]],
    group_lookup: dict[str, list[str]] | None,
) -> dict[str, list[str]]:
    if group_lookup is not None:
        return group_lookup
    return consumable.get(int(row.get("page_no", 0) or 0), {})


def _rescue_group(
    group: dict[str, Any],
    consumable: dict[int, dict[str, Any]],
    diagnostics: dict[str, int],
    group_context: dict[str, Any],
) -> None:
    rows = group.get("rows", []) or []
    if not rows:
        return
    group_lookup = _next_group_rescue_lookup(group, group_context)
    anchor_value = rows[0].get("value", "")
    if _is_terminal_group_value(anchor_value):
        return
    if _promote_group_anchor_value(group, rows, consumable, diagnostics, group_lookup):
        return
    raw_text_lookup = _next_group_raw_text_lookup(group, group_context)
    if _shift_override_eligible(group, rows, consumable, group_lookup):
        _apply_shift_override(rows, consumable, diagnostics, group_lookup)
        _apply_raw_text_fallback(rows, raw_text_lookup, diagnostics)
        return
    for row in rows:
        if _is_diagnostic_row(row):
            continue
        key = _norm_label_key(row.get("label", ""))
        if not key:
            continue
        page_lookup = _row_rescue_lookup(row, consumable, group_lookup)
        candidates = page_lookup.get(key)
        if not candidates:
            continue
        current_value = parsing.normalize_space(str(row.get("value") or ""))
        if current_value:
            if _norm_label_key(candidates[0]) == _norm_label_key(current_value):
                candidates.pop(0)
            continue
        row["value"] = candidates.pop(0)
        row["source_method"] = "pdfplumber_text_rescue"
    _apply_raw_text_fallback(rows, raw_text_lookup, diagnostics)


def _promote_group_anchor_value(
    group: dict[str, Any],
    rows: list[dict[str, Any]],
    consumable: dict[int, dict[str, Any]],
    diagnostics: dict[str, int],
    group_lookup: dict[str, list[str]] | None,
) -> bool:
    group_label = parsing.normalize_space(str(group.get("group_label", "") or ""))
    if not group_label or rows[0].get("is_group_anchor"):
        return False

    first_row = rows[0]
    page_lookup = _row_rescue_lookup(first_row, consumable, group_lookup)
    group_key = _norm_label_key(group_label)
    group_candidates = page_lookup.get(group_key)
    if not group_candidates:
        return False

    group_value = parsing.normalize_space(group_candidates[0])
    value_lines = [parsing.normalize_space(value) for value in group.get("value_lines", []) or [] if parsing.normalize_space(value)]
    first_value = parsing.normalize_space(str(first_row.get("value") or ""))
    if not group_value or not value_lines:
        return False
    if group_value != value_lines[0] or group_value != first_value:
        return False

    child_rows = [
        row
        for row in rows
        if not row.get("is_group_anchor")
        and not _is_note_row(row, "")
        and not _is_diagnostic_row(row)
        and _norm_label_key(row.get("label", "")) != "continuation"
    ]
    if not child_rows:
        return False

    meta = group.get("_decompose_meta", {}) or {}
    labels_count = int(meta.get("labels_count", 0) or 0)
    values_count = int(meta.get("values_count", 0) or 0)
    if labels_count and values_count and values_count < labels_count + 1:
        return False

    group_candidates.pop(0)
    anchor_row = dict(first_row)
    anchor_row["label"] = group_label
    anchor_row["value"] = group_value
    anchor_row["is_group_anchor"] = True
    anchor_row["source_method"] = "pdfplumber_text_rescue"

    realigned_rows = [anchor_row]
    for row in child_rows:
        key = _norm_label_key(row.get("label", ""))
        candidates = page_lookup.get(key) if key else None
        old_value = parsing.normalize_space(str(row.get("value") or ""))
        value = candidates.pop(0) if candidates else ""
        if old_value != parsing.normalize_space(value):
            diagnostics["anchor_value_child_realignments"] += 1
        row["value"] = value
        row["source_method"] = "pdfplumber_text_rescue"
        realigned_rows.append(row)

    group["rows"] = realigned_rows
    diagnostics["anchor_value_groups"] += 1
    diagnostics["anchor_values_promoted"] += 1
    return True


def _shift_override_eligible(
    group: dict[str, Any],
    rows: list[dict[str, Any]],
    consumable: dict[int, dict[str, Any]],
    group_lookup: dict[str, list[str]] | None,
) -> bool:
    meta = group.get("_decompose_meta", {}) or {}
    labels_count = int(meta.get("labels_count", 0) or 0)
    values_count = int(meta.get("values_count", 0) or 0)
    if not labels_count > values_count > 0:
        return False
    for row in rows:
        if _is_diagnostic_row(row):
            continue
        if row.get("is_group_anchor"):
            continue
        old_value = parsing.normalize_space(str(row.get("value") or ""))
        if not old_value:
            continue
        key = _norm_label_key(row.get("label", ""))
        page_lookup = _row_rescue_lookup(row, consumable, group_lookup)
        if key and not page_lookup.get(key):
            return True
    return False


def _apply_shift_override(
    rows: list[dict[str, Any]],
    consumable: dict[int, dict[str, Any]],
    diagnostics: dict[str, int],
    group_lookup: dict[str, list[str]] | None,
) -> None:
    """Correct under-supplied groups using text-grid lookup in source order.

    This assumes repeated labels on a page appear in the text-grid lookup in the
    same order as structured groups are processed. If a future multi-column PDF
    violates that, this should become bbox-bounded instead of page-wide.
    """
    diagnostics["shift_override_groups"] += 1
    for row in rows:
        if _is_diagnostic_row(row):
            continue
        if row.get("is_group_anchor"):
            continue
        key = _norm_label_key(row.get("label", ""))
        old_value = parsing.normalize_space(str(row.get("value") or ""))
        page_lookup = _row_rescue_lookup(row, consumable, group_lookup)
        candidates = page_lookup.get(key) if key else None
        if candidates:
            value = candidates.pop(0)
            if old_value != parsing.normalize_space(value):
                diagnostics["shift_overrides_applied"] += 1
            row["value"] = value
            row["source_method"] = "pdfplumber_text_rescue"
            continue
        if old_value:
            diagnostics["shift_clears_applied"] += 1
        row["value"] = ""
        row["source_method"] = "pdfplumber_text_rescue"


def _dedupe_page_sections(document: dict[str, Any]) -> None:
    for page in document.get("pages", []) or []:
        seen: set[str] = set()
        unique: list[str] = []
        for title in page.get("sections_detected", []) or []:
            if title in seen:
                continue
            seen.add(title)
            unique.append(title)
        page["sections_detected"] = unique


def _update_statistics(document: dict[str, Any]) -> None:
    sections = list(document.get("sections", []) or [])
    room_count = 0
    group_count = 0
    row_count = 0
    for section in sections:
        group_count += len(section.get("groups", []) or [])
        row_count += sum(len(group.get("rows", []) or []) for group in section.get("groups", []) or [])
        for room in section.get("rooms", []) or []:
            room_count += 1
            group_count += len(room.get("groups", []) or [])
            row_count += sum(len(group.get("rows", []) or []) for group in room.get("groups", []) or [])
            row_count += len(room.get("notes", []) or [])
        row_count += len(section.get("notes", []) or [])
    document["statistics"] = {
        "page_count": len(document.get("pages", []) or []),
        "section_count": len(sections),
        "room_count": room_count,
        "group_count": group_count,
        "row_count": row_count,
    }


def _is_terminal_group_value(value: Any) -> bool:
    return parsing.normalize_space(str(value or "")).strip().lower() in TERMINAL_GROUP_VALUES


def _is_known_group_boundary_label(label: Any) -> bool:
    return _norm_label_key(label) in KNOWN_GROUP_BOUNDARY_LABELS


def _match_group_boundary_key(words: list[dict[str, Any]], group_keys: set[str]) -> str:
    if not words:
        return ""
    first_token = _norm_word_token(str(words[0].get("text", "") or ""))
    has_dash_marker = first_token == ""
    return _match_line_label_key(words, group_keys, allow_extra_words=has_dash_marker)


def _match_line_label_key(words: list[dict[str, Any]], keys: set[str], *, allow_extra_words: bool) -> str:
    if not words or not keys:
        return ""
    tokens = [_norm_word_token(str(word.get("text", "") or "")) for word in words]
    if tokens and tokens[0] == "":
        tokens = tokens[1:]
    if not tokens:
        return ""
    specs = _label_token_specs(list(keys))
    for key, label_tokens in specs:
        if len(tokens) < len(label_tokens):
            continue
        if tuple(tokens[: len(label_tokens)]) != label_tokens:
            continue
        if allow_extra_words or len(tokens) == len(label_tokens):
            return key
    return ""


def _match_room_boundary_key(line: dict[str, Any], room_keys: set[str]) -> str:
    return _match_line_label_key(line.get("words", []) or [], room_keys, allow_extra_words=False)


def _norm_word_token(value: Any) -> str:
    return _norm_label_key(value).strip("-")


def _norm_label_key(label: Any) -> str:
    """Normalize labels for text-grid rescue; wrapped PDF labels may still miss in v0."""
    normalized = parsing.normalize_space(str(label or ""))
    normalized = re.sub(r"\*+", "", normalized)
    normalized = re.sub(r"\s+", " ", normalized)
    return normalized.strip().lower()


def _cell(row: list[str], index: int) -> str:
    if 0 <= index < len(row):
        return parsing.normalize_space(row[index])
    return ""


def _value_text(row: list[str]) -> str:
    values = [_cell(row, index) for index in range(2, len(row))]
    return "\n".join(value for value in values if value)


def _split_lines(value: Any) -> list[str]:
    return [parsing.normalize_space(line) for line in str(value or "").splitlines() if parsing.normalize_space(line)]


def _clean_label_lines(value: Any) -> list[str]:
    lines = [_clean_label(line) for line in _split_lines(value) if _clean_label(line)]
    merged: list[str] = []
    for line in lines:
        if merged and _label_wraps_to_next_line(merged[-1]):
            merged[-1] = parsing.normalize_space(f"{merged[-1]} {line}")
        else:
            merged.append(line)
    return merged


def _label_wraps_to_next_line(label: str) -> bool:
    stripped = parsing.normalize_space(label)
    return stripped.endswith("&") or stripped.count("(") > stripped.count(")")


def _clean_label(value: Any) -> str:
    return parsing.normalize_space(str(value or "")).strip(" -*:")


def _row_joined(row: list[str]) -> str:
    return parsing.normalize_space(" ".join(cell for cell in row if cell))


def _row_has_text(row: list[str]) -> bool:
    return any(parsing.normalize_space(cell) for cell in row)


def _sheet_name(value: str) -> str:
    aliases = {
        "15 CABINETS": "15_CABINETS",
        "16 ELECTRICAL / ALARM SYSTEM / CCTV / SOLAR PV SYSTEM": "16_ELECTRICAL",
        "17 APPLIANCES, ACCESSORIES & HOT WATER UNIT": "17_APPLIANCES",
        "18 AIR-CONDITIONING": "18_AIR-CONDITIONING",
        "19 PLUMBING & GAS": "19_PLUMBING_GAS",
        "20 PLUMBING FIXTURES & TAPWARE": "20_PLUMBING",
        "21 MIRRORS": "21_MIRRORS",
        "22 WINDOW FURNISHINGS": "22_WINDOW_FURNISHINGS",
        "23 TILING / HARD FLOORING": "23_FLOORING",
        "24 GLASS SPLASHBACK": "24_SPLASHBACK",
        "25 CARPET": "25_CARPET",
    }
    raw = parsing.normalize_space(value)
    if raw in aliases:
        return aliases[raw]
    cleaned = re.sub(r"[\[\]:*?/\\]", " ", raw)
    cleaned = re.sub(r"\s+", "_", cleaned).strip("_") or "Evoca"
    return cleaned[:31]


def _unique_sheet_name(base: str, used: set[str]) -> str:
    name = base[:31] or "Sheet"
    if name not in used:
        used.add(name)
        return name
    counter = 2
    while True:
        suffix = f" {counter}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        if candidate not in used:
            used.add(candidate)
            return candidate
        counter += 1
