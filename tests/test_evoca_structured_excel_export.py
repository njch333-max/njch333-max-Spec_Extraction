from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from App.services.evoca_structured_adapter import build_evoca_snapshot_from_structured
from App.services.export_service import _write_evoca_structured_review_excel


def test_evoca_structured_spec_list_excel_uses_source_native_section_tabs(tmp_path: Path) -> None:
    fixture_path = next(Path("tests/fixtures/evoca_structured_section_filter").glob("*EVOC447*.json"))
    structured = json.loads(fixture_path.read_text(encoding="utf-8"))
    snapshot = build_evoca_snapshot_from_structured(
        structured,
        job_no="EVOC447",
        source_document=str(fixture_path),
    ).model_dump()

    workbook_path = tmp_path / "evoca_structured.xlsx"
    _write_evoca_structured_review_excel(workbook_path, "EVOC447", snapshot)

    workbook = load_workbook(workbook_path)

    assert "Summary" in workbook.sheetnames
    assert "By Section" not in workbook.sheetnames
    assert "Special Sections" not in workbook.sheetnames
    assert "Unretained_Evoca_source_rows" not in workbook.sheetnames
    assert "15_CABINETS" in workbook.sheetnames
    appliance_sheet_name = next(name for name in workbook.sheetnames if name.startswith("17_APPLIANCES"))
    assert any(name.startswith("23_") for name in workbook.sheetnames)
    assert any(name.startswith("24_") for name in workbook.sheetnames)
    assert any(name.startswith("25_") for name in workbook.sheetnames)

    cabinets = workbook["15_CABINETS"]
    headers = [cabinets.cell(row=1, column=column).value for column in range(1, 9)]
    assert headers == ["Page", "Order", "Room", "Group", "Label", "Value", "Anchor", "Source Text"]

    cabinet_values = [
        cabinets.cell(row=row, column=column).value
        for row in range(1, cabinets.max_row + 1)
        for column in range(1, cabinets.max_column + 1)
    ]
    assert "room: Kitchen" in cabinet_values
    assert "Quantum Quartz" in cabinet_values
    plumbing = workbook["20_PLUMBING_FIXTURES_&_TAPWARE"]
    plumbing_values = [
        plumbing.cell(row=row, column=column).value
        for row in range(1, plumbing.max_row + 1)
        for column in range(1, plumbing.max_column + 1)
    ]
    assert "Alora Gloss White Wall Faced Toilet Suite" in plumbing_values

    appliances = workbook[appliance_sheet_name]
    appliance_values = [
        appliances.cell(row=row, column=column).value
        for row in range(1, appliances.max_row + 1)
        for column in range(1, appliances.max_column + 1)
    ]
    assert "Appliances" in appliance_values
    assert "Fisher & Paykel 900mm Oven OB90S9LEX2 (Electric)" in appliance_values
