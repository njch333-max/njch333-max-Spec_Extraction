from __future__ import annotations

import importlib
import io
import json
import os
import shutil
import tempfile
import types
import unittest
import urllib.error
from datetime import datetime
from pathlib import Path
from unittest import mock

TEST_DATA_DIR = Path(tempfile.mkdtemp(prefix="spec-extraction-test-data-"))
os.environ["SPEC_EXTRACTION_DATA_DIR"] = str(TEST_DATA_DIR)
os.environ["SPEC_EXTRACTION_ENABLE_OPENAI"] = "0"
os.environ["SPEC_EXTRACTION_ENABLE_OPENAI_VISION"] = "0"

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from App.main import _build_material_summary, _flatten_rooms, _format_brisbane_time, _format_run_duration, _run_duration_display, app
from App.services import cleaning_rules, extraction_service, parsing as parsing_module, store
from App.services import appliance_official
from App.services.appliance_official import _build_direct_product_candidates, _extract_size_from_text, _primary_model_token
from App.services.export_service import build_spec_list_excel
from App.services.parsing import enrich_snapshot_rooms, parse_documents
from App.services.runtime import ensure_job_dirs, utc_now_iso


class SmokeTest(unittest.TestCase):
    @classmethod
    def tearDownClass(cls) -> None:
        shutil.rmtree(TEST_DATA_DIR, ignore_errors=True)

    def setUp(self) -> None:
        self._reset_db()

    def tearDown(self) -> None:
        self._reset_db()

    def _reset_db(self) -> None:
        with store.connect() as conn:
            conn.execute("DELETE FROM auth_events")
            conn.execute("DELETE FROM reviews")
            conn.execute("DELETE FROM snapshot_verifications")
            conn.execute("DELETE FROM snapshots")
            conn.execute("DELETE FROM runs")
            conn.execute("DELETE FROM job_files")
            conn.execute("DELETE FROM jobs")
            conn.execute("DELETE FROM builder_templates")
            conn.execute("DELETE FROM builders")

    def test_health(self) -> None:
        client = TestClient(app)
        response = client.get("/api/health")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["status"], "ok")

    def test_runtime_loads_env_file_before_constants(self) -> None:
        import App.services.runtime as runtime

        env_path = runtime.ENV_PATH
        original_text = env_path.read_text(encoding="utf-8") if env_path.exists() else None
        env_keys = ["SPEC_EXTRACTION_ENABLE_OPENAI", "SPEC_EXTRACTION_OPENAI_MODEL", "OPENAI_API_KEY"]
        original_env = {key: os.environ.get(key) for key in env_keys}
        for key in env_keys:
            os.environ.pop(key, None)
        try:
            env_path.write_text(
                "SPEC_EXTRACTION_ENABLE_OPENAI=1\n"
                "SPEC_EXTRACTION_OPENAI_MODEL=gpt-4.1-mini\n"
                "OPENAI_API_KEY=test-openai-key\n",
                encoding="utf-8",
            )
            runtime = importlib.reload(runtime)
            self.assertTrue(runtime.OPENAI_ENABLED)
            self.assertEqual(runtime.OPENAI_MODEL, "gpt-4.1-mini")
            self.assertEqual(runtime.OPENAI_API_KEY, "test-openai-key")
        finally:
            if original_text is None:
                env_path.unlink(missing_ok=True)
            else:
                env_path.write_text(original_text, encoding="utf-8")
            for key, value in original_env.items():
                if value is None:
                    os.environ.pop(key, None)
                else:
                    os.environ[key] = value
            importlib.reload(runtime)

    def test_parser_extracts_room_and_basic_appliance(self) -> None:
        snapshot = parse_documents(
            job_no="37529",
            builder_name="Clarendon",
            source_kind="spec",
            documents=[
                {
                    "file_name": "sample.txt",
                    "role": "spec",
                    "pages": [
                        {
                            "page_no": 1,
                            "text": (
                                "Kitchen\n"
                                "Bench Tops 20mm stone by builder\n"
                                "Door Colour Polytec Classic White Matt\n"
                                "Kickboard Matching White\n"
                                "Handles Hettich 9070585 Chrome\n"
                                "Drawers Soft Close\n"
                                "Hinges Not Soft Close\n"
                                "Splashback Tiled by others\n"
                                "Flooring Hybrid flooring\n"
                                "Cooktop: Westinghouse WHC943BD 90cm\n"
                            ),
                            "needs_ocr": False,
                        }
                    ],
                }
            ],
        )
        self.assertEqual(snapshot["rooms"][0]["room_key"], "kitchen")
        self.assertEqual(snapshot["rooms"][0]["drawers_soft_close"], "Soft Close")
        self.assertEqual(snapshot["rooms"][0]["hinges_soft_close"], "Not Soft Close")
        self.assertEqual(snapshot["appliances"][0]["model_no"], "WHC943BD")
        self.assertEqual(snapshot["analysis"]["mode"], "heuristic_only")
        self.assertEqual(snapshot["analysis"]["parser_strategy"], "global_conservative")

    def test_parser_extracts_explicit_appliance_model_numbers(self) -> None:
        snapshot = parse_documents(
            job_no="37529",
            builder_name="Clarendon",
            source_kind="spec",
            documents=[
                {
                    "file_name": "appliance-schedule.pdf",
                    "role": "spec",
                    "pages": [
                        {
                            "page_no": 7,
                            "text": (
                                "Under Bench Oven: WESTINGHOUSE 2 X WVE6515SDA 60CM ELECTRIC OVEN S/S ELECTRIC\n"
                                "Dishwasher Make: WESTINGHOUSE Freestanding (WSF6608X) 600mm S/S\n"
                                "Fridge: N/A CLIENT TO CHECK\n"
                                "Integrated Fridge/Freezer: FISHER & PAYKEL 2 X RB60V18\n"
                            ),
                            "needs_ocr": False,
                        }
                    ],
                }
            ],
        )
        appliances = {row["appliance_type"]: row for row in snapshot["appliances"]}
        self.assertEqual(appliances["Oven"]["make"], "Westinghouse")
        self.assertEqual(appliances["Oven"]["model_no"], "2 x WVE6515SDA")
        self.assertEqual(appliances["Dishwasher"]["model_no"], "WSF6608X")
        self.assertEqual(appliances["Fridge"]["make"], "Fisher & Paykel")
        self.assertEqual(appliances["Fridge"]["model_no"], "2 x RB60V18")
        models = " ".join(row["model_no"] for row in snapshot["appliances"])
        self.assertNotIn("WESTINGHOUSE", models)
        self.assertNotIn("CLIENT", models)
        self.assertNotIn("OPENING", models)
        self.assertNotIn("ELECTRIC", models)

    def test_structure_first_parse_uses_layout_room_blocks_for_evoca(self) -> None:
        snapshot = parse_documents(
            job_no="38225",
            builder_name="Evoca",
            source_kind="spec",
            documents=[
                {
                    "file_name": "evoca-layout.pdf",
                    "role": "spec",
                    "pages": [
                        {
                            "page_no": 1,
                            "raw_text": "Robe Sliding Type Frame Colour Belgian Oak Matt",
                            "text": "Robe Sliding Type Frame Colour Belgian Oak Matt",
                            "needs_ocr": False,
                            "page_layout": {
                                "page_type": "joinery",
                                "section_label": "ROBE JOINERY SELECTION SHEET",
                                "room_label": "ROBE",
                                "room_blocks": [
                                    {
                                        "room_label": "ROBE",
                                        "rows": [
                                            {
                                                "row_label": "Base Cabinetry Colour",
                                                "value_region_text": "Belgian Oak Matt",
                                                "supplier_region_text": "Polytec",
                                                "notes_region_text": "",
                                                "row_kind": "material",
                                            },
                                            {
                                                "row_label": "Handles",
                                                "value_region_text": "4062-128-TG",
                                                "supplier_region_text": "",
                                                "notes_region_text": "",
                                                "row_kind": "handle",
                                            },
                                        ],
                                    }
                                ],
                                "rows": [],
                            },
                        }
                    ],
                }
            ],
        )
        rooms = {row["room_key"]: row for row in snapshot["rooms"]}
        self.assertIn("robe", rooms)
        self.assertNotIn("robe_sliding_type_frame_colour", rooms)
        self.assertEqual(rooms["robe"]["original_room_label"], "ROBE")

    def test_structure_first_parse_uses_layout_room_blocks_for_simonds(self) -> None:
        snapshot = parse_documents(
            job_no="s1",
            builder_name="Simonds",
            source_kind="spec",
            documents=[
                {
                    "file_name": "simonds-layout.pdf",
                    "role": "spec",
                    "pages": [
                        {
                            "page_no": 1,
                            "raw_text": "Kitchen Wall Run Base Cabinet Panels Manufacturer Laminex",
                            "text": "Kitchen Wall Run Base Cabinet Panels Manufacturer Laminex",
                            "needs_ocr": False,
                            "page_layout": {
                                "page_type": "joinery",
                                "section_label": "KITCHEN COLOUR SCHEDULE",
                                "room_label": "KITCHEN",
                                "room_blocks": [
                                    {
                                        "room_label": "KITCHEN",
                                        "rows": [
                                            {
                                                "row_label": "Base Cabinetry Colour",
                                                "value_region_text": "Classic White Matt",
                                                "supplier_region_text": "Laminex",
                                                "notes_region_text": "",
                                                "row_kind": "material",
                                            }
                                        ],
                                    }
                                ],
                                "rows": [],
                            },
                        }
                    ],
                }
            ],
        )
        rooms = {row["room_key"]: row for row in snapshot["rooms"]}
        self.assertIn("kitchen", rooms)
        self.assertNotIn("kitchen_wall_run_base_cabinet_panels_manufacturer_laminex", rooms)
        self.assertEqual(rooms["kitchen"]["original_room_label"], "KITCHEN")

    def test_coerce_layout_room_blocks_keeps_default_room_when_property_row_mentions_other_room(self) -> None:
        layout = {
            "page_type": "joinery",
            "section_label": "",
            "room_label": "",
            "room_blocks": [
                {
                    "room_label": "",
                    "rows": [
                        {
                            "row_label": "Underbench including Island",
                            "value_region_text": "",
                            "supplier_region_text": "",
                            "notes_region_text": "",
                            "row_kind": "material",
                        },
                        {
                            "row_label": "Manufacturer",
                            "value_region_text": "Polytec",
                            "supplier_region_text": "",
                            "notes_region_text": "",
                            "row_kind": "material",
                        },
                        {
                            "row_label": "Pantry Door Handle",
                            "value_region_text": "** Pantry Door Handle**",
                            "supplier_region_text": "#N/A",
                            "notes_region_text": "",
                            "row_kind": "handle",
                        },
                    ],
                }
            ],
            "rows": [],
        }
        blocks = parsing_module._coerce_layout_room_blocks(
            layout,
            section_label="",
            room_label="",
            raw_page_text="Kitchen\nUnderbench including Island",
            page_type="joinery",
        )
        self.assertEqual(len(blocks), 1)
        self.assertEqual(blocks[0]["room_label"], "Kitchen")

    def test_coerce_layout_room_blocks_splits_inline_room_heading_after_default_room(self) -> None:
        layout = {
            "page_type": "joinery",
            "section_label": "",
            "room_label": "",
            "room_blocks": [
                {
                    "room_label": "",
                    "rows": [
                        {
                            "row_label": "Benchtops",
                            "value_region_text": "",
                            "supplier_region_text": "",
                            "notes_region_text": "",
                            "row_kind": "material",
                        },
                        {
                            "row_label": "Manufacturer",
                            "value_region_text": "Quantum Quartz",
                            "supplier_region_text": "",
                            "notes_region_text": "",
                            "row_kind": "material",
                        },
                        {
                            "row_label": "Bulters/WIP",
                            "value_region_text": "",
                            "supplier_region_text": "",
                            "notes_region_text": "",
                            "row_kind": "other",
                        },
                        {
                            "row_label": "Benchtops",
                            "value_region_text": "",
                            "supplier_region_text": "",
                            "notes_region_text": "",
                            "row_kind": "material",
                        },
                        {
                            "row_label": "Manufacturer",
                            "value_region_text": "Caesarstone",
                            "supplier_region_text": "",
                            "notes_region_text": "",
                            "row_kind": "material",
                        },
                    ],
                }
            ],
            "rows": [],
        }
        blocks = parsing_module._coerce_layout_room_blocks(
            layout,
            section_label="",
            room_label="",
            raw_page_text="Kitchen\nBulters/WIP",
            page_type="joinery",
        )
        self.assertEqual([block["room_label"] for block in blocks], ["Kitchen", "Butlers/WIP"])
        self.assertEqual(blocks[0]["rows"][0]["row_label"], "Benchtops")
        self.assertEqual(blocks[1]["rows"][0]["row_label"], "Benchtops")

    def test_coerce_layout_room_blocks_ignores_metadata_block_when_explicit_room_blocks_exist(self) -> None:
        layout = {
            "page_type": "joinery",
            "section_label": "Client Initials: ______",
            "room_label": "Initials",
            "room_blocks": [
                {
                    "room_label": "Initials",
                    "rows": [
                        {
                            "row_label": "Manufacturer",
                            "value_region_text": "Noise",
                            "supplier_region_text": "",
                            "notes_region_text": "",
                            "row_kind": "material",
                        }
                    ],
                },
                {
                    "room_label": "Butlers",
                    "rows": [
                        {
                            "row_label": "Benchtops Manufacturer",
                            "value_region_text": "Not Applicable",
                            "supplier_region_text": "",
                            "notes_region_text": "",
                            "row_kind": "material",
                        }
                    ],
                },
                {
                    "room_label": "Laundry",
                    "rows": [
                        {
                            "row_label": "Benchtops Manufacturer",
                            "value_region_text": "Quantum Quartz",
                            "supplier_region_text": "",
                            "notes_region_text": "",
                            "row_kind": "material",
                        }
                    ],
                },
            ],
            "rows": [],
        }
        blocks = parsing_module._coerce_layout_room_blocks(
            layout,
            section_label="Client Initials: ______",
            room_label="Initials",
            raw_page_text="Client Initials\nButlers\nLaundry",
            page_type="joinery",
        )
        self.assertEqual([block["room_label"] for block in blocks], ["Butlers", "Laundry"])
        self.assertTrue(all(block["room_label"] != "Initials" for block in blocks))
        self.assertIn("Benchtops Manufacturer", [row["row_label"] for row in blocks[0]["rows"]])

    def test_infer_default_layout_room_label_ignores_accessory_value_room_words(self) -> None:
        inferred = parsing_module._infer_default_layout_room_label(
            room_label="",
            section_label="",
            raw_page_text="",
            rows=[
                {
                    "row_label": "Accessories",
                    "value_region_text": "ALDER SACHI ROBE HOOK IN MATT BLACK",
                    "supplier_region_text": "",
                    "notes_region_text": "",
                    "row_kind": "accessory",
                },
                {
                    "row_label": "Kitchen Sink",
                    "value_region_text": "",
                    "supplier_region_text": "",
                    "notes_region_text": "",
                    "row_kind": "sink",
                },
            ],
            page_type="joinery",
        )
        self.assertEqual(inferred, "Kitchen")

    def test_structure_first_parse_trims_builder_noise_from_room_titles(self) -> None:
        snapshot = parse_documents(
            job_no="e2",
            builder_name="Evoca",
            source_kind="spec",
            documents=[
                {
                    "file_name": "evoca-noisy-room-labels.pdf",
                    "role": "spec",
                    "pages": [
                        {
                            "page_no": 1,
                            "raw_text": "Robe Sliding Type Standard Frame",
                            "text": "Robe Sliding Type Standard Frame",
                            "needs_ocr": False,
                            "page_layout": {
                                "page_type": "joinery",
                                "section_label": "ROBE SLIDING TYPE STANDARD FRAME",
                                "room_label": "Robe Sliding Type Standard Frame",
                                "room_blocks": [
                                    {
                                        "room_label": "Robe Sliding Type Standard Frame",
                                        "rows": [
                                            {
                                                "row_label": "Base Cabinetry Colour",
                                                "value_region_text": "Belgian Oak Matt",
                                                "supplier_region_text": "Polytec",
                                                "notes_region_text": "",
                                                "row_kind": "material",
                                            }
                                        ],
                                    }
                                ],
                                "rows": [],
                            },
                        },
                        {
                            "page_no": 2,
                            "raw_text": "Study Desk",
                            "text": "Study Desk",
                            "needs_ocr": False,
                            "page_layout": {
                                "page_type": "joinery",
                                "section_label": "STUDY DESK",
                                "room_label": "Study Desk",
                                "room_blocks": [
                                    {
                                        "room_label": "Study Desk",
                                        "rows": [
                                            {
                                                "row_label": "Base Cabinetry Colour",
                                                "value_region_text": "Classic White",
                                                "supplier_region_text": "Polytec",
                                                "notes_region_text": "",
                                                "row_kind": "material",
                                            }
                                        ],
                                    }
                                ],
                                "rows": [],
                            },
                        },
                    ],
                }
            ],
        )
        rooms = {row["room_key"]: row for row in snapshot["rooms"]}
        self.assertIn("robe", rooms)
        self.assertIn("study_desk", rooms)
        self.assertNotIn("robe_sliding_type_standard_frame", rooms)
        self.assertEqual(rooms["study_desk"]["original_room_label"], "Study Desk")

    def test_imperial_tap_overlay_prefers_layout_rows_and_ignores_client_metadata(self) -> None:
        documents = [
            {
                "file_name": "imperial-tap.pdf",
                "role": "spec",
                "pages": [
                    {
                        "page_no": 8,
                        "text": "Client Name: Eloise Cawcutt-Foxover\nTAPWARE (KITCHEN) Franke Eos Neo pull out tap copper TA9601CP",
                        "raw_text": "Client Name: Eloise Cawcutt-Foxover\nTAPWARE (KITCHEN) Franke Eos Neo pull out tap copper TA9601CP",
                        "needs_ocr": False,
                        "page_layout": {
                            "page_type": "unknown",
                            "section_label": "TAPWARE (KITCHEN)",
                            "room_label": "KITCHEN",
                            "room_blocks": [
                                {
                                    "room_label": "KITCHEN",
                                    "rows": [
                                        {
                                            "row_label": "Tap",
                                            "value_region_text": "Franke Eos Neo pull out tap copper TA9601CP",
                                            "supplier_region_text": "BY CLIENT",
                                            "notes_region_text": "",
                                            "row_kind": "tap",
                                        }
                                    ],
                                }
                            ],
                            "rows": [
                                {
                                    "row_label": "client name",
                                    "value_region_text": "Eloise Cawcutt-Foxover",
                                    "supplier_region_text": "",
                                    "notes_region_text": "",
                                    "row_kind": "metadata",
                                }
                            ],
                        },
                    }
                ],
            }
        ]
        overlays = parsing_module._collect_imperial_room_overlays(documents)
        self.assertEqual(overlays["kitchen"]["tap_info"], "Franke Eos Neo pull out tap copper TA9601CP")

    def test_clean_handle_entries_merges_supplier_only_prefix_with_description(self) -> None:
        cleaned = parsing_module._clean_handle_entries(
            [
                "Furnware",
                "Finger Pull on Uppers",
                "Momo Barrington Eclipse Plain 96mm in Matt Brass Part Number:BEPL96.MBR - Horizontal on Drawers and Vertical on Doors",
            ]
        )
        self.assertIn(
            "Furnware - Momo Barrington Eclipse Plain 96mm in Matt Brass Part Number:BEPL96.MBR - Horizontal on Drawers and Vertical on Doors",
            cleaned,
        )
        self.assertIn("Finger Pull on Uppers", cleaned)

    def test_clean_handle_entries_merges_slash_split_drawers_fragment(self) -> None:
        cleaned = parsing_module._clean_handle_entries(
            [
                "Hettich - Matane 9113228 Brushed Stainless Steel Look 105MM Long - Horizontal Mount Doors/",
                "Drawers - Base Cabinets * Drawer Location - Centre to Profile* Door Location - 25MM down and 50MM in",
            ]
        )
        self.assertEqual(
            cleaned,
            [
                "Hettich - Matane 9113228 Brushed Stainless Steel Look 105MM Long - Horizontal Mount Doors/Drawers - Base Cabinets * Drawer Location - Centre to Profile* Door Location - 25MM down and 50MM in"
            ],
        )

    def test_collect_field_does_not_treat_sink_mixer_or_basin_waste_as_sink_or_basin(self) -> None:
        lines = [
            "Sink Mixer Spin Tall Basin Mixer Chrome Highgrove",
            "Basin Byron Bench Mount Basin White Gloss Highgrove",
            "Basin Waste Pop Up Waste Without Overflow 32x80mm Chrome Highgrove",
        ]
        self.assertEqual(parsing_module._collect_field(lines, ["Sink"]), [])
        self.assertEqual(parsing_module._collect_field(lines, ["Basin"]), ["Byron Bench Mount Basin White Gloss Highgrove"])

    def test_dedupe_appliances_drops_drawing_dimension_noise_when_real_model_exists(self) -> None:
        rows = [
            parsing_module.ApplianceRow(
                appliance_type="Fridge",
                make="Fisher & Paykel",
                model_no="30120018EQ740900740EQ30",
                source_file="drawings-and-colours.pdf",
                page_refs="1",
                evidence_snippet="Fridge 30120018EQ740900740EQ30",
            ),
            parsing_module.ApplianceRow(
                appliance_type="Fridge",
                make="Fisher & Paykel",
                model_no="2 x RB60V18",
                source_file="drawings-and-colours.pdf",
                page_refs="5",
                evidence_snippet="Fridge 2 x RB60V18",
            ),
        ]
        deduped = parsing_module._dedupe_appliances(rows)
        self.assertEqual([(row.make, row.model_no) for row in deduped], [("Fisher & Paykel", "2 x RB60V18")])

    def test_imperial_toe_kick_cleaner_preserves_match_above_materials(self) -> None:
        cleaned = parsing_module._imperial_clean_toe_kick_value(
            [
                "MATCH ABOVE Polytec Classic White Matt or Laminex Gumnut Natural Finish 2606 or Laminex Blackbutt Truescale Natural Finish 2618",
                "Polytec + Laminex",
                "",
            ]
        )
        self.assertEqual(
            cleaned,
            "Match Above Polytec Classic White Matt / Laminex Gumnut Natural Finish 2606 / Laminex Blackbutt Truescale Natural Finish 2618",
        )

    def test_openai_request_retries_after_http_429(self) -> None:
        response_payload = {"output_text": "{\"page_type\": \"joinery\"}"}

        class _FakeResponse:
            def __enter__(self):
                return self

            def __exit__(self, exc_type, exc, tb):
                return False

            def read(self) -> bytes:
                import json

                return json.dumps(response_payload).encode("utf-8")

        too_many = urllib.error.HTTPError(
            url="https://api.openai.com/v1/responses",
            code=429,
            msg="Too Many Requests",
            hdrs={"Retry-After": "0"},
            fp=io.BytesIO(b""),
        )
        with mock.patch.object(extraction_service.runtime, "OPENAI_REQUEST_MIN_INTERVAL_SECONDS", 0.0), mock.patch.object(
            extraction_service.runtime, "OPENAI_REQUEST_MAX_RETRIES", 2
        ), mock.patch.object(extraction_service.runtime, "OPENAI_REQUEST_RETRY_BASE_SECONDS", 0.0), mock.patch(
            "App.services.extraction_service.time.sleep", return_value=None
        ), mock.patch(
            "App.services.extraction_service.urllib.request.urlopen",
            side_effect=[too_many, _FakeResponse()],
        ) as mocked_urlopen:
            extraction_service._LAST_OPENAI_REQUEST_AT = 0.0
            payload = extraction_service._post_responses_api_content([{"type": "input_text", "text": "{}"}], model="gpt-test")
        self.assertEqual(payload, response_payload)
        self.assertEqual(mocked_urlopen.call_count, 2)

    def test_openai_request_fails_fast_on_insufficient_quota(self) -> None:
        quota_error = urllib.error.HTTPError(
            url="https://api.openai.com/v1/responses",
            code=429,
            msg="Too Many Requests",
            hdrs={},
            fp=io.BytesIO(
                b'{'
                b'"error":{'
                b'"message":"You exceeded your current quota, please check your plan and billing details.",'
                b'"type":"insufficient_quota",'
                b'"param":null,'
                b'"code":"insufficient_quota"'
                b"}}"
            ),
        )
        with mock.patch.object(extraction_service.runtime, "OPENAI_REQUEST_MIN_INTERVAL_SECONDS", 0.0), mock.patch.object(
            extraction_service.runtime, "OPENAI_REQUEST_MAX_RETRIES", 4
        ), mock.patch(
            "App.services.extraction_service.urllib.request.urlopen",
            side_effect=[quota_error],
        ) as mocked_urlopen:
            extraction_service._LAST_OPENAI_REQUEST_AT = 0.0
            with self.assertRaisesRegex(RuntimeError, "insufficient_quota"):
                extraction_service._post_responses_api_content([{"type": "input_text", "text": "{}"}], model="gpt-test")
        self.assertEqual(mocked_urlopen.call_count, 1)

    def test_extract_site_address_from_glued_simonds_header(self) -> None:
        raw_text = (
            "Colour Selections\n"
            "Project Administrator:Lyn SpicerClient Name:Lot 4269, 2 BELLTHORPE STREET REDLAND BAY QLD 4165"
            "Job Address: Colour Consultant:Jess McMahon"
        )
        self.assertEqual(
            parsing_module._extract_site_address_from_text(raw_text),
            "Lot 4269, 2 BELLTHORPE STREET REDLAND BAY QLD 4165",
        )

    def test_generic_sinkware_overlay_does_not_pollute_tall_cabinetry(self) -> None:
        section = {
            "page_type": "sinkware_tapware",
            "original_section_label": "Ensuite",
            "file_name": "sample.pdf",
            "page_nos": [13],
            "text": "Ensuite sinkware page",
            "layout_rows": [
                {"row_label": "Basin", "value_text": "", "supplier_text": "", "notes_text": ""},
                {"row_label": "Type", "value_text": "Overmount", "supplier_text": "", "notes_text": ""},
                {
                    "row_label": "Floor Waste",
                    "value_text": "Spin Gun Metal Tall Basin Mixer (SP110-GM)",
                    "supplier_text": "",
                    "notes_text": "",
                },
            ],
        }
        overlay = extraction_service._extract_generic_layout_overlay(section)
        self.assertEqual(overlay["door_colours_tall"], "")
        self.assertEqual(overlay["basin_info"], "Overmount")

    def test_structure_first_parse_uses_sink_tap_pages_as_overlays_not_rooms(self) -> None:
        snapshot = parse_documents(
            job_no="s2",
            builder_name="Simonds",
            source_kind="spec",
            documents=[
                {
                    "file_name": "simonds-layout.pdf",
                    "role": "spec",
                    "pages": [
                        {
                            "page_no": 1,
                            "raw_text": "KITCHEN COLOUR SCHEDULE",
                            "text": "KITCHEN COLOUR SCHEDULE",
                            "needs_ocr": False,
                            "page_layout": {
                                "page_type": "joinery",
                                "section_label": "KITCHEN COLOUR SCHEDULE",
                                "room_label": "KITCHEN",
                                "room_blocks": [
                                    {
                                        "room_label": "KITCHEN",
                                        "rows": [
                                            {
                                                "row_label": "Base Cabinetry Colour",
                                                "value_region_text": "Classic White Matt",
                                                "supplier_region_text": "Laminex",
                                                "notes_region_text": "",
                                                "row_kind": "material",
                                            }
                                        ],
                                    }
                                ],
                                "rows": [],
                            },
                        },
                        {
                            "page_no": 2,
                            "raw_text": "SINKWARE (Kitchen Sink Range) Franke Sirius",
                            "text": "SINKWARE (Kitchen Sink Range) Franke Sirius",
                            "needs_ocr": False,
                            "page_layout": {
                                "page_type": "sinkware_tapware",
                                "section_label": "SINKWARE & TAPWARE",
                                "room_label": "",
                                "room_blocks": [
                                    {
                                        "room_label": "Kitchen Sink Range",
                                        "rows": [
                                            {
                                                "row_label": "SINKWARE (Kitchen Sink Range)",
                                                "value_region_text": "Franke Sirius",
                                                "supplier_region_text": "",
                                                "notes_region_text": "",
                                                "row_kind": "sink",
                                            }
                                        ],
                                    }
                                ],
                                "rows": [],
                            },
                        },
                    ],
                }
            ],
        )
        rooms = {row["room_key"]: row for row in snapshot["rooms"]}
        self.assertEqual(set(rooms), {"kitchen"})

    def test_structure_first_parse_rejects_sink_tap_room_blocks_that_are_not_real_rooms(self) -> None:
        snapshot = parse_documents(
            job_no="s2-bad-room",
            builder_name="Simonds",
            source_kind="spec",
            documents=[
                {
                    "file_name": "simonds-layout.pdf",
                    "role": "spec",
                    "pages": [
                        {
                            "page_no": 0,
                            "raw_text": "BATHROOM COLOUR SCHEDULE",
                            "text": "BATHROOM COLOUR SCHEDULE",
                            "needs_ocr": False,
                            "page_layout": {
                                "page_type": "joinery",
                                "section_label": "BATHROOM COLOUR SCHEDULE",
                                "room_label": "Bathroom",
                                "room_blocks": [
                                    {
                                        "room_label": "Bathroom",
                                        "rows": [
                                            {
                                                "row_label": "Base Cabinetry Colour",
                                                "value_region_text": "Classic White Matt",
                                                "supplier_region_text": "Laminex",
                                                "notes_region_text": "",
                                                "row_kind": "material",
                                            }
                                        ],
                                    }
                                ],
                                "rows": [],
                            },
                        },
                        {
                            "page_no": 1,
                            "raw_text": "SINKWARE & TAPWARE",
                            "text": "SINKWARE & TAPWARE",
                            "needs_ocr": False,
                            "page_layout": {
                                "page_type": "sinkware_tapware",
                                "section_label": "SINKWARE & TAPWARE",
                                "room_label": "",
                                "room_blocks": [
                                    {
                                        "room_label": "2No Bath towel hooks + 1No Hand towel hook",
                                        "rows": [
                                            {
                                                "row_label": "Toilet Roll Holder",
                                                "value_region_text": "Matt Black",
                                                "supplier_region_text": "",
                                                "notes_region_text": "",
                                                "row_kind": "accessory",
                                            }
                                        ],
                                    },
                                    {
                                        "room_label": "Bathroom",
                                        "rows": [
                                            {
                                                "row_label": "Basin Mixer",
                                                "value_region_text": "Matt Black",
                                                "supplier_region_text": "",
                                                "notes_region_text": "",
                                                "row_kind": "tap",
                                            }
                                        ],
                                    },
                                ],
                                "rows": [],
                            },
                        }
                    ],
                }
            ],
        )
        room_labels = [row["original_room_label"] for row in snapshot["rooms"]]
        self.assertEqual(room_labels, ["Bathroom"])

    def test_structure_first_parse_trims_sinkware_suffixes_from_room_blocks(self) -> None:
        self.assertEqual(parsing_module._clean_layout_room_label("Master Ensuite Shower"), "Master Ensuite")
        self.assertEqual(parsing_module._clean_layout_room_label("Bathroom Bath/Spa"), "Bathroom")

    def test_parser_normalizes_soft_close_states(self) -> None:
        snapshot = parse_documents(
            job_no="38111",
            builder_name="Clarendon",
            source_kind="spec",
            documents=[
                {
                    "file_name": "soft-close.txt",
                    "role": "spec",
                    "pages": [
                        {
                            "page_no": 1,
                            "text": (
                                "Kitchen\n"
                                "Drawers Hettich Multitech standard construction runners - NOT soft close\n"
                                "Hinges Blumotion soft-close hinges\n"
                            ),
                            "needs_ocr": False,
                        }
                    ],
                }
            ],
        )
        room = snapshot["rooms"][0]
        self.assertEqual(room["drawers_soft_close"], "Not Soft Close")
        self.assertEqual(room["hinges_soft_close"], "Soft Close")

    def test_parser_normalizes_brand_casing_and_preserves_benchtop_text(self) -> None:
        snapshot = parse_documents(
            job_no="37016",
            builder_name="Clarendon",
            source_kind="spec",
            documents=[
                {
                    "file_name": "clarendon.txt",
                    "role": "spec",
                    "pages": [
                        {
                            "page_no": 1,
                            "text": (
                                "Kitchen\n"
                                "Bench Tops Quantum Zero Midnight Black - 20mm pencil round edge to cooktop run; "
                                "Quantum Zero Venatino Statuario - 40mm mitred apron edge to island bench\n"
                                "Door Colour PolYTEC Classic White Matt Finish Thermolaminate\n"
                            ),
                            "needs_ocr": False,
                        }
                    ],
                }
            ],
        )
        room = snapshot["rooms"][0]
        self.assertIn("Quantum Zero Midnight Black - 20mm pencil round edge", room["bench_tops_wall_run"])
        self.assertIn("Quantum Zero Venatino Statuario - 40mm mitred apron edge", room["bench_tops_island"])
        self.assertEqual(room["door_panel_colours"][0], "Polytec Classic White Matt Finish Thermolaminate")

    def test_parser_splits_inline_kitchen_benchtop_sentence_into_wall_run_and_island(self) -> None:
        snapshot = parse_documents(
            job_no="37050",
            builder_name="Clarendon",
            source_kind="spec",
            documents=[
                {
                    "file_name": "clarendon-kitchen.txt",
                    "role": "spec",
                    "pages": [
                        {
                            "page_no": 1,
                            "text": (
                                "Kitchen\n"
                                "Benchtop - Quantum Zero White Swirl - 20MM Pencil Round Edge - TO Cooktop Run / "
                                "40MM Mitred Apron Edge - TO Island Bench\n"
                            ),
                            "needs_ocr": False,
                        }
                    ],
                }
            ],
        )
        room = snapshot["rooms"][0]
        self.assertEqual(room["bench_tops_wall_run"], "Quantum Zero White Swirl - 20MM Pencil Round Edge")
        self.assertEqual(room["bench_tops_island"], "Quantum Zero White Swirl - 40MM Mitred Apron Edge")
        self.assertIn("Back Benchtops Quantum Zero White Swirl - 20MM Pencil Round Edge", room["bench_tops"])
        self.assertIn("Island Benchtop Quantum Zero White Swirl - 40MM Mitred Apron Edge", room["bench_tops"])

    def test_parser_stops_benchtop_before_doors_panels_field(self) -> None:
        snapshot = parse_documents(
            job_no="37051",
            builder_name="Clarendon",
            source_kind="spec",
            documents=[
                {
                    "file_name": "vanities.txt",
                    "role": "spec",
                    "pages": [
                        {
                            "page_no": 1,
                            "text": (
                                "Vanities\n"
                                "Benchtop Quantum Zero Luna White - 20MM Pencil Round Edge / 140MM Mitred Apron Edge - to Powder Room 2\n"
                                "Doors/Panels - Polytec Jamaican Oak Matt Finish Melamine with Matching 1MM ABS Edges (Vertical Grain Direction)\n"
                            ),
                            "needs_ocr": False,
                        }
                    ],
                }
            ],
        )
        room = snapshot["rooms"][0]
        self.assertNotIn("Doors/Panels", " ".join(room["bench_tops"]))
        self.assertNotIn("Polytec Jamaican Oak", " ".join(room["bench_tops"]))
        self.assertIn("Polytec Jamaican Oak Matt Finish Melamine with Matching 1MM ABS Edges", room["door_colours_base"])
        self.assertEqual(room["door_colours_overheads"], "")

    def test_clarendon_job25_address_extraction_stops_before_schedule_text(self) -> None:
        documents = [
            {
                "file_name": "49906622 AMENDED Signed Drawings and Colours REV B 20-10-25.pdf",
                "pages": [
                    {
                        "page_no": 2,
                        "raw_text": (
                            "Site Address: Lot 8 (#25) Lake Serenity Boulevard Helensvale REV B "
                            "BENCHTOP - QUANTUM ZERO WHITE SWIRL - 20MM PENCIL ROUND EDGE"
                        ),
                        "text": "",
                    }
                ],
            }
        ]
        self.assertEqual(
            parsing_module._extract_site_address_from_documents(documents),
            "Lot 8 (#25) Lake Serenity Boulevard Helensvale",
        )

    def test_site_address_extraction_salvages_suffix_after_private_prefix(self) -> None:
        documents = [
            {
                "file_name": "SIGNED FINAL COLOURS_FOXOVER 21 Shadowood st KENMORE 23 3 26.pdf",
                "pages": [
                    {
                        "page_no": 1,
                        "raw_text": "Address:2510-076 - Private - 21 Shadowood Street, Kenmore Hills",
                        "text": "",
                    }
                ],
            }
        ]
        self.assertEqual(
            parsing_module._extract_site_address_from_documents(documents),
            "21 Shadowood Street, Kenmore Hills",
        )

    def test_site_address_extraction_stops_before_job_number(self) -> None:
        documents = [
            {
                "file_name": "evoca-example.pdf",
                "pages": [
                    {
                        "page_no": 1,
                        "raw_text": "Lot Address Lot 1038, Oyster St, Worongary QLD 4213 Job Number EVOC467",
                        "text": "",
                    }
                ],
            }
        ]
        self.assertEqual(
            parsing_module._extract_site_address_from_documents(documents),
            "Lot 1038, Oyster St, Worongary QLD 4213",
        )

    def test_site_address_extraction_falls_back_to_pdf_path_when_loaded_text_is_sparse(self) -> None:
        class _FakePage:
            def extract_text(self) -> str:
                return "Site Address : LOT 12 No. (#22) GEORGIA CLOSE, BIRKDALE Job No: 49906613"

        with tempfile.TemporaryDirectory() as tmp_dir:
            fake_pdf_path = Path(tmp_dir) / "clarendon-afc.pdf"
            fake_pdf_path.write_bytes(b"%PDF-1.4\n%test\n")
            documents = [
                {
                    "file_name": "clarendon-afc.pdf",
                    "path": str(fake_pdf_path),
                    "pages": [
                        {
                            "page_no": 1,
                            "raw_text": "Site Address :\nJob No:\nHome:",
                            "text": "Site Address :\nJob No:\nHome:",
                        }
                    ],
                }
            ]
            with (
                mock.patch.object(parsing_module, "pdfplumber") as mock_pdfplumber,
                mock.patch.object(parsing_module, "PdfReader", return_value=types.SimpleNamespace(pages=[_FakePage()])),
            ):
                mock_pdfplumber.open.side_effect = RuntimeError("force reader fallback")
                self.assertEqual(
                    parsing_module._extract_site_address_from_documents(documents),
                    "LOT 12 No. (#22) GEORGIA CLOSE, BIRKDALE",
                )

    def test_clarendon_job25_polish_prefers_raw_text_and_preserves_summary_detail(self) -> None:
        afc_raw = "LOT 8 No. (#25) Lake Serenity Boulevard, HELENSVALE"
        kitchen_raw = (
            "Site Address: Lot 8 (#25) Lake Serenity Boulevard Helensvale REV B "
            "KITCHEN COLOUR SCHEDULE "
            "BENCHTOP - QUANTUM ZERO WHITE SWIRL - 20MM PENCIL ROUND EDGE - TO COOKTOP + SIDE BENCHTOP/ "
            "40MM MITRED APRON EDGE - TO ISLAND BENCH "
            "DOOR COLOUR - POLYTEC ASTON WHITE SMOOTH FINISH THERMOLAMINATE - HAMPTON EM9 PROFILE"
            "PLAIN GLASS DISPLAY CABINET WITH STANDARD WHITE INTERNALS AND MELAMINE SHELF "
            "THERMOLAMINATE NOTES : * BULKHEAD SHADOWLINE : MATCHING MELAMINE FINISH * KICKBOARDS : MATCHING MELAMINE FINISH "
            "HANDLE 1 - HETTICH BELLUNO 9995772 200MM LONG BRUSHED STAINLESS STEEL LOOK - TO BASE CABINETS "
            "HANDLE 2 - HETTICH SALEMI 9113368 30MM BRUSHED STAINLESS STEEL LOOK - TO UPPER CABINETS "
            "DOOR HINGES - HETTICH SOFT CLOSE "
            "DRAWER RUNNERS - HETTICH INNOTECH ATIRA SOFT CLOSE RUNNERS"
        )
        vanities_raw = (
            "VANITIES COLOUR SCHEDULE "
            "BENCHTOP - QUANTUM ZERO WHITE SWIRL - 20MM PENCIL ROUND EDGE "
            "DOOR COLOUR - POLYTEC ASTON WHITE SMOOTH FINISH THERMOLAMINATE - HAMPTON EM9 PROFILE "
            "THERMOLAMINATE NOTES : * KICKBOARDS : N/A FLOATING "
            "HANDLES - HETTICH SALEMI 9113368 30MM BRUSHED STAINLESS STEEL LOOK - DOOR LOCATION - 25MM IN AND 50MM DOWNDRAWER LOCATION - CENTRE TO PROFILE "
            "DOOR HINGES - HETTICH SOFT CLOSE "
            "DRAWER RUNNERS - HETTICH INNOTECH ATIRA SOFT CLOSE RUNNERS"
        )
        source_documents = [
            {
                "file_name": "49906622 - COLOURS - AFC.pdf",
                "role": "spec",
                "pages": [{"page_no": 1, "text": afc_raw, "raw_text": afc_raw, "needs_ocr": False}],
            },
            {
                "file_name": "49906622 AMENDED Signed Drawings and Colours REV B 20-10-25.pdf",
                "role": "spec",
                "pages": [
                    {"page_no": 2, "text": kitchen_raw, "raw_text": kitchen_raw, "needs_ocr": False},
                    {"page_no": 12, "text": vanities_raw, "raw_text": vanities_raw, "needs_ocr": False},
                ],
            },
        ]
        snapshot = parse_documents(
            job_no="37796",
            builder_name="Clarendon",
            source_kind="spec",
            documents=source_documents,
        )
        vision_docs = [
            {
                "file_name": "49906622 - COLOURS - AFC.pdf",
                "role": "spec",
                "pages": [{"page_no": 1, "text": afc_raw, "raw_text": afc_raw, "needs_ocr": False}],
            },
            {
                "file_name": "49906622 AMENDED Signed Drawings and Colours REV B 20-10-25.pdf",
                "role": "spec",
                "pages": [
                    {
                        "page_no": 2,
                        "text": (
                            "KITCHEN COLOUR SCHEDULE "
                            "BENCHTOP - QUANTUM ZERO WHITE SWIRL - 20MM PENCIL ROUND EDGE - TO COOKTOP + SIDE BENCHTOP/ "
                            "40MM MITRED APRON EDGE - TO ISLAND BENCH "
                            "DOOR COLOUR - POLYTEC ASTON WHITE SMOOTH FINISH THERMOLAMINATE - HAMPTON EM9 PROFILEPLAIN GLASS DISPLAY CABINET WITH STANDARD WHITE INTERNALS AND MELAMINE SHELF"
                        ),
                        "raw_text": kitchen_raw,
                        "needs_ocr": False,
                    },
                    {"page_no": 12, "text": vanities_raw, "raw_text": vanities_raw, "needs_ocr": False},
                ],
            },
        ]
        polished = extraction_service._apply_clarendon_reference_polish(
            snapshot,
            vision_docs,
            builder_name="Clarendon",
            parser_strategy=cleaning_rules.global_parser_strategy(),
            rule_flags=cleaning_rules.global_rule_flags(),
        )
        rooms = {row["room_key"]: row for row in polished["rooms"]}
        kitchen = rooms["kitchen"]
        self.assertEqual(polished["site_address"], "Lot 8 (#25) Lake Serenity Boulevard Helensvale")
        self.assertEqual(kitchen["toe_kick"], ["Matching Melamine finish"])
        self.assertEqual(kitchen["bulkheads"], ["Bulkhead shadowline as matching Melamine finish"])
        self.assertEqual(kitchen["bench_tops_wall_run"], "Quantum Zero White Swirl - 20MM Pencil Round Edge")
        self.assertEqual(kitchen["bench_tops_island"], "Quantum Zero White Swirl - 40MM Mitred Apron Edge")
        self.assertNotIn("Display Cabinet", kitchen["door_colours_base"])
        self.assertTrue(any("Belluno 9995772" in handle for handle in kitchen["handles"]))
        self.assertTrue(any("Salemi 9113368" in handle for handle in kitchen["handles"]))

        material_summary = _build_material_summary(polished)
        door_entries = [entry["display_text"] for entry in material_summary["door_colours"]["entries"]]
        handle_entries = [entry["display_text"] for entry in material_summary["handles"]["entries"]]
        self.assertTrue(
            any(
                entry.startswith("Polytec Aston White Smooth Finish Thermolaminate - Hampton EM9 Profile (")
                and "KITCHEN" in entry
                for entry in door_entries
            )
        )
        self.assertTrue(all("Display Cabinet" not in entry for entry in door_entries))
        self.assertTrue(
            any(
                entry.startswith("Hettich - Salemi 9113368 30MM Brushed Stainless Steel Look (")
                and "VANITIES" in entry
                for entry in handle_entries
            )
        )

    def test_multi_file_parse_keeps_room_materials_from_room_master_only(self) -> None:
        snapshot = parse_documents(
            job_no="37052",
            builder_name="Clarendon",
            source_kind="spec",
            documents=[
                {
                    "file_name": "supplement.pdf",
                    "role": "spec",
                    "pages": [
                        {
                            "page_no": 1,
                            "text": (
                                "Kitchen\n"
                                "Benchtop Wrong Stone 40mm\n"
                                "Sink Type/Model Franke Box Sink\n"
                            ),
                            "needs_ocr": False,
                        }
                    ],
                },
                {
                    "file_name": "master.pdf",
                    "role": "spec",
                    "pages": [
                        {
                            "page_no": 1,
                            "text": (
                                "KITCHEN COLOUR SCHEDULE\n"
                                "Benchtop Master Stone 20mm Pencil Round Edge\n"
                            ),
                            "needs_ocr": False,
                        }
                    ],
                },
            ],
        )
        room = snapshot["rooms"][0]
        self.assertIn("Master Stone 20mm Pencil Round Edge", " ".join(room["bench_tops"]))
        self.assertNotIn("Wrong Stone", " ".join(room["bench_tops"]))
        self.assertIn("Franke Box", room["sink_info"])

    def test_labeled_appliance_details_do_not_jump_to_next_row_context(self) -> None:
        snapshot = parse_documents(
            job_no="37053",
            builder_name="Clarendon",
            source_kind="spec",
            documents=[
                {
                    "file_name": "appliance-table.txt",
                    "role": "spec",
                    "pages": [
                        {
                            "page_no": 1,
                            "text": (
                                "Rangehood:\n"
                                "Westinghouse\n"
                                "HP280L\n"
                            ),
                            "needs_ocr": False,
                        }
                    ],
                }
            ],
        )
        self.assertEqual(snapshot["appliances"][0]["appliance_type"], "Rangehood")
        self.assertEqual(snapshot["appliances"][0]["make"], "Westinghouse")
        self.assertEqual(snapshot["appliances"][0]["model_no"], "")
        self.assertEqual(parsing_module._limit_appliance_details_to_local_context("Westinghouse\nHP280L\n"), "Westinghouse")

    def test_match_loose_appliance_type_ignores_mid_line_spillover(self) -> None:
        self.assertEqual(parsing_module._match_loose_appliance_type("model: gg-6c dishwasher ( kitchen )"), "")
        self.assertEqual(parsing_module._match_loose_appliance_type("rangehood"), "Rangehood")

    def test_build_appliance_row_prefers_explicit_model_over_url_slug(self) -> None:
        row = parsing_module._build_appliance_row(
            appliance_type="Rangehood",
            details="https://www.thegoodguys.com.au/schweigen-60cm-undermount-rangehood-gg-6c\n600mm Undermount - Schweigen 60cm",
            evidence="RANGEHOOD https://www.thegoodguys.com.au/schweigen-60cm-undermount-rangehood-gg-6c 600mm Undermount - Schweigen 60cm Undermount Rangehood Model: GG-6C",
            file_name="imperial-appliances.pdf",
            pages=[{"page_no": 1, "text": "", "raw_text": ""}],
            confidence=0.5,
        )
        self.assertIsNotNone(row)
        assert row is not None
        self.assertEqual(row.make, "Schweigen")
        self.assertEqual(row.model_no, "GG-6C")

    def test_loose_appliance_rows_do_not_bind_fridge_model_to_rangehood(self) -> None:
        snapshot = parse_documents(
            job_no="38251",
            builder_name="Imperial",
            source_kind="spec",
            documents=[
                {
                    "file_name": "imperial-appliances.pdf",
                    "role": "spec",
                    "pages": [
                        {
                            "page_no": 5,
                            "raw_text": (
                                "Address: 2510-076 - Private - 21 Shadowood Street, Kenmore Hills\n"
                                "Client: Eloise Cawcutt Foxover\n"
                                "Date: 20.3.26\n"
                                "APPLIANCES\n"
                                "AREA / ITEM SPECS / DESCRIPTION IMAGE SUPPLIER NOTES\n"
                                "Bertazzoni 90cm\n"
                                "Professional Series\n"
                                "OVEN BY CLIENT\n"
                                "Freestanding Dual Fuel Oven/Stove\n"
                                "PRO906MFESXE\n"
                                "COOKTOP as above BY CLIENT\n"
                                "ASKO\n"
                                "60cm classic built under dishwasher\n"
                                "DISHWASHER BY CLIENT\n"
                                "stainless steel\n"
                                "DBI364ID.S.AU\n"
                                "Document Ref: VVQAD-QWDJF-RPUJB-WQZDZ Page 5 of 8\n"
                            ),
                            "text": (
                                "Address: 2510-076 - Private - 21 Shadowood Street, Kenmore Hills APPLIANCES "
                                "Bertazzoni 90cm Professional Series OVEN BY CLIENT Freestanding Dual Fuel Oven/Stove "
                                "PRO906MFESXE COOKTOP as above BY CLIENT ASKO 60cm classic built under dishwasher "
                                "DISHWASHER BY CLIENT stainless steel DBI364ID.S.AU"
                            ),
                            "needs_ocr": False,
                        },
                        {
                            "page_no": 6,
                            "raw_text": (
                                "Whispar 90cm Monte Carlo Undermount\n"
                                "RANGEHOOD Rangehood with Power on board Motor BY CLIENT\n"
                                "X3md09s5.op/t\n"
                                "MICROWAVE LEAVE STANDARD SPACE BY CLIENT\n"
                                "Fisher and Paykel\n"
                                "Series 7 569L French Door Fridge\n"
                                "FRIDGE BY CLIENT PLUMBED IN FRIDGE\n"
                                "with Ice and Water plumbed in\n"
                                "RF610ANUX5\n"
                                "DESIGNER: MELISSA COAKES CLIENT NAME: SIGNATURE: SIGNED DATE:\n"
                                "Document Ref: VVQAD-QWDJF-RPUJB-WQZDZ Page 6 of 8\n"
                            ),
                            "text": (
                                "Whispar 90cm Monte Carlo Undermount\n"
                                "RANGEHOOD Rangehood with Power on board Motor BY CLIENT\n"
                                "X3md09s5.op/t\n"
                                "MICROWAVE LEAVE STANDARD SPACE BY CLIENT\n"
                                "Fisher and Paykel\n"
                                "Series 7 569L French Door Fridge\n"
                                "FRIDGE BY CLIENT PLUMBED IN FRIDGE\n"
                                "with Ice and Water plumbed in\n"
                                "RF610ANUX5\n"
                                "DESIGNER: MELISSA COAKES CLIENT NAME: SIGNATURE: SIGNED DATE:\n"
                                "Document Ref: VVQAD-QWDJF-RPUJB-WQZDZ Page 6 of 8\n"
                            ),
                            "needs_ocr": False,
                        }
                    ],
                }
            ],
        )
        appliances = snapshot["appliances"]
        oven_rows = [row for row in appliances if row["appliance_type"] == "Oven"]
        self.assertTrue(oven_rows)
        dishwasher_rows = [row for row in appliances if row["appliance_type"] == "Dishwasher"]
        self.assertTrue(dishwasher_rows)
        self.assertEqual(dishwasher_rows[0]["model_no"], "DBI364ID.S.AU")
        cooktop_rows = [row for row in appliances if row["appliance_type"] == "Cooktop"]
        self.assertTrue(cooktop_rows)
        self.assertEqual(cooktop_rows[0]["make"], "")
        self.assertEqual(cooktop_rows[0]["model_no"], "as above BY CLIENT")
        fridge_rows = [row for row in appliances if row["appliance_type"] == "Fridge"]
        self.assertTrue(fridge_rows)
        self.assertEqual(fridge_rows[0]["model_no"], "RF610ANUX5")
        microwave_rows = [row for row in appliances if row["appliance_type"] == "Microwave"]
        self.assertEqual(len(microwave_rows), 1)
        self.assertEqual(microwave_rows[0]["model_no"], "LEAVE STANDARD SPACE BY CLIENT")
        self.assertFalse(
            any(
                row["appliance_type"] == "Rangehood"
                and row.get("make") == "Fisher & Paykel"
                and row.get("model_no") == "RF610ANUX5"
                for row in appliances
            )
        )

    def test_appliance_rows_handle_pypdf_reordered_appliance_pages(self) -> None:
        snapshot = parse_documents(
            job_no="38251",
            builder_name="Imperial",
            source_kind="spec",
            documents=[
                {
                    "file_name": "imperial-appliances-pypdf.pdf",
                    "role": "spec",
                    "pages": [
                        {
                            "page_no": 5,
                            "raw_text": (
                                "DISHWASHER\n"
                                "ASKO \n"
                                "60cm classic built under dishwasher \n"
                                "stainless steel \n"
                                "DBI364ID.S.AU\n"
                                "BY CLIENT\n"
                                "OVEN BY CLIENT\n"
                                "COOKTOP BY CLIENT\n"
                                "Bertazzoni 90cm \n"
                                "Professional Series \n"
                                "Freestanding Dual Fuel Oven/Stove \n"
                                "PRO906MFESXE\n"
                                "as above\n"
                                "AREA / ITEM SPECS / DESCRIPTION IMAGE SUPPLIER NOTES\n"
                                "APPLIANCES\n"
                                "Address:2510-076 - Private - 21 Shadowood Street, Kenmore Hills\n"
                                "Client:Eloise Cawcutt Foxover\n"
                                "Date:20.3.26\n"
                                "Document Ref: VVQAD-QWDJF-RPUJB-WQZDZ Page 5 of 8\n"
                            ),
                            "text": (
                                "DISHWASHER ASKO 60cm classic built under dishwasher stainless steel "
                                "DBI364ID.S.AU BY CLIENT OVEN BY CLIENT COOKTOP BY CLIENT Bertazzoni "
                                "90cm Professional Series Freestanding Dual Fuel Oven/Stove PRO906MFESXE as above"
                            ),
                            "needs_ocr": False,
                        },
                        {
                            "page_no": 6,
                            "raw_text": (
                                "MICROWAVE LEAVE STANDARD SPACE BY CLIENT\n"
                                "RANGEHOOD\n"
                                "Whispar 90cm Monte Carlo Undermount \n"
                                "Rangehood with Power on board Motor \n"
                                "X3md09s5.op/t \n"
                                "BY CLIENT\n"
                                "FRIDGE\n"
                                "Fisher and Paykel \n"
                                "Series 7 569L French Door Fridge \n"
                                "with Ice and Water plumbed in \n"
                                "RF610ANUX5 \n"
                                "BY CLIENT PLUMBED IN FRIDGE\n"
                                "SIGNATURE: SIGNED DATE:DESIGNER: MELISSA COAKES CLIENT NAME:\n"
                                "Document Ref: VVQAD-QWDJF-RPUJB-WQZDZ Page 6 of 8\n"
                            ),
                            "text": (
                                "MICROWAVE LEAVE STANDARD SPACE BY CLIENT RANGEHOOD Whispar 90cm Monte Carlo "
                                "Undermount Rangehood with Power on board Motor X3md09s5.op/t BY CLIENT "
                                "FRIDGE Fisher and Paykel Series 7 569L French Door Fridge with Ice and Water "
                                "plumbed in RF610ANUX5 BY CLIENT PLUMBED IN FRIDGE"
                            ),
                            "needs_ocr": False,
                        },
                    ],
                }
            ],
        )
        appliances = snapshot["appliances"]
        oven_rows = [row for row in appliances if row["appliance_type"] == "Oven"]
        self.assertEqual(len(oven_rows), 1)
        self.assertEqual(oven_rows[0]["model_no"], "BY CLIENT")
        dishwasher_rows = [row for row in appliances if row["appliance_type"] == "Dishwasher"]
        self.assertEqual(len(dishwasher_rows), 1)
        self.assertEqual(dishwasher_rows[0]["make"], "ASKO")
        cooktop_rows = [row for row in appliances if row["appliance_type"] == "Cooktop"]
        self.assertEqual(len(cooktop_rows), 1)
        self.assertIn(cooktop_rows[0]["model_no"], {"", "BY CLIENT"})
        fridge_rows = [row for row in appliances if row["appliance_type"] == "Fridge"]
        self.assertEqual(len(fridge_rows), 1)
        self.assertIn(fridge_rows[0]["model_no"], {"", "RF610ANUX5", "BY CLIENT PLUMBED IN FRIDGE"})
        microwave_rows = [row for row in appliances if row["appliance_type"] == "Microwave"]
        self.assertEqual(len(microwave_rows), 1)
        self.assertEqual(microwave_rows[0]["model_no"], "LEAVE STANDARD SPACE BY CLIENT")
        rangehood_rows = [row for row in appliances if row["appliance_type"] == "Rangehood"]
        self.assertTrue(rangehood_rows)
        self.assertTrue(
            any(
                row.get("make") == "Whispar" or row.get("model_no") == "X3MD09S5.OP/T"
                for row in rangehood_rows
            )
        )

    def test_clarendon_appliance_parser_ignores_drawing_noise_and_preserves_na_by_others(self) -> None:
        snapshot = parse_documents(
            job_no="38211",
            builder_name="Imperial",
            source_kind="spec",
            documents=[
                {
                    "file_name": "Colour Selections.pdf",
                    "role": "spec",
                    "pages": [
                        {
                            "page_no": 2,
                            "raw_text": (
                                "RANGEHOOD (KITCHEN)\n"
                                "https://www.thegoodguys.com.au/schweigen-60cm-undermount-rangehood-gg-6c\n"
                                "By others\n"
                                "Recirculating\n"
                                "600mm Undermount - Schweigen 60cm \n"
                                "Undermount Rangehood\n"
                                "Model: GG-6C\n"
                                "DISHWASHER ( KITCHEN )\n"
                                "N / A - By others\n"
                                "By others\n"
                                "600mm Standard - Specs TBC\n"
                                "COOKTOP (KITCHEN)\n"
                                "https://www.thegoodguys.com.au/westinghouse-60cm-induction-cooktop-whi643be\n"
                                "By others\n"
                                "600mm Electric - Westinghouse 60cm \n"
                                "Induction Cooktop\n"
                                "Model: WHI643BE\n"
                                "OVEN (KITCHEN)\n"
                                "https://www.thegoodguys.com.au/westinghouse-60cm-electric-oven-wve6516dd\n"
                                "By others\n"
                                "600mm - Westinghouse 60cm Electric \n"
                                "Oven\n"
                                "Model: WVE6516DD\n"
                            ),
                            "text": "",
                            "needs_ocr": False,
                        }
                    ],
                }
            ],
        )
        appliances = {row["appliance_type"]: row for row in snapshot["appliances"]}
        self.assertNotIn("GG-6C", appliances["Dishwasher"]["model_no"])
        self.assertEqual(appliances["Cooktop"]["model_no"], "WHI643BE")
        self.assertEqual(appliances["Oven"]["model_no"], "WVE6516DD")
        self.assertEqual(appliances["Dishwasher"]["model_no"], "N / A - By others")

    def test_clarendon_reference_polish_rebuilds_clean_room_fields(self) -> None:
        snapshot = {
            "job_no": "37031",
            "builder_name": "Clarendon",
            "source_kind": "spec",
            "generated_at": "2026-03-22T00:00:00+00:00",
            "analysis": {"mode": "openai_merged", "parser_strategy": "global_conservative"},
            "rooms": [
                {
                    "room_key": "kitchen",
                    "original_room_label": "Kitchen",
                    "bench_tops": ["SINKCUT OUTCENTRE", "QUANTUM ZERO MIDNIGHT - 20MM PENCIL ROUND EDGE", "60MM UP/ DOWN TO DOORS DOOR"],
                    "bench_tops_wall_run": "",
                    "bench_tops_island": "",
                    "bench_tops_other": "SINKCUT OUTCENTRE | QUANTUM ZERO MIDNIGHT - 20MM PENCIL ROUND EDGE | 60MM UP/ DOWN TO DOORS DOOR",
                    "door_panel_colours": [],
                    "door_colours_overheads": "",
                    "door_colours_base": "",
                    "door_colours_island": "",
                    "door_colours_bar_back": "",
                    "toe_kick": ["MATCHING MELAMINE FINISH / MATCHING THERMO FINISH"],
                    "bulkheads": ["BY BUILDERBULKHEAD SHADOWLINE AS MATCHING MELAMINE"],
                    "handles": [],
                    "drawers_soft_close": "Soft Close",
                    "hinges_soft_close": "Not Soft Close",
                    "splashback": "BY OTHERS20MM STONE",
                    "flooring": "",
                    "sink_info": "& TAP TOCABINET UNDER CUT OUT DETAIL FOR PARISI QUADRO PK8644 DOUBLE BOWL UNDERMOUNT",
                    "basin_info": "",
                    "tap_info": "",
                    "source_file": "schedule.pdf",
                    "page_refs": "1",
                    "evidence_snippet": "",
                    "confidence": 0.6,
                },
                {
                    "room_key": "butlers_pantry",
                    "original_room_label": "Butler's Pantry",
                    "bench_tops": ["bad"],
                    "bench_tops_wall_run": "",
                    "bench_tops_island": "",
                    "bench_tops_other": "bad",
                    "door_panel_colours": [],
                    "door_colours_overheads": "",
                    "door_colours_base": "",
                    "door_colours_island": "",
                    "door_colours_bar_back": "",
                    "toe_kick": ["bad"],
                    "bulkheads": ["bad"],
                    "handles": [],
                    "drawers_soft_close": "Not Soft Close",
                    "hinges_soft_close": "",
                    "splashback": "bad",
                    "flooring": "",
                    "sink_info": "bad",
                    "basin_info": "bad",
                    "tap_info": "bad",
                    "source_file": "schedule.pdf",
                    "page_refs": "2",
                    "evidence_snippet": "",
                    "confidence": 0.6,
                },
                {
                    "room_key": "vanities",
                    "original_room_label": "Vanities",
                    "bench_tops": ["bad"],
                    "bench_tops_wall_run": "",
                    "bench_tops_island": "",
                    "bench_tops_other": "bad",
                    "door_panel_colours": [],
                    "door_colours_overheads": "",
                    "door_colours_base": "",
                    "door_colours_island": "",
                    "door_colours_bar_back": "",
                    "toe_kick": ["bad"],
                    "bulkheads": ["bad"],
                    "handles": [],
                    "drawers_soft_close": "",
                    "hinges_soft_close": "",
                    "splashback": "",
                    "flooring": "",
                    "sink_info": "",
                    "basin_info": "",
                    "tap_info": "",
                    "source_file": "schedule.pdf",
                    "page_refs": "3",
                    "evidence_snippet": "",
                    "confidence": 0.6,
                },
                {
                    "room_key": "laundry",
                    "original_room_label": "Laundry",
                    "bench_tops": ["bad"],
                    "bench_tops_wall_run": "",
                    "bench_tops_island": "",
                    "bench_tops_other": "bad",
                    "door_panel_colours": [],
                    "door_colours_overheads": "",
                    "door_colours_base": "",
                    "door_colours_island": "",
                    "door_colours_bar_back": "",
                    "toe_kick": ["bad"],
                    "bulkheads": ["bad"],
                    "handles": [],
                    "drawers_soft_close": "",
                    "hinges_soft_close": "",
                    "splashback": "BY OTHERS20MM STONE BENCHTOP",
                    "flooring": "",
                    "sink_info": "bad",
                    "basin_info": "",
                    "tap_info": "bad",
                    "source_file": "schedule.pdf",
                    "page_refs": "4",
                    "evidence_snippet": "",
                    "confidence": 0.6,
                },
            ],
            "appliances": [],
            "others": {},
            "warnings": [],
        }
        documents = [
            {
                "file_name": "drawings.pdf",
                "role": "spec",
                "pages": [
                    {
                        "page_no": 1,
                        "text": (
                            "BENCHTOP COLOUR 1 - QUANTUM ZERO MIDNIGHT BLACK - 20MM PENCIL ROUND EDGE - TO COOKTOP RUN"
                            "BENCHTOP COLOUR 2 - QUANTUM ZERO VENATINO STATUARIO - 40MM MITRED APRON EDGE - TO ISLAND BENCH"
                            "DOOR COLOUR 1 - POL YTEC CLASSIC WHITE MATT FINISH THERMOLAMINATE - ATLANTA EM2 PROFILE - TO COOKTOP RUN BASE, UPPER + TALL CABINETRY "
                            "DOOR COLOUR 2 - POL YTEC TEMPEST WOODGRAIN FINISH THERMOLAMINATE - ATLANTA EM2 PROFILE (VERTICAL GRAIN DIRECTION) - ISLAND BASE CABINETRY + BAR BACK PANELS "
                            "HANDLE 1 - HETTICH CIPRI 9070585 GLOSS CHROME PLATED 30MM KNOB - 30MM IN AND 60MM UP/ DOWN TO DOORS "
                            "HANDLE 2 - MOMO FLORENCIA CTCP .CP .FG CHROME PLATED 104MM LONG - CENTRE TO DRAWER PROFILE "
                            "DOOR HINGES - HETTICH STANDARD HINGES - NOT SOFT CLOSE "
                            "DRAWER RUNNERS - HETTICH INNOTECH ATIRA SOFT CLOSE RUNNERS "
                            "KITCHEN COLOUR SCHEDULE THERMOLAMINATE NOTES : * BULKHEAD SHADOWLINE : MATCHING MELAMINE FINISH* KICKBOARDS : MATCHING MELAMINE FINISH / MATCHING THERMO FINISH"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 2,
                        "text": (
                            "BENCHTOP - QUANTUM ZERO MIDNIGHT - 20MM PENCIL ROUND EDGE "
                            "DOOR COLOUR - POL YTEC CLASSIC WHITE MATT FINISH THERMOLAMINATE - ATLANTA EM2 PROFILE "
                            "THERMOLAMINATE NOTES : * BULKHEAD SHADOWLINE : MATCHING MELAMINE FINISH* KICKBOARDS : MATCHING MELAMINE FINISH "
                            "HANDLES - HETTICH CIPRI 9070585 GLOSS CHROME PLATED 30MM KNOB - 30MM IN AND 60MM UP/ DOWN TO DOORS "
                            "DOOR HINGES - HETTICH STANDARD HINGES - NOT SOFT CLOSE "
                            "BUTLERS PANTRY COLOUR SCHEDULE"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 3,
                        "text": (
                            "BENCHTOP - QUANTUM ZERO MIDNIGHT - 20MM PENCIL ROUND EDGE "
                            "DOOR COLOUR - POL YTEC CLASSIC WHITE MATT FINISH THERMOLAMINATE - ATLANTA EM2 PROFILE "
                            "THERMOLAMINATE NOTES : * KICKBOARDS : N/A FLOATING "
                            "HANDLES - HETTICH CIPRI 9070585 GLOSS CHROME PLATED 30MM KNOB - DOOR LOCATION : 30MM IN AND 60MM UP/ DOWN TO DOORS DRAWER LOCATION : CTR TO PROFILE "
                            "DOOR HINGES - HETTICH STANDARD HINGES - NOT SOFT CLOSE "
                            "DRAWER RUNNERS - HETTICH MUL TITECH STANDARD CONSTRUCTION RUNNERS - NOT SOFT CLOSE "
                            "VANITIES COLOUR SCHEDULE"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 4,
                        "text": (
                            "BENCHTOP - QUANTUM ZERO MIDNIGHT - 20MM PENCIL ROUND EDGE "
                            "DOOR COLOUR - POL YTEC CLASSIC WHITE MATT FINISH THERMOLAMINATE - ATLANTA EM2 PROFILE "
                            "THERMOLAMINATE NOTES : * BULKHEAD SHADOWLINE : MATCHING MELAMINE FINISH* KICKBOARDS : MATCHING MELAMINE FINISH "
                            "HANDLES - HETTICH CIPRI 9070585 GLOSS CHROME PLATED 30MM KNOB - 30MM IN AND 60MM UP/ DOWN TO DOORS "
                            "DOOR HINGES - HETTICH STANDARD HINGES - NOT SOFT CLOSE "
                            "LAUNDRY COLOUR SCHEDULE"
                        ),
                        "needs_ocr": False,
                    },
                ],
            },
            {
                "file_name": "fixtures.pdf",
                "role": "spec",
                "pages": [
                    {
                        "page_no": 5,
                        "text": (
                            "KITCHEN SUPPLIER DESCRIPTION DESIGN COMMENTS "
                            "Sink Type: PARISI QUADRO_DOUBLE BOWL_STAINLESS STEEL (PK8644) UNDERMOUNT "
                            "Tap Type: PHOENIX Nostalgia Twin Handle Sink Mixer 230mm Shepherds Crook NS714-62 CHROME & WHITE "
                            "Splashback: "
                            "Sink Type/Model: PARISI UNDERMOUNT - QUADRO SINGLE BOWL STAINLESS STEEL (PK4444) UNDERMOUNT "
                            "Tap Type: PHOENIX Nostalgia Sink Mixer 220mm Shepherds Crook NS738-62 CHROME & WHITE "
                            "BUTLERS PANTRY"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 6,
                        "text": (
                            "Vanity Inset Basin JOHNSON SUISSE Emilia Rectangular Undercounter Basin (JBSE250.PW6) WHITE "
                            "Vanity Tap Style: PHOENIX NOSTALGIA BASIN MIXER 160MM SHEPHERDS CROOK (NS748-62) CHROME & WHITE "
                            "Vanity Waste Colour: CHROME"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 7,
                        "text": (
                            "LAUNDRY SUPPLIER DESCRIPTION DESIGN COMMENTS "
                            "Drop in Tub: EVERHARD "
                            "Splashback: "
                            "EVERHARD INDUSTRIES CLASSIC 45L UTILITY SINK (71245) "
                            "PINA SINK MIXER GOOSENECK 200MM_CHROME (153-7330-00) "
                            "15MM CP QUARTER TURN WASHING MACHINE COCK (60822)"
                        ),
                        "needs_ocr": False,
                    },
                ],
            },
        ]
        polished = extraction_service._apply_clarendon_reference_polish(
            snapshot,
            documents,
            builder_name="Clarendon",
            parser_strategy="global_conservative",
        )
        rooms = {row["room_key"]: row for row in polished["rooms"]}
        self.assertNotIn("SINKCUT", " ".join(rooms["kitchen"]["bench_tops"]))
        self.assertEqual(rooms["kitchen"]["bench_tops_wall_run"], "Quantum Zero Midnight Black - 20MM Pencil Round Edge")
        self.assertEqual(rooms["kitchen"]["bench_tops_island"], "Quantum Zero Venatino Statuario - 40MM Mitred Apron Edge")
        self.assertEqual(rooms["kitchen"]["splashback"], "Tiled splashback by others")
        self.assertEqual(
            rooms["kitchen"]["handles"],
            [
                "Hettich - Cipri 9070585 Gloss Chrome Plated 30MM Knob",
                "Momo Florencia Ctcp .Cp .Fg Chrome Plated 104MM Long - Centre to Drawer Profile",
            ],
        )
        self.assertEqual(rooms["kitchen"]["drawers_soft_close"], "Soft Close")
        self.assertEqual(rooms["kitchen"]["hinges_soft_close"], "Not Soft Close")
        self.assertNotIn("\n", str(rooms["kitchen"]["sink_info"]))
        self.assertEqual(rooms["butlers_pantry"]["door_colours_base"], "Polytec Classic White Matt Finish Thermolaminate - Atlanta EM2 Profile")
        self.assertEqual(rooms["butlers_pantry"]["drawers_soft_close"], "Not Soft Close")
        self.assertEqual(rooms["vanities"]["toe_kick"], ["N/A floating - no kickboard"])
        self.assertTrue(str(rooms["vanities"]["basin_info"]).startswith("Johnson Suisse Emilia"))
        self.assertEqual(rooms["laundry"]["splashback"], "Tiled splashback by others")
        self.assertTrue(str(rooms["laundry"]["tap_info"]).startswith("Pina Sink Mixer Gooseneck"))

    def test_clarendon_reference_polish_handles_luxe_single_line_schedule_family(self) -> None:
        snapshot = {
            "job_no": "37868",
            "builder_name": "Clarendon",
            "source_kind": "spec",
            "generated_at": "2026-03-22T00:00:00+00:00",
            "analysis": {"mode": "openai_merged", "parser_strategy": "global_conservative"},
            "rooms": [
                {"room_key": "kitchen", "original_room_label": "Kitchen", "bench_tops": ["noisy"], "bench_tops_wall_run": "", "bench_tops_island": "", "bench_tops_other": "noisy", "door_panel_colours": [], "door_colours_overheads": "", "door_colours_base": "", "door_colours_island": "", "door_colours_bar_back": "", "toe_kick": [], "bulkheads": [], "handles": [], "drawers_soft_close": "", "hinges_soft_close": "", "splashback": "", "flooring": "", "sink_info": "", "basin_info": "", "tap_info": "", "source_file": "schedule.pdf", "page_refs": "1", "evidence_snippet": "", "confidence": 0.6},
                {"room_key": "butlers_pantry", "original_room_label": "Butler's Pantry", "bench_tops": ["noisy"], "bench_tops_wall_run": "", "bench_tops_island": "", "bench_tops_other": "noisy", "door_panel_colours": [], "door_colours_overheads": "", "door_colours_base": "", "door_colours_island": "", "door_colours_bar_back": "", "toe_kick": [], "bulkheads": [], "handles": [], "drawers_soft_close": "", "hinges_soft_close": "", "splashback": "", "flooring": "", "sink_info": "", "basin_info": "", "tap_info": "", "source_file": "schedule.pdf", "page_refs": "2", "evidence_snippet": "", "confidence": 0.6},
                {"room_key": "vanities", "original_room_label": "Vanities", "bench_tops": ["noisy"], "bench_tops_wall_run": "", "bench_tops_island": "", "bench_tops_other": "noisy", "door_panel_colours": [], "door_colours_overheads": "", "door_colours_base": "", "door_colours_island": "", "door_colours_bar_back": "", "toe_kick": [], "bulkheads": [], "handles": [], "drawers_soft_close": "", "hinges_soft_close": "", "splashback": "", "flooring": "", "sink_info": "", "basin_info": "", "tap_info": "", "source_file": "schedule.pdf", "page_refs": "3", "evidence_snippet": "", "confidence": 0.6},
                {"room_key": "laundry", "original_room_label": "Laundry", "bench_tops": ["noisy"], "bench_tops_wall_run": "", "bench_tops_island": "", "bench_tops_other": "noisy", "door_panel_colours": [], "door_colours_overheads": "", "door_colours_base": "", "door_colours_island": "", "door_colours_bar_back": "", "toe_kick": [], "bulkheads": [], "handles": [], "drawers_soft_close": "", "hinges_soft_close": "", "splashback": "", "flooring": "", "sink_info": "", "basin_info": "", "tap_info": "", "source_file": "schedule.pdf", "page_refs": "4", "evidence_snippet": "", "confidence": 0.6},
            ],
            "appliances": [],
            "others": {},
            "warnings": [],
        }
        documents = [
            {
                "file_name": "luxe-drawings.pdf",
                "role": "spec",
                "pages": [
                    {
                        "page_no": 1,
                        "text": (
                            "KITCHEN COLOUR SCHEDULE "
                            "BENCHTOP - QUANTUM ZERO BELLA CARRARA - 20MM PENCIL ROUND EDGE - TO COOKTOP RUN / 40MM MITRED APRON EDGE - TO ISLAND BENCHTOP "
                            "WATERFALL ENDS (MITRED JOIN) "
                            "MIRROR SPLASHBACK - BRONZE MIRRORKOTE "
                            "DOOR COLOUR - POL YTEC SOFT WALNUT MATT FINISH MELAMINE WITH MATCHING 1MM ABS EDGES (VERTICAL GRAIN DIRECTION) "
                            "KICKBOARDS - AS POL YTEC CLASSIC WHITE 'MATT' FINISH MELAMINE "
                            "SQUARE EDGE RAILS - AS POL YTEC CLASSIC WHITE 'MATT' FINISH MELAMINE "
                            "BULKHEAD SHADOWLINE - AS POL YTEC CLASSIC WHITE 'MATT' FINISH MELAMINE "
                            "HANDLES - SQUARE EDGE HANDLELESS* NOTE : 10MM DOOR OVERHANG TO UPPER CABINETS "
                            "DOOR HINGES - HETTICH SOFT CLOSE "
                            "DRAWER RUNNERS - HETTICH INNOTECH ATIRA SOFT CLOSE RUNNERS"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 2,
                        "text": (
                            "BUTLERS PANTRY COLOUR SCHEDULE "
                            "BENCHTOP - POL YTEC ARGENTO STONE MATT FINISH - 21MM TIGHTFORM EDGE LAMINATE "
                            "DOOR COLOUR - POL YTEC SOFT WALNUT MATT FINISH MELAMINE WITH MATCHING 1MM ABS EDGES (VERTICAL GRAIN DIRECTION) "
                            "KICKBOARDS - AS POL YTEC CLASSIC WHITE 'MATT' FINISH MELAMINE "
                            "BULKHEAD SHADOWLINE - AS POL YTEC CLASSIC WHITE 'MATT' FINISH MELAMINE "
                            "HANDLES - SQUARE EDGE HANDLELESS* NOTE : 10MM DOOR OVERHANG TO UPPER CABINETS "
                            "DOOR HINGES - HETTICH SOFT CLOSE"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 3,
                        "text": (
                            "VANITIES COLOUR SCHEDULE "
                            "BENCHTOP - QUANTUM ZERO LUNA WHITE - 20MM PENCIL ROUND EDGE / 140MM MITRED APRON EDGE (POWDER ROOM 3) "
                            "DOOR/PANEL COLOUR - POL YTEC SOFT WALNUT MATT FINISH MELAMINE WITH MATCHING 1MM ABS EDGES (VERTICAL GRAIN DIRECTION) "
                            "KICKBOARDS - N/A FLOATING "
                            "SQUARE EDGE RAILS - AS POL YTEC CLASSIC WHITE 'MATT' FINISH MELAMINE "
                            "HANDLES - SQUARE EDGE HANDLELESS "
                            "DOOR HINGES - HETTICH STANDARD HINGES - NOT SOFT CLOSE"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 4,
                        "text": (
                            "LAUNDRY COLOUR SCHEDULE "
                            "BENCHTOP - POL YTEC ARGENTO STONE MATT FINISH - 21MM TIGHTFORM EDGE LAMINATE "
                            "DOOR COLOUR - POL YTEC SOFT WALNUT MATT FINISH MELAMINE WITH MATCHING 1MM ABS EDGES (VERTICAL GRAIN DIRECTION) "
                            "KICKBOARDS - AS POL YTEC CLASSIC WHITE MATT FINISH MELAMINE "
                            "HANDLES - SQUARE EDGE HANDLELESS* NOTE : 10MM DOOR OVERHANG TO UPPER CABINETS "
                            "DOOR HINGES - HETTICH STANDARD HINGES - NOT SOFT CLOSE "
                            "SPLASHBACK - TILED SPLASHBACK BY OTHERS"
                        ),
                        "needs_ocr": False,
                    },
                ],
            }
        ]
        polished = extraction_service._apply_clarendon_reference_polish(
            snapshot,
            documents,
            builder_name="Clarendon",
            parser_strategy="global_conservative",
        )
        rooms = {row["room_key"]: row for row in polished["rooms"]}
        self.assertEqual(rooms["kitchen"]["bench_tops_wall_run"], "Quantum Zero Bella Carrara - 20MM Pencil Round Edge")
        self.assertEqual(rooms["kitchen"]["bench_tops_island"], "Quantum Zero Bella Carrara - 40MM Mitred Apron Edge Waterfall Ends (Mitred Join)")
        self.assertEqual(rooms["kitchen"]["splashback"], "Mirror Splashback - Bronze Mirrorkote")
        self.assertEqual(rooms["kitchen"]["handles"], ["Square Edge Handleless"])
        self.assertEqual(rooms["kitchen"]["door_colours_base"], "Polytec Soft Walnut Matt Finish Melamine with Matching 1MM ABS Edges (Vertical Grain Direction)")
        self.assertEqual(rooms["kitchen"]["hinges_soft_close"], "Soft Close")
        self.assertEqual(rooms["kitchen"]["drawers_soft_close"], "Soft Close")
        self.assertEqual(rooms["butlers_pantry"]["bench_tops"], ["Polytec Argento Stone Matt Finish - 21MM Tightform Edge Laminate"])
        self.assertEqual(rooms["butlers_pantry"]["handles"], ["Square Edge Handleless"])
        self.assertEqual(rooms["vanities"]["bench_tops"], ["Quantum Zero Luna White - 20MM Pencil Round Edge / 140MM Mitred Apron Edge (Powder Room 3)"])
        self.assertEqual(rooms["vanities"]["door_colours_base"], "Polytec Soft Walnut Matt Finish Melamine with Matching 1MM ABS Edges (Vertical Grain Direction)")
        self.assertEqual(rooms["vanities"]["handles"], ["Square Edge Handleless"])
        self.assertEqual(rooms["laundry"]["bench_tops"], ["Polytec Argento Stone Matt Finish - 21MM Tightform Edge Laminate"])
        self.assertEqual(rooms["laundry"]["splashback"], "Tiled splashback by others")
        self.assertEqual(rooms["laundry"]["hinges_soft_close"], "Not Soft Close")

    def test_room_fixture_enrichment_splits_door_colours_and_filters_plumbing_appliances(self) -> None:
        documents = [
            {
                "file_name": "fixtures.txt",
                "role": "spec",
                "pages": [
                    {
                        "page_no": 1,
                        "text": (
                            "Kitchen\n"
                            "Door Colour 1 - Polytec Classic White Matt - to cooktop run base, upper + tall cabinetry\n"
                            "Door Colour 2 - Polytec Tempest Woodgrain - island base cabinetry + bar back panels\n"
                            "Sink Type: PARISI Quadro Double Bowl (PK8644)\n"
                            "Tap Type: PHOENIX Nostalgia Sink Mixer NS714-62\n"
                            "Cooktop: Westinghouse WHC943BD 90cm\n"
                            "Butler's Pantry\n"
                            "Sink Type/Model: PARISI Quadro Single Bowl (PK4444)\n"
                            "Tap Type: PHOENIX Nostalgia Sink Mixer NS738-62\n"
                            "Vanities\n"
                            "Vanity Inset Basin JOHNSON SUISSE Emilia Rectangular Undercounter Basin (JBSE250.PW6)\n"
                            "Vanity Tap Style: PHOENIX Nostalgia Basin Mixer NS748-62\n"
                        ),
                        "needs_ocr": False,
                    }
                ],
            }
        ]
        snapshot = parse_documents(job_no="37529", builder_name="Clarendon", source_kind="spec", documents=documents)
        enriched = enrich_snapshot_rooms(snapshot, documents)
        rooms = {row["room_key"]: row for row in enriched["rooms"]}
        self.assertEqual(rooms["kitchen"]["sink_info"], "Parisi Quadro Double Bowl (PK8644)")
        self.assertEqual(rooms["kitchen"]["tap_info"], "Phoenix Nostalgia Sink Mixer NS714-62")
        self.assertEqual(rooms["kitchen"]["door_colours_overheads"], "Polytec Classic White Matt")
        self.assertEqual(rooms["kitchen"]["door_colours_base"], "Polytec Classic White Matt")
        self.assertEqual(rooms["kitchen"]["door_colours_island"], "Polytec Tempest Woodgrain")
        self.assertEqual(rooms["kitchen"]["door_colours_bar_back"], "Polytec Tempest Woodgrain")
        self.assertEqual(rooms["butlers_pantry"]["sink_info"], "Parisi Quadro Single Bowl (PK4444)")
        self.assertEqual(
            rooms["vanities"]["basin_info"],
            "Johnson Suisse Emilia Rectangular Undercounter Basin (JBSE250.PW6)",
        )
        self.assertEqual(rooms["vanities"]["tap_info"], "Phoenix Nostalgia Basin Mixer NS748-62")
        appliance_types = [row["appliance_type"].lower() for row in enriched["appliances"]]
        self.assertIn("cooktop", appliance_types)
        self.assertNotIn("sink", appliance_types)

    def test_build_spec_snapshot_marks_openai_fallback_when_request_fails(self) -> None:
        with (
            mock.patch.object(extraction_service.runtime, "OPENAI_ENABLED", True),
            mock.patch.object(extraction_service.runtime, "SPEC_OPENAI_MERGE_ENABLED", True),
            mock.patch.object(extraction_service.runtime, "OPENAI_API_KEY", "test-key"),
            mock.patch.object(extraction_service.runtime, "OPENAI_MODEL", "gpt-4.1-mini"),
            mock.patch("App.services.extraction_service._post_responses_api", side_effect=RuntimeError("boom")),
        ):
            snapshot = extraction_service.build_spec_snapshot(
                job={"job_no": "37529"},
                builder={"name": "Clarendon"},
                files=[],
                template_files=[],
        )
        self.assertEqual(snapshot["analysis"]["mode"], "openai_fallback")
        self.assertEqual(snapshot["analysis"]["parser_strategy"], "global_conservative")
        self.assertTrue(snapshot["analysis"]["openai_attempted"])
        self.assertFalse(snapshot["analysis"]["openai_succeeded"])
        self.assertIn("normalize_brand_casing", snapshot["analysis"]["rule_flags"])

    def test_build_spec_snapshot_accepts_fenced_openai_json(self) -> None:
        openai_payload = {
            "output_text": """```json
{"rooms":[{"room_key":"kitchen","original_room_label":"Kitchen"}],"appliances":[],"others":{},"warnings":[]}
```"""
        }
        with (
            mock.patch.object(extraction_service.runtime, "OPENAI_ENABLED", True),
            mock.patch.object(extraction_service.runtime, "SPEC_OPENAI_MERGE_ENABLED", True),
            mock.patch.object(extraction_service.runtime, "OPENAI_API_KEY", "test-key"),
            mock.patch.object(extraction_service.runtime, "OPENAI_MODEL", "gpt-4.1-mini"),
            mock.patch("App.services.extraction_service._post_responses_api", return_value=openai_payload),
        ):
            snapshot = extraction_service.build_spec_snapshot(
                job={"job_no": "37529"},
                builder={"name": "Clarendon"},
                files=[],
                template_files=[],
            )
        self.assertEqual(snapshot["analysis"]["mode"], "openai_merged")
        self.assertEqual(snapshot["analysis"]["parser_strategy"], "global_conservative")
        self.assertTrue(snapshot["analysis"]["openai_succeeded"])
        self.assertEqual(snapshot["rooms"][0]["room_key"], "kitchen")

    def test_build_spec_snapshot_handles_invalid_openai_field_shapes(self) -> None:
        openai_payload = {
            "output_text": '{"rooms": [], "appliances": [], "others": "bad-shape", "warnings": "bad-shape"}'
        }
        with (
            mock.patch.object(extraction_service.runtime, "OPENAI_ENABLED", True),
            mock.patch.object(extraction_service.runtime, "SPEC_OPENAI_MERGE_ENABLED", True),
            mock.patch.object(extraction_service.runtime, "OPENAI_API_KEY", "test-key"),
            mock.patch.object(extraction_service.runtime, "OPENAI_MODEL", "gpt-4.1-mini"),
            mock.patch("App.services.extraction_service._post_responses_api", return_value=openai_payload),
        ):
            snapshot = extraction_service.build_spec_snapshot(
                job={"job_no": "37529"},
                builder={"name": "Clarendon", "parser_strategy": "stable_hybrid"},
                files=[],
                template_files=[],
            )
        self.assertEqual(snapshot["analysis"]["mode"], "openai_merged")
        self.assertEqual(snapshot["others"]["flooring_notes"], "")
        self.assertEqual(snapshot["others"]["splashback_notes"], "")
        self.assertEqual(snapshot["warnings"], [])

    def test_build_spec_snapshot_skips_openai_merge_by_default_for_spec_runs(self) -> None:
        with (
            mock.patch.object(extraction_service.runtime, "OPENAI_ENABLED", True),
            mock.patch.object(extraction_service.runtime, "OPENAI_API_KEY", "test-key"),
            mock.patch.object(extraction_service.runtime, "SPEC_OPENAI_MERGE_ENABLED", False),
            mock.patch("App.services.extraction_service._post_responses_api") as post_responses_api,
        ):
            snapshot = extraction_service.build_spec_snapshot(
                job={"job_no": "37529"},
                builder={"name": "Clarendon"},
                files=[],
                template_files=[],
            )
        self.assertEqual(snapshot["analysis"]["mode"], "heuristic_only")
        self.assertFalse(snapshot["analysis"]["openai_attempted"])
        self.assertIn("disabled by default for spec runs", snapshot["analysis"]["note"])
        post_responses_api.assert_not_called()

    def test_build_spec_snapshot_limits_docling_to_policy_builders_by_default(self) -> None:
        documents = [
            {
                "file_name": "clarendon.pdf",
                "path": "clarendon.pdf",
                "role": "spec",
                "pages": [
                    {
                        "page_no": 1,
                        "text": "KITCHEN COLOUR SCHEDULE",
                        "raw_text": "KITCHEN COLOUR SCHEDULE",
                        "needs_ocr": True,
                    }
                ],
            }
        ]
        heuristic_snapshot = {
            "job_no": "37529",
            "builder_name": "Clarendon",
            "source_kind": "spec",
            "generated_at": "2026-03-29T10:00:00+00:00",
            "site_address": "",
            "rooms": [{"room_key": "kitchen", "original_room_label": "KITCHEN"}],
            "special_sections": [],
            "appliances": [],
            "others": {"flooring_notes": "", "splashback_notes": "", "manual_notes": ""},
            "warnings": [],
            "source_documents": [],
            "analysis": {"mode": "heuristic_only"},
        }
        with (
            mock.patch.object(extraction_service.runtime, "OPENAI_ENABLED", True),
            mock.patch.object(extraction_service.runtime, "OPENAI_API_KEY", "test-key"),
            mock.patch.object(extraction_service.runtime, "OPENAI_VISION_ENABLED", True),
            mock.patch.object(extraction_service.runtime, "SPEC_HEAVY_VISION_ENABLED", False),
            mock.patch.object(extraction_service.runtime, "SPEC_DOCLING_BUILDERS", {"imperial", "simonds", "evoca"}),
            mock.patch("App.services.extraction_service._load_documents", return_value=documents),
            mock.patch("App.services.extraction_service._page_requires_vision", return_value=True),
            mock.patch("App.services.extraction_service._docling_available", return_value=True),
            mock.patch("App.services.extraction_service._request_docling_page_layout") as request_docling_page_layout,
            mock.patch("App.services.extraction_service.parsing.parse_documents", return_value=heuristic_snapshot),
            mock.patch(
                "App.services.extraction_service.parsing.enrich_snapshot_rooms",
                side_effect=lambda payload, _documents, rule_flags=None: payload,
            ),
            mock.patch("App.services.extraction_service._stabilize_snapshot_layout", side_effect=lambda payload, **_: payload),
            mock.patch("App.services.extraction_service._apply_builder_specific_polish", side_effect=lambda payload, *_args, **_kwargs: payload),
            mock.patch("App.services.extraction_service._enrich_snapshot_appliances", side_effect=lambda payload, *_args, **_kwargs: payload),
        ):
            snapshot = extraction_service.build_spec_snapshot(
                job={"job_no": "37529"},
                builder={"name": "Clarendon"},
                files=[{"path": "clarendon.pdf", "original_name": "clarendon.pdf"}],
                template_files=[],
            )
        self.assertEqual(snapshot["analysis"]["layout_provider"], "heuristic")
        self.assertFalse(snapshot["analysis"]["docling_attempted"])
        self.assertFalse(snapshot["analysis"]["vision_attempted"])
        self.assertIn("Docling is disabled by builder policy", snapshot["analysis"]["docling_note"])
        request_docling_page_layout.assert_not_called()

    def test_build_spec_snapshot_uses_docling_without_heavy_vision_by_default_for_yellowwood(self) -> None:
        documents = [
            {
                "file_name": "yellowwood.pdf",
                "path": "yellowwood.pdf",
                "role": "spec",
                "pages": [
                    {
                        "page_no": 17,
                        "text": "KITCHEN\nBASE CUPBOARDS & DRAWERS\nOVERHEAD CUPBOARDS\nPANTRY DOOR HANDLES",
                        "raw_text": "KITCHEN\nBASE CUPBOARDS & DRAWERS\nOVERHEAD CUPBOARDS\nPANTRY DOOR HANDLES",
                        "needs_ocr": False,
                    }
                ],
            }
        ]
        heuristic_snapshot = {
            "job_no": "38095",
            "builder_name": "Yellowwood",
            "source_kind": "spec",
            "generated_at": "2026-03-29T10:00:00+00:00",
            "site_address": "",
            "rooms": [{"room_key": "kitchen", "original_room_label": "KITCHEN", "room_name": "KITCHEN"}],
            "special_sections": [],
            "appliances": [],
            "others": {"flooring_notes": "", "splashback_notes": "", "manual_notes": ""},
            "warnings": [],
            "source_documents": [],
            "analysis": {"mode": "heuristic_only"},
        }
        docling_layout = {
            "page_type": "joinery",
            "section_label": "KITCHEN",
            "room_label": "KITCHEN",
            "room_blocks": [
                {
                    "room_label": "KITCHEN",
                    "rows": [
                        {
                            "row_label": "Base Cupboards & Drawers",
                            "value_region_text": "White Melamine",
                            "supplier_region_text": "",
                            "notes_region_text": "",
                            "row_kind": "material",
                        }
                    ],
                }
            ],
        }
        with (
            mock.patch.object(extraction_service.runtime, "OPENAI_ENABLED", False),
            mock.patch.object(extraction_service.runtime, "SPEC_HEAVY_VISION_ENABLED", False),
            mock.patch("App.services.extraction_service._load_documents", return_value=documents),
            mock.patch("App.services.extraction_service._docling_available", return_value=True),
            mock.patch("App.services.extraction_service._request_docling_page_layout", return_value=docling_layout),
            mock.patch("App.services.extraction_service.parsing.parse_documents", return_value=heuristic_snapshot),
            mock.patch(
                "App.services.extraction_service.parsing.enrich_snapshot_rooms",
                side_effect=lambda payload, _documents, rule_flags=None: payload,
            ),
            mock.patch("App.services.extraction_service._stabilize_snapshot_layout", side_effect=lambda payload, **_: payload),
            mock.patch("App.services.extraction_service._apply_builder_specific_polish", side_effect=lambda payload, *_args, **_kwargs: payload),
            mock.patch("App.services.extraction_service._enrich_snapshot_appliances", side_effect=lambda payload, *_args, **_kwargs: payload),
        ):
            snapshot = extraction_service.build_spec_snapshot(
                job={"job_no": "38095"},
                builder={"name": "Yellowwood"},
                files=[{"path": "yellowwood.pdf", "original_name": "yellowwood.pdf"}],
                template_files=[],
            )
        self.assertEqual(snapshot["analysis"]["layout_provider"], "docling")
        self.assertTrue(snapshot["analysis"]["docling_attempted"])
        self.assertFalse(snapshot["analysis"]["vision_attempted"])

    def test_spec_docling_enabled_matches_builder_aliases(self) -> None:
        with mock.patch.object(extraction_service.runtime, "SPEC_DOCLING_BUILDERS", {"imperial", "simonds", "evoca", "yellowwood"}):
            self.assertTrue(extraction_service._spec_docling_enabled("Yellowwood Homes", "spec"))
            self.assertTrue(extraction_service._spec_docling_enabled("Imperial Kitchens", "spec"))
            self.assertFalse(extraction_service._spec_docling_enabled("Clarendon", "spec"))

    def test_page_requires_vision_for_yellowwood_schedule_pages(self) -> None:
        page = {
            "page_no": 17,
            "text": "KITCHEN\nBASE CUPBOARDS & DRAWERS\nOVERHEAD CUPBOARDS\nPANTRY DOOR HANDLES\nISLAND BAR BACK",
            "raw_text": "KITCHEN\nBASE CUPBOARDS & DRAWERS\nOVERHEAD CUPBOARDS\nPANTRY DOOR HANDLES\nISLAND BAR BACK",
            "needs_ocr": False,
        }
        self.assertTrue(
            extraction_service._page_requires_vision(
                builder_name="Yellowwood",
                source_kind="spec",
                file_name="yellowwood.pdf",
                page=page,
                heuristic={},
            )
        )

    def test_build_spec_snapshot_uses_docling_without_heavy_vision_by_default_for_imperial(self) -> None:
        documents = [
            {
                "file_name": "imperial.pdf",
                "path": "imperial.pdf",
                "role": "spec",
                "pages": [
                    {
                        "page_no": 1,
                        "text": "BASE CABINETRY COLOURClassic White Matt Polytec",
                        "raw_text": "BASE CABINETRY COLOURClassic White Matt Polytec",
                        "needs_ocr": True,
                    }
                ],
            }
        ]
        heuristic_after_docling = {
            "job_no": "38211",
            "builder_name": "Imperial",
            "source_kind": "spec",
            "generated_at": "2026-03-29T10:00:00+00:00",
            "site_address": "",
            "rooms": [{"room_key": "kitchen", "original_room_label": "KITCHEN", "door_colours_base": "Polytec - Classic White Matt"}],
            "special_sections": [],
            "appliances": [],
            "others": {"flooring_notes": "", "splashback_notes": "", "manual_notes": ""},
            "warnings": [],
            "source_documents": [],
            "analysis": {"mode": "heuristic_only", "room_master_file": "imperial.pdf"},
        }
        with (
            mock.patch.object(extraction_service.runtime, "OPENAI_ENABLED", True),
            mock.patch.object(extraction_service.runtime, "OPENAI_API_KEY", "test-key"),
            mock.patch.object(extraction_service.runtime, "OPENAI_VISION_ENABLED", True),
            mock.patch.object(extraction_service.runtime, "SPEC_HEAVY_VISION_ENABLED", False),
            mock.patch.object(extraction_service.runtime, "SPEC_DOCLING_BUILDERS", {"imperial", "simonds", "evoca"}),
            mock.patch("App.services.extraction_service._load_documents", return_value=documents),
            mock.patch("App.services.extraction_service._page_requires_vision", return_value=True),
            mock.patch("App.services.extraction_service._docling_available", return_value=True),
            mock.patch(
                "App.services.extraction_service._request_docling_page_layout",
                return_value={
                    "page_type": "joinery",
                    "section_label": "KITCHEN JOINERY SELECTION SHEET",
                    "room_label": "KITCHEN",
                    "rows": [
                        {
                            "row_label": "BASE CABINETRY COLOUR",
                            "value_region_text": "Classic White Matt",
                            "supplier_region_text": "Polytec",
                            "notes_region_text": "",
                            "row_kind": "material",
                        }
                    ],
                },
            ),
            mock.patch("App.services.extraction_service.parsing.parse_documents", return_value=heuristic_after_docling),
            mock.patch(
                "App.services.extraction_service.parsing.enrich_snapshot_rooms",
                side_effect=lambda payload, _documents, rule_flags=None: payload,
            ),
            mock.patch("App.services.extraction_service._stabilize_snapshot_layout", side_effect=lambda payload, **_: payload),
            mock.patch("App.services.extraction_service._apply_builder_specific_polish", side_effect=lambda payload, *_args, **_kwargs: payload),
            mock.patch("App.services.extraction_service._enrich_snapshot_appliances", side_effect=lambda payload, *_args, **_kwargs: payload),
        ):
            snapshot = extraction_service.build_spec_snapshot(
                job={"job_no": "38211"},
                builder={"name": "Imperial"},
                files=[{"path": "imperial.pdf", "original_name": "imperial.pdf"}],
                template_files=[],
            )
        self.assertEqual(snapshot["analysis"]["layout_mode"], "docling")
        self.assertEqual(snapshot["analysis"]["layout_provider"], "docling")
        self.assertTrue(snapshot["analysis"]["docling_attempted"])
        self.assertTrue(snapshot["analysis"]["docling_succeeded"])
        self.assertFalse(snapshot["analysis"]["vision_attempted"])
        self.assertIn("disabled by default for spec runs", snapshot["analysis"]["vision_note"])

    def test_build_spec_snapshot_applies_vision_layout_before_final_snapshot(self) -> None:
        documents = [
            {
                "file_name": "imperial.pdf",
                "path": "imperial.pdf",
                "role": "spec",
                "pages": [
                    {
                        "page_no": 1,
                        "text": "BASE CABINETRY COLOURClassic White Matt Polytec",
                        "raw_text": "BASE CABINETRY COLOURClassic White Matt Polytec",
                        "needs_ocr": True,
                    }
                ],
            }
        ]
        heuristic_initial = {
            "job_no": "38211",
            "builder_name": "Imperial",
            "source_kind": "spec",
            "generated_at": "2026-03-29T10:00:00+00:00",
            "site_address": "",
            "rooms": [{"room_key": "kitchen", "original_room_label": "KITCHEN", "door_colours_base": ""}],
            "special_sections": [],
            "appliances": [],
            "others": {"flooring_notes": "", "splashback_notes": "", "manual_notes": ""},
            "warnings": [],
            "source_documents": [],
            "analysis": {"mode": "heuristic_only"},
        }
        heuristic_after_vision = {
            **heuristic_initial,
            "rooms": [{"room_key": "kitchen", "original_room_label": "KITCHEN", "door_colours_base": "Polytec - Classic White Matt"}],
            "analysis": {"mode": "heuristic_only", "room_master_file": "imperial.pdf"},
        }
        with (
            mock.patch.object(extraction_service.runtime, "OPENAI_ENABLED", True),
            mock.patch.object(extraction_service.runtime, "OPENAI_API_KEY", "test-key"),
            mock.patch.object(extraction_service.runtime, "OPENAI_VISION_ENABLED", True),
            mock.patch.object(extraction_service.runtime, "SPEC_HEAVY_VISION_ENABLED", True),
            mock.patch.object(extraction_service.runtime, "OPENAI_VISION_MAX_PAGES", 4),
            mock.patch("App.services.extraction_service._load_documents", return_value=documents),
            mock.patch("App.services.extraction_service._page_requires_vision", return_value=True),
            mock.patch("App.services.extraction_service._docling_available", return_value=False),
            mock.patch("App.services.extraction_service._render_pdf_page_png", return_value=b"png"),
            mock.patch(
                "App.services.extraction_service._request_page_layout",
                return_value={
                    "page_type": "joinery",
                    "section_label": "KITCHEN JOINERY SELECTION SHEET",
                    "room_label": "KITCHEN",
                    "rows": [
                        {
                            "row_label": "BASE CABINETRY COLOUR",
                            "value_region_text": "Classic White Matt",
                            "supplier_region_text": "Polytec",
                            "notes_region_text": "",
                            "row_kind": "material",
                        }
                    ],
                },
            ),
            mock.patch(
                "App.services.extraction_service.parsing.parse_documents",
                return_value=heuristic_after_vision,
            ),
            mock.patch(
                "App.services.extraction_service.parsing.enrich_snapshot_rooms",
                side_effect=lambda payload, _documents, rule_flags=None: payload,
            ),
            mock.patch("App.services.extraction_service._stabilize_snapshot_layout", side_effect=lambda payload, **_: payload),
            mock.patch("App.services.extraction_service._apply_builder_specific_polish", side_effect=lambda payload, *_args, **_kwargs: payload),
            mock.patch("App.services.extraction_service._enrich_snapshot_appliances", side_effect=lambda payload, *_args, **_kwargs: payload),
            mock.patch(
                "App.services.extraction_service._try_openai",
                return_value=(
                    None,
                    {
                        "mode": "heuristic_only",
                        "parser_strategy": "global_conservative",
                        "openai_attempted": False,
                        "openai_succeeded": False,
                        "openai_model": "gpt-4.1-mini",
                        "vision_attempted": False,
                        "vision_succeeded": False,
                        "vision_pages": [],
                        "vision_page_count": 0,
                        "vision_note": "",
                        "note": "",
                    },
                ),
            ),
        ):
            snapshot = extraction_service.build_spec_snapshot(
                job={"job_no": "38211"},
                builder={"name": "Imperial"},
                files=[{"path": "imperial.pdf", "original_name": "imperial.pdf"}],
                template_files=[],
        )
        self.assertEqual(snapshot["rooms"][0]["door_colours_base"], "Polytec - Classic White Matt")
        self.assertTrue(snapshot["analysis"]["layout_attempted"])
        self.assertTrue(snapshot["analysis"]["layout_succeeded"])
        self.assertEqual(snapshot["analysis"]["layout_mode"], "heavy_vision")
        self.assertEqual(snapshot["analysis"]["layout_provider"], "heavy_vision")
        self.assertEqual(snapshot["analysis"]["layout_pages"], [1])
        self.assertEqual(snapshot["analysis"]["heavy_vision_pages"], [1])
        self.assertFalse(snapshot["analysis"]["docling_attempted"])
        self.assertTrue(snapshot["analysis"]["vision_attempted"])
        self.assertTrue(snapshot["analysis"]["vision_succeeded"])
        self.assertEqual(snapshot["analysis"]["vision_pages"], [1])
        self.assertEqual(snapshot["analysis"]["vision_page_count"], 1)

    def test_build_spec_snapshot_dual_runs_docling_and_heavy_vision_for_key_tables(self) -> None:
        documents = [
            {
                "file_name": "imperial.pdf",
                "path": "imperial.pdf",
                "role": "spec",
                "pages": [
                    {
                        "page_no": 1,
                        "text": "BASE CABINETRY COLOURClassic White Matt Polytec",
                        "raw_text": "BASE CABINETRY COLOURClassic White Matt Polytec",
                        "needs_ocr": True,
                    }
                ],
            }
        ]
        heuristic_after_docling = {
            "job_no": "38211",
            "builder_name": "Imperial",
            "source_kind": "spec",
            "generated_at": "2026-03-29T10:00:00+00:00",
            "site_address": "",
            "rooms": [{"room_key": "kitchen", "original_room_label": "KITCHEN", "door_colours_base": "Polytec - Classic White Matt"}],
            "special_sections": [],
            "appliances": [],
            "others": {"flooring_notes": "", "splashback_notes": "", "manual_notes": ""},
            "warnings": [],
            "source_documents": [],
            "analysis": {"mode": "heuristic_only", "room_master_file": "imperial.pdf"},
        }
        with (
            mock.patch.object(extraction_service.runtime, "OPENAI_ENABLED", True),
            mock.patch.object(extraction_service.runtime, "OPENAI_API_KEY", "test-key"),
            mock.patch.object(extraction_service.runtime, "OPENAI_VISION_ENABLED", True),
            mock.patch.object(extraction_service.runtime, "SPEC_HEAVY_VISION_ENABLED", True),
            mock.patch.object(extraction_service.runtime, "OPENAI_VISION_MAX_PAGES", 4),
            mock.patch("App.services.extraction_service._load_documents", return_value=documents),
            mock.patch("App.services.extraction_service._page_requires_vision", return_value=True),
            mock.patch("App.services.extraction_service._docling_available", return_value=True),
            mock.patch(
                "App.services.extraction_service._request_docling_page_layout",
                return_value={
                    "page_type": "joinery",
                    "section_label": "KITCHEN JOINERY SELECTION SHEET",
                    "room_label": "KITCHEN",
                    "rows": [
                        {
                            "row_label": "BASE CABINETRY COLOUR",
                            "value_region_text": "Classic White Matt",
                            "supplier_region_text": "Polytec",
                            "notes_region_text": "",
                            "row_kind": "material",
                        }
                    ],
                },
            ),
            mock.patch("App.services.extraction_service._render_pdf_page_png", return_value=b"png"),
            mock.patch(
                "App.services.extraction_service._request_page_layout",
                return_value={
                    "page_type": "joinery",
                    "section_label": "NA KITCHEN JOINERY SELECTION SHEET",
                    "room_label": "NA KITCHEN",
                    "rows": [
                        {
                            "row_label": "HANDLES",
                            "value_region_text": "Belluno 9995772",
                            "supplier_region_text": "Hettich",
                            "notes_region_text": "Base cabs",
                            "row_kind": "handle",
                        }
                    ],
                },
            ),
            mock.patch(
                "App.services.extraction_service.parsing.parse_documents",
                return_value=heuristic_after_docling,
            ),
            mock.patch(
                "App.services.extraction_service.parsing.enrich_snapshot_rooms",
                side_effect=lambda payload, _documents, rule_flags=None: payload,
            ),
            mock.patch("App.services.extraction_service._stabilize_snapshot_layout", side_effect=lambda payload, **_: payload),
            mock.patch("App.services.extraction_service._apply_builder_specific_polish", side_effect=lambda payload, *_args, **_kwargs: payload),
            mock.patch("App.services.extraction_service._enrich_snapshot_appliances", side_effect=lambda payload, *_args, **_kwargs: payload),
            mock.patch(
                "App.services.extraction_service._try_openai",
                return_value=(
                    None,
                    {
                        "mode": "heuristic_only",
                        "parser_strategy": "global_conservative",
                        "openai_attempted": False,
                        "openai_succeeded": False,
                        "openai_model": "gpt-4.1-mini",
                        "vision_attempted": False,
                        "vision_succeeded": False,
                        "vision_pages": [],
                        "vision_page_count": 0,
                        "vision_note": "",
                        "note": "",
                    },
                ),
            ),
        ):
            snapshot = extraction_service.build_spec_snapshot(
                job={"job_no": "38211"},
                builder={"name": "Imperial"},
                files=[{"path": "imperial.pdf", "original_name": "imperial.pdf"}],
                template_files=[],
            )
        self.assertEqual(snapshot["rooms"][0]["door_colours_base"], "Polytec - Classic White Matt")
        self.assertEqual(snapshot["analysis"]["layout_mode"], "mixed")
        self.assertEqual(snapshot["analysis"]["layout_provider"], "mixed")
        self.assertTrue(snapshot["analysis"]["docling_attempted"])
        self.assertTrue(snapshot["analysis"]["docling_succeeded"])
        self.assertEqual(snapshot["analysis"]["docling_pages"], [1])
        self.assertTrue(snapshot["analysis"]["vision_attempted"])
        self.assertTrue(snapshot["analysis"]["vision_succeeded"])
        self.assertEqual(snapshot["analysis"]["vision_pages"], [1])
        self.assertEqual(snapshot["analysis"]["heavy_vision_pages"], [1])

    def test_docling_markdown_to_layout_restores_joinery_rows(self) -> None:
        markdown = (
            "## KITCHEN JOINERY SELECTION SHEET\n\n"
            "| AREA / ITEM | SPECS / DESCRIPTION | SUPPLIER | NOTES |\n"
            "|---|---|---|---|\n"
            "| BENCHTOP | Caesarstone Fresh Concrete 20mm Pencil Round Edge | Caesarstone | |\n"
            "| BASE CABINETRY COLOUR | Polytec Classic White Matt | Polytec | |\n"
            "| KICKBOARDS | MATCH ABOVE Polytec Classic White Matt | Polytec + Laminex | |\n"
        )
        layout = extraction_service._docling_markdown_to_layout(
            markdown,
            builder_name="Imperial",
            source_kind="spec",
            file_name="imperial.pdf",
            raw_page_text="KITCHEN JOINERY SELECTION SHEET",
        )
        self.assertEqual(layout["page_type"], "joinery")
        self.assertEqual(layout["room_label"], "KITCHEN")
        rows = layout["room_blocks"][0]["rows"]
        labels = [row["row_label"] for row in rows]
        self.assertIn("Benchtop", labels)
        self.assertIn("Base Cabinetry Colour", labels)
        self.assertIn("Kickboards", labels)
        base_row = next(row for row in rows if row["row_label"] == "Base Cabinetry Colour")
        self.assertEqual(base_row["row_kind"], "material")

    def test_merge_page_layouts_combines_rows_and_drops_footer_noise(self) -> None:
        docling_layout = {
            "page_type": "joinery",
            "section_label": "KITCHEN JOINERY SELECTION SHEET",
            "room_label": "KITCHEN",
            "room_blocks": [
                {
                    "room_label": "KITCHEN",
                    "rows": [
                        {
                            "row_label": "BASE CABINETRY COLOUR",
                            "value_region_text": "Classic White Matt",
                            "supplier_region_text": "",
                            "notes_region_text": "",
                            "row_kind": "material",
                        }
                    ],
                }
            ],
            "rows": [],
        }
        vision_layout = {
            "page_type": "joinery",
            "section_label": "NA KITCHEN JOINERY SELECTION SHEET",
            "room_label": "NA KITCHEN",
            "room_blocks": [
                {
                    "room_label": "NA KITCHEN",
                    "rows": [
                        {
                            "row_label": "BASE CABINETRY COLOUR",
                            "value_region_text": "Classic White Matt",
                            "supplier_region_text": "Polytec",
                            "notes_region_text": "",
                            "row_kind": "material",
                        },
                        {
                            "row_label": "HANDLES",
                            "value_region_text": "Belluno 9995772",
                            "supplier_region_text": "Hettich",
                            "notes_region_text": "Base cabs",
                            "row_kind": "handle",
                        },
                        {
                            "row_label": "Client",
                            "value_region_text": "Jason",
                            "supplier_region_text": "",
                            "notes_region_text": "",
                            "row_kind": "metadata",
                        },
                    ],
                }
            ],
            "rows": [],
        }
        merged = extraction_service._merge_page_layouts(docling_layout, vision_layout, builder_name="Imperial", raw_page_text="KITCHEN JOINERY SELECTION SHEET")
        self.assertEqual(merged["room_label"], "KITCHEN")
        rows = merged["room_blocks"][0]["rows"]
        labels = [row["row_label"] for row in rows]
        self.assertIn("BASE CABINETRY COLOUR", labels)
        self.assertIn("HANDLES", labels)
        self.assertNotIn("Client", labels)
        base_row = next(row for row in rows if row["row_label"] == "BASE CABINETRY COLOUR")
        self.assertEqual(base_row["supplier_region_text"], "Polytec")

    def test_build_spec_snapshot_keeps_docling_when_vision_cross_check_fails(self) -> None:
        documents = [
            {
                "file_name": "imperial.pdf",
                "path": "imperial.pdf",
                "role": "spec",
                "pages": [
                    {
                        "page_no": 1,
                        "text": "BASE CABINETRY COLOURClassic White Matt Polytec",
                        "raw_text": "BASE CABINETRY COLOURClassic White Matt Polytec",
                        "needs_ocr": True,
                    }
                ],
            }
        ]
        heuristic_after_docling = {
            "job_no": "38211",
            "builder_name": "Imperial",
            "source_kind": "spec",
            "generated_at": "2026-03-29T10:00:00+00:00",
            "site_address": "",
            "rooms": [{"room_key": "kitchen", "original_room_label": "KITCHEN", "door_colours_base": "Polytec - Classic White Matt"}],
            "special_sections": [],
            "appliances": [],
            "others": {"flooring_notes": "", "splashback_notes": "", "manual_notes": ""},
            "warnings": [],
            "source_documents": [],
            "analysis": {"mode": "heuristic_only", "room_master_file": "imperial.pdf"},
        }
        with (
            mock.patch.object(extraction_service.runtime, "OPENAI_ENABLED", True),
            mock.patch.object(extraction_service.runtime, "OPENAI_API_KEY", "test-key"),
            mock.patch.object(extraction_service.runtime, "OPENAI_VISION_ENABLED", True),
            mock.patch.object(extraction_service.runtime, "SPEC_HEAVY_VISION_ENABLED", True),
            mock.patch.object(extraction_service.runtime, "OPENAI_VISION_MAX_PAGES", 4),
            mock.patch("App.services.extraction_service._load_documents", return_value=documents),
            mock.patch("App.services.extraction_service._page_requires_vision", return_value=True),
            mock.patch("App.services.extraction_service._docling_available", return_value=True),
            mock.patch(
                "App.services.extraction_service._request_docling_page_layout",
                return_value={
                    "page_type": "joinery",
                    "section_label": "KITCHEN JOINERY SELECTION SHEET",
                    "room_label": "KITCHEN",
                    "rows": [
                        {
                            "row_label": "BASE CABINETRY COLOUR",
                            "value_region_text": "Classic White Matt",
                            "supplier_region_text": "Polytec",
                            "notes_region_text": "",
                            "row_kind": "material",
                        }
                    ],
                },
            ),
            mock.patch("App.services.extraction_service._render_pdf_page_png", return_value=b"png"),
            mock.patch("App.services.extraction_service._request_page_layout", side_effect=RuntimeError("429 test failure")),
            mock.patch(
                "App.services.extraction_service.parsing.parse_documents",
                return_value=heuristic_after_docling,
            ),
            mock.patch(
                "App.services.extraction_service.parsing.enrich_snapshot_rooms",
                side_effect=lambda payload, _documents, rule_flags=None: payload,
            ),
            mock.patch("App.services.extraction_service._stabilize_snapshot_layout", side_effect=lambda payload, **_: payload),
            mock.patch("App.services.extraction_service._apply_builder_specific_polish", side_effect=lambda payload, *_args, **_kwargs: payload),
            mock.patch("App.services.extraction_service._enrich_snapshot_appliances", side_effect=lambda payload, *_args, **_kwargs: payload),
            mock.patch(
                "App.services.extraction_service._try_openai",
                return_value=(
                    None,
                    {
                        "mode": "heuristic_only",
                        "parser_strategy": "global_conservative",
                        "openai_attempted": False,
                        "openai_succeeded": False,
                        "openai_model": "gpt-4.1-mini",
                        "vision_attempted": False,
                        "vision_succeeded": False,
                        "vision_pages": [],
                        "vision_page_count": 0,
                        "vision_note": "",
                        "note": "",
                    },
                ),
            ),
        ):
            snapshot = extraction_service.build_spec_snapshot(
                job={"job_no": "38211"},
                builder={"name": "Imperial"},
                files=[{"path": "imperial.pdf", "original_name": "imperial.pdf"}],
                template_files=[],
            )
        self.assertEqual(snapshot["analysis"]["layout_mode"], "docling")
        self.assertEqual(snapshot["analysis"]["layout_provider"], "docling")
        self.assertTrue(snapshot["analysis"]["docling_attempted"])
        self.assertTrue(snapshot["analysis"]["docling_succeeded"])
        self.assertTrue(snapshot["analysis"]["vision_attempted"])
        self.assertFalse(snapshot["analysis"]["vision_succeeded"])
        self.assertIn("429 test failure", snapshot["analysis"]["vision_note"])

    def test_apply_final_layout_pages_prefers_heuristic_when_docling_room_titles_are_noise(self) -> None:
        documents = [
            {
                "file_name": "evoca.pdf",
                "pages": [
                    {
                        "page_no": 8,
                        "raw_text": "Kitchen colour schedule",
                        "text": "Kitchen colour schedule",
                        "page_layout": {
                            "page_type": "joinery",
                            "section_label": "KITCHEN",
                            "room_label": "KITCHEN",
                            "room_blocks": [
                                {
                                    "room_label": "KITCHEN",
                                    "rows": [
                                        {
                                            "row_label": "Benchtops",
                                            "value_region_text": "",
                                            "supplier_region_text": "",
                                            "notes_region_text": "",
                                            "row_kind": "material",
                                        },
                                        {
                                            "row_label": "Manufacturer",
                                            "value_region_text": "Quantum Quartz",
                                            "supplier_region_text": "",
                                            "notes_region_text": "",
                                            "row_kind": "material",
                                        },
                                        {
                                            "row_label": "Colour",
                                            "value_region_text": "Champagne",
                                            "supplier_region_text": "",
                                            "notes_region_text": "",
                                            "row_kind": "material",
                                        },
                                    ],
                                }
                            ],
                            "rows": [],
                        },
                    }
                ],
            }
        ]
        docling_layouts = {
            (0, 0): {
                "page_type": "joinery",
                "section_label": "Client Initials: __________",
                "room_label": "Initials",
                "room_blocks": [{"room_label": "", "rows": [{"row_label": "Manufacturer", "value_region_text": "Quantum Quartz", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"}]}],
                "rows": [],
            }
        }
        mixed_pages, docling_pages, vision_pages = extraction_service._apply_final_layout_pages(
            documents,
            [(0, 0)],
            docling_layouts=docling_layouts,
            vision_layouts={},
            builder_name="Evoca",
        )
        page = documents[0]["pages"][0]
        self.assertEqual(mixed_pages, [])
        self.assertEqual(docling_pages, [8])
        self.assertEqual(vision_pages, [])
        self.assertEqual(page["layout_mode"], "lightweight")
        self.assertEqual(page["page_layout"]["room_blocks"][0]["room_label"], "KITCHEN")

    def test_layout_completeness_penalizes_joinery_header_noise(self) -> None:
        clean_layout = {
            "page_type": "joinery",
            "section_label": "KITCHEN",
            "room_label": "KITCHEN",
            "rows": [
                {
                    "row_label": "Benchtops",
                    "value_region_text": "20mm Arissed",
                    "supplier_region_text": "Quantum Quartz",
                    "notes_region_text": "Champagne",
                    "row_kind": "material",
                }
            ],
        }
        noisy_layout = {
            "page_type": "joinery",
            "section_label": "KITCHEN",
            "room_label": "KITCHEN",
            "rows": [
                {
                    "row_label": "Benchtops",
                    "value_region_text": "15 CABINETS All Cabinets include Soft Close Hinges & Runners, lined internally with White Melamine",
                    "supplier_region_text": "",
                    "notes_region_text": "Page 8 of 83 Client Initials",
                    "row_kind": "material",
                }
            ],
        }
        self.assertGreater(
            extraction_service._layout_completeness_score(clean_layout),
            extraction_service._layout_completeness_score(noisy_layout),
        )

    def test_generic_sanitizers_strip_property_noise_and_accessory_tails(self) -> None:
        self.assertEqual(
            extraction_service._sanitize_generic_material_field(
                "Polytec | Belgian Oak Matt | Laminate | 4062-128-TG",
                field_name="door_colours_base",
            ),
            "Polytec | Belgian Oak Matt",
        )
        self.assertEqual(
            extraction_service._sanitize_generic_fixture_field(
                "Alder - Maxx (WELS 6 stars) - Rectangle Sink Mixer - Matt Black + ALDER SACHI ROBE HOOK IN MATT BLACK",
                kind="tap",
            ),
            "Alder - Maxx (WELS 6 stars) - Rectangle Sink Mixer - Matt Black",
        )
        self.assertEqual(
            extraction_service._sanitize_generic_fixture_field(
                "Alder - Milano - Vegie Mixer - Chrome - ALDER SACHI | Alder - Milano - Vegie Mixer - Chrome",
                kind="tap",
            ),
            "Alder - Milano - Vegie Mixer - Chrome",
        )
        self.assertEqual(
            extraction_service._sanitize_generic_material_entries(
                ["Laminex - Blackened Legno", "Blackened Legno", "Laminex - Blackened Legno"],
                field_name="toe_kick",
                room_key="kitchen",
            ),
            ["Laminex - Blackened Legno"],
        )
        self.assertEqual(
            extraction_service._sanitize_generic_handle_entries(
                ["N/A, Category 6, C137 Black 100m, Horizontal", "up to 20mm Drop Down - No Handle N/A Category 6, Horizontal, Soft Close"]
            ),
            ["C137 Black 100m, Horizontal", "up to 20mm Drop Down - No Handle N/A, Horizontal"],
        )

    def test_polish_generic_layout_room_sanitizes_toe_kick_before_material_cleanup(self) -> None:
        polished = extraction_service._polish_generic_layout_room(
            {"room_key": "kitchen", "original_room_label": "Kitchen"},
            {"toe_kick": ["Laminex - Blackened Legno", "Blackened Legno"]},
        )
        self.assertEqual(polished["toe_kick"], ["Laminex - Blackened Legno"])

    def test_imperial_material_and_flooring_cleaners_strip_meta_noise(self) -> None:
        self.assertEqual(
            parsing_module._imperial_clean_flooring_value(
                [": Tiles with Floating TBC Soft close Caesarstone NOTES SUPPLIER 1 x Waterfall End 20mm Stone"]
            ),
            "Tiles with Floating TBC",
        )
        self.assertEqual(
            parsing_module._imperial_layout_row_material_text(
                {
                    "value_text": "9 Greenland Court, Springfield PRIVATE - Yasantha Warawita 12/03/2026",
                    "supplier_text": "Caesarstone",
                    "notes_text": "20mm Stone",
                }
            ),
            "20mm Caesarstone - Stone",
        )

    def test_imperial_extract_soft_close_and_flooring_swaps_reversed_na_and_soft_close(self) -> None:
        soft_close, flooring = parsing_module._imperial_extract_soft_close_and_flooring(
            "Hinges & Drawer Runners: NAFloor Type & Kick refacing required: Soft Close",
            ["Hinges & Drawer Runners: NAFloor Type & Kick refacing required: Soft Close"],
        )
        self.assertEqual(soft_close, "Soft Close")
        self.assertEqual(flooring, "NA")

    def test_vision_layout_to_text_preserves_row_boundaries(self) -> None:
        layout = {
            "page_type": "joinery",
            "section_label": "KITCHEN JOINERY SELECTION SHEET",
            "room_label": "KITCHEN",
            "rows": [
                {
                    "row_label": "BENCHTOP",
                    "value_region_text": "20mm Frosty Carrina (5141)",
                    "supplier_region_text": "Caesarstone",
                    "notes_region_text": "Cooktop Run",
                    "row_kind": "material",
                },
                {
                    "row_label": "GPO'S",
                    "value_region_text": "Drawer GPO by builder",
                    "supplier_region_text": "",
                    "notes_region_text": "",
                    "row_kind": "accessory",
                },
                {
                    "row_label": "ALL COLOURS SHOWN",
                    "value_region_text": "Disclaimer",
                    "supplier_region_text": "",
                    "notes_region_text": "",
                    "row_kind": "footer",
                },
            ],
        }
        normalized = extraction_service._vision_layout_to_text(layout)
        self.assertIn("KITCHEN JOINERY SELECTION SHEET", normalized)
        self.assertIn("BENCHTOP 20mm Frosty Carrina (5141) Caesarstone Cooktop Run", normalized)
        self.assertIn("GPO'S Drawer GPO by builder", normalized)
        self.assertNotIn("Disclaimer", normalized)

    def test_build_spec_snapshot_enriches_official_appliance_resources(self) -> None:
        documents = [
            {
                "stored_name": "sample.pdf",
                "original_name": "sample.pdf",
                "path": "sample.pdf",
            }
        ]
        mock_snapshot = {
            "job_no": "37529",
            "builder_name": "Clarendon",
            "source_kind": "spec",
            "generated_at": "2026-03-22T10:00:00+00:00",
            "rooms": [],
            "appliances": [
                {
                    "appliance_type": "Cooktop",
                    "make": "Westinghouse",
                    "model_no": "WHC943BD",
                    "website_url": "",
                    "overall_size": "",
                    "source_file": "sample.pdf",
                    "page_refs": "1",
                    "evidence_snippet": "Cooktop row",
                    "confidence": 0.82,
                }
            ],
            "others": {},
            "warnings": [],
            "source_documents": [],
            "analysis": {"mode": "heuristic_only"},
        }
        with (
            mock.patch("App.services.extraction_service._load_documents", return_value=[]),
            mock.patch("App.services.extraction_service.parsing.parse_documents", return_value=mock_snapshot),
            mock.patch(
                "App.services.extraction_service.parsing.enrich_snapshot_rooms",
                side_effect=lambda payload, _documents, rule_flags=None: payload,
            ),
            mock.patch("App.services.appliance_official.lookup_official_appliance_resources", return_value={
                "product_url": "https://official.example/product/WHC943BD",
                "spec_url": "https://official.example/spec/WHC943BD.pdf",
                "manual_url": "https://official.example/manual/WHC943BD.pdf",
                "website_url": "https://official.example/product/WHC943BD",
                "overall_size": "900 x 510 x 60 mm",
            }),
        ):
            snapshot = extraction_service.build_spec_snapshot(
                job={"job_no": "37529"},
                builder={"name": "Clarendon"},
                files=documents,
                template_files=[],
            )
        appliance = snapshot["appliances"][0]
        self.assertEqual(appliance["product_url"], "https://official.example/product/WHC943BD")
        self.assertEqual(appliance["spec_url"], "https://official.example/spec/WHC943BD.pdf")
        self.assertEqual(appliance["manual_url"], "https://official.example/manual/WHC943BD.pdf")
        self.assertEqual(appliance["website_url"], "https://official.example/product/WHC943BD")
        self.assertEqual(appliance["overall_size"], "900 x 510 x 60 mm")

    def test_extract_size_from_product_page_hwd_lines(self) -> None:
        self.assertEqual(
            _extract_size_from_text("Dimension 51 mm (H) 900 mm (W) 520 mm (D)"),
            "51 mm (H) x 900 mm (W) x 520 mm (D)",
        )

    def test_primary_model_token_and_direct_westinghouse_candidates(self) -> None:
        self.assertEqual(_primary_model_token("2 x WVE6515SDA"), "WVE6515SDA")
        self.assertEqual(
            _build_direct_product_candidates("Westinghouse", "Cooktop", "WHC943BD"),
            ["https://www.westinghouse.com.au/cooking/cooktops/whc943bd/"],
        )
        self.assertIn(
            "https://www.aegaustralia.com.au/cooking/cooktops/induction-cooktops/nik95i00fz/",
            _build_direct_product_candidates("AEG", "Cooktop", "NIK95I00FZ"),
        )

    def test_yellowwood_overlay_maps_back_benchtops_and_filters_non_cabinet_colours(self) -> None:
        documents = [
            {
                "file_name": "yellowwood.pdf",
                "role": "spec",
                "pages": [
                    {
                        "page_no": 20,
                        "text": (
                            "JOINERY - REFER TO CABINETRY PLANS FOR ALL FURTHER DETAIL\n"
                            "KITCHEN Back Benchtops 40mm\n"
                            "YDL Giusto Polished\n"
                            "Island\n"
                            "Benchtop 40mm +\n"
                            "Waterfall Ends\n"
                            "YDL Giusto Polished\n"
                            "Overhead Cupboards\n"
                            "Polytec Blossom White Matt\n"
                            "Base Cupboards &\n"
                            "Drawers\n"
                            "Polytec Blossom White Matt\n"
                            "Island Bar Back Polytec Topiary Matt\n"
                            "Island Bench\n"
                            "Base Cupboards &\n"
                            "Drawers\n"
                            "Polytec Topiary Matt\n"
                            "LAUNDRY\n"
                            "Benchtop 20mm\n"
                            "YDL Classic White Polished\n"
                            "Overhead Cupboards\n"
                            "Polytec Blossom White Matt\n"
                            "Base Cupboards & Drawers\n"
                            "Polytec Blossom White Matt\n"
                            "BATHROOM VANITY\n"
                            "Benchtop 20mm\n"
                            "YDL Classic White Polished\n"
                            "Floor Mounted Vanity\n"
                            "Polytec Nouveau Grey Matt\n"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 3,
                        "text": (
                            "EXTERNAL FINISHES\n"
                            "GUTTER Colorbond Dover White\n"
                            "ENTRY DOOR Painted Titanium White\n"
                        ),
                        "needs_ocr": False,
                    },
                ],
            }
        ]
        snapshot = {
            "job_no": "37014",
            "builder_name": "Yellowwood",
            "source_kind": "spec",
            "generated_at": "2026-03-22T10:00:00+00:00",
            "rooms": [
                {
                    "room_key": "kitchen",
                    "original_room_label": "KITCHEN",
                    "bench_tops": ["{'back': 'YDL Giusto Polished 40mm', 'island': 'YDL Giusto Polished 40mm + Waterfall Ends'}", "40mm +"],
                    "door_panel_colours": [],
                    "door_colours_overheads": "",
                    "door_colours_base": "",
                    "door_colours_island": "",
                    "door_colours_bar_back": "",
                    "handles": [],
                    "drawers_soft_close": "",
                    "hinges_soft_close": "",
                },
                {
                    "room_key": "laundry",
                    "original_room_label": "LAUNDRY",
                    "bench_tops": ["20mm YDL Classic White Polished"],
                    "door_panel_colours": ["Dover White"],
                    "door_colours_overheads": "",
                    "door_colours_base": "Dover White",
                    "door_colours_island": "",
                    "door_colours_bar_back": "",
                    "handles": [],
                    "drawers_soft_close": "",
                    "hinges_soft_close": "",
                },
                {
                    "room_key": "bathroom",
                    "original_room_label": "BATHROOM",
                    "bench_tops": ["20mm YDL Classic White Polished"],
                    "door_panel_colours": ["Painted Titanium White"],
                    "door_colours_overheads": "",
                    "door_colours_base": "Painted Titanium White",
                    "door_colours_island": "",
                    "door_colours_bar_back": "",
                    "handles": [],
                    "drawers_soft_close": "",
                    "hinges_soft_close": "",
                },
            ],
            "appliances": [],
            "others": {},
            "warnings": [],
            "source_documents": [],
            "analysis": {"mode": "heuristic_only"},
        }
        enriched = enrich_snapshot_rooms(snapshot, documents)
        rooms = {row["room_key"]: row for row in enriched["rooms"]}
        self.assertEqual(rooms["kitchen"]["bench_tops_wall_run"], "40mm YDL Giusto Polished")
        self.assertEqual(rooms["kitchen"]["bench_tops_island"], "40mm + Waterfall Ends YDL Giusto Polished")
        self.assertEqual(rooms["kitchen"]["door_colours_overheads"], "Polytec Blossom White Matt")
        self.assertEqual(rooms["kitchen"]["door_colours_base"], "Polytec Blossom White Matt")
        self.assertEqual(rooms["kitchen"]["door_colours_island"], "Polytec Topiary Matt")
        self.assertEqual(rooms["kitchen"]["door_colours_bar_back"], "Polytec Topiary Matt")
        self.assertEqual(rooms["laundry"]["door_colours_base"], "Polytec Blossom White Matt")
        self.assertNotIn("Dover White", rooms["laundry"]["door_panel_colours"])
        self.assertEqual(rooms["bathroom"]["door_colours_base"], "Polytec Nouveau Grey Matt")
        self.assertNotIn("Painted Titanium White", rooms["bathroom"]["door_panel_colours"])

    def test_collect_layout_sections_for_yellowwood_inherits_previous_room_for_continuation_pages(self) -> None:
        document = {
            "file_name": "yellowwood.pdf",
            "builder_name": "Yellowwood",
            "pages": [
                {
                    "page_no": 16,
                    "text": "KITCHEN\nBack Benchtops 20mm\nOverhead Cupboards",
                    "raw_text": "KITCHEN\nBack Benchtops 20mm\nOverhead Cupboards",
                    "page_layout": {
                        "page_type": "joinery",
                        "section_label": "KITCHEN",
                        "room_label": "KITCHEN",
                        "room_blocks": [
                            {
                                "room_label": "KITCHEN",
                                "rows": [
                                    {
                                        "row_label": "Back Benchtops",
                                        "value_region_text": "20mm",
                                        "supplier_region_text": "",
                                        "notes_region_text": "",
                                        "row_kind": "material",
                                    }
                                ],
                            }
                        ],
                    },
                },
                {
                    "page_no": 17,
                    "text": "Base Cupboards & Drawers\nPantry X4 Shelves White Melamine\nexcluding Pantry)",
                    "raw_text": "Base Cupboards & Drawers\nPantry X4 Shelves White Melamine\nexcluding Pantry)",
                    "page_layout": {
                        "page_type": "joinery",
                        "section_label": "Pantry X4 Shelves White Melamine",
                        "room_label": "Pantry X4 Shelves White Melamine",
                        "room_blocks": [
                            {
                                "room_label": "Pantry X4 Shelves White Melamine",
                                "rows": [
                                    {
                                        "row_label": "Base Cupboards & Drawers",
                                        "value_region_text": "Polytec Classic White Matt",
                                        "supplier_region_text": "",
                                        "notes_region_text": "",
                                        "row_kind": "material",
                                    }
                                ],
                            }
                        ],
                    },
                },
            ],
        }
        sections = parsing_module._collect_layout_sections_for_document(document)
        labels = [section["original_section_label"] for section in sections]
        self.assertEqual(labels, ["KITCHEN", "KITCHEN"])
        self.assertEqual([section["page_nos"] for section in sections], [[16], [17]])

    def test_collect_spec_sections_for_yellowwood_uses_builder_name_without_document_builder_field(self) -> None:
        document = {
            "file_name": "yellowwood.pdf",
            "pages": [
                {
                    "page_no": 16,
                    "text": "KITCHEN\nBack Benchtops 20mm\nYDL Classic White Polished",
                    "raw_text": "KITCHEN\nBack Benchtops 20mm\nYDL Classic White Polished",
                },
                {
                    "page_no": 17,
                    "text": "Base Cupboards & Drawers\nPolytec Classic White Matt\nKickboard Polytec Classic White Matt",
                    "raw_text": "Base Cupboards & Drawers\nPolytec Classic White Matt\nKickboard Polytec Classic White Matt",
                },
            ],
        }
        sections = parsing_module._collect_spec_sections_for_document("Yellowwood Homes", document)
        kitchen_sections = [section for section in sections if section["section_key"] == "kitchen"]
        self.assertEqual([section["page_nos"] for section in kitchen_sections], [[16], [17]])

    def test_flatten_rooms_only_shows_split_benchtops_for_kitchen(self) -> None:
        rows = _flatten_rooms(
            {
                "rooms": [
                    {
                        "room_key": "kitchen",
                        "original_room_label": "KITCHEN",
                        "bench_tops": [],
                        "bench_tops_wall_run": "40mm YDL Giusto Polished",
                        "bench_tops_island": "40mm + Waterfall Ends YDL Giusto Polished",
                        "bench_tops_other": "",
                    },
                    {
                        "room_key": "laundry",
                        "original_room_label": "LAUNDRY",
                        "bench_tops": ["20mm YDL Classic White Polished"],
                        "bench_tops_wall_run": "20mm YDL Classic White Polished",
                        "bench_tops_island": "",
                        "bench_tops_other": "20mm YDL Classic White Polished",
                    },
                ]
            }
        )
        flattened = {row["room_key"]: row for row in rows}
        self.assertTrue(flattened["kitchen"]["show_split_benchtops"])
        self.assertFalse(flattened["laundry"]["show_split_benchtops"])
        self.assertEqual(flattened["laundry"]["bench_tops_other"], "20mm YDL Classic White Polished")

    def test_build_spec_snapshot_keeps_global_conservative_room_shape_when_openai_splits_rooms(self) -> None:
        base_snapshot = {
            "job_no": "37529",
            "builder_name": "Clarendon",
            "source_kind": "spec",
            "generated_at": "2026-03-22T10:00:00+00:00",
            "rooms": [
                {
                    "room_key": "kitchen",
                    "original_room_label": "Kitchen",
                    "bench_tops": ["20mm stone"],
                    "door_panel_colours": ["Polytec White"],
                    "toe_kick": [],
                    "bulkheads": [],
                    "handles": [],
                    "drawers_soft_close": "",
                    "hinges_soft_close": "",
                    "splashback": "",
                    "flooring": "",
                    "source_file": "sample.pdf",
                    "page_refs": "1",
                    "evidence_snippet": "Kitchen line",
                    "confidence": 0.7,
                },
                {
                    "room_key": "vanities",
                    "original_room_label": "Vanities",
                    "bench_tops": ["20mm stone"],
                    "door_panel_colours": ["Polytec White"],
                    "toe_kick": [],
                    "bulkheads": [],
                    "handles": [],
                    "drawers_soft_close": "",
                    "hinges_soft_close": "",
                    "splashback": "",
                    "flooring": "",
                    "source_file": "sample.pdf",
                    "page_refs": "2",
                    "evidence_snippet": "Vanities line",
                    "confidence": 0.7,
                },
            ],
            "appliances": [],
            "others": {"flooring_notes": "", "splashback_notes": ""},
            "warnings": [],
            "source_documents": [],
            "analysis": {"mode": "heuristic_only"},
        }
        openai_payload = {
            "output_text": (
                '{"rooms": ['
                '{"room_key": "kitchen", "original_room_label": "Kitchen", "bench_tops": ["AI bench"]},'
                '{"room_key": "main_bathroom", "original_room_label": "Main Bathroom", "bench_tops": ["AI vanity bench"]},'
                '{"room_key": "ensuite_1", "original_room_label": "Ensuite 1", "bench_tops": ["AI vanity bench"]}'
                '], "appliances": [], "others": {}, "warnings": []}'
            )
        }
        with (
            mock.patch.object(extraction_service.runtime, "OPENAI_ENABLED", True),
            mock.patch.object(extraction_service.runtime, "OPENAI_API_KEY", "test-key"),
            mock.patch.object(extraction_service.runtime, "OPENAI_MODEL", "gpt-4.1-mini"),
            mock.patch("App.services.extraction_service._load_documents", return_value=[]),
            mock.patch("App.services.extraction_service.parsing.parse_documents", return_value=base_snapshot),
            mock.patch(
                "App.services.extraction_service.parsing.enrich_snapshot_rooms",
                side_effect=lambda payload, _documents, rule_flags=None: payload,
            ),
            mock.patch("App.services.extraction_service._post_responses_api", return_value=openai_payload),
        ):
            snapshot = extraction_service.build_spec_snapshot(
                job={"job_no": "37529"},
                builder={"name": "Clarendon", "parser_strategy": "stable_hybrid"},
                files=[],
                template_files=[],
            )
        self.assertEqual([row["room_key"] for row in snapshot["rooms"]], ["kitchen", "vanities"])
        self.assertEqual(snapshot["analysis"]["parser_strategy"], "global_conservative")
        self.assertEqual(snapshot["rooms"][0]["bench_tops"], ["20MM Stone"])

    def test_build_spec_snapshot_keeps_source_driven_rooms_under_global_conservative(self) -> None:
        base_snapshot = {
            "job_no": "37017",
            "builder_name": "Clarendon",
            "source_kind": "spec",
            "generated_at": "2026-03-22T10:00:00+00:00",
            "rooms": [
                {"room_key": "powder", "original_room_label": "Powder", "bench_tops": [], "door_panel_colours": [], "toe_kick": [], "bulkheads": [], "handles": [], "sink_info": "", "basin_info": "", "tap_info": "", "drawers_soft_close": "", "hinges_soft_close": "", "splashback": "", "flooring": "", "source_file": "sample.pdf", "page_refs": "1", "evidence_snippet": "", "confidence": 0.4},
                {"room_key": "bathroom", "original_room_label": "Bathroom", "bench_tops": [], "door_panel_colours": [], "toe_kick": [], "bulkheads": [], "handles": [], "sink_info": "", "basin_info": "Bathroom basin", "tap_info": "", "drawers_soft_close": "", "hinges_soft_close": "", "splashback": "", "flooring": "", "source_file": "sample.pdf", "page_refs": "2", "evidence_snippet": "", "confidence": 0.4},
                {"room_key": "ensuite", "original_room_label": "Ensuite", "bench_tops": [], "door_panel_colours": [], "toe_kick": [], "bulkheads": [], "handles": [], "sink_info": "", "basin_info": "", "tap_info": "Ensuite tap", "drawers_soft_close": "", "hinges_soft_close": "", "splashback": "", "flooring": "", "source_file": "sample.pdf", "page_refs": "3", "evidence_snippet": "", "confidence": 0.4},
                {"room_key": "vanity", "original_room_label": "Vanity", "bench_tops": ["20mm stone"], "door_panel_colours": ["Polytec White"], "toe_kick": [], "bulkheads": [], "handles": [], "sink_info": "", "basin_info": "Primary vanity basin", "tap_info": "Primary vanity tap", "drawers_soft_close": "", "hinges_soft_close": "", "splashback": "", "flooring": "", "source_file": "sample.pdf", "page_refs": "4", "evidence_snippet": "", "confidence": 0.7},
                {"room_key": "laundry", "original_room_label": "Laundry", "bench_tops": ["Laminate"], "door_panel_colours": [], "toe_kick": [], "bulkheads": [], "handles": [], "sink_info": "", "basin_info": "", "tap_info": "", "drawers_soft_close": "", "hinges_soft_close": "", "splashback": "", "flooring": "", "source_file": "sample.pdf", "page_refs": "5", "evidence_snippet": "", "confidence": 0.6},
                {"room_key": "butlers_pantry", "original_room_label": "Butler's Pantry", "bench_tops": ["Stone"], "door_panel_colours": [], "toe_kick": [], "bulkheads": [], "handles": [], "sink_info": "", "basin_info": "", "tap_info": "", "drawers_soft_close": "", "hinges_soft_close": "", "splashback": "", "flooring": "", "source_file": "sample.pdf", "page_refs": "6", "evidence_snippet": "", "confidence": 0.6},
                {"room_key": "wip", "original_room_label": "WIP", "bench_tops": [], "door_panel_colours": [], "toe_kick": [], "bulkheads": [], "handles": [], "sink_info": "", "basin_info": "", "tap_info": "", "drawers_soft_close": "", "hinges_soft_close": "", "splashback": "", "flooring": "", "source_file": "sample.pdf", "page_refs": "7", "evidence_snippet": "", "confidence": 0.2},
                {"room_key": "theatre", "original_room_label": "Theatre", "bench_tops": ["Laminate"], "door_panel_colours": [], "toe_kick": [], "bulkheads": [], "handles": [], "sink_info": "", "basin_info": "", "tap_info": "", "drawers_soft_close": "", "hinges_soft_close": "", "splashback": "", "flooring": "", "source_file": "sample.pdf", "page_refs": "8", "evidence_snippet": "", "confidence": 0.6},
                {"room_key": "rumpus", "original_room_label": "Rumpus", "bench_tops": ["Laminate"], "door_panel_colours": [], "toe_kick": [], "bulkheads": [], "handles": [], "sink_info": "", "basin_info": "", "tap_info": "", "drawers_soft_close": "", "hinges_soft_close": "", "splashback": "", "flooring": "", "source_file": "sample.pdf", "page_refs": "9", "evidence_snippet": "", "confidence": 0.6},
                {"room_key": "kitchen", "original_room_label": "Kitchen", "bench_tops": ["Stone"], "door_panel_colours": [], "toe_kick": [], "bulkheads": [], "handles": [], "sink_info": "", "basin_info": "", "tap_info": "", "drawers_soft_close": "", "hinges_soft_close": "", "splashback": "", "flooring": "", "source_file": "sample.pdf", "page_refs": "10", "evidence_snippet": "", "confidence": 0.8},
                {"room_key": "pantry", "original_room_label": "Pantry", "bench_tops": [], "door_panel_colours": [], "toe_kick": [], "bulkheads": [], "handles": [], "sink_info": "", "basin_info": "", "tap_info": "", "drawers_soft_close": "", "hinges_soft_close": "", "splashback": "", "flooring": "", "source_file": "sample.pdf", "page_refs": "11", "evidence_snippet": "", "confidence": 0.2},
                {"room_key": "wir", "original_room_label": "WIR", "bench_tops": [], "door_panel_colours": [], "toe_kick": [], "bulkheads": [], "handles": [], "sink_info": "", "basin_info": "", "tap_info": "", "drawers_soft_close": "", "hinges_soft_close": "", "splashback": "", "flooring": "Carpet", "source_file": "sample.pdf", "page_refs": "12", "evidence_snippet": "", "confidence": 0.2},
            ],
            "appliances": [],
            "others": {},
            "warnings": [],
            "source_documents": [],
            "analysis": {"mode": "heuristic_only"},
        }
        with (
            mock.patch.object(extraction_service.runtime, "OPENAI_ENABLED", False),
            mock.patch("App.services.extraction_service._load_documents", return_value=[]),
            mock.patch("App.services.extraction_service.parsing.parse_documents", return_value=base_snapshot),
            mock.patch(
                "App.services.extraction_service.parsing.enrich_snapshot_rooms",
                side_effect=lambda payload, _documents, rule_flags=None: payload,
            ),
        ):
            snapshot = extraction_service.build_spec_snapshot(
                job={"job_no": "37017"},
                builder={"name": "Clarendon", "parser_strategy": "stable_hybrid"},
                files=[],
                template_files=[],
            )
        self.assertEqual(
            [row["room_key"] for row in snapshot["rooms"]],
            ["powder", "bathroom", "ensuite", "vanity", "laundry", "butlers_pantry", "walk_in_pantry", "theatre", "rumpus", "kitchen", "pantry", "wir"],
        )
        bathroom = next(row for row in snapshot["rooms"] if row["room_key"] == "bathroom")
        vanity = next(row for row in snapshot["rooms"] if row["room_key"] == "vanity")
        self.assertEqual(bathroom["basin_info"], "Bathroom Basin")
        self.assertEqual(vanity["bench_tops"], ["20MM Stone"])
        self.assertEqual(vanity["basin_info"], "Primary Vanity Basin")

    def test_parse_documents_keeps_bathroom_ensuite_and_powder_separate(self) -> None:
        snapshot = parse_documents(
            job_no="39001",
            builder_name="Clarendon",
            source_kind="spec",
            documents=[
                {
                    "file_name": "rooms.txt",
                    "role": "spec",
                    "pages": [
                        {
                            "page_no": 1,
                            "text": (
                                "Main Bathroom\n"
                                "Bench Tops 20mm stone\n"
                                "Ensuite 1\n"
                                "Bench Tops 30mm stone\n"
                                "Powder Room 3\n"
                                "Bench Tops 40mm stone\n"
                            ),
                            "needs_ocr": False,
                        }
                    ],
                }
            ],
        )
        rooms = {row["room_key"]: row for row in snapshot["rooms"]}
        self.assertIn("main_bathroom", rooms)
        self.assertIn("ensuite_1", rooms)
        self.assertIn("powder_room_3", rooms)
        self.assertEqual(rooms["main_bathroom"]["bench_tops"], ["20mm stone"])
        self.assertEqual(rooms["ensuite_1"]["bench_tops"], ["30mm stone"])
        self.assertEqual(rooms["powder_room_3"]["bench_tops"], ["40mm stone"])

    def test_source_driven_room_matching_does_not_bleed_vanity_overlay_between_rooms(self) -> None:
        snapshot = {
            "job_no": "39002",
            "builder_name": "Clarendon",
            "source_kind": "spec",
            "generated_at": "2026-03-22T10:00:00+00:00",
            "rooms": [
                {
                    "room_key": "main_bathroom",
                    "original_room_label": "Main Bathroom",
                    "bench_tops": ["20mm stone"],
                    "door_panel_colours": [],
                    "toe_kick": [],
                    "bulkheads": [],
                    "handles": [],
                    "drawers_soft_close": "",
                    "hinges_soft_close": "",
                    "splashback": "",
                    "flooring": "",
                    "sink_info": "",
                    "basin_info": "Bathroom basin",
                    "tap_info": "",
                    "source_file": "sample.pdf",
                    "page_refs": "1",
                    "evidence_snippet": "",
                    "confidence": 0.7,
                },
                {
                    "room_key": "powder_room_3",
                    "original_room_label": "Powder Room 3",
                    "bench_tops": ["30mm stone"],
                    "door_panel_colours": [],
                    "toe_kick": [],
                    "bulkheads": [],
                    "handles": [],
                    "drawers_soft_close": "",
                    "hinges_soft_close": "",
                    "splashback": "",
                    "flooring": "",
                    "sink_info": "",
                    "basin_info": "",
                    "tap_info": "Powder tap",
                    "source_file": "sample.pdf",
                    "page_refs": "2",
                    "evidence_snippet": "",
                    "confidence": 0.7,
                },
            ],
            "appliances": [],
            "others": {},
            "warnings": [],
            "source_documents": [],
            "analysis": {"mode": "heuristic_only"},
        }
        with (
            mock.patch.object(extraction_service.runtime, "OPENAI_ENABLED", False),
            mock.patch("App.services.extraction_service._load_documents", return_value=[]),
            mock.patch("App.services.extraction_service.parsing.parse_documents", return_value=snapshot),
            mock.patch(
                "App.services.extraction_service.parsing.enrich_snapshot_rooms",
                side_effect=lambda payload, _documents, rule_flags=None: payload,
            ),
        ):
            result = extraction_service.build_spec_snapshot(
                job={"job_no": "39002"},
                builder={"name": "Clarendon"},
                files=[],
                template_files=[],
            )
        rooms = {row["room_key"]: row for row in result["rooms"]}
        self.assertIn("main_bathroom", rooms)
        self.assertIn("powder_room_3", rooms)
        self.assertEqual(rooms["main_bathroom"]["basin_info"], "Bathroom Basin")
        self.assertEqual(rooms["powder_room_3"]["tap_info"], "Powder Tap")

    def test_parse_documents_uses_room_master_file_for_multifile_clarendon(self) -> None:
        snapshot = parse_documents(
            job_no="37868",
            builder_name="Clarendon",
            source_kind="spec",
            documents=[
                {
                    "file_name": "drawings-and-colours.pdf",
                    "role": "spec",
                    "pages": [
                        {
                            "page_no": 1,
                            "text": (
                                "KITCHEN COLOUR SCHEDULE\n"
                                "Bench Tops Quantum Zero Bella Carrara - 20MM Pencil Round Edge\n"
                                "VANITIES COLOUR SCHEDULE\n"
                                "Bench Tops Quantum Zero Luna White - 20MM Pencil Round Edge\n"
                            ),
                            "needs_ocr": False,
                        }
                    ],
                },
                {
                    "file_name": "colours-afc.pdf",
                    "role": "spec",
                    "pages": [
                        {
                            "page_no": 1,
                            "text": (
                                "Main Bathroom\n"
                                "Vanity Inset Basin JOHNSON SUISSE Emilia Basin (JBSE250.PW6)\n"
                                "Laundry Door Glazing: CLEAR GLAZING\n"
                                "Vanity Waste Colour: CHROME POP UP\n"
                                "Powder Room 3: TRANSLUCENT LAMINATE P/C Windows & Doors\n"
                            ),
                            "needs_ocr": False,
                        }
                    ],
                },
            ],
        )
        rooms = {row["room_key"]: row for row in snapshot["rooms"]}
        analysis = snapshot["analysis"]
        warning_text = " | ".join(snapshot["warnings"])
        self.assertEqual(set(rooms.keys()), {"kitchen", "vanities"})
        self.assertEqual(analysis["room_master_file"], "drawings-and-colours.pdf")
        self.assertIn("Drawings and Colours filename match", analysis["room_master_reason"])
        self.assertEqual(analysis["supplement_files"], ["colours-afc.pdf"])
        self.assertGreaterEqual(analysis["ignored_room_like_lines_count"], 1)
        self.assertTrue(str(rooms["vanities"]["basin_info"]).startswith("Johnson Suisse Emilia"))
        self.assertIn("Ignored room-like section", warning_text)
        self.assertNotIn("Laundry Door", " ".join(rooms.keys()))
        self.assertNotIn("powder_room_3", rooms)

    def test_parse_documents_clarendon_drawings_master_recovers_glued_room_titles(self) -> None:
        snapshot = parse_documents(
            job_no="37031",
            builder_name="Clarendon",
            source_kind="spec",
            documents=[
                {
                    "file_name": "49906511 AMENDED Drawings and Colours REV B 02-09-25.pdf",
                    "role": "spec",
                    "pages": [
                        {
                            "page_no": 1,
                            "text": (
                                "KITCHEN COLOUR SCHEDULE\n"
                                "BUTLERS PANTRY COLOUR SCHEDULE\n"
                                "VanitiesDate: 02-09-25 VANITIES COLOUR SCHEDULE\n"
                                "LaundryDate: 02-09-25 LAUNDRY COLOUR SCHEDULE\n"
                                "TheatreDate: 02-09-25 THEATRE ROOM COLOUR SCHEDULE\n"
                            ),
                            "needs_ocr": False,
                        }
                    ],
                },
                {
                    "file_name": "49906511 COLOURS AFC .pdf",
                    "role": "spec",
                    "pages": [
                        {
                            "page_no": 1,
                            "text": (
                                "Main Bathroom\n"
                                "Vanity Basin Johnson Suisse Emilia Basin\n"
                                "Ensuite\n"
                                "Vanity Tap Phoenix Mixer\n"
                                "Powder Room 3\n"
                                "Door/Panel Colour Blossom White\n"
                            ),
                            "needs_ocr": False,
                        }
                    ],
                },
            ],
        )
        rooms = {row["room_key"]: row for row in snapshot["rooms"]}
        labels = {row["original_room_label"] for row in snapshot["rooms"]}
        self.assertEqual(
            set(rooms.keys()),
            {"kitchen", "butlers_pantry", "vanities", "laundry", "theatre"},
        )
        self.assertIn("VANITIES", labels)
        self.assertIn("LAUNDRY", labels)
        self.assertTrue(any("THEATRE" in label for label in labels))
        self.assertNotIn("main_bathroom", rooms)
        self.assertNotIn("ensuite", rooms)
        self.assertNotIn("powder_room_3", rooms)

    def test_collect_spec_sections_for_clarendon_prefers_raw_titles_over_layout_schema(self) -> None:
        document = {
            "file_name": "49906511 AMENDED Drawings and Colours REV B 02-09-25.pdf",
            "builder_name": "Clarendon",
            "pages": [
                {
                    "page_no": 1,
                    "text": (
                        "KITCHEN COLOUR SCHEDULE\n"
                        "VanitiesDate: 02-09-25 VANITIES COLOUR SCHEDULE\n"
                        "LaundryDate: 02-09-25 LAUNDRY COLOUR SCHEDULE\n"
                    ),
                    "raw_text": (
                        "KITCHEN COLOUR SCHEDULE\n"
                        "VanitiesDate: 02-09-25 VANITIES COLOUR SCHEDULE\n"
                        "LaundryDate: 02-09-25 LAUNDRY COLOUR SCHEDULE\n"
                    ),
                    "page_layout": {
                        "page_type": "joinery",
                        "section_label": "KITCHEN COLOUR SCHEDULE",
                        "room_label": "KITCHEN",
                        "room_blocks": [
                            {
                                "room_label": "KITCHEN",
                                "rows": [
                                    {
                                        "row_label": "Bench Tops",
                                        "value_region_text": "Quantum Zero Bella Carrara",
                                        "supplier_region_text": "",
                                        "notes_region_text": "",
                                        "row_kind": "material",
                                    }
                                ],
                            }
                        ],
                    },
                }
            ],
        }
        sections = parsing_module._collect_spec_sections_for_document("Clarendon", document)
        labels = {section["original_section_label"] for section in sections}
        self.assertIn("KITCHEN", labels)
        self.assertIn("VANITIES", labels)
        self.assertIn("LAUNDRY", labels)

    def test_parse_documents_prefers_colour_schedule_file_as_room_master_for_multifile_clarendon(self) -> None:
        snapshot = parse_documents(
            job_no="49906613",
            builder_name="Clarendon",
            source_kind="spec",
            documents=[
                {
                    "file_name": "49906613 Amended Signed Drawings REV C 04-09-25.pdf",
                    "role": "spec",
                    "pages": [
                        {
                            "page_no": 1,
                            "text": (
                                "KITCHEN COLOUR SCHEDULE\n"
                                "VANITIES COLOUR SCHEDULE\n"
                                "LAUNDRY COLOUR SCHEDULE\n"
                                "BUTLERS PANTRY COLOUR SCHEDULE\n"
                            ),
                            "needs_ocr": False,
                        }
                    ],
                },
                {
                    "file_name": "49906613 COLOURS AFC AMENDED .pdf",
                    "role": "spec",
                    "pages": [
                        {
                            "page_no": 6,
                            "text": (
                                "VANITIES COLOUR SCHEDULE\n"
                                "Door/Panel Colour - Polytec Blossom White Matt Finish Thermolaminate - Hamptons EM9 Profile\n"
                                "Door/Panel Colour 2 - Polytec Habitit Smooth Finish Thermolaminate - Hamptons EM9 Profile\n"
                                "Floor Mounted Vanity - Polytec Habitit Smooth Finish Thermolaminate - Hamptons EM9 Profile\n"
                                "Back Benchtops Quantum Zero Luna White - 20MM Pencil Round Edge\n"
                            ),
                            "needs_ocr": False,
                        }
                    ],
                },
            ],
        )
        analysis = snapshot["analysis"]
        vanities = next(room for room in snapshot["rooms"] if room["room_key"] == "vanities")
        self.assertEqual(analysis["room_master_file"], "49906613 COLOURS AFC AMENDED .pdf")
        self.assertEqual(vanities["source_file"], "49906613 COLOURS AFC AMENDED .pdf")
        self.assertEqual(vanities["page_refs"], "6")

    def test_parse_documents_keeps_clarendon_wet_area_rooms_from_afc_supplement(self) -> None:
        snapshot = parse_documents(
            job_no="49906622",
            builder_name="Clarendon",
            source_kind="spec",
            documents=[
                {
                    "file_name": "49906622 AMENDED Signed Drawings and Colours REV B 20-10-25.pdf",
                    "role": "spec",
                    "pages": [
                        {
                            "page_no": 1,
                            "text": (
                                "KITCHEN COLOUR SCHEDULE\n"
                                "VANITIES COLOUR SCHEDULE\n"
                                "Forstan Pty Ltd. T/A YourHome Kitchens P.O. Box 8248 Baulkham HillsBCNSW 2153 Phone : 02 8824 7771\n"
                            ),
                            "needs_ocr": False,
                        }
                    ],
                },
                {
                    "file_name": "49906622 - COLOURS - AFC.pdf",
                    "role": "spec",
                    "pages": [
                        {
                            "page_no": 1,
                            "text": (
                                "Main Bathroom\n"
                                "Vanity Inset Basin JOHNSON SUISSE Emilia Basin (JBSE250.PW6)\n"
                                "WC\n"
                                "Toilet Suite JOHNSON SUISSE Venezia Comfort\n"
                                "Ensuite\n"
                                "Vanity Tap Style PHOENIX PINA WALL BASIN/BATH MIXER SET 180MM\n"
                            ),
                            "needs_ocr": False,
                        }
                    ],
                },
            ],
        )
        room_keys = {row["room_key"] for row in snapshot["rooms"]}
        self.assertIn("kitchen", room_keys)
        self.assertIn("vanities", room_keys)
        vanities = next(row for row in snapshot["rooms"] if row["room_key"] == "vanities")
        self.assertIn("Johnson Suisse Emilia Basin", vanities["basin_info"])
        self.assertIn("Phoenix PINA", vanities["tap_info"])
        self.assertNotIn("main_bathroom", room_keys)
        self.assertNotIn("ensuite", room_keys)
        self.assertNotIn("forstan_pty_ltd_t_a_yourhome_kitchens_p_o_box_8248_baulkham_hillsbcnsw_2153_phone", room_keys)

    def test_source_room_label_recovers_glued_clarendon_schedule_headers(self) -> None:
        self.assertEqual(parsing_module.source_room_label("VanitiesDate: 02-09-25 VANITIES COLOUR SCHEDULE"), "VANITIES")
        self.assertEqual(parsing_module.source_room_label("LaundryDate: 02-09-25 LAUNDRY COLOUR SCHEDULE"), "LAUNDRY")
        self.assertEqual(parsing_module.source_room_key("TheatreDate: 02-09-25 THEATRE ROOM COLOUR SCHEDULE"), "theatre")
        self.assertEqual(parsing_module.source_room_label("KITCHEN SPLASHBACK"), "KITCHEN")
        self.assertEqual(parsing_module.source_room_label("BED 1 ENSUITE VANITY"), "BED 1 ENSUITE VANITY")
        self.assertEqual(parsing_module.source_room_label("BATHOOM VANITY").lower(), "bathroom vanity")
        self.assertEqual(parsing_module.source_room_label("BED 1 WALK IN ROBE FIT OUT"), "BED 1 WALK IN ROBE")
        self.assertEqual(parsing_module.source_room_label("BED 3 ROBE FIT OUT"), "BED 3 ROBE")
        self.assertEqual(parsing_module.source_room_label("ROBE FIT OUT TO BED 3"), "BED 3 ROBE")
        self.assertEqual(parsing_module.source_room_key("BATHOOM VANITY"), "bathroom")
        self.assertEqual(parsing_module.source_room_key("BED 1 WALK IN ROBE FIT OUT"), "bed_1_wir")
        self.assertEqual(parsing_module.source_room_key("BED 3 ROBE FIT OUT"), "bed_3_robe")

    def test_yellowwood_page_room_hint_ignores_fake_room_labels_and_keeps_real_titles(self) -> None:
        self.assertEqual(
            parsing_module._yellowwood_page_room_hint(
                "BASE CUPBOARDS & DRAWERS\nPANTRY X4 SHELVES WHITE MELAMINE\nexcluding Pantry)"
            ),
            "",
        )
        self.assertEqual(
            parsing_module._yellowwood_page_room_hint(
                "BED 1 ENSUITE VANITY\nFLOOR MOUNTED VANITY\nDRAWERS"
            ),
            "BED 1 ENSUITE VANITY",
        )
        self.assertEqual(
            parsing_module._yellowwood_page_room_hint(
                "KITCHEN SPLASHBACK\nSUBWAY TILE"
            ),
            "KITCHEN",
        )

    def test_collect_spec_sections_for_yellowwood_uses_raw_text_split_for_multi_room_pages(self) -> None:
        document = {
            "file_name": "yellowwood.pdf",
            "builder_name": "Yellowwood",
            "pages": [
                {
                    "page_no": 18,
                    "text": (
                        "Pantry Door Handles  C137 Caloundra Lip Pull\n"
                        "Matt Black 200mm\n"
                        "Handle House\n"
                        "LAUNDRY\n"
                        "Freestanding Laundry Cabinet\n"
                        "WC\n"
                        "Wall Hung Basin Only Refer to Plumbing section below N/A\n"
                        "BED 1 ENSUITE VANITY\n"
                        "Benchtop 20mm\n"
                        "YDL Classic White Polished\n"
                    ),
                    "raw_text": (
                        "Pantry Door Handles  C137 Caloundra Lip Pull\n"
                        "Matt Black 200mm\n"
                        "Handle House\n"
                        "LAUNDRY\n"
                        "Freestanding Laundry Cabinet\n"
                        "WC\n"
                        "Wall Hung Basin Only Refer to Plumbing section below N/A\n"
                        "BED 1 ENSUITE VANITY\n"
                        "Benchtop 20mm\n"
                        "YDL Classic White Polished\n"
                    ),
                    "page_layout": {
                        "page_type": "joinery",
                        "section_label": "Pantry Door Handles",
                        "room_label": "Pantry Door Handles",
                        "room_blocks": [
                            {
                                "room_label": "Pantry Door Handles",
                                "rows": [
                                    {
                                        "row_label": "Pantry Door Handles",
                                        "value_region_text": "C137 Caloundra Lip Pull",
                                        "supplier_region_text": "Handle House",
                                        "notes_region_text": "",
                                        "row_kind": "handle",
                                    }
                                ],
                            }
                        ],
                    },
                }
            ],
        }
        sections = parsing_module._collect_spec_sections_for_document("Yellowwood", document)
        labels = {section["original_section_label"] for section in sections}
        self.assertNotIn("Pantry", labels)
        self.assertIn("Laundry", labels)
        self.assertIn("WC", labels)
        self.assertIn("BED 1 ENSUITE VANITY", labels)

    def test_yellowwood_section_filter_drops_material_name_rooms_but_keeps_material_driven_rooms(self) -> None:
        fake_material = {
            "section_key": "revival_hoya_powder_matt_200x200mm",
            "original_section_label": "Revival Hoya Powder Matt 200x200mm",
            "text": "Revival Hoya Powder Matt 200x200mm\nNational Tiles",
        }
        robe_room = {
            "section_key": "bed_1_wir",
            "original_section_label": "BED 1 + WIR",
            "text": "BED 1 + WIR\nCarpet Mapelton Falls\nColour: Elanus Grey Cut Pile\nFlooring Xtra",
        }
        media_room = {
            "section_key": "media_room",
            "original_section_label": "MEDIA ROOM",
            "text": "MEDIA ROOM\nCarpet Mapelton Falls\nColour: Elanus Grey Cut Pile\nFlooring Xtra",
        }
        wir_room = {
            "section_key": "bed_1_wir",
            "original_section_label": "BED 1 WALK IN ROBE",
            "text": "BED 1 WALK IN ROBE\nFIT OUT\nBoard Colour: Standard White\nAs supplied by builder",
        }
        self.assertFalse(parsing_module._yellowwood_should_keep_section(fake_material))
        self.assertFalse(parsing_module._yellowwood_should_keep_section(robe_room))
        self.assertFalse(parsing_module._yellowwood_should_keep_section(media_room))
        self.assertTrue(parsing_module._yellowwood_should_keep_section(wir_room))

    def test_yellowwood_section_filter_drops_contents_page_noise(self) -> None:
        contents_like = {
            "section_key": "ensuite",
            "original_section_label": "Ensuite",
            "text": (
                "ENSUITE\nOTHER THAN TILING TO WET AREAS 15-16 "
                "JOINERY - REFER TO CABINETRY PLANS 16-20 "
                "TILING SCHEDULE 21-25 EXTERNAL PAINTING REFERENCE 26 "
                "BATHWARE & FIXTURES 27-32 APPLIANCES 32-33"
            ),
        }
        self.assertFalse(parsing_module._yellowwood_should_keep_section(contents_like))

    def test_collect_yellowwood_text_room_sections_inherits_previous_room_on_continuation_page(self) -> None:
        document = {
            "file_name": "yellowwood.pdf",
            "pages": [
                {
                    "page_no": 16,
                    "raw_text": (
                        "INTERNAL FINISHES\nJOINERY - REFER TO CABINETRY PLANS FOR ALL FURTHER DETAIL\n"
                        "KITCHEN  Back Benchtops 20mm\nYDL Classic White Polished\nYellowwood supplier\n"
                    ),
                    "text": (
                        "INTERNAL FINISHES\nJOINERY - REFER TO CABINETRY PLANS FOR ALL FURTHER DETAIL\n"
                        "KITCHEN  Back Benchtops 20mm\nYDL Classic White Polished\nYellowwood supplier\n"
                    ),
                },
                {
                    "page_no": 17,
                    "raw_text": (
                        "Base Cupboards & Drawers\nPolytec Classic White Matt\nAs supplied by cabinetmaker\n"
                        "Kickboard Polytec Classic White Matt\nAs supplied by cabinetmaker\n"
                    ),
                    "text": (
                        "Base Cupboards & Drawers\nPolytec Classic White Matt\nAs supplied by cabinetmaker\n"
                        "Kickboard Polytec Classic White Matt\nAs supplied by cabinetmaker\n"
                    ),
                },
            ],
        }
        sections = parsing_module._collect_yellowwood_text_room_sections_for_document(document)
        kitchen_sections = [section for section in sections if section["section_key"] == "kitchen"]
        self.assertEqual([section["page_nos"] for section in kitchen_sections], [[16], [17]])
        self.assertIn("Base Cupboards & Drawers", kitchen_sections[1]["text"])

    def test_collect_yellowwood_text_room_sections_attaches_preamble_to_previous_kitchen_room(self) -> None:
        document = {
            "file_name": "yellowwood.pdf",
            "pages": [
                {
                    "page_no": 17,
                    "raw_text": "KITCHEN\nBase Cupboards & Drawers\nPolytec Classic White Matt\n",
                    "text": "KITCHEN\nBase Cupboards & Drawers\nPolytec Classic White Matt\n",
                },
                {
                    "page_no": 18,
                    "raw_text": (
                        "INTERNAL FINISHES\n"
                        "JOINERY - REFER TO CABINETRY PLANS FOR ALL FURTHER DETAIL\n"
                        "Pantry Door Handles C137 Caloundra Lip Pull Matt Black 200mm\n"
                        "Handle House\n"
                        "LAUNDRY\n"
                        "Freestanding Laundry Cabinet Kitset Only Refer to Plumbing section below N/A\n"
                    ),
                    "text": (
                        "INTERNAL FINISHES\n"
                        "JOINERY - REFER TO CABINETRY PLANS FOR ALL FURTHER DETAIL\n"
                        "Pantry Door Handles C137 Caloundra Lip Pull Matt Black 200mm\n"
                        "Handle House\n"
                        "LAUNDRY\n"
                        "Freestanding Laundry Cabinet Kitset Only Refer to Plumbing section below N/A\n"
                    ),
                },
            ],
        }
        sections = parsing_module._collect_yellowwood_text_room_sections_for_document(document)
        labels = {section["original_section_label"] for section in sections}
        kitchen_section = next(section for section in sections if section["section_key"] == "kitchen" and section["page_nos"] == [18])
        self.assertIn("Pantry Door", kitchen_section["text"])
        self.assertIn("Handles C137 Caloundra Lip Pull Matt Black 200mm", kitchen_section["text"])
        self.assertNotIn("Pantry", labels)

    def test_collect_yellowwood_text_room_sections_recovers_split_bed_robe_fit_out_headings(self) -> None:
        document = {
            "file_name": "yellowwood.pdf",
            "pages": [
                {
                    "page_no": 20,
                    "raw_text": (
                        "BED 3\n"
                        "ROBE FIT OUT\n"
                        "Robe Fit Out X1 Single Shelf with Hanging Rail\n"
                        "White Melamine\n"
                        "Yellowwood supplier\n"
                        "BED 4\n"
                        "ROBE FIT OUT\n"
                        "Robe Fit Out X1 Single Shelf with Hanging Rail\n"
                        "White Melamine\n"
                        "Yellowwood supplier\n"
                        "LAUNDRY LINEN FIT OUT\n"
                        "White Melamine\n"
                        "PASSAGE LINEN FIT OUT\n"
                        "White Melamine\n"
                    ),
                    "text": (
                        "BED 3\n"
                        "ROBE FIT OUT\n"
                        "Robe Fit Out X1 Single Shelf with Hanging Rail\n"
                        "White Melamine\n"
                        "Yellowwood supplier\n"
                        "BED 4\n"
                        "ROBE FIT OUT\n"
                        "Robe Fit Out X1 Single Shelf with Hanging Rail\n"
                        "White Melamine\n"
                        "Yellowwood supplier\n"
                        "LAUNDRY LINEN FIT OUT\n"
                        "White Melamine\n"
                        "PASSAGE LINEN FIT OUT\n"
                        "White Melamine\n"
                    ),
                },
            ],
        }
        sections = parsing_module._collect_yellowwood_text_room_sections_for_document(document)
        labels = [section["original_section_label"] for section in sections]
        self.assertIn("BED 3 ROBE", labels)
        self.assertIn("BED 4 ROBE", labels)
        self.assertNotIn("ROBE FIT OUT", labels)
        bed4 = next(section for section in sections if section["original_section_label"] == "BED 4 ROBE")
        self.assertNotIn("LAUNDRY LINEN FIT OUT", bed4["text"])
        self.assertNotIn("PASSAGE LINEN FIT OUT", bed4["text"])

    def test_yellowwood_helpers_keep_island_bench_ignore_bulkhead_note_and_require_explicit_led(self) -> None:
        lines = [
            "Overhead Cupboards *To Builders",
            "Bulkhead above rear Kitchen, Pantry and Fridge* Polytec Classic White Matt",
            "Island Bench 20mm YDL Classic White Polished",
            "Island Bench Kickboard Polytec Classic White Matt",
            "installed by builder",
        ]
        self.assertEqual(parsing_module._collect_island_benchtop_values(lines), ["20mm YDL Classic White Polished"])
        self.assertEqual(parsing_module._collect_explicit_bulkhead_values(lines), [])
        self.assertFalse(parsing_module._has_explicit_led_field(lines))
        self.assertFalse(parsing_module._has_explicit_led_field(["LED Topmount*"]))

    def test_yellowwood_material_probe_uses_evidence_snippet_for_real_robe_rooms(self) -> None:
        row = {
            "room_key": "bed_3_robe",
            "original_room_label": "BED 3 ROBE",
            "bench_tops": [],
            "door_panel_colours": [],
            "toe_kick": [],
            "bulkheads": [],
            "evidence_snippet": "BED 3 ROBE FIT OUT White Melamine Yellowwood supplier",
        }
        self.assertTrue(parsing_module._yellowwood_should_keep_final_room(row))

    def test_yellowwood_material_driven_rooms_keep_real_media_and_robe_when_joinery_material_exists(self) -> None:
        media_section = {
            "section_key": "media_room",
            "original_section_label": "MEDIA ROOM",
            "text": "MEDIA ROOM\nTV UNIT\nPolytec Boston Oak Woodmatt\nAs supplied by cabinetmaker",
        }
        robe_section = {
            "section_key": "bed_2_robe",
            "original_section_label": "BED 2 ROBE",
            "text": "BED 2 ROBE\nFIT OUT\nLaminex Classic White\nYellowwood supplier",
        }
        self.assertTrue(parsing_module._yellowwood_should_keep_section(media_section))
        self.assertTrue(parsing_module._yellowwood_should_keep_section(robe_section))

    def test_polish_generic_layout_room_prefers_specific_yellowwood_joinery_title(self) -> None:
        polished = extraction_service._polish_generic_layout_room(
            {
                "room_key": "bathroom",
                "original_room_label": "BATHROOM VANITY",
                "room_name": "BATHROOM VANITY",
            },
            {
                "original_room_label": "BATHROOM",
            },
        )
        self.assertEqual(polished["original_room_label"], "BATHROOM VANITY")

    def test_apply_snapshot_cleaning_rules_drops_yellowwood_rooms_without_joinery_material_evidence(self) -> None:
        snapshot = {
            "job_no": "38095",
            "builder_name": "Yellowwood",
            "source_kind": "spec",
            "generated_at": "2026-04-03T10:00:00+10:00",
            "rooms": [
                {
                    "room_key": "kitchen",
                    "original_room_label": "KITCHEN",
                    "bench_tops_wall_run": "20mm YDL Classic White Polished",
                    "door_colours_base": "Polytec Classic White Matt",
                    "door_panel_colours": [],
                    "toe_kick": [],
                    "bulkheads": [],
                    "handles": [],
                },
                {
                    "room_key": "media_room",
                    "original_room_label": "MEDIA ROOM",
                    "flooring": "Carpet Mapelton Falls",
                    "door_panel_colours": [],
                    "toe_kick": [],
                    "bulkheads": [],
                    "handles": [],
                },
                {
                    "room_key": "laundry",
                    "original_room_label": "LAUNDRY",
                    "sink_info": "Wall Hung Basin",
                    "tap_info": "Highgrove mixer",
                    "door_panel_colours": [],
                    "toe_kick": [],
                    "bulkheads": [],
                    "handles": [],
                },
                {
                    "room_key": "ensuite_1",
                    "original_room_label": "BED 1 ENSUITE VANITY",
                    "bench_tops_wall_run": "20mm YDL Classic White Polished",
                    "door_colours_base": "Polytec Classic White Matt",
                    "door_panel_colours": [],
                    "toe_kick": [],
                    "bulkheads": [],
                    "handles": [],
                },
                {
                    "room_key": "ensuite_1",
                    "original_room_label": "BED 1 ENSUITE",
                    "door_panel_colours": [],
                    "toe_kick": [],
                    "bulkheads": [],
                    "handles": [],
                    "evidence_snippet": "BED 1 ENSUITE MIRROR Frameless Mirror As supplied by builder",
                },
            ],
            "appliances": [],
            "others": {},
            "warnings": [],
            "source_documents": [],
            "analysis": {"mode": "heuristic_only"},
        }
        cleaned = parsing_module.apply_snapshot_cleaning_rules(snapshot)
        kept = {row["original_room_label"] for row in cleaned["rooms"]}
        self.assertIn("KITCHEN", kept)
        self.assertIn("BED 1 ENSUITE VANITY", kept)
        self.assertNotIn("MEDIA ROOM", kept)
        self.assertNotIn("LAUNDRY", kept)

    def test_apply_snapshot_cleaning_rules_purifies_yellowwood_wet_area_fixtures(self) -> None:
        snapshot = {
            "job_no": "38095",
            "builder_name": "Yellowwood",
            "source_kind": "spec",
            "generated_at": "2026-04-03T10:00:00+10:00",
            "rooms": [
                {
                    "room_key": "bathroom",
                    "original_room_label": "BATHROOM VANITY",
                    "bench_tops_wall_run": "20mm YDL Classic White Polished",
                    "door_colours_base": "Polytec Classic White Matt",
                    "sink_info": "Mixer Spin Tall Basin Mixer Chrome Highgrove",
                    "basin_info": "Byron Bench Mount Basin White Gloss Highgrove Basin Waste Pop Up Waste Without Overflow 32x80mm Chrome Highgrove Toilet Meldon Rimless Wall Faced Toilet Suite Highgrove",
                    "tap_info": "Spin Tall Basin Mixer Chrome Highgrove Basin Waste Pop Up Waste Without Overflow 32x80mm Chrome Highgrove",
                    "door_panel_colours": [],
                    "toe_kick": [],
                    "bulkheads": [],
                    "handles": [],
                }
            ],
            "appliances": [],
            "others": {},
            "warnings": [],
            "source_documents": [],
            "analysis": {"mode": "heuristic_only"},
        }
        cleaned = parsing_module.apply_snapshot_cleaning_rules(snapshot)
        room = cleaned["rooms"][0]
        self.assertEqual(room["sink_info"], "")
        self.assertEqual(room["basin_info"], "Byron Bench Mount Basin White Gloss Highgrove")
        self.assertEqual(room["tap_info"], "Spin Tall Basin Mixer Chrome Highgrove")

    def test_apply_snapshot_cleaning_rules_drops_yellowwood_duplicate_plumbing_other_items(self) -> None:
        snapshot = {
            "job_no": "38095",
            "builder_name": "Yellowwood",
            "source_kind": "spec",
            "generated_at": "2026-04-03T10:00:00+10:00",
            "rooms": [
                {
                    "room_key": "kitchen",
                    "original_room_label": "KITCHEN",
                    "bench_tops_wall_run": "20mm YDL Classic White Polished",
                    "door_colours_base": "Polytec Classic White Matt",
                    "tap_info": "Spin Gooseneck Matte Black Highgrove",
                    "other_items": [
                        {"label": "Mixer", "value": "Spin Gooseneck Matte Black Highgrove"},
                        {"label": "Pull-Out Mixer", "value": "Spin Gooseneck Chrome Highgrove"},
                        {"label": "RAIL", "value": "Square Edge recessed rail in black"},
                    ],
                    "door_panel_colours": [],
                    "toe_kick": [],
                    "bulkheads": [],
                    "handles": [],
                }
            ],
            "appliances": [],
            "others": {},
            "warnings": [],
            "source_documents": [],
            "analysis": {"mode": "heuristic_only"},
        }
        cleaned = parsing_module.apply_snapshot_cleaning_rules(snapshot)
        room = cleaned["rooms"][0]
        labels = {item["label"] for item in room["other_items"]}
        self.assertNotIn("Mixer", labels)
        self.assertNotIn("Pull-Out Mixer", labels)
        self.assertIn("RAIL", labels)

    def test_apply_snapshot_cleaning_rules_trims_yellowwood_vanity_other_item_tails(self) -> None:
        snapshot = {
            "job_no": "38095",
            "builder_name": "Yellowwood",
            "source_kind": "spec",
            "generated_at": "2026-04-03T10:00:00+10:00",
            "rooms": [
                {
                    "room_key": "bathroom",
                    "original_room_label": "BATHROOM VANITY",
                    "bench_tops_wall_run": "20mm YDL Classic White Polished",
                    "door_colours_base": "Polytec Classic White Matt",
                    "tap_info": "Spin Tall Basin Mixer Chrome Highgrove",
                    "other_items": [
                        {
                            "label": "Accessories 1",
                            "value": "Towel Rail - Spin Double Towel Rail Chrome 600mm Highgrove Shower Floor Waste LIDO Square tile grate 125mm with DN100 outlet Chrome Highgrove BATHROOM",
                        }
                    ],
                    "door_panel_colours": [],
                    "toe_kick": [],
                    "bulkheads": [],
                    "handles": [],
                }
            ],
            "appliances": [],
            "others": {},
            "warnings": [],
            "source_documents": [],
            "analysis": {"mode": "heuristic_only"},
        }
        cleaned = parsing_module.apply_snapshot_cleaning_rules(snapshot)
        room = cleaned["rooms"][0]
        self.assertEqual(
            room["other_items"],
            [{"label": "Accessories 1", "value": "Towel Rail - Spin Double Towel Rail Chrome 600mm Highgrove"}],
        )

    def test_apply_snapshot_cleaning_rules_trims_yellowwood_vanity_accessory_tails(self) -> None:
        snapshot = {
            "job_no": "38095",
            "builder_name": "Yellowwood",
            "source_kind": "spec",
            "generated_at": "2026-04-03T10:00:00+10:00",
            "rooms": [
                {
                    "room_key": "bathroom",
                    "original_room_label": "BATHROOM VANITY",
                    "bench_tops_wall_run": "20mm YDL Classic White Polished",
                    "door_colours_base": "Polytec Classic White Matt",
                    "accessories": [
                        "Towel Rail - Spin Double Towel Rail Chrome 600mm Highgrove Shower Floor Waste LIDO Square tile grate 125mm with DN100 outlet Chrome Highgrove BATHROOM"
                    ],
                    "door_panel_colours": [],
                    "toe_kick": [],
                    "bulkheads": [],
                    "handles": [],
                }
            ],
            "appliances": [],
            "others": {},
            "warnings": [],
            "source_documents": [],
            "analysis": {"mode": "heuristic_only"},
        }
        cleaned = parsing_module.apply_snapshot_cleaning_rules(snapshot)
        room = cleaned["rooms"][0]
        self.assertEqual(room["accessories"], ["Towel Rail - Spin Double Towel Rail Chrome 600mm Highgrove"])

    def test_apply_snapshot_cleaning_rules_trims_generic_yellowwood_fixture_waste_and_accessory_tails(self) -> None:
        snapshot = {
            "job_no": "38095",
            "builder_name": "Yellowwood",
            "source_kind": "spec",
            "generated_at": "2026-04-04T10:00:00+10:00",
            "rooms": [
                {
                    "room_key": "bathroom",
                    "original_room_label": "BATHROOM VANITY",
                    "bench_tops_wall_run": "20mm YDL Classic White Polished",
                    "door_colours_base": "Polytec Classic White Matt",
                    "sink_info": "Byron Bench Mount Basin White Gloss Waste Chrome Highgrove BATHROOM",
                    "basin_info": "Byron Bench Mount Basin White Gloss Waste Chrome Highgrove Toilet Suite Valencia Back To Wall Highgrove",
                    "tap_info": "Spin Tall Basin Mixer Chrome Highgrove Towel Rail - Spin Double Towel Rail Chrome 600mm Highgrove",
                    "door_panel_colours": [],
                    "toe_kick": [],
                    "bulkheads": [],
                    "handles": [],
                }
            ],
            "appliances": [],
            "others": {},
            "warnings": [],
            "source_documents": [],
            "analysis": {"mode": "heuristic_only"},
        }
        cleaned = parsing_module.apply_snapshot_cleaning_rules(snapshot)
        room = cleaned["rooms"][0]
        self.assertEqual(room["sink_info"], "Byron Bench Mount Basin White Gloss")
        self.assertEqual(room["basin_info"], "Byron Bench Mount Basin White Gloss")
        self.assertEqual(room["tap_info"], "Spin Tall Basin Mixer Chrome Highgrove")

    def test_dedupe_appliances_drops_gibberish_fridge_when_real_fridge_model_exists(self) -> None:
        rows = [
            parsing_module.ApplianceRow(
                appliance_type="Fridge",
                make="Fisher & Paykel",
                model_no="RB60V18",
                source_file="clarendon.pdf",
                evidence="Integrated Fridge/Freezer: Fisher & Paykel 2 X RB60V18",
            ),
            parsing_module.ApplianceRow(
                appliance_type="Fridge",
                make="",
                model_no="30120018EQ740900740EQ30",
                source_file="clarendon.pdf",
                evidence="30120018EQ740900740EQ30 44102X INTEGRATED FRIDGE/ FREEZERFISHER & PAYKEL RB60V18",
            ),
        ]
        deduped = parsing_module._dedupe_appliances(rows)
        self.assertEqual([(row.appliance_type, row.model_no) for row in deduped], [("Fridge", "RB60V18")])

    def test_dedupe_appliances_drops_client_to_check_placeholder_when_real_model_exists(self) -> None:
        rows = [
            parsing_module.ApplianceRow(
                appliance_type="Fridge",
                make="Fisher & Paykel",
                model_no="RB60V18",
                source_file="clarendon.pdf",
                evidence="Integrated Fridge/Freezer: Fisher & Paykel 2 X RB60V18",
            ),
            parsing_module.ApplianceRow(
                appliance_type="Fridge",
                make="",
                model_no="N/A CLIENT TO CHECK",
                source_file="clarendon.pdf",
                evidence="Fridge: N/A CLIENT TO CHECK",
            ),
        ]
        deduped = parsing_module._dedupe_appliances(rows)
        self.assertEqual([(row.appliance_type, row.model_no) for row in deduped], [("Fridge", "RB60V18")])

    def test_clarendon_schedule_room_key_preserves_rumpus_desk(self) -> None:
        self.assertEqual(
            extraction_service._clarendon_schedule_room_key("RUMPUS - DESK JOINERY COLOUR SCHEDULE"),
            "rumpus_desk",
        )

    def test_collect_spec_sections_for_clarendon_separates_rumpus_room_and_rumpus_desk(self) -> None:
        text = (
            "RUMPUS ROOM COLOUR SCHEDULE\n"
            "Door Colour - Polytec Classic White Matt Finish Thermolaminate - Atlanta EM2 Profile\n"
            "RUMPUS - DESK JOINERY COLOUR SCHEDULE\n"
            "Door Colour 2 - Laminex Classic White - TO TALL OPEN SHELVES\n"
        )
        document = {
            "file_name": "clarendon.pdf",
            "role": "spec",
            "pages": [
                {"page_no": 1, "text": text, "raw_text": text, "needs_ocr": False},
            ],
        }
        sections = parsing_module._collect_spec_sections_for_document("Clarendon", document)
        labels = {
            str(section.get("original_section_label", "")): str(section.get("text", ""))
            for section in sections
            if section.get("section_kind") == "room"
        }
        self.assertIn("RUMPUS ROOM", labels)
        self.assertIn("RUMPUS - DESK", labels)
        self.assertNotIn("Tall Open Shelves", labels["RUMPUS ROOM"])
        self.assertIn("TALL OPEN SHELVES", labels["RUMPUS - DESK"])

    def test_enrich_snapshot_rooms_applies_clarendon_afc_flooring_overlay(self) -> None:
        snapshot = {
            "job_no": "49906511",
            "builder_name": "Clarendon",
            "source_kind": "spec",
            "generated_at": "2026-04-04T00:00:00+10:00",
            "rooms": [
                {"room_key": "kitchen", "original_room_label": "KITCHEN", "flooring": ""},
                {"room_key": "butlers_pantry", "original_room_label": "BUTLERS PANTRY", "flooring": ""},
                {"room_key": "theatre", "original_room_label": "THEATRE ROOM", "flooring": ""},
                {"room_key": "rumpus_room", "original_room_label": "RUMPUS ROOM", "flooring": ""},
                {"room_key": "rumpus_desk", "original_room_label": "RUMPUS - DESK", "flooring": ""},
                {"room_key": "laundry", "original_room_label": "LAUNDRY", "flooring": ""},
            ],
            "appliances": [],
            "others": {"flooring_notes": "CARPET & MAIN FLOOR TILE"},
            "warnings": [],
            "source_documents": [],
            "analysis": {"mode": "heuristic_only"},
        }
        text = (
            "CARPET & MAIN FLOOR TILE\n"
            "Kitchen/Pantry/Family/Meals: TILED\n"
            "Theatre: CARPET\n"
            "Rumpus: CARPET\n"
            "WIL/Linen/s Ground Floor: TILED\n"
            "WIR/S & Robes: CARPET\n"
        )
        documents = [
            {
                "file_name": "job1_afc.pdf",
                "role": "spec",
                "pages": [
                    {"page_no": 14, "text": text, "raw_text": text, "needs_ocr": False},
                ],
            }
        ]
        enriched = parsing_module.enrich_snapshot_rooms(snapshot, documents)
        rooms = {row["room_key"]: row for row in enriched["rooms"]}
        self.assertEqual(rooms["kitchen"]["flooring"], "TILED")
        self.assertEqual(rooms["butlers_pantry"]["flooring"], "TILED")
        self.assertEqual(rooms["theatre"]["flooring"], "CARPET")
        self.assertEqual(rooms["rumpus_room"]["flooring"], "CARPET")
        self.assertEqual(rooms["rumpus_desk"]["flooring"], "")
        self.assertEqual(rooms["laundry"]["flooring"], "TILED")
        self.assertEqual(enriched["others"]["flooring_notes"], "")

    def test_enrich_snapshot_rooms_moves_clarendon_rumpus_desk_tall_out_of_rumpus_room(self) -> None:
        snapshot = {
            "job_no": "49906511",
            "builder_name": "Clarendon",
            "source_kind": "spec",
            "generated_at": "2026-04-04T00:00:00+10:00",
            "rooms": [
                {
                    "room_key": "rumpus_room",
                    "original_room_label": "RUMPUS ROOM",
                    "door_colours_tall": "Polytec CLASSIC WHITE MATT FINSIH MELAMINE WITH MATCHING 1MM ABSE EDGES - TO TALL OPEN SHELVES",
                    "has_explicit_tall": True,
                },
                {
                    "room_key": "rumpus_desk",
                    "original_room_label": "RUMPUS - DESK",
                    "door_colours_tall": "",
                    "has_explicit_tall": False,
                },
            ],
            "appliances": [],
            "others": {"flooring_notes": ""},
            "warnings": [],
            "source_documents": [],
            "analysis": {"mode": "heuristic_only"},
        }
        enriched = parsing_module.enrich_snapshot_rooms(snapshot, [])
        rooms = {row["room_key"]: row for row in enriched["rooms"]}
        self.assertEqual(rooms["rumpus_room"]["door_colours_tall"], "")
        self.assertFalse(rooms["rumpus_room"]["has_explicit_tall"])
        self.assertIn("TALL OPEN SHELVES", rooms["rumpus_desk"]["door_colours_tall"])
        self.assertTrue(rooms["rumpus_desk"]["has_explicit_tall"])

    def test_enrich_snapshot_rooms_clears_clarendon_appliance_noise_from_splashback_notes(self) -> None:
        snapshot = {
            "job_no": "49906511",
            "builder_name": "Clarendon",
            "source_kind": "spec",
            "generated_at": "2026-04-04T00:00:00+10:00",
            "rooms": [],
            "appliances": [],
            "others": {
                "flooring_notes": "",
                "splashback_notes": "APPLIANCES Freestanding Cooker: N/A Under Bench Oven: Westinghouse 2 X WVE6515SDA 60CM ELECTRIC",
            },
            "warnings": [],
            "source_documents": [],
            "analysis": {"mode": "heuristic_only"},
        }
        enriched = parsing_module.enrich_snapshot_rooms(snapshot, [])
        self.assertEqual(enriched["others"]["splashback_notes"], "")

    def test_clean_handle_entries_drops_orphan_handle_house_fragment(self) -> None:
        self.assertEqual(
            parsing_module._clean_handle_entries(
                [
                    "C137 Caloundra Lip Pull Matt Black 200mm (Laid Vertical)",
                    "House",
                ]
            ),
            ["C137 Caloundra Lip Pull Matt Black 200mm (Laid Vertical)"],
        )

    def test_merge_yellowwood_layout_and_text_sections_filters_fake_material_rooms(self) -> None:
        layout_sections = [
            {
                "section_key": "kitchen",
                "original_section_label": "KITCHEN",
                "section_kind": "room",
                "page_nos": [17],
                "text": "KITCHEN\nBase Cupboards & Drawers\nPolytec Blossom White Matt",
            },
            {
                "section_key": "revival_hoya_powder_matt_200x200mm",
                "original_section_label": "Revival Hoya Powder Matt 200x200mm",
                "section_kind": "room",
                "page_nos": [28],
                "text": "Revival Hoya Powder Matt 200x200mm\nNational Tiles",
            },
        ]
        merged = parsing_module._merge_yellowwood_layout_and_text_sections(layout_sections, [])
        labels = {section["original_section_label"] for section in merged}
        self.assertIn("KITCHEN", labels)
        self.assertNotIn("Revival Hoya Powder Matt 200x200mm", labels)

    def test_enrich_snapshot_rooms_applies_yellowwood_flooring_overlay_and_clears_contents_noise(self) -> None:
        snapshot = {
            "job_no": "37",
            "builder_name": "Yellowwood",
            "source_kind": "spec",
            "generated_at": "2026-04-04T00:00:00+10:00",
            "rooms": [
                {"room_key": "kitchen", "original_room_label": "KITCHEN", "door_colours_base": "Polytec Classic White Matt", "flooring": ""},
                {"room_key": "bed_1_wir", "original_room_label": "BED 1 WALK IN ROBE", "door_colours_base": "Polytec Standard White", "flooring": ""},
                {"room_key": "bed_2_robe", "original_room_label": "BED 2 ROBE", "door_colours_base": "Polytec Standard White", "flooring": ""},
                {"room_key": "bed_3_robe", "original_room_label": "BED 3 ROBE", "door_colours_base": "Polytec Standard White", "flooring": ""},
                {"room_key": "bed_4_robe", "original_room_label": "BED 4 ROBE", "door_colours_base": "Polytec Standard White", "flooring": ""},
                {"room_key": "ensuite_1", "original_room_label": "BED 1 ENSUITE VANITY", "bench_tops_wall_run": "20mm YDL Classic White Polished", "flooring": ""},
                {"room_key": "bathroom", "original_room_label": "BATHROOM VANITY", "bench_tops_wall_run": "20mm YDL Classic White Polished", "flooring": ""},
            ],
            "appliances": [],
            "others": {
                "flooring_notes": "OTHER THAN TILING TO WET AREAS............................................................................................................................15-16"
            },
            "warnings": [],
            "source_documents": [],
            "analysis": {"mode": "heuristic_only"},
        }
        documents = [
            {
                "file_name": "job37.pdf",
                "role": "spec",
                "builder_name": "Yellowwood",
                "pages": [
                    {
                        "page_no": 2,
                        "text": "TABLE OF CONTENTS FLOORING - OTHER THAN TILING TO WET AREAS 15-16 TILING SCHEDULE 21-25",
                        "raw_text": "TABLE OF CONTENTS FLOORING - OTHER THAN TILING TO WET AREAS 15-16 TILING SCHEDULE 21-25",
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 15,
                        "text": (
                            "FLOORING - OTHER THAN TILING TO WET AREAS\n"
                            "ENTRY, HALLWAY, LIVING, RUMPUS, PASSAGE, PASSAGE LINEN, KITCHEN & DINING\n"
                            "Hybrid Flooring Admired Grand Hybrid\n"
                            "Colour: Spotted Gum\n"
                            "1530x228x6.5mm\n"
                            "*Matching trims as required*\n"
                            "Flooring Xtra\n"
                            "BED 1 + WIR\n"
                            "Carpet Silverwood\n"
                            "Colour: Putty\n"
                            "8mm Underlay\n"
                            "Flooring Xtra\n"
                            "BED 2 + ROBE\n"
                            "Carpet Silverwood\n"
                            "Colour: Putty\n"
                            "8mm Underlay\n"
                            "Flooring Xtra\n"
                            "BED 3 + ROBE\n"
                            "Carpet Silverwood\n"
                            "Colour: Putty\n"
                            "8mm Underlay\n"
                            "Flooring Xtra\n"
                            "BED 4 + ROBE\n"
                            "Carpet Silverwood\n"
                            "Colour: Putty\n"
                            "8mm Underlay\n"
                        ),
                        "raw_text": (
                            "FLOORING - OTHER THAN TILING TO WET AREAS\n"
                            "ENTRY, HALLWAY, LIVING, RUMPUS, PASSAGE, PASSAGE LINEN, KITCHEN & DINING\n"
                            "Hybrid Flooring Admired Grand Hybrid\n"
                            "Colour: Spotted Gum\n"
                            "1530x228x6.5mm\n"
                            "*Matching trims as required*\n"
                            "Flooring Xtra\n"
                            "BED 1 + WIR\n"
                            "Carpet Silverwood\n"
                            "Colour: Putty\n"
                            "8mm Underlay\n"
                            "Flooring Xtra\n"
                            "BED 2 + ROBE\n"
                            "Carpet Silverwood\n"
                            "Colour: Putty\n"
                            "8mm Underlay\n"
                            "Flooring Xtra\n"
                            "BED 3 + ROBE\n"
                            "Carpet Silverwood\n"
                            "Colour: Putty\n"
                            "8mm Underlay\n"
                            "Flooring Xtra\n"
                            "BED 4 + ROBE\n"
                            "Carpet Silverwood\n"
                            "Colour: Putty\n"
                            "8mm Underlay\n"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 16,
                        "text": (
                            "Flooring Xtra\n"
                            "LAUNDRY Tiles Refer to Tiling section below N/A\n"
                            "BED 1 ENSUITE Tiles Refer to Tiling section below N/A\n"
                            "BATHROOM Tiles Refer to Tiling section below N/A\n"
                        ),
                        "raw_text": (
                            "Flooring Xtra\n"
                            "LAUNDRY Tiles Refer to Tiling section below N/A\n"
                            "BED 1 ENSUITE Tiles Refer to Tiling section below N/A\n"
                            "BATHROOM Tiles Refer to Tiling section below N/A\n"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 22,
                        "text": (
                            "TILING SCHEDULE\n"
                            "BATHROOM\n"
                            "Floor Tile Regina Grey Matt 450x450mm\n"
                            "Lay Pattern - Square\n"
                            "Grout: Mapei 110 Manhattan\n"
                            "National Tiles\n"
                            "Wall Tile Regina Grey Matt 450x450mm\n"
                        ),
                        "raw_text": (
                            "TILING SCHEDULE\n"
                            "BATHROOM\n"
                            "Floor Tile Regina Grey Matt 450x450mm\n"
                            "Lay Pattern - Square\n"
                            "Grout: Mapei 110 Manhattan\n"
                            "National Tiles\n"
                            "Wall Tile Regina Grey Matt 450x450mm\n"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 23,
                        "text": (
                            "BED 1 ENSUITE\n"
                            "Floor Tile Regina Grey Matt 450x450mm\n"
                            "Lay Pattern - Square\n"
                            "Grout: Mapei 110 Manhattan\n"
                            "National Tiles\n"
                            "Wall Tile Regina Grey Matt 450x450mm\n"
                        ),
                        "raw_text": (
                            "BED 1 ENSUITE\n"
                            "Floor Tile Regina Grey Matt 450x450mm\n"
                            "Lay Pattern - Square\n"
                            "Grout: Mapei 110 Manhattan\n"
                            "National Tiles\n"
                            "Wall Tile Regina Grey Matt 450x450mm\n"
                        ),
                        "needs_ocr": False,
                    },
                ],
            }
        ]
        enriched = parsing_module.enrich_snapshot_rooms(snapshot, documents)
        rooms = {row["room_key"]: row for row in enriched["rooms"]}
        self.assertIn("Hybrid Flooring Admired Grand Hybrid", rooms["kitchen"]["flooring"])
        self.assertIn("Spotted Gum", rooms["kitchen"]["flooring"])
        self.assertIn("Carpet Silverwood", rooms["bed_1_wir"]["flooring"])
        self.assertIn("Carpet Silverwood", rooms["bed_2_robe"]["flooring"])
        self.assertIn("Flooring Xtra", rooms["bed_4_robe"]["flooring"])
        self.assertIn("Floor Tile Regina Grey Matt 450x450mm", rooms["bathroom"]["flooring"])
        self.assertIn("Floor Tile Regina Grey Matt 450x450mm", rooms["ensuite_1"]["flooring"])
        self.assertEqual(enriched["others"]["flooring_notes"], "")

    def test_enrich_snapshot_rooms_merges_yellowwood_bath_family_into_bathroom_vanity(self) -> None:
        snapshot = {
            "job_no": "37",
            "builder_name": "Yellowwood Homes",
            "source_kind": "spec",
            "generated_at": "2026-04-04T00:00:00+10:00",
            "rooms": [
                {
                    "room_key": "bathroom",
                    "original_room_label": "BATHROOM VANITY",
                    "room_name": "BATHROOM VANITY",
                    "bench_tops_wall_run": "20mm YDL Classic White Polished",
                    "bench_tops_island": "",
                    "bench_tops_other": "",
                    "bench_tops": [],
                    "door_panel_colours": [],
                    "door_colours_overheads": "",
                    "door_colours_base": "Polytec Classic White Matt",
                    "door_colours_tall": "",
                    "door_colours_island": "",
                    "door_colours_bar_back": "",
                    "toe_kick": ["Polytec Classic White Matt"],
                    "bulkheads": [],
                    "handles": [],
                    "floating_shelf": "",
                    "led": "",
                    "accessories": [],
                    "other_items": [],
                    "sink_info": "",
                    "basin_info": "Byron Bench Mount Basin White Gloss Highgrove",
                    "tap_info": "Spin Tall Basin Mixer Chrome Highgrove",
                    "drawers_soft_close": "",
                    "hinges_soft_close": "",
                    "splashback": "",
                    "flooring": "",
                    "source_file": "yellowwood.pdf",
                    "page_refs": "19",
                    "evidence_snippet": "",
                    "confidence": 0.55,
                }
            ],
            "appliances": [],
            "others": {"flooring_notes": "", "splashback_notes": ""},
            "warnings": [],
            "source_documents": [],
            "analysis": {"mode": "docling"},
        }
        documents = [
            {
                "file_name": "yellowwood.pdf",
                "role": "spec",
                "builder_name": "Yellowwood Homes",
                "pages": [
                    {
                        "page_no": 28,
                        "text": (
                            "BATHROOM\n"
                            "Sink Mixer Spin Tall Basin Mixer Chrome Highgrove\n"
                            "Basin Byron Bench Mount Basin White Gloss Highgrove\n"
                        ),
                        "raw_text": (
                            "LOT 13 ZACHARY COURT, BEACHMERE\n"
                            "BATHROOM\n"
                            "Sink Mixer Spin Tall Basin Mixer Chrome Highgrove\n"
                            "Basin Byron Bench Mount Basin White Gloss Highgrove\n"
                            "page 28/33\n"
                            "Tone Interior Design Consulting | tone.interiorco@gmail.com |0450026294\n"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 29,
                        "text": (
                            "Bath Eden Freestanding Back to Wall Bath with Overflow Gloss White Size as per plan Highgrove\n"
                            "Bath Waste Turn Down waste Without Overflow Chrome 40x60mm Highgrove\n"
                        ),
                        "raw_text": (
                            "LOT 13 ZACHARY COURT, BEACHMERE\n"
                            "Bath Eden Freestanding Back to Wall Bath with Overflow Gloss White Size as per plan Highgrove\n"
                            "Bath Waste Turn Down waste Without Overflow Chrome 40x60mm Highgrove\n"
                            "page 29/33\n"
                            "Tone Interior Design Consulting | tone.interiorco@gmail.com |0450026294\n"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 30,
                        "text": (
                            "Bath Mixer Spin Mixer Set Chrome Highgrove\n"
                            "Bath Spout Spin In Wall Bath Spout 220mm Chrome Highgrove\n"
                            "WC\n"
                        ),
                        "raw_text": (
                            "LOT 13 ZACHARY COURT, BEACHMERE\n"
                            "Bath Mixer Spin Mixer Set Chrome Highgrove\n"
                            "Bath Spout Spin In Wall Bath Spout 220mm Chrome Highgrove\n"
                            "WC\n"
                            "page 30/33\n"
                            "Tone Interior Design Consulting | tone.interiorco@gmail.com |0450026294\n"
                        ),
                        "needs_ocr": False,
                    }
                ],
            }
        ]
        enriched = parsing_module.enrich_snapshot_rooms(snapshot, documents)
        rooms = {row["room_key"]: row for row in enriched["rooms"]}
        bath_items = {item["label"]: item["value"] for item in rooms["bathroom"]["other_items"]}
        self.assertIn("BATH", bath_items)
        self.assertIn("Eden Freestanding Back to Wall Bath", bath_items["BATH"])
        self.assertIn("BATH MIXER", bath_items)
        self.assertIn("Spin Mixer Set Chrome", bath_items["BATH MIXER"])
        self.assertIn("BATH SPOUT", bath_items)
        self.assertIn("Bath Spout 220mm Chrome", bath_items["BATH SPOUT"])
        self.assertIn("BATH WASTE", bath_items)
        self.assertIn("Turn Down waste Without Overflow", bath_items["BATH WASTE"])

    def test_enrich_snapshot_rooms_moves_clarendon_laundry_accessories_out_of_vanities(self) -> None:
        snapshot = {
            "job_no": "49906613",
            "builder_name": "Clarendon",
            "source_kind": "spec",
            "generated_at": "2026-04-04T00:00:00+10:00",
            "rooms": [
                {
                    "room_key": "vanities",
                    "original_room_label": "VANITIES",
                    "accessories": [
                        "LINCOLN SENTRY 7104019 CENTRE PILLAR, CHROME",
                        "LINCOLN SENTRY 7104020 FINISTA END PILLAR, CHROME",
                    ],
                },
                {
                    "room_key": "laundry",
                    "original_room_label": "LAUNDRY",
                    "accessories": [],
                },
            ],
            "appliances": [],
            "others": {"flooring_notes": "", "splashback_notes": ""},
            "warnings": [],
            "source_documents": [],
            "analysis": {"mode": "heuristic_only"},
        }
        enriched = parsing_module.enrich_snapshot_rooms(snapshot, [])
        rooms = {row["room_key"]: row for row in enriched["rooms"]}
        self.assertEqual(rooms["vanities"]["accessories"], [])
        self.assertIn("Lincoln Sentry - 7104019 CENTRE PILLAR, CHROME", rooms["laundry"]["accessories"])
        self.assertIn("Lincoln Sentry - 7104020 FINISTA END PILLAR, CHROME", rooms["laundry"]["accessories"])

    def test_merge_yellowwood_layout_and_text_sections_trims_override_pages_from_merged_layout_room(self) -> None:
        layout_sections = [
            {
                "section_key": "bathroom",
                "original_section_label": "Bathroom",
                "section_kind": "room",
                "page_nos": [25, 28],
                "page_texts": [
                    {"page_no": 25, "text": "BATHROOM VANITY\nHandles\nHandle House\nBED 1 WALK IN ROBE FIT OUT"},
                    {"page_no": 28, "text": "BATHROOM\nFloor Tile Regina Grey Matt 450x450mm"},
                ],
                "raw_page_texts": [
                    {"page_no": 25, "text": "BATHROOM VANITY\nHandles\nHandle House\nBED 1 WALK IN ROBE FIT OUT"},
                    {"page_no": 28, "text": "BATHROOM\nFloor Tile Regina Grey Matt 450x450mm"},
                ],
                "layout_rows": [
                    {"page_no": 25, "row_label": "Handles", "value_text": "C185", "supplier_text": "Handle House", "notes_text": ""},
                    {"page_no": 28, "row_label": "Floor Tile", "value_text": "Regina Grey Matt 450x450mm", "supplier_text": "National Tiles", "notes_text": ""},
                ],
                "text_parts": [
                    "BATHROOM VANITY\nHandles\nHandle House\nBED 1 WALK IN ROBE FIT OUT",
                    "BATHROOM\nFloor Tile Regina Grey Matt 450x450mm",
                ],
                "text": "BATHROOM VANITY\nHandles\nHandle House\nBED 1 WALK IN ROBE FIT OUT\nBATHROOM\nFloor Tile Regina Grey Matt 450x450mm",
            }
        ]
        text_sections = [
            {
                "section_key": "bathroom",
                "original_section_label": "Bathroom",
                "section_kind": "room",
                "page_nos": [25],
                "page_texts": [{"page_no": 25, "text": "BATHROOM VANITY\nHandles\nC185 Esperance Pulls Satin Nickel"}],
                "raw_page_texts": [{"page_no": 25, "text": "BATHROOM VANITY\nHandles\nC185 Esperance Pulls Satin Nickel"}],
                "layout_rows": [],
                "text_parts": ["BATHROOM VANITY\nHandles\nC185 Esperance Pulls Satin Nickel"],
                "text": "BATHROOM VANITY\nHandles\nC185 Esperance Pulls Satin Nickel",
            },
            {
                "section_key": "bed_1_wir",
                "original_section_label": "BED 1 WALK IN ROBE",
                "section_kind": "room",
                "page_nos": [25],
                "page_texts": [{"page_no": 25, "text": "BED 1 WALK IN ROBE\nBoard Colour: Standard White"}],
                "raw_page_texts": [{"page_no": 25, "text": "BED 1 WALK IN ROBE\nBoard Colour: Standard White"}],
                "layout_rows": [],
                "text_parts": ["BED 1 WALK IN ROBE\nBoard Colour: Standard White"],
                "text": "BED 1 WALK IN ROBE\nBoard Colour: Standard White",
            },
        ]
        merged = parsing_module._merge_yellowwood_layout_and_text_sections(layout_sections, text_sections)
        bathroom = next(section for section in merged if section["section_key"] == "bathroom")
        self.assertEqual(bathroom["page_nos"], [28, 25])
        self.assertNotIn("BED 1 WALK IN ROBE FIT OUT", bathroom["text"])
        self.assertIn("Floor Tile Regina Grey Matt 450x450mm", bathroom["text"])

    def test_collect_room_sections_for_yellowwood_ignores_fake_material_layout_rooms(self) -> None:
        document = {
            "file_name": "yellowwood.pdf",
            "builder_name": "Yellowwood",
            "pages": [
                {
                    "page_no": 28,
                    "text": "Revival Hoya Powder Matt 200x200mm\nNational Tiles",
                    "raw_text": "Revival Hoya Powder Matt 200x200mm\nNational Tiles",
                    "page_layout": {
                        "page_type": "joinery",
                        "section_label": "Revival Hoya Powder Matt 200x200mm",
                        "room_label": "Revival Hoya Powder Matt 200x200mm",
                        "room_blocks": [
                            {
                                "room_label": "Revival Hoya Powder Matt 200x200mm",
                                "rows": [
                                    {
                                        "row_label": "Wall Tile",
                                        "value_region_text": "Revival Hoya Powder Matt 200x200mm",
                                        "supplier_region_text": "National Tiles",
                                        "notes_region_text": "",
                                        "row_kind": "material",
                                    }
                                ],
                            }
                        ],
                    },
                }
            ],
        }
        sections = parsing_module._collect_room_sections_for_document(document)
        self.assertEqual(sections, [])

    def test_spec_room_label_noise_filters_clarendon_company_heading(self) -> None:
        self.assertTrue(
            parsing_module._looks_like_spec_room_label_noise(
                "Forstan Pty Ltd. T/A YourHome Kitchens P.O. Box 8248 Baulkham HillsBCNSW 2153 Phone"
            )
        )
        self.assertTrue(parsing_module._looks_like_spec_room_label_noise("Location"))
        self.assertTrue(parsing_module._looks_like_spec_room_label_noise("Additional Wet Area"))
        self.assertFalse(parsing_module._looks_like_spec_room_label_noise("KITCHEN"))

    def test_parse_documents_defaults_grouped_vanities_door_colours_to_base_without_explicit_overheads(self) -> None:
        snapshot = parse_documents(
            job_no="37061",
            builder_name="Clarendon",
            source_kind="spec",
            documents=[
                {
                    "file_name": "vanities-schedule.pdf",
                    "role": "spec",
                    "pages": [
                        {
                            "page_no": 1,
                            "text": (
                                "VANITIES COLOUR SCHEDULE\n"
                                "Door/Panel Colour - Polytec Blossom White Matt Finish Thermolaminate - Hamptons EM9 Profile\n"
                                "Door/Panel Colour 2 - Polytec Habitit Smooth Finish Thermolaminate - Hamptons EM9 Profile\n"
                                "Benchtop - Quantum Zero Luna White - 20MM Pencil Round Edge / 140MM Mitred Apron Edge - to Powder Room 2\n"
                            ),
                            "needs_ocr": False,
                        }
                    ],
                }
            ],
        )
        vanities = snapshot["rooms"][0]
        self.assertEqual(vanities["room_key"], "vanities")
        self.assertEqual(vanities["door_colours_overheads"], "")
        self.assertIn("Polytec Blossom White Matt Finish Thermolaminate - Hamptons EM9 Profile", vanities["door_colours_base"])
        self.assertIn("Polytec Habitit Smooth Finish Thermolaminate - Hamptons EM9 Profile", vanities["door_colours_base"])

    def test_clarendon_schedule_overlay_defaults_doors_panels_to_base_only_without_explicit_overheads(self) -> None:
        overlay = extraction_service._extract_clarendon_schedule_overlay(
            "vanities",
            (
                "VANITIES COLOUR SCHEDULE "
                "BENCHTOP - QUANTUM ZERO LUNA WHITE - 20MM PENCIL ROUND EDGE / 140MM MITRED APRON EDGE - TO POWDER ROOM 2 "
                "DOORS/PANELS - POL YTEC JAMAICAN OAK MATT FINISH MELAMINE WITH MATCHING 1MM ABS EDGES (VERTICAL GRAIN DIRECTION) "
                "KICKBOARDS - N/A FLOATING"
            ),
        )
        self.assertFalse(overlay["has_explicit_overheads"])
        self.assertEqual(overlay["door_colours_overheads"], "")
        self.assertEqual(
            overlay["door_colours_base"].lower(),
            "polytec jamaican oak matt finish melamine with matching 1mm abs edges (vertical grain direction)",
        )

    def test_clarendon_vanities_fixture_fallback_does_not_merge_material_fields(self) -> None:
        overlay = extraction_service._select_clarendon_room_overlay(
            {"room_key": "vanities", "original_room_label": "Vanities"},
            {
                "vanities": {
                    "door_colours_base": "Polytec Jamaican Oak Matt Finish Melamine with Matching 1MM ABS Edges (Vertical Grain Direction)",
                    "basin_info": "",
                    "tap_info": "",
                },
                "vanity": {
                    "door_colours_base": "Polytec Habitit Smooth Finish Thermolaminate - Hamptons EM9 Profile",
                    "basin_info": "Johnson Suisse Emilia Rectangular Undercounter Basin (JBSE250.PW6) White",
                    "tap_info": "Phoenix Nostalgia Basin Mixer NS748-62",
                },
            },
        )
        self.assertEqual(
            overlay["door_colours_base"],
            "Polytec Jamaican Oak Matt Finish Melamine with Matching 1MM ABS Edges (Vertical Grain Direction)",
        )
        self.assertEqual(overlay["basin_info"], "Johnson Suisse Emilia Rectangular Undercounter Basin (JBSE250.PW6) White")
        self.assertEqual(overlay["tap_info"], "Phoenix Nostalgia Basin Mixer NS748-62")

    def test_clarendon_polish_prefers_same_room_vanities_materials_over_contaminated_row_values(self) -> None:
        row = {
            "room_key": "vanities",
            "original_room_label": "VANITIES",
            "bench_tops": ["Quantum Zero Luna White - 20MM Pencil Round Edge / 140MM Mitred Apron Edge - to Powder Room 2"],
            "bench_tops_wall_run": "",
            "bench_tops_island": "",
            "bench_tops_other": "Quantum Zero Luna White - 20MM Pencil Round Edge / 140MM Mitred Apron Edge - to Powder Room 2",
            "door_panel_colours": [
                "Polytec Habitit Smooth Finish Thermolaminate - Hamptons EM9 Profile",
                "Polytec Blossom White Matt Finish Thermolaminate - Hamptons EM9 Profile",
            ],
            "door_colours_overheads": "",
            "door_colours_base": "Polytec Habitit Smooth Finish Thermolaminate - Hamptons EM9 Profile | Polytec Blossom White Matt Finish Thermolaminate - Hamptons EM9 Profile",
            "door_colours_island": "",
            "door_colours_bar_back": "",
            "has_explicit_overheads": False,
            "has_explicit_base": False,
            "has_explicit_island": False,
            "has_explicit_bar_back": False,
            "toe_kick": ["N/A floating"],
            "bulkheads": [],
            "handles": [],
            "sink_info": "",
            "basin_info": "",
            "tap_info": "",
            "drawers_soft_close": "",
            "hinges_soft_close": "",
            "splashback": "",
            "flooring": "",
            "source_file": "49906613 Amended Signed Drawings REV C 04-09-25.pdf",
            "page_refs": "12",
            "evidence_snippet": "",
            "confidence": 0.6,
        }
        overlay = extraction_service._select_clarendon_room_overlay(
            {"room_key": "vanities", "original_room_label": "VANITIES"},
            {
                "vanities": extraction_service._extract_clarendon_schedule_overlay(
                    "vanities",
                    (
                        "VANITIES COLOUR SCHEDULE "
                        "BENCHTOP - QUANTUM ZERO LUNA WHITE - 20MM PENCIL ROUND EDGE / 140MM MITRED APRON EDGE - TO POWDER ROOM 2 "
                        "DOORS/PANELS - POL YTEC JAMAICAN OAK MATT FINISH MELAMINE WITH MATCHING 1MM ABS EDGES (VERTICAL GRAIN DIRECTION) "
                        "KICKBOARDS - N/A FLOATING"
                    ),
                ),
                "vanity": {
                    "basin_info": "Johnson Suisse Emilia Rectangular Undercounter Basin (JBSE250.PW6) White",
                    "tap_info": "Phoenix Nostalgia Basin Mixer NS748-62",
                },
            },
        )
        polished = extraction_service._polish_clarendon_room(row, overlay)
        self.assertEqual(
            polished["door_colours_base"],
            "Polytec Jamaican Oak Matt Finish Melamine with Matching 1MM ABS Edges (Vertical Grain Direction)",
        )
        self.assertEqual(polished["door_colours_overheads"], "")
        self.assertEqual(polished["basin_info"], "Johnson Suisse Emilia Rectangular Undercounter Basin (JBSE250.PW6) White")
        self.assertEqual(polished["tap_info"], "Phoenix Nostalgia Basin Mixer NS748-62")

    def test_clarendon_schedule_overlay_recovers_glued_vanities_doors_panels_from_realistic_ocr_text(self) -> None:
        text = (
            "NOTE: ALL PLUMBING SETOUT DIMENSIONS ARE FROM THE TIMBER FRAME\n"
            "REV C 29/07/25\n"
            "BENCHTOP - QUANTUM ZERO LUNA WHITE - 20MM PENCIL ROUND EDGE / 140MM MITRED APRON EDGE - TO POWDER ROOM 2"
            "DOORS/PANELS - POL YTEC JAMAICAN OAK MATT FINISH MELAMINE WITH MATCHING 1MM ABS EDGES (VERTICAL GRAIN DIRECTION) "
            "KICKBOARDS - N/A FLOATING BENCHTOP SHADOWLINE - AS DOOR COLOUR (HORIZONTAL GRAIN DIRECTION)"
            "CARCASS & SHELF EDGES - STANDARD WHITEHANDLES - HETTICH NARNI 9995574 BRUSHED STAINLESS STEEL 37MM LONG - PROUD MOUNTED DOORS "
            "WITH MIN' REVEALS TO DRAWERSDOOR HINGES - HETTICH STANDARD HINGES - NOT SOFT CLOSEDRAWER RUNNERS - HETTICH INNOTECH ATIRA SOFT CLOSE RUNNERS "
            "VANITIES COLOUR SCHEDULE"
        )
        overlay = extraction_service._extract_clarendon_schedule_overlay("vanities", text)
        self.assertEqual(overlay["door_colours_overheads"], "")
        self.assertEqual(
            overlay["door_colours_base"].lower(),
            "polytec jamaican oak matt finish melamine with matching 1mm abs edges (vertical grain direction)",
        )
        self.assertEqual(
            overlay["bench_tops_other"],
            "Quantum Zero Luna White - 20MM Pencil Round Edge / 140MM Mitred Apron Edge - to Powder Room 2",
        )

    def test_material_summary_keeps_distinct_benchtop_thickness_and_edge_variants(self) -> None:
        summary = _build_material_summary(
            {
                "rooms": [
                    {
                        "room_key": "kitchen",
                        "original_room_label": "Kitchen",
                        "bench_tops_wall_run": "Quantum Zero White Swirl - 20MM Pencil Round Edge",
                        "bench_tops_island": "Quantum Zero White Swirl - 40MM Mitred Apron Edge - to Island Bench",
                        "bench_tops_other": "",
                        "door_panel_colours": [],
                    }
                ]
            }
        )
        self.assertEqual(summary["bench_tops"]["count"], 2)
        texts = [entry["text"] for entry in summary["bench_tops"]["entries"]]
        displays = [entry["display_text"] for entry in summary["bench_tops"]["entries"]]
        self.assertIn("Quantum Zero White Swirl - 20MM Pencil Round Edge", texts)
        self.assertIn("Quantum Zero White Swirl - 40MM Mitred Apron Edge", texts)
        self.assertIn("Quantum Zero White Swirl - 20MM Pencil Round Edge (Kitchen)", displays)

    def test_material_summary_includes_floating_shelf_material(self) -> None:
        summary = _build_material_summary(
            {
                "rooms": [
                    {
                        "room_key": "office",
                        "original_room_label": "OFFICE",
                        "bench_tops_other": "",
                        "floating_shelf": "Polytec Boston Oak Woodmatt 33mm pencil round edge",
                    }
                ]
            }
        )
        self.assertIn(
            "Polytec Boston Oak Woodmatt 33mm pencil round edge",
            [entry["text"] for entry in summary["bench_tops"]["entries"]],
        )

    def test_material_summary_includes_room_labels_for_shared_values(self) -> None:
        summary = _build_material_summary(
            {
                "rooms": [
                    {
                        "room_key": "kitchen",
                        "original_room_label": "Kitchen",
                        "bench_tops_wall_run": "40mm stone - Arissed - By Builder",
                        "handles": ["No handles - Bronte Handle"],
                        "door_colours_base": "Polytec Classic White Matt",
                    },
                    {
                        "room_key": "wip",
                        "original_room_label": "WIP",
                        "bench_tops_other": "40mm stone - Arissed - By Builder",
                        "handles": ["No handles - Bronte Handle"],
                        "door_colours_base": "Polytec Classic White Matt",
                    },
                ]
            }
        )
        bench_display = [entry["display_text"] for entry in summary["bench_tops"]["entries"]]
        handle_display = [entry["display_text"] for entry in summary["handles"]["entries"]]
        door_display = [entry["display_text"] for entry in summary["door_colours"]["entries"]]
        self.assertIn("40mm stone - Arissed - By Builder (Kitchen / WIP)", bench_display)
        self.assertIn("No handles - Bronte Handle (Kitchen / WIP)", handle_display)
        self.assertIn("Polytec Classic White Matt (Kitchen / WIP)", door_display)

    def test_parse_documents_recovers_glued_schedule_headings_from_room_master(self) -> None:
        snapshot = parse_documents(
            job_no="37869",
            builder_name="Clarendon",
            source_kind="spec",
            documents=[
                {
                    "file_name": "drawings-and-colours.pdf",
                    "role": "spec",
                    "pages": [
                        {
                            "page_no": 1,
                            "text": (
                                "KITCHEN COLOUR SCHEDULEBENCHTOP - QUANTUM ZERO BELLA CARRARA - 20MM PENCIL ROUND EDGE\n"
                                "ROOM BUTLERS PANTRY COLOUR SCHEDULEPANTRY\n"
                                "VANITIES COLOUR SCHEDULENOTE : ALL PLUMBING SETOUT DIMENSIONS ARE FROM THE TIMBER FRAME\n"
                                "LAUNDRY COLOUR SCHEDULEBENCHTOP - POLYTEC ARGENTO STONE - 21MM TIGHTFORM EDGE LAMINATE\n"
                            ),
                            "needs_ocr": False,
                        }
                    ],
                },
                {
                    "file_name": "colours-afc.pdf",
                    "role": "spec",
                    "pages": [
                        {
                            "page_no": 1,
                            "text": (
                                "Study\n"
                                "Hybrid flooring\n"
                                "Main Bathroom\n"
                                "Vanity Inset Basin JOHNSON SUISSE Emilia Basin (JBSE250.PW6)\n"
                                "Laundry door: ALUMINUM SLIDING 2340MM HT\n"
                            ),
                            "needs_ocr": False,
                        }
                    ],
                },
            ],
        )
        rooms = {row["room_key"]: row for row in snapshot["rooms"]}
        analysis = snapshot["analysis"]
        warning_text = " | ".join(snapshot["warnings"])
        self.assertEqual(set(rooms.keys()), {"kitchen", "butlers_pantry", "vanities", "laundry"})
        self.assertEqual(analysis["room_master_file"], "drawings-and-colours.pdf")
        self.assertEqual(analysis["supplement_files"], ["colours-afc.pdf"])
        self.assertGreaterEqual(analysis["ignored_room_like_lines_count"], 1)
        self.assertTrue(str(rooms["vanities"]["basin_info"]).startswith("Johnson Suisse Emilia"))
        self.assertIn("Study", warning_text)
        self.assertNotIn("KITCHEN COLOUR SCHEDULEBENCHTOP", " ".join(room["original_room_label"] for room in snapshot["rooms"]))

    def test_parse_documents_prefilters_supplement_rooms_even_when_master_file_is_second(self) -> None:
        snapshot = parse_documents(
            job_no="37736",
            builder_name="Clarendon",
            source_kind="spec",
            documents=[
                {
                    "file_name": "colours-afc.pdf",
                    "role": "spec",
                    "pages": [
                        {
                            "page_no": 1,
                            "text": (
                                "Kitchen/Pantry/Family/Meals\n"
                                "Theatre\n"
                                "Rumpus\n"
                                "WIR/S & Robes\n"
                                "Main Bathroom\n"
                                "Vanity Inset Basin JOHNSON SUISSE Emilia Basin (JBSE250.PW6)\n"
                            ),
                            "needs_ocr": False,
                        }
                    ],
                },
                {
                    "file_name": "drawings-and-colours.pdf",
                    "role": "spec",
                    "pages": [
                        {
                            "page_no": 1,
                            "text": (
                                "KITCHEN COLOUR SCHEDULEBENCHTOP - QUANTUM ZERO BELLA CARRARA - 20MM PENCIL ROUND EDGE\n"
                                "ROOM BUTLERS PANTRY COLOUR SCHEDULEPANTRY\n"
                                "VANITIES COLOUR SCHEDULENOTE : ALL PLUMBING SETOUT DIMENSIONS ARE FROM THE TIMBER FRAME\n"
                                "LAUNDRY COLOUR SCHEDULEBENCHTOP - POLYTEC ARGENTO STONE - 21MM TIGHTFORM EDGE LAMINATE\n"
                            ),
                            "needs_ocr": False,
                        }
                    ],
                },
            ],
        )
        rooms = {row["room_key"]: row for row in snapshot["rooms"]}
        analysis = snapshot["analysis"]
        warning_text = " | ".join(snapshot["warnings"])
        self.assertEqual(set(rooms.keys()), {"kitchen", "butlers_pantry", "vanities", "laundry"})
        self.assertEqual(analysis["room_master_file"], "drawings-and-colours.pdf")
        self.assertEqual(analysis["supplement_files"], ["colours-afc.pdf"])
        self.assertGreaterEqual(analysis["ignored_room_like_lines_count"], 1)
        self.assertTrue(str(rooms["vanities"]["basin_info"]).startswith("Johnson Suisse Emilia"))
        self.assertEqual(rooms["vanities"]["original_room_label"], "VANITIES")
        self.assertTrue(bool(warning_text))
        self.assertIn("Theatre", warning_text)

    def test_source_room_label_preserves_walk_in_pantry_full_name(self) -> None:
        self.assertEqual(
            parsing_module.source_room_label("WALK-IN-PANTRY COLOUR SCHEDULE Pantry"),
            "WALK-IN-PANTRY",
        )
        self.assertEqual(
            parsing_module.source_room_key("WALK-IN-PANTRY COLOUR SCHEDULE Pantry"),
            "walk_in_pantry",
        )

    def test_parse_documents_keeps_walk_in_pantry_and_meals_room_from_master_schedule(self) -> None:
        snapshot = parse_documents(
            job_no="37825",
            builder_name="Clarendon",
            source_kind="spec",
            documents=[
                {
                    "file_name": "drawings-and-colours.pdf",
                    "role": "spec",
                    "pages": [
                        {
                            "page_no": 1,
                            "text": (
                                "BUTLERS PANTRY COLOUR SCHEDULE Pantry\n"
                                "BENCHTOP - QUANTUM ZERO MONTE BIANCO - 20MM PENCIL ROUND EDGE\n"
                                "DOOR COLOUR - POL YTEC BLOSSOM WHITE SMOOTH THERMOLAMINATE FINISH - CLASSIC SQUARE EM4 PROFILE\n"
                            ),
                            "needs_ocr": False,
                        },
                        {
                            "page_no": 2,
                            "text": (
                                "WALK-IN-PANTRY COLOUR SCHEDULE Pantry\n"
                                "BENCHTOP - QUANTUM ZERO MONTE BIANCO - 20MM PENCIL ROUND EDGE\n"
                                "DOOR COLOUR 1 - POL YTEC BLOSSOM WHITE SMOOTH THERMOLAMINATE FINISH - CLASSIC SQUARE EM4 PROFILE\n"
                            ),
                            "needs_ocr": False,
                        },
                        {
                            "page_no": 3,
                            "text": (
                                "MEALS ROOM COLOUR SCHEDULEBENCHTOP - POL YTEC 'PLANTATION ASH' WOODMATT FINISH 33MM SQUARE EDGE LAMINATE\n"
                                "DOOR COLOUR - POL YTEC BLOSSOM WHITE SMOOTH THERMOLAMINATE FINISH - CLASSIC SQUARE EM4 PROFILE\n"
                                "DOOR HINGES - HETTICH SOFT CLOSE\n"
                            ),
                            "needs_ocr": False,
                        },
                    ],
                },
                {
                    "file_name": "colours-afc.pdf",
                    "role": "spec",
                    "pages": [
                        {
                            "page_no": 1,
                            "text": (
                                "Kitchen/Pantry/Family/Meals\n"
                                "HYBRID FLOORING\n"
                            ),
                            "needs_ocr": False,
                        }
                    ],
                },
            ],
        )
        rooms = {row["room_key"]: row for row in snapshot["rooms"]}
        warning_text = " | ".join(snapshot["warnings"])
        analysis = snapshot["analysis"]
        self.assertEqual(analysis["room_master_file"], "drawings-and-colours.pdf")
        self.assertEqual(rooms["butlers_pantry"]["original_room_label"], "BUTLERS PANTRY")
        self.assertEqual(rooms["walk_in_pantry"]["original_room_label"], "WALK-IN-PANTRY")
        self.assertEqual(rooms["meals_room"]["original_room_label"], "MEALS ROOM")
        self.assertTrue(str(rooms["meals_room"]["hinges_soft_close"]).startswith("Soft Close"))
        self.assertTrue(bool(rooms["meals_room"]["bench_tops"]))
        self.assertTrue(bool(warning_text))

    def test_builder_defaults_to_global_conservative_for_all_builders(self) -> None:
        clarendon_id = store.create_builder("Clarendon", "clarendon", "")
        yellowwood_id = store.create_builder("Yellowwood", "yellowwood", "")
        self.assertEqual(store.get_builder(clarendon_id)["parser_strategy"], "global_conservative")
        self.assertEqual(store.get_builder(yellowwood_id)["parser_strategy"], "global_conservative")

    def test_spec_list_page_requires_login(self) -> None:
        builder_id = store.create_builder("Clarendon", "clarendon", "")
        job_id = store.create_job("37529", builder_id, "Spec List Test", "")
        client = TestClient(app)
        response = client.get(f"/jobs/{job_id}/spec-list", follow_redirects=False)
        self.assertEqual(response.status_code, 303)
        self.assertEqual(response.headers["location"], "/login")

    def test_job_detail_shows_parse_actions_analysis_and_auto_upload_inputs(self) -> None:
        builder_id = store.create_builder("Clarendon", "clarendon", "")
        job_id = store.create_job("37529", builder_id, "Diagnostics Test", "")
        raw_snapshot = {
            "job_no": "37529",
            "builder_name": "Clarendon",
            "source_kind": "spec",
            "generated_at": "2026-03-22T10:00:00+00:00",
            "analysis": {
                "mode": "openai_merged",
                "parser_strategy": "global_conservative",
                "openai_attempted": True,
                "openai_succeeded": True,
                "openai_model": "gpt-4.1-mini",
                "note": "OpenAI result merged with heuristic parsing.",
                "worker_pid": 4242,
                "app_build_id": "build-test",
                "room_master_file": "drawings-and-colours.pdf",
                "room_master_reason": "drawings-and-colours.pdf selected as room master by schedule density.",
                "supplement_files": ["colours-afc.pdf"],
                "ignored_room_like_lines_count": 7,
            },
            "rooms": [],
            "appliances": [],
            "others": {},
            "warnings": [],
            "source_documents": [],
        }
        store.upsert_snapshot(job_id, "raw_spec", raw_snapshot)
        client = TestClient(app)
        self._login(client)
        response = client.get(f"/jobs/{job_id}")
        self.assertEqual(response.status_code, 200)
        self.assertIn("Parse Spec Files", response.text)
        self.assertIn("Parse Drawing Files", response.text)
        self.assertIn("OpenAI merged", response.text)
        self.assertIn("Global Conservative", response.text)
        self.assertIn("build-test", response.text)
        self.assertIn("gpt-4.1-mini", response.text)
        self.assertIn("drawings-and-colours.pdf", response.text)
        self.assertIn("Ignored unmatched room-like lines", response.text)
        self.assertIn("requestSubmit()", response.text)
        self.assertNotIn("Upload Specs", response.text)
        self.assertNotIn("Upload Drawings", response.text)
        self.assertNotIn("Builder Cleaning Rules", response.text)

    def test_job_detail_handles_scalar_room_fields_in_review_snapshot(self) -> None:
        builder_id = store.create_builder("Clarendon", "clarendon", "")
        job_id = store.create_job("37006", builder_id, "Legacy Snapshot Shapes", "")
        snapshot = {
            "job_no": "37006",
            "builder_name": "Clarendon",
            "source_kind": "spec",
            "generated_at": "2026-03-22T10:00:00+00:00",
            "analysis": {
                "mode": "openai_merged",
                "openai_attempted": True,
                "openai_succeeded": True,
                "openai_model": "gpt-4.1-mini",
                "note": "Legacy scalar room fields should still render.",
            },
            "rooms": [
                {
                    "room_key": "Kitchen",
                    "original_room_label": "Kitchen",
                    "bench_tops": "20mm stone bench",
                    "door_panel_colours": "Polytec White",
                    "toe_kick": "Matching melamine",
                    "bulkheads": None,
                    "handles": "Hettich knob",
                    "drawers_soft_close": "Yes",
                    "hinges_soft_close": "No",
                    "splashback": "Tiled by others",
                    "flooring": "Tile",
                    "source_file": "legacy.pdf",
                    "page_refs": ["PAGE 2", "PAGE 3"],
                    "evidence_snippet": "Legacy mixed field shapes",
                    "confidence": "High",
                }
            ],
            "appliances": [],
            "others": {},
            "warnings": [],
            "source_documents": [],
        }
        store.upsert_snapshot(job_id, "raw_spec", snapshot)
        store.upsert_review(job_id, snapshot)
        client = TestClient(app)
        self._login(client)
        response = client.get(f"/jobs/{job_id}")
        self.assertEqual(response.status_code, 200)
        self.assertNotIn("<span class=\"eyebrow\">Review</span>", response.text)
        self.assertIn("Legacy scalar room fields should still render.", response.text)
        self.assertIn("OpenAI merged", response.text)
        self.assertIn("Open Spec List", response.text)

    def test_run_history_partial_shows_live_stage_and_message(self) -> None:
        builder_id = store.create_builder("Clarendon", "clarendon", "")
        job_id = store.create_job("37529", builder_id, "Run Status Test", "")
        run_id = store.create_run(job_id, "spec")
        with mock.patch("App.services.store._pid_is_running", side_effect=lambda pid: int(pid or 0) == 4242):
            self.assertTrue(store.acquire_worker_lease("worker-a", 4242, "build-test"))
            self.assertFalse(store.acquire_worker_lease("worker-b", 9898, "build-other"))
        claimed = store.claim_next_run(worker_pid=4242, app_build_id="build-test", worker_token="worker-a")
        self.assertEqual(claimed["id"], run_id)
        store.update_run_runtime_metadata(run_id, "global_conservative", 4242, "build-test")
        store.update_run_progress(run_id, "official_size_extraction", "Extracting official dimensions from spec PDF for Westinghouse WHC943BD", worker_token="worker-a")
        client = TestClient(app)
        self._login(client)
        response = client.get(f"/jobs/{job_id}/run-history")
        self.assertEqual(response.status_code, 200)
        self.assertIn("Official size extraction", response.text)
        self.assertIn("Extracting official dimensions from spec PDF for Westinghouse WHC943BD", response.text)
        self.assertIn("Global Conservative", response.text)
        self.assertIn("Worker / Build", response.text)
        self.assertIn("PID 4242 | build-test", response.text)
        self.assertIn("Duration", response.text)
        self.assertIn("hx-trigger=\"load, every 2s\"", response.text)

    def test_run_history_partial_shows_open_result_for_succeeded_spec_run(self) -> None:
        builder_id = store.create_builder("Clarendon", "clarendon", "")
        job_id = store.create_job("37530", builder_id, "Run Result Test", "")
        run_id = store.create_run(job_id, "spec")
        payload = {
            "job_no": "37530",
            "builder_name": "Clarendon",
            "source_kind": "spec",
            "generated_at": "2026-04-03T10:00:00+10:00",
            "analysis": {"mode": "heuristic_only", "parser_strategy": "global_conservative"},
            "rooms": [{"room_key": "kitchen", "original_room_label": "KITCHEN"}],
            "appliances": [],
            "others": {},
            "warnings": [],
            "source_documents": [],
        }
        with store.connect() as conn:
            conn.execute(
                "UPDATE runs SET status = 'succeeded', started_at = ?, finished_at = ?, result_json = ? WHERE id = ?",
                ("2026-04-03T00:00:00+00:00", "2026-04-03T00:01:30+00:00", json.dumps(payload), run_id),
            )
        client = TestClient(app)
        self._login(client)
        response = client.get(f"/jobs/{job_id}/run-history")
        self.assertEqual(response.status_code, 200)
        self.assertIn("1m 30s", response.text)
        self.assertIn(f"/jobs/{job_id}/runs/{run_id}/spec-list", response.text)

    def test_historical_spec_list_page_uses_run_result_json_read_only(self) -> None:
        builder_id = store.create_builder("Clarendon", "clarendon", "")
        job_id = store.create_job("37531", builder_id, "Historical Spec Result", "")
        store.upsert_snapshot(
            job_id,
            "raw_spec",
            {
                "job_no": "37531",
                "builder_name": "Clarendon",
                "source_kind": "spec",
                "generated_at": "2026-04-03T09:00:00+10:00",
                "analysis": {"mode": "heuristic_only", "parser_strategy": "global_conservative"},
                "rooms": [{"room_key": "pantry", "original_room_label": "PANTRY"}],
                "appliances": [],
                "others": {},
                "warnings": [],
                "source_documents": [],
            },
        )
        run_id = store.create_run(job_id, "spec")
        historical_payload = {
            "job_no": "37531",
            "builder_name": "Clarendon",
            "source_kind": "spec",
            "generated_at": "2026-04-03T10:00:00+10:00",
            "analysis": {"mode": "heuristic_only", "parser_strategy": "global_conservative"},
            "rooms": [{"room_key": "kitchen", "original_room_label": "KITCHEN"}],
            "appliances": [],
            "others": {},
            "warnings": [],
            "source_documents": [],
        }
        with store.connect() as conn:
            conn.execute(
                "UPDATE runs SET status = 'succeeded', started_at = ?, finished_at = ?, result_json = ? WHERE id = ?",
                ("2026-04-03T00:00:00+00:00", "2026-04-03T00:00:45+00:00", json.dumps(historical_payload), run_id),
            )
        client = TestClient(app)
        self._login(client)
        response = client.get(f"/jobs/{job_id}/runs/{run_id}/spec-list")
        self.assertEqual(response.status_code, 200)
        self.assertIn("Historical spec result", response.text)
        self.assertIn("KITCHEN", response.text)
        self.assertNotIn("PANTRY", response.text)
        self.assertIn("Export Disabled", response.text)
        self.assertIn("Open Latest Spec List", response.text)

    def test_jobs_page_renders_open_as_button(self) -> None:
        builder_id = store.create_builder("Clarendon", "clarendon", "")
        job_id = store.create_job("37532", builder_id, "Jobs Page Button", "")
        client = TestClient(app)
        self._login(client)
        response = client.get("/jobs")
        self.assertEqual(response.status_code, 200)
        self.assertIn(f'href="/jobs/{job_id}" class="ghost-button"', response.text)

    def test_run_duration_display_handles_running_and_missing_started_at(self) -> None:
        running = {
            "status": "running",
            "started_at": "2026-04-03T00:00:00+00:00",
            "finished_at": "",
        }
        queued = {"status": "queued", "started_at": "", "finished_at": ""}
        self.assertEqual(
            _run_duration_display(running, now=datetime.fromisoformat("2026-04-03T10:00:30+10:00")),
            "30s",
        )
        self.assertEqual(_run_duration_display(queued), "-")

    def test_parse_request_requires_uploaded_files(self) -> None:
        builder_id = store.create_builder("Clarendon", "clarendon", "")
        job_id = store.create_job("37529", builder_id, "Missing Files Test", "")
        client = TestClient(app)
        self._login(client)
        job_page = client.get(f"/jobs/{job_id}")
        csrf = job_page.text.split('name="csrf_token" value="', 1)[1].split('"', 1)[0]
        response = client.post(
            f"/jobs/{job_id}/runs/start",
            data={"run_kind": "spec", "csrf_token": csrf},
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 303)
        self.assertEqual(response.headers["location"], f"/jobs/{job_id}")
        self.assertEqual(store.list_runs(job_id), [])

    def test_jobs_page_filters_by_job_number_query(self) -> None:
        builder_id = store.create_builder("Clarendon", "clarendon", "")
        store.create_job("37529", builder_id, "Kitchen Spec", "")
        store.create_job("47001", builder_id, "Laundry Spec", "")
        client = TestClient(app)
        self._login(client)
        response = client.get("/jobs?q=375")
        self.assertEqual(response.status_code, 200)
        self.assertIn("37529", response.text)
        self.assertNotIn("47001", response.text)
        self.assertIn('value="375"', response.text)

    def test_jobs_page_sorts_by_created_or_updated_time(self) -> None:
        builder_id = store.create_builder("Clarendon", "clarendon", "")
        older_job = store.create_job("11111", builder_id, "Older Job", "")
        newer_job = store.create_job("22222", builder_id, "Newer Job", "")
        with store.connect() as conn:
            conn.execute(
                "UPDATE jobs SET created_at = ?, updated_at = ? WHERE id = ?",
                ("2026-04-01T00:00:00+00:00", "2026-04-03T00:00:00+00:00", older_job),
            )
            conn.execute(
                "UPDATE jobs SET created_at = ?, updated_at = ? WHERE id = ?",
                ("2026-04-02T00:00:00+00:00", "2026-04-02T00:00:00+00:00", newer_job),
            )
        client = TestClient(app)
        self._login(client)
        created_response = client.get("/jobs?sort=created_desc")
        self.assertEqual(created_response.status_code, 200)
        self.assertLess(created_response.text.index("22222"), created_response.text.index("11111"))
        updated_response = client.get("/jobs?sort=updated_desc")
        self.assertEqual(updated_response.status_code, 200)
        self.assertIn("Sort by", updated_response.text)
        self.assertLess(updated_response.text.index("11111"), updated_response.text.index("22222"))

    def test_jobs_page_shows_room_count_from_latest_raw_snapshot(self) -> None:
        builder_id = store.create_builder("Clarendon", "clarendon", "")
        job_with_rooms = store.create_job("37529", builder_id, "Kitchen Spec", "")
        job_without_rooms = store.create_job("47001", builder_id, "Laundry Spec", "")
        store.upsert_snapshot(
            job_with_rooms,
            "raw_spec",
            {
                "job_no": "37529",
                "builder_name": "Clarendon",
                "source_kind": "spec",
                "generated_at": "2026-03-22T10:00:00+00:00",
                "analysis": {"mode": "heuristic_only", "parser_strategy": "global_conservative"},
                "rooms": [{"room_key": "kitchen"}, {"room_key": "laundry"}],
                "appliances": [],
                "others": {},
                "warnings": [],
                "source_documents": [],
            },
        )
        client = TestClient(app)
        self._login(client)
        response = client.get("/jobs")
        self.assertEqual(response.status_code, 200)
        self.assertIn("Last Updated", response.text)
        self.assertIn('name="sort" value="created_desc"', response.text)
        self.assertIn('name="sort" value="updated_desc"', response.text)
        self.assertIn('data-label="Rooms">2<', response.text)
        self.assertIn('data-label="Last Updated">', response.text)
        self.assertIn("Delete", response.text)
        self.assertIn(f'/jobs/{job_with_rooms}/delete', response.text)

    def test_delete_job_removes_database_records_and_job_directory(self) -> None:
        builder_id = store.create_builder("Clarendon", "clarendon", "")
        job_id = store.create_job("37529", builder_id, "Delete Me", "")
        dirs = ensure_job_dirs("37529")
        marker = dirs["job_root"] / "marker.txt"
        marker.write_text("delete me", encoding="utf-8")
        store.create_job_file(job_id, "spec", "sample.pdf", "sample.pdf", "application/pdf", 123)
        store.upsert_snapshot(
            job_id,
            "raw_spec",
            {
                "job_no": "37529",
                "builder_name": "Clarendon",
                "source_kind": "spec",
                "generated_at": "2026-03-22T10:00:00+00:00",
                "analysis": {"mode": "heuristic_only", "parser_strategy": "global_conservative"},
                "rooms": [],
                "appliances": [],
                "others": {},
                "warnings": [],
                "source_documents": [],
            },
        )
        client = TestClient(app)
        self._login(client)
        jobs_page = client.get("/jobs")
        csrf = jobs_page.text.split('name="csrf_token" value="', 1)[1].split('"', 1)[0]
        response = client.post(f"/jobs/{job_id}/delete", data={"csrf_token": csrf}, follow_redirects=False)
        self.assertEqual(response.status_code, 303)
        self.assertEqual(response.headers["location"], "/jobs")
        self.assertIsNone(store.get_job(job_id))
        self.assertFalse(dirs["job_root"].exists())

    def test_delete_job_refuses_when_run_is_active(self) -> None:
        builder_id = store.create_builder("Clarendon", "clarendon", "")
        job_id = store.create_job("37529", builder_id, "Delete Me", "")
        store.create_run(job_id, "spec")
        with store.connect() as conn:
            conn.execute("UPDATE runs SET status = 'running' WHERE job_id = ?", (job_id,))
        dirs = ensure_job_dirs("37529")
        marker = dirs["job_root"] / "marker.txt"
        marker.write_text("keep me", encoding="utf-8")
        client = TestClient(app)
        self._login(client)
        jobs_page = client.get("/jobs")
        csrf = jobs_page.text.split('name="csrf_token" value="', 1)[1].split('"', 1)[0]
        response = client.post(f"/jobs/{job_id}/delete", data={"csrf_token": csrf}, follow_redirects=False)
        self.assertEqual(response.status_code, 303)
        self.assertEqual(response.headers["location"], "/jobs")
        self.assertIsNotNone(store.get_job(job_id))
        self.assertTrue(dirs["job_root"].exists())

    def test_builder_rules_routes_redirect_and_builders_page_uses_global_profile(self) -> None:
        builder_a = store.create_builder("Clarendon", "clarendon", "")
        builder_b = store.create_builder("Yellowwood", "yellowwood", "")
        job_id = store.create_job("37016", builder_a, "Rules Job", "")
        store.upsert_snapshot(
            job_id,
            "raw_spec",
            {
                "job_no": "37016",
                "builder_name": "Clarendon",
                "source_kind": "spec",
                "generated_at": "2026-03-22T10:00:00+00:00",
                "analysis": {
                    "mode": "heuristic_only",
                    "parser_strategy": "global_conservative",
                    "openai_attempted": False,
                    "openai_succeeded": False,
                    "openai_model": "",
                    "note": "",
                },
                "rooms": [],
                "appliances": [],
                "others": {},
                "warnings": [],
                "source_documents": [],
            },
        )
        client = TestClient(app)
        self._login(client)

        builders_page = client.get("/builders")
        self.assertEqual(builders_page.status_code, 200)
        self.assertIn("global conservative extraction profile", builders_page.text.lower())
        self.assertNotIn("Cleaning Rules", builders_page.text)
        csrf = builders_page.text.split('name="csrf_token" value="', 1)[1].split('"', 1)[0]

        page = client.get(f"/builders/{builder_a}/rules", follow_redirects=False)
        self.assertEqual(page.status_code, 303)
        self.assertEqual(page.headers["location"], "/builders")

        response = client.post(
            f"/builders/{builder_a}/rules",
            data={"csrf_token": csrf},
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 303)
        self.assertEqual(response.headers["location"], "/builders")
        builder_a_row = store.get_builder(builder_a)
        builder_b_row = store.get_builder(builder_b)
        self.assertEqual(builder_a_row["parser_strategy"], "global_conservative")
        self.assertEqual(builder_b_row["parser_strategy"], "global_conservative")
        self.assertTrue(builder_a_row["rule_flags"]["normalize_brand_casing"])
        self.assertTrue(builder_b_row["rule_flags"]["normalize_brand_casing"])

        job_page = client.get(f"/jobs/{job_id}")
        self.assertEqual(job_page.status_code, 200)
        self.assertIn("Global Extraction Profile", job_page.text)
        self.assertNotIn("Builder Cleaning Rules", job_page.text)

    def test_spec_list_page_shows_material_summary_official_links_and_unicode_excel(self) -> None:
        builder_id = store.create_builder("Clarendon", "clarendon", "")
        job_id = store.create_job("37529", builder_id, "Unicode Test", "")
        raw_snapshot = {
            "job_no": "37529",
            "builder_name": "Clarendon",
            "source_kind": "spec",
            "generated_at": "2026-03-22T10:00:00+00:00",
            "analysis": {
                "mode": "heuristic_only",
                "openai_attempted": False,
                "openai_succeeded": False,
                "openai_model": "gpt-4.1-mini",
                "note": "OpenAI is disabled in runtime settings.",
            },
            "rooms": [
                {
                    "room_key": "kitchen",
                    "original_room_label": "Kitchen \u4e2d\u6587",
                    "bench_tops": [
                        "Quantum Zero Midnight Black 20mm pencil round edge to cooktop run and Quantum Zero Venatino Statuario 40mm mitred apron edge to island bench",
                    ],
                    "door_panel_colours": [
                        "Polytec Blossom White Matt Finish - overhead cabinetry",
                        "Polytec Blossom White Matt Finish - base cabinetry",
                        "Polytec Tempest Woodgrain - island base cabinetry",
                    ],
                    "door_colours_overheads": "Polytec Blossom White Matt Finish - overhead cabinetry",
                    "door_colours_base": "Polytec Blossom White Matt Finish - base cabinetry",
                    "door_colours_island": "Polytec Tempest Woodgrain - island base cabinetry",
                    "door_colours_bar_back": "",
                    "toe_kick": ["Matching finish"],
                    "bulkheads": ["Bulkhead A"],
                    "handles": [
                        "Hettich Cipri 9070585 Gloss Chrome Plated 30mm Knob - door location: 30mm in and 60mm up/down",
                        "Hettich Cipri 9070585 Gloss Chrome Plated 30mm Knob - drawer location: centre to profile",
                    ],
                    "sink_info": "PARISI Quadro Double Bowl (PK8644)",
                    "basin_info": "",
                    "tap_info": "PHOENIX Nostalgia Sink Mixer NS714-62",
                    "drawers_soft_close": "Soft Close",
                    "hinges_soft_close": "No",
                    "splashback": "Glass / \u4e2d\u6587",
                    "flooring": "Hybrid flooring",
                    "source_file": "sample spec.pdf",
                    "page_refs": "12",
                    "evidence_snippet": "Kitchen evidence with \u4e2d\u6587 and symbols",
                    "confidence": 0.91,
                }
            ],
            "appliances": [
                {
                    "appliance_type": "Cooktop",
                    "make": "Westinghouse",
                    "model_no": "WHC943BD",
                    "product_url": "https://official.example/product/WHC943BD",
                    "spec_url": "https://official.example/spec/WHC943BD.pdf",
                    "manual_url": "https://official.example/manual/WHC943BD.pdf",
                    "website_url": "https://official.example/product/WHC943BD",
                    "overall_size": "900 x 510 x 60 mm",
                    "source_file": "sample spec.pdf",
                    "page_refs": "13",
                    "evidence_snippet": "Cooktop evidence \u4e2d\u6587",
                    "confidence": 0.87,
                },
                {
                    "appliance_type": "Sink",
                    "make": "Parisi",
                    "model_no": "PK8644",
                    "product_url": "https://official.example/sink",
                    "spec_url": "",
                    "manual_url": "",
                    "website_url": "https://official.example/sink",
                    "overall_size": "860 x 440 x 210 mm",
                    "source_file": "sample spec.pdf",
                    "page_refs": "12",
                    "evidence_snippet": "Sink evidence",
                    "confidence": 0.81,
                },
            ],
            "others": {
                "flooring_notes": "Hybrid flooring / \u4e2d\u6587",
                "splashback_notes": "Glass splashback",
            },
            "warnings": ["Low-text page detected in template.pdf page 8."],
            "source_documents": [{"file_name": "template.pdf", "role": "spec", "page_count": "20"}],
        }
        reviewed_snapshot = {
            **raw_snapshot,
            "rooms": [{**raw_snapshot["rooms"][0], "splashback": "Reviewed value should not appear"}],
        }
        store.upsert_snapshot(job_id, "raw_spec", raw_snapshot)
        store.upsert_review(job_id, reviewed_snapshot)

        client = TestClient(app)
        self._login(client)

        page = client.get(f"/jobs/{job_id}/spec-list")
        self.assertEqual(page.status_code, 200)
        self.assertIn("Material Summary", page.text)
        self.assertIn("2 distinct items", page.text)
        self.assertIn("Hettich Cipri 9070585 Gloss Chrome Plated 30mm Knob", page.text)
        self.assertIn("Quantum Zero Midnight Black 20mm pencil round edge", page.text)
        self.assertIn("Quantum Zero Venatino Statuario 40mm mitred apron edge", page.text)
        self.assertIn("Kitchen \u4e2d\u6587", page.text)
        self.assertIn("PARISI Quadro Double Bowl (PK8644)", page.text)
        self.assertIn("PHOENIX Nostalgia Sink Mixer NS714-62", page.text)
        self.assertIn("Wall Run Bench Top", page.text)
        self.assertIn("Island Bench Top", page.text)
        self.assertIn("https://official.example/product/WHC943BD", page.text)
        self.assertNotIn("Reviewed value should not appear", page.text)
        self.assertNotIn("<td>Sink</td>", page.text)
        self.assertIn("Not Soft Close", page.text)

        self._mark_raw_spec_qa_passed(job_id)
        export_response = client.get(f"/jobs/{job_id}/spec-list.xlsx")
        self.assertEqual(export_response.status_code, 200)
        workbook = load_workbook(io.BytesIO(export_response.content))
        rooms_sheet = workbook["Rooms"]
        appliances_sheet = workbook["Appliances"]
        warnings_sheet = workbook["Warnings"]
        meta_sheet = workbook["Meta"]
        room_headers = {cell.value: index + 1 for index, cell in enumerate(next(rooms_sheet.iter_rows(min_row=1, max_row=1))[0:])}
        self.assertEqual(rooms_sheet["B2"].value, "Kitchen \u4e2d\u6587")
        self.assertEqual(rooms_sheet.cell(row=2, column=room_headers["bench_tops_wall_run"]).value, "Quantum Zero Midnight Black 20mm pencil round edge")
        self.assertEqual(rooms_sheet.cell(row=2, column=room_headers["bench_tops_island"]).value, "Quantum Zero Venatino Statuario 40mm mitred apron edge")
        self.assertEqual(rooms_sheet.cell(row=2, column=room_headers["door_colours_overheads"]).value, "Polytec Blossom White Matt Finish - overhead cabinetry")
        self.assertEqual(rooms_sheet.cell(row=2, column=room_headers["sink_info"]).value, "PARISI Quadro Double Bowl (PK8644)")
        self.assertEqual(rooms_sheet.cell(row=2, column=room_headers["tap_info"]).value, "PHOENIX Nostalgia Sink Mixer NS714-62")
        self.assertEqual(rooms_sheet.cell(row=2, column=room_headers["drawers_soft_close"]).value, "Soft Close")
        self.assertEqual(rooms_sheet.cell(row=2, column=room_headers["hinges_soft_close"]).value, "Not Soft Close")
        self.assertEqual(appliances_sheet["A2"].value, "Cooktop")
        self.assertEqual(appliances_sheet["D2"].value, "https://official.example/product/WHC943BD")
        self.assertEqual(appliances_sheet["F2"].value, "900 x 510 x 60 mm")
        self.assertIsNotNone(appliances_sheet["D2"].hyperlink)
        self.assertIsNone(appliances_sheet["A3"].value)
        self.assertEqual(warnings_sheet["A2"].value, "Low-text page detected in template.pdf page 8.")
        meta_rows = [row[0] for row in meta_sheet.iter_rows(min_row=2, values_only=True)]
        self.assertIn("analysis_mode", meta_rows)
        self.assertIn("analysis_rule_flags", meta_rows)

    def test_raw_spec_snapshot_creates_pending_pdf_qa_and_blocks_exports(self) -> None:
        builder_id = store.create_builder("Imperial", "imperial", "")
        job_id = store.create_job("38251", builder_id, "Pending QA", "")
        store.upsert_snapshot(
            job_id,
            "raw_spec",
            {
                "job_no": "38251",
                "builder_name": "Imperial",
                "source_kind": "spec",
                "generated_at": utc_now_iso(),
                "analysis": {},
                "rooms": [
                    {
                        "room_key": "kitchen",
                        "original_room_label": "KITCHEN",
                        "bench_tops_wall_run": "20mm Caesarstone - Fresh Concrete",
                        "page_refs": "1-3",
                    }
                ],
                "appliances": [],
                "special_sections": [],
                "others": {},
                "warnings": [],
                "source_documents": [],
            },
        )
        verification = store.get_job_snapshot_verification(job_id, "raw_spec")
        self.assertIsNotNone(verification)
        self.assertEqual(verification["status"], "pending")
        self.assertGreater(len(verification["checklist"]), 0)

        client = TestClient(app)
        self._login(client)

        spec_page = client.get(f"/jobs/{job_id}/spec-list")
        self.assertEqual(spec_page.status_code, 200)
        self.assertIn("Pending PDF QA", spec_page.text)

        export_response = client.get(f"/jobs/{job_id}/spec-list.xlsx", follow_redirects=False)
        self.assertEqual(export_response.status_code, 303)
        self.assertEqual(export_response.headers["location"], f"/jobs/{job_id}/pdf-qa")

    def test_pdf_qa_page_can_mark_snapshot_passed_and_unlock_exports(self) -> None:
        builder_id = store.create_builder("Clarendon", "clarendon", "")
        job_id = store.create_job("37796", builder_id, "QA Pass", "")
        store.upsert_snapshot(
            job_id,
            "raw_spec",
            {
                "job_no": "37796",
                "builder_name": "Clarendon",
                "source_kind": "spec",
                "generated_at": utc_now_iso(),
                "analysis": {},
                "rooms": [
                    {
                        "room_key": "kitchen",
                        "original_room_label": "KITCHEN",
                        "bench_tops_wall_run": "Quantum Zero White Swirl - 20MM Pencil Round Edge",
                        "door_colours_base": "Polytec Aston White Smooth Finish Thermolaminate - Hampton EM9 Profile",
                        "page_refs": "2",
                    }
                ],
                "appliances": [],
                "special_sections": [],
                "others": {},
                "warnings": [],
                "source_documents": [],
            },
        )

        client = TestClient(app)
        self._login(client)
        qa_page = client.get(f"/jobs/{job_id}/pdf-qa")
        self.assertEqual(qa_page.status_code, 200)
        csrf = qa_page.text.split('name="csrf_token" value="', 1)[1].split('"', 1)[0]
        verification = store.get_job_snapshot_verification(job_id, "raw_spec")
        self.assertIsNotNone(verification)
        payload = self._qa_form_payload(verification, csrf, item_status="pass")
        response = client.post(f"/jobs/{job_id}/pdf-qa/mark-pass", data=payload, follow_redirects=False)
        self.assertEqual(response.status_code, 303)
        updated = store.get_job_snapshot_verification(job_id, "raw_spec")
        self.assertEqual(updated["status"], "passed")

        export_response = client.get(f"/jobs/{job_id}/spec-list.xlsx", follow_redirects=False)
        self.assertEqual(export_response.status_code, 200)

    def test_spec_list_page_hides_non_kitchen_island_bar_back_and_implicit_overheads(self) -> None:
        builder_id = store.create_builder("Clarendon", "clarendon", "")
        job_id = store.create_job("37588", builder_id, "Non Kitchen Door Groups", "")
        store.upsert_snapshot(
            job_id,
            "raw_spec",
            {
                "job_no": "37588",
                "builder_name": "Clarendon",
                "source_kind": "spec",
                "generated_at": "2026-03-23T10:00:00+00:00",
                "analysis": {"mode": "heuristic_only", "openai_attempted": False, "openai_succeeded": False, "openai_model": "gpt-4.1-mini"},
                "rooms": [
                    {
                        "room_key": "vanities",
                        "original_room_label": "VANITIES",
                        "bench_tops": ["Quantum Zero Luna White - 20MM Pencil Round Edge"],
                        "door_panel_colours": [
                            "Polytec Blossom White Matt Finish Thermolaminate - Hamptons EM9 Profile",
                            "Polytec Habitit Smooth Finish Thermolaminate - Hamptons EM9 Profile",
                        ],
                        "door_colours_overheads": "Polytec Blossom White Matt Finish Thermolaminate - Hamptons EM9 Profile",
                        "door_colours_base": "Polytec Habitit Smooth Finish Thermolaminate - Hamptons EM9 Profile",
                        "door_colours_island": "",
                        "door_colours_bar_back": "",
                        "has_explicit_overheads": False,
                        "has_explicit_base": False,
                        "has_explicit_island": False,
                        "has_explicit_bar_back": False,
                        "toe_kick": [],
                        "bulkheads": [],
                        "handles": [],
                        "drawers_soft_close": "",
                        "hinges_soft_close": "",
                        "splashback": "",
                        "flooring": "",
                        "source_file": "schedule.pdf",
                        "page_refs": "3",
                        "evidence_snippet": "",
                        "confidence": 0.6,
                    }
                ],
                "appliances": [],
                "others": {},
                "warnings": [],
                "source_documents": [],
            },
        )
        client = TestClient(app)
        self._login(client)
        response = client.get(f"/jobs/{job_id}/spec-list")
        self.assertEqual(response.status_code, 200)
        self.assertNotIn("<strong>Island</strong>", response.text)
        self.assertNotIn("<strong>Bar Back</strong>", response.text)
        self.assertNotIn("<strong>Overheads</strong>", response.text)
        self.assertIn("<strong>Base</strong>", response.text)

    def test_spec_list_page_shows_empty_message_without_raw_snapshot(self) -> None:
        builder_id = store.create_builder("Yellowwood", "yellowwood", "")
        job_id = store.create_job("37974", builder_id, "No Snapshot", "")
        client = TestClient(app)
        self._login(client)
        response = client.get(f"/jobs/{job_id}/spec-list")
        self.assertEqual(response.status_code, 200)
        self.assertIn("No raw spec snapshot yet", response.text)

    def test_parse_documents_imperial_uses_title_boundaries_and_special_sections(self) -> None:
        documents = [
            {
                "file_name": "imperial-colours.pdf",
                "role": "spec",
                "pages": [
                    {
                        "page_no": 1,
                        "text": (
                            "Address:82/1 Goodwin St KANGAROO POINT\n"
                            "Client:Tracey Godfrey\n"
                            "Date:29.8.25\n"
                            "KITCHEN JOINERY SELECTION SHEET\n"
                            "Ceiling height:2200mm builder's bulkhead Cabinetry Height:2170mm 55mm Cove Cornice (to be checked)\n"
                            "Bulkhead:MDF Bulkhead 80mm high (to have cornice applied) Shadowline:Base under benchtop for Bronte Handle\n"
                            "Hinges & Drawer Runners:Soft Close Floor Type & Kick refacing required:\n"
                            "AREA / ITEM SPECS / DESCRIPTION IMAGE SUPPLIER NOTES\n"
                            "SPLASHBACK COLOUR\n"
                            "Caesarstone\n"
                            "Organic White\n"
                            "20mm Pencil Round Edge\n"
                            "Caesarstone\n"
                            "Up to overheads on cooktop run and\n"
                            "same height on all other walls\n"
                            "as per plans\n"
                            "BENCHTOPS COLOUR\n"
                            "Caesarstone\n"
                            "Organic White\n"
                            "20mm with 40mm Double Mitred\n"
                            "Pencil Round Edge\n"
                            "Caesarstone NOTE: Undermount Sink\n"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 2,
                        "text": (
                            "UPPER CABINETRY COLOUR + TALL CABINETS\n"
                            "Polytec\n"
                            "Valla Profile Door in\n"
                            "Boston Oak Woodmatt\n"
                            "EM0\n"
                            "Polytec\n"
                            "BASE CABINETRY COLOUR\n"
                            "Polytec\n"
                            "Ascot Profile Door\n"
                            "in Gossamer White Smooth\n"
                            "EM0\n"
                            "Polytec\n"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 3,
                        "text": (
                            "KICKBOARDS\n"
                            "MATCH ABOVE:\n"
                            "Polytec\n"
                            "Gossamer White Smooth\n"
                            "Polytec\n"
                            "Boston Oak Woodmatt under talls.\n"
                            "Polytec\n"
                            "HANDLES to OVERHEADS\n"
                            "NO HANDLE for OVERHEADS - RECESSED FINGER SPACE\n"
                            "Touch catch above ovens\n"
                            "Polytec\n"
                            "HANDLES BASE CABS NO HANDLES - BRONTE HANDLE Polytec\n"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 4,
                        "text": (
                            "CUSTOM HANDLES\n"
                            "Polytec\n"
                            "Boston Oak Woodmatt Melamine - Custom Made Handles - 1200mm high x 50mm wide outset 41mm\n"
                            "Polytec VERTICAL\n"
                            "DESIGNER: MELISSA COAKES CLIENT NAME: SIGNATURE: SIGNED DATE:\n"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 5,
                        "text": (
                            "FEATURE TALL DOORS JOINERY SELECTION SHEET\n"
                            "Bulkhead:NA Shadowline:NA\n"
                            "Hinges & Drawer Runners:Soft Close Floor Type & Kick refacing required:NA\n"
                            "TALL DOORS\n"
                            "Polytec\n"
                            "Valla Profile Door in\n"
                            "Thermolaminated Vinyl Wrap\n"
                            "Boston Oak Woodmatt\n"
                            "EM0 Edge\n"
                            "KICKBOARDS Polytec\n"
                            "BOSTON OAK WOODMATT Polytec\n"
                        ),
                        "needs_ocr": False,
                    },
                ],
            }
        ]
        snapshot = parse_documents(
            job_no="37647",
            builder_name="Imperial",
            source_kind="spec",
            documents=documents,
        )
        rooms = {row["room_key"]: row for row in snapshot["rooms"]}
        self.assertIn("kitchen", rooms)
        kitchen = rooms["kitchen"]
        self.assertEqual(kitchen["bench_tops_wall_run"], "20mm Caesarstone - Organic White - Pencil Round Edge - with 40mm Double Mitred")
        self.assertEqual(kitchen["splashback"], "20mm Caesarstone - Organic White - Pencil Round Edge")
        self.assertEqual(kitchen["door_colours_overheads"], "Polytec - EM0 - Valla Profile Door in - Boston Oak Woodmatt")
        self.assertEqual(kitchen["door_colours_tall"], "Polytec - EM0 - Valla Profile Door in - Boston Oak Woodmatt")
        self.assertEqual(kitchen["door_colours_base"], "Polytec - EM0 - Ascot Profile Door - in Gossamer White Smooth")
        self.assertEqual(kitchen["drawers_soft_close"], "Soft Close")
        self.assertEqual(kitchen["hinges_soft_close"], "Soft Close")
        self.assertTrue(kitchen["toe_kick"])
        self.assertIn("Gossamer White Smooth", " ".join(kitchen["toe_kick"]))
        self.assertIn("Boston Oak Woodmatt", " ".join(kitchen["toe_kick"]))
        self.assertEqual(
            kitchen["handles"],
            [
                "NO HANDLE for OVERHEADS - RECESSED FINGER SPACE",
                "Touch catch above ovens",
                "NO HANDLES - BRONTE HANDLE",
                "Custom Made Handles - Polytec Boston Oak Woodmatt Melamine - 1200mm high x 50mm wide outset 41mm - VERTICAL",
            ],
        )
        self.assertTrue(any("FEATURE TALL DOORS" == row["original_section_label"] for row in snapshot["special_sections"]))

        enriched = enrich_snapshot_rooms(snapshot, documents)
        kitchen_after_enrichment = {row["room_key"]: row for row in enriched["rooms"]}["kitchen"]
        self.assertEqual(
            kitchen_after_enrichment["bench_tops_wall_run"],
            "20mm Caesarstone - Organic White - Pencil Round Edge - with 40mm Double Mitred",
        )
        self.assertEqual(len(kitchen_after_enrichment["bench_tops"]), 1)
        self.assertIn(
            "20mm Caesarstone - Organic White - Pencil Round Edge - with 40mm Double Mitred",
            kitchen_after_enrichment["bench_tops"][0],
        )
        self.assertEqual(
            kitchen_after_enrichment["door_colours_base"],
            "Polytec - EM0 - Ascot Profile Door - in Gossamer White Smooth",
        )
        self.assertEqual(
            kitchen_after_enrichment["door_colours_overheads"],
            "Polytec - EM0 - Valla Profile Door in - Boston Oak Woodmatt",
        )

    def test_parse_documents_imperial_job27_handles_stay_within_handle_rows(self) -> None:
        documents = [
            {
                "file_name": "imperial-godfrey.pdf",
                "role": "spec",
                "pages": [
                    {
                        "page_no": 1,
                        "text": (
                            "Address:82/1 Goodwin St KANGAROO POINT\n"
                            "Client:Tracey Godfrey\n"
                            "Date:29.8.25\n"
                            "KITCHEN JOINERY SELECTION SHEET\n"
                            "Ceiling height:2200mm builder's bulkhead Cabinetry Height:2170mm 55mm Cove Cornice (to be checked)\n"
                            "Bulkhead:MDF Bulkhead 80mm high (to have cornice applied) Shadowline:Base under benchtop for Bronte Handle\n"
                            "Hinges & Drawer Runners:Soft Close Floor Type & Kick refacing required:\n"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 3,
                        "text": (
                            "KICKBOARDS\n"
                            "MATCH ABOVE:\n"
                            "Polytec\n"
                            "Gossamer White Smooth\n"
                            "Polytec\n"
                            "Boston Oak Woodmatt under talls.\n"
                            "Polytec\n"
                            "HANDLES to OVERHEADS\n"
                            "NO HANDLE for OVERHEADS - RECESSED FINGER SPACE\n"
                            "Touch catch above ovens\n"
                            "Polytec\n"
                            "HANDLES BASE CABS NO HANDLES - BRONTE HANDLE Polytec\n"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 4,
                        "text": (
                            "CUSTOM HANDLES\n"
                            "Polytec\n"
                            "Boston Oak Woodmatt\n"
                            "Melamine - Custom Made\n"
                            "Handles - 1200mm high x 50mm wide outset 41mm\n"
                            "Polytec VERTICAL\n"
                            "ALL COLOURS SHOWN ARE APPROXIMATE REPRESENTATIONS ONLY AND CANNOT BE RELIED ON COME INSTALLATION. IMPERIAL KITCHENS CAN NOT GUARENTEE STONE, LAMINATE, OR ANY BOARD COLOUR AVAILABILITY AT THE TIME OF THE CONSULTATION. PRODUCT\n"
                            "AVAILABILITY IS SUBJECT TO SUPPLIER AT TIME OF INSTALL. BY SIGNING THIS I ACCEPT THE ABOVE SELECTIONS ARE CORRECT AND I AM HAPPY TO PROCEED.\n"
                            "DESIGNER: MELISSA COAKES CLIENT NAME: SIGNATURE: SIGNED DATE:\n"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 7,
                        "text": (
                            "BATH + ENSUITE JOINERY SELECTION SHEET\n"
                            "Bulkhead:NA - open above overheads Shadowline:recessed rails for Bronte finger pulls\n"
                            "Hinges & Drawer Runners:Soft Close Floor Type & Kick refacing required:recessed kicks to just cover plumbing\n"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 8,
                        "text": (
                            "HANDLES NO HANDLES - BRONTE HANDLE Polytec\n"
                            "DESIGNER: MELISSA COAKES CLIENT NAME: SIGNATURE: SIGNED DATE:\n"
                            "ALL COLOURS SHOWN ARE APPROXIMATE REPRESENTATIONS ONLY AND CANNOT BE RELIED ON COME INSTALLATION. IMPERIAL KITCHENS CAN NOT GUARENTEE STONE, LAMINATE, OR ANY BOARD COLOUR AVAILABILITY AT THE TIME OF THE CONSULTATION. PRODUCT\n"
                            "AVAILABILITY IS SUBJECT TO SUPPLIER AT TIME OF INSTALL. BY SIGNING THIS I ACCEPT THE ABOVE SELECTIONS ARE CORRECT AND I AM HAPPY TO PROCEED.\n"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 9,
                        "text": (
                            "BAR JOINERY SELECTION SHEET\n"
                            "Shadowline:Under benchtop for recessed finger pull bronte handlesBulkhead:NO BULKHEAD HERE - Open above\n"
                            "Caesarstone\n"
                            "Hinges & Drawer Runners: NAFloor Type & Kick refacing required:Soft Close\n"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 11,
                        "text": (
                            "Polytec\n"
                            "PolytecKICKBOARDS\n"
                            "HANDLES to OVERHEADS\n"
                            "Polytec\n"
                            "Boston Oak Woodmatt\n"
                            "NO HANDLES TO OVERHEADS - RECESSED FINGER SPACE\n"
                            "HANDLES NO HANDLES - BRONTE HANDLE \n"
                            "ALL COLOURS SHOWN ARE APPROXIMATE REPRESENTATIONS ONLY AND CANNOT BE RELIED ON COME INSTALLATION. IMPERIAL KITCHENS CAN NOT GUARENTEE STONE, LAMINATE, OR ANY BOARD COLOUR AVAILABILITY AT THE TIME OF THE \n"
                            "CONSULTATION. PRODUCT AVAILABILITY IS SUBJECT TO SUPPLIER AT TIME OF INSTALL. BY SIGNING THIS I ACCEPT THE ABOVE SELECTIONS ARE CORRECT AND I AM HAPPY TO PROCEED.\n"
                            "DESIGNER: MELISSA COAKES CLIENT NAME: SIGNATURE: SIGNED DATE:\n"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 12,
                        "text": (
                            "LAUNDRY JOINERY SELECTION SHEET\n"
                            "Bulkhead:30mm colour board to ceiling - vertical grain Shadowline:Under bench for Bronte Handle\n"
                            "Hinges & Drawer Runners:Soft Close Floor Type & Kick refacing required:NA\n"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 14,
                        "text": (
                            "HANDLES Bronte Handle -Base cabs only\n"
                            "Recessed finger space - OHs only\n"
                            "ALL COLOURS SHOWN ARE APPROXIMATE REPRESENTATIONS ONLY AND CANNOT BE RELIED ON COME INSTALLATION. IMPERIAL KITCHENS CAN NOT GUARENTEE STONE, LAMINATE, OR ANY BOARD COLOUR AVAILABILITY AT THE TIME OF THE CONSULTATION.\n"
                            "PRODUCT AVAILABILITY IS SUBJECT TO SUPPLIER AT TIME OF INSTALL. BY SIGNING THIS I ACCEPT THE ABOVE SELECTIONS ARE CORRECT AND I AM HAPPY TO PROCEED.\n"
                        ),
                        "needs_ocr": False,
                    },
                ],
            }
        ]
        snapshot = parse_documents("37647", "Imperial", "spec", documents)
        rooms = {row["room_key"]: row for row in snapshot["rooms"]}
        self.assertEqual(
            rooms["kitchen"]["handles"],
            [
                "NO HANDLE for OVERHEADS - RECESSED FINGER SPACE",
                "Touch catch above ovens",
                "NO HANDLES - BRONTE HANDLE",
                "Custom Made Handles - Polytec Boston Oak Woodmatt Melamine - 1200mm high x 50mm wide outset 41mm - VERTICAL",
            ],
        )
        self.assertEqual(
            rooms["bar"]["handles"],
            [
                "NO HANDLES - BRONTE HANDLE",
                "NO HANDLES TO OVERHEADS - RECESSED FINGER SPACE",
            ],
        )
        self.assertEqual(
            rooms["laundry"]["handles"],
            [
                "Bronte Handle -Base cabs only",
                "Recessed finger space - OHs only",
            ],
        )
        self.assertEqual(rooms["bath_ensuite"]["handles"], ["NO HANDLES - BRONTE HANDLE"])

    def test_parse_documents_imperial_keeps_body_before_title_and_continuation_accessories(self) -> None:
        documents = [
            {
                "file_name": "imperial-office.pdf",
                "role": "spec",
                "pages": [
                    {
                        "page_no": 1,
                        "text": (
                            "BENCHTOP\n"
                            "Tasmanian Oak Matt Laminate Benchtop 33mm square edge\n"
                            "BASE CABINETRY COLOUR Polytec Classic White Matt\n"
                            "KICKBOARDS Polytec Classic White Matt\n"
                            "Hinges & Drawer Runners: NAFloor Type & Kick refacing required:SOFT CLOSE\n"
                            "NOTESSUPPLIERAREA / ITEM SPECS / DESCRIPTION IMAGE\n"
                            "Shadowline:NABulkhead:NA\n"
                            "Ceiling height:NA Cabinetry Height:760mm TO TOP OF BENCHTOP\n"
                            "LIVING & OFFICE JOINERY SELECTION SHEET\n"
                            "Address:16 Dovedale Cres ASHGROVE\n"
                            "Client:Phill Deacon\n"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 2,
                        "text": (
                            "ALL COLOURS SHOWN ARE APPROXIMATE REPRESENTATIONS ONLY AND CANNOT BE RELIED ON COME INSTALLATION.\n"
                            "DESIGNER: MELISSA COAKES CLIENT NAME: SIGNATURE: SIGNED DATE:\n"
                            "ACCESSORIES\n"
                            "Safe Desk Prodigy Cable Basket 950mm Black\n"
                            "Product Code: 7112195\n"
                            "ACCESSORIES\n"
                            "2 x Black Cable Grommet in black 80mm diameter\n"
                            "LED STRIP LIGHTING\n"
                            "Warm white strip light\n"
                            "HANDLES Square Edge recessed rail on drawers and door.\n"
                            "RAIL\n"
                            "Square Edge recessed rail in black\n"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 3,
                        "text": (
                            "APPLIANCES\n"
                            "OVEN (KITCHEN) N / A - By others\n"
                            "RANGEHOOD (KITCHEN) N / A - By others\n"
                            "SINKWARE & TAPWARE\n"
                            "SINKWARE (KITCHEN)\n"
                            "2 x Abey Schock Soho Large Single Bowl\n"
                        ),
                        "needs_ocr": False,
                    },
                ],
            }
        ]
        snapshot = parse_documents(
            job_no="37642",
            builder_name="Imperial",
            source_kind="spec",
            documents=documents,
        )
        rooms = {row["room_key"]: row for row in snapshot["rooms"]}
        self.assertIn("living_and_office", rooms)
        office = rooms["living_and_office"]
        self.assertEqual(office["original_room_label"], "LIVING & OFFICE")
        self.assertEqual(office["bench_tops_other"], "Tasmanian Oak Matt - Laminate Benchtop - 33mm square edge")
        self.assertEqual(office["door_colours_base"], "Polytec - Classic White Matt")
        self.assertEqual(office["led"], "Yes")
        self.assertEqual(
            office["accessories"],
            [
                "Safe Desk Prodigy Cable Basket 950mm Black",
                "2 x Black Cable Grommet in black 80mm diameter",
            ],
        )
        self.assertEqual(
            office["other_items"],
            [{"label": "RAIL", "value": "Square Edge recessed rail in black"}],
        )
        self.assertNotIn("OVEN (KITCHEN)", office["handles"])
        self.assertNotIn("SINKWARE", office["evidence_snippet"])

    def test_format_brisbane_time_and_run_duration(self) -> None:
        self.assertEqual(_format_brisbane_time("2026-03-24T10:00:00+00:00"), "2026-03-24 20:00 AEST")
        self.assertEqual(
            _format_run_duration(
                {
                    "started_at": "2026-03-24T10:00:00+00:00",
                    "finished_at": "2026-03-24T10:02:05+00:00",
                }
            ),
            "2m 5s",
        )

    def test_spec_list_page_renders_tall_and_special_sections(self) -> None:
        builder_id = store.create_builder("Imperial", "imperial", "")
        job_id = store.create_job("37647", builder_id, "Imperial Test", "")
        run_id = store.create_run(job_id, "spec")
        with store.connect() as conn:
            conn.execute(
                """
                UPDATE runs
                SET status = 'succeeded', stage = 'done', started_at = ?, finished_at = ?, parser_strategy = ?, app_build_id = ?, message = 'Completed'
                WHERE id = ?
                """,
                ("2026-03-24T10:00:00+00:00", "2026-03-24T10:02:05+00:00", "global_conservative", "test-build-001", run_id),
            )
        store.upsert_snapshot(
            job_id,
            "raw_spec",
            {
                "job_no": "37647",
                "builder_name": "Imperial",
                "source_kind": "spec",
                "generated_at": "2026-03-24T10:00:00+00:00",
                "site_address": "92 Haldham Crescent, Regents Park",
                "analysis": {"mode": "heuristic_only", "parser_strategy": "global_conservative"},
                "rooms": [
                    {
                        "room_key": "kitchen",
                        "original_room_label": "KITCHEN",
                        "bench_tops": ["Caesarstone Organic White 20mm with 40mm Double Mitred Pencil Round Edge"],
                        "bench_tops_other": "Caesarstone Organic White 20mm with 40mm Double Mitred Pencil Round Edge",
                        "floating_shelf": "Polytec Boston Oak Woodmatt 33mm pencil round edge",
                        "door_panel_colours": [],
                        "door_colours_overheads": "Polytec Valla Profile Door in Boston Oak Woodmatt EM0",
                        "door_colours_base": "Polytec Ascot Profile Door in Gossamer White Smooth EM0",
                        "door_colours_tall": "Polytec Valla Profile Door in Boston Oak Woodmatt EM0",
                        "door_colours_island": "",
                        "door_colours_bar_back": "",
                        "has_explicit_overheads": True,
                        "has_explicit_base": True,
                        "has_explicit_tall": True,
                        "toe_kick": [],
                        "bulkheads": [],
                        "handles": [],
                        "led": "Yes",
                        "accessories": ["Safe Desk Prodigy Cable Basket 950mm Black", "2 x Black Cable Grommet"],
                        "other_items": [{"label": "RAIL", "value": "Square Edge recessed rail in black"}],
                        "drawers_soft_close": "",
                        "hinges_soft_close": "",
                        "splashback": "",
                        "flooring": "",
                        "source_file": "imperial-colours.pdf",
                        "page_refs": "1, 2",
                        "evidence_snippet": "",
                        "confidence": 0.7,
                    }
                ],
                "special_sections": [
                    {
                        "section_key": "feature_tall_doors",
                        "original_section_label": "FEATURE TALL DOORS",
                        "fields": {"Tall": "Polytec Valla Profile Door in Thermolaminated Vinyl Wrap Boston Oak Woodmatt EM0 Edge"},
                        "source_file": "imperial-colours.pdf",
                        "page_refs": "5, 6",
                        "evidence_snippet": "TALL DOORS Polytec ...",
                        "confidence": 0.7,
                    }
                ],
                "appliances": [],
                "others": {},
                "warnings": [],
                "source_documents": [],
            },
        )
        client = TestClient(app)
        self._login(client)
        response = client.get(f"/jobs/{job_id}/spec-list")
        self.assertEqual(response.status_code, 200)
        self.assertIn("<strong>Tall</strong>", response.text)
        self.assertIn("FEATURE TALL DOORS", response.text)
        self.assertIn("2026-03-24 20:00 AEST", response.text)
        self.assertIn("Extraction duration:</strong> 2m 5s", response.text)
        self.assertIn("<strong>Floating Shelf</strong>", response.text)
        self.assertIn("<strong>LED</strong>", response.text)
        self.assertIn("Accessories 1", response.text)
        self.assertIn("RAIL", response.text)
        self.assertIn("Spec List for 37647 - 92 Haldham Crescent, Regents Park", response.text)

    def test_spec_list_excel_includes_tall_and_special_sections_sheet(self) -> None:
        snapshot = {
            "job_no": "37647",
            "builder_name": "Imperial",
            "source_kind": "spec",
            "generated_at": "2026-03-24T10:00:00+00:00",
            "analysis": {"mode": "heuristic_only", "parser_strategy": "global_conservative"},
            "rooms": [
                {
                    "room_key": "kitchen",
                    "original_room_label": "KITCHEN",
                    "bench_tops": ["Caesarstone Organic White 20mm with 40mm Double Mitred Pencil Round Edge"],
                    "floating_shelf": "Polytec Boston Oak Woodmatt 33mm pencil round edge",
                    "door_panel_colours": [],
                    "door_colours_overheads": "Polytec Valla Profile Door in Boston Oak Woodmatt EM0",
                    "door_colours_base": "Polytec Ascot Profile Door in Gossamer White Smooth EM0",
                    "door_colours_tall": "Polytec Valla Profile Door in Boston Oak Woodmatt EM0",
                    "toe_kick": [],
                    "bulkheads": [],
                    "handles": [],
                    "led": "Yes",
                    "accessories": ["Safe Desk Prodigy Cable Basket 950mm Black"],
                    "other_items": [{"label": "RAIL", "value": "Square Edge recessed rail in black"}],
                    "drawers_soft_close": "",
                    "hinges_soft_close": "",
                    "splashback": "",
                    "flooring": "",
                    "source_file": "imperial-colours.pdf",
                    "page_refs": "1, 2",
                    "evidence_snippet": "",
                    "confidence": 0.7,
                }
            ],
            "special_sections": [
                {
                    "section_key": "feature_tall_doors",
                    "original_section_label": "FEATURE TALL DOORS",
                    "fields": {"Tall": "Polytec Valla Profile Door in Thermolaminated Vinyl Wrap Boston Oak Woodmatt EM0 Edge"},
                    "source_file": "imperial-colours.pdf",
                    "page_refs": "5, 6",
                    "evidence_snippet": "TALL DOORS Polytec ...",
                    "confidence": 0.7,
                }
            ],
            "appliances": [],
            "others": {},
            "warnings": [],
            "source_documents": [],
        }
        excel_path = Path(build_spec_list_excel("37647", snapshot))
        workbook = load_workbook(excel_path)
        self.assertIn("Special Sections", workbook.sheetnames)
        rooms_sheet = workbook["Rooms"]
        headers = [cell.value for cell in next(rooms_sheet.iter_rows(min_row=1, max_row=1))]
        self.assertIn("door_colours_tall", headers)
        self.assertIn("floating_shelf", headers)
        self.assertIn("accessories", headers)
        self.assertIn("other_items", headers)
        special_sheet = workbook["Special Sections"]
        self.assertEqual(special_sheet["A2"].value, "feature_tall_doors")
        self.assertEqual(special_sheet["C2"].value, "Tall")

    def test_jobs_open_links_use_new_tab(self) -> None:
        builder_id = store.create_builder("Imperial", "imperial", "")
        store.create_job("37642", builder_id, "Imperial Live", "")
        client = TestClient(app)
        self._login(client)
        response = client.get("/jobs")
        self.assertEqual(response.status_code, 200)
        self.assertIn('target="_blank"', response.text)
        self.assertIn('rel="noopener"', response.text)
        self.assertIn('class="table-cardify"', response.text)
        self.assertIn('data-label="Job No."', response.text)

    def test_workspace_and_spec_list_sidebar_start_hidden(self) -> None:
        builder_id = store.create_builder("Imperial", "imperial", "")
        job_id = store.create_job("37642", builder_id, "Imperial Sidebar", "")
        client = TestClient(app)
        self._login(client)

        jobs_response = client.get("/jobs")
        self.assertEqual(jobs_response.status_code, 200)
        self.assertNotIn("shell-rail-collapsed", jobs_response.text)

        detail_response = client.get(f"/jobs/{job_id}")
        self.assertEqual(detail_response.status_code, 200)
        self.assertIn("shell-rail-collapsed", detail_response.text)
        self.assertIn("Show Sidebar", detail_response.text)

        spec_list_response = client.get(f"/jobs/{job_id}/spec-list")
        self.assertEqual(spec_list_response.status_code, 200)
        self.assertIn("shell-rail-collapsed", spec_list_response.text)
        self.assertIn("Show Sidebar", spec_list_response.text)

    def test_parse_documents_imperial_job31_keeps_room_names_and_same_section_values(self) -> None:
        documents = [
            {
                "file_name": "37642 Signed Variation COLOURS_Deacon 20 10 25.pdf",
                "role": "spec",
                "pages": [
                    {
                        "page_no": 1,
                        "text": (
                            "BENCHTOP+ SPLASHBACK\n"
                            "Frosty Carrina (5141)\n"
                            "20mm \n"
                            "Pencil Round Edge\n"
                            "Benchtop + Waterfall ends to island plus Splashback \n"
                            "up to overheads.\n"
                            "Caesarstone\n"
                            "BENCHTOP ON ISLAND TO HAVE 2 X \n"
                            "WATERFALL ENDS MITRED JOINS\n"
                            "BASE CABINETRY COLOUR Polytec\n"
                            "Classic White Matt Polytec\n"
                            "AREA / ITEM SPECS / DESCRIPTION IMAGE SUPPLIER NOTES\n"
                            "Bulkhead:MDF Bulkhead flush with carcass Shadowline:Yes shadowline under benchtops - recessed square edge finger pulls\n"
                            "Hinges & Drawer Runners:Soft Close Floor Type & Kick refacing required:\n"
                            "KITCHEN JOINERY SELECTION SHEET\n"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 2,
                        "text": (
                            "TOUCH CATCH HANDLES TO TALL PANTRY DOORS ONLY Polytec\n"
                            "HANDLES Square Edge recessed rail \n"
                            "finger pull doors and drawers Polytec\n"
                            "NO HANDLES to OVERHEADS Recessed finger space overheads. Polytec\n"
                            "FEATURE CABINETRY COLOUR\n"
                            "Polytec \n"
                            "COVE PROFILE 25\n"
                            "Classic White Matt Finish\n"
                            "Polytec Overheads and \n"
                            "Bar Back only\n"
                            "KICKBOARDS Polytec\n"
                            "Classic White Matt Polytec\n"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 3,
                        "text": (
                            "KICKBOARDS Polytec\n"
                            "Classic White Matt Polytec\n"
                            "BENCHTOP+ SPLASHBACK\n"
                            "Frosty Carrina (5141)\n"
                            "20mm \n"
                            "Pencil Round Edge\n"
                            "SPLASHBACK - 200MM HIGH AT REAR AND ON \n"
                            "RIGHT HAND WALL ONLY.\n"
                            "Caesarstone\n"
                            "BENCHTOP AREA WITH COOKTOP TO \n"
                            "BE REPLACED WITH FROSTY CARIN \n"
                            "BUT SPLASHBACK TO REMAIN AS \n"
                            "CURRENT TILES IF POSSIBLE.\n"
                            "BASE CABINETRY COLOUR Polytec\n"
                            "Classic White Matt Polytec\n"
                            "AREA / ITEM SPECS / DESCRIPTION IMAGE SUPPLIER NOTES\n"
                            "Bulkhead:MDF Bulkhead flush with carcass Shadowline:Yes shadowline under benchtops - recessed square edge finger pulls\n"
                            "Hinges & Drawer Runners:Soft Close Floor Type & Kick refacing required:\n"
                            "WALK-BEHIND PANTRY JOINERY SELECTION SHEET\n"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 4,
                        "text": "HANDLES Square Edge recessed rail \nfinger pull doors and drawers Polytec\n",
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 5,
                        "text": (
                            "KICKBOARDS Polytec\n"
                            "Classic White Matt Polytec 100mm high kicks only\n"
                            "BENCHTOP STANDARD 16mm Panel \n"
                            "Classic White Matt - Flush with doors Polytec\n"
                            "BASE CABINETRY COLOUR\n"
                            "Polytec\n"
                            "COVE PROFILE 25\n"
                            "THERMOLANINATED DOORS \n"
                            "AND END PANELS\n"
                            "Classic White Matt\n"
                            "Polytec\n"
                            "AREA / ITEM SPECS / DESCRIPTION IMAGE SUPPLIER NOTES\n"
                            "Bulkhead:NA Shadowline:NA\n"
                            "Hinges & Drawer Runners:STD Floor Type & Kick refacing required:NA\n"
                            "BENCH SEAT JOINERY SELECTION SHEET\n"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 7,
                        "text": (
                            "BENCHTOP\n"
                            "KICKBOARDS Polytec\n"
                            "Classic White Matt\n"
                            "Tasmanian Oak Matt \n"
                            "Laminate Benchtop \n"
                            "33mm square edge\n"
                            "BASE CABINETRY COLOUR Polytec\n"
                            "Classic White Matt\n"
                            "Hinges & Drawer Runners: NAFloor Type & Kick refacing required:SOFT CLOSE\n"
                            "OFFICE JOINERY SELECTION SHEET\n"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 8,
                        "text": (
                            "ACCESSORIES\n"
                            "OE ELSAFE DESK PRODIGY \n"
                            "CABLE BASKET 950MM BLACK\n"
                            "Product Code: 7112195\n"
                            "ACCESSORIES\n"
                            " 2 x Black Cable Grommet\n"
                            "in black 80mm diameter\n"
                            "Installed rear underside of desk on \n"
                            "right next to cupboardLincoln Sentry\n"
                            "Furnware\n"
                            "HANDLES Square Edge recessed rail on drawers and door. Furnware\n"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 10,
                        "text": (
                            "SINKWARE & TAPWARE\n"
                            "AREA / ITEM SPECS / DESCRIPTION IMAGE SUPPLIER NOTES\n"
                            "Taphole location: Centred to back\n"
                            "SINKS TO BE INSTAL\n"
                            "LED UNDERMOUTNED\n"
                            "SINKWARE (KITCHEN)\n"
                            "2 x ABEY Schock SOHO\n"
                            "Large Single Bowl Puro\n"
                            "N120P\n"
                            "Puro Black Cristadur\n"
                            "ABEY BY IMPERIAL\n"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 11,
                        "text": (
                            "2 x 304 Gooseneck Pull Out with Dual Spray \n"
                            "Function Kitchen Mixer \n"
                            "in Eureka Gold finish\n"
                            "KTA014-G\n"
                            "TAPWARE (KITCHEN) ABEY BY IMPERIAL\n"
                        ),
                        "needs_ocr": False,
                    },
                ],
            }
        ]
        snapshot = enrich_snapshot_rooms(parse_documents("37642", "Imperial", "spec", documents), documents)
        rooms = {row["room_key"]: row for row in snapshot["rooms"]}
        self.assertIn("walk_behind_pantry", rooms)
        self.assertEqual(rooms["walk_behind_pantry"]["original_room_label"], "WALK-BEHIND PANTRY")
        kitchen = rooms["kitchen"]
        self.assertEqual(kitchen["bench_tops_wall_run"], "20mm Caesarstone - Frosty Carrina (5141) - Pencil Round Edge")
        self.assertIn("BENCHTOP ON ISLAND TO HAVE 2 X WATERFALL ENDS MITRED JOINS", kitchen["bench_tops_island"])
        self.assertEqual(kitchen["splashback"], "20mm Caesarstone - Frosty Carrina (5141) - Pencil Round Edge - up to overheads.")
        self.assertIn("Taphole location: Centred to back", kitchen["sink_info"])
        self.assertIn("Undermounted", kitchen["sink_info"])
        self.assertEqual(kitchen["tap_info"], "2 x 304 Gooseneck Pull Out with Dual Spray Function Kitchen Mixer in Eureka Gold finish KTA014-G")
        office = rooms["office"]
        self.assertEqual(office["bench_tops_other"], "Tasmanian Oak Matt - Laminate Benchtop - 33mm square edge")
        self.assertEqual(
            office["accessories"],
            [
                "OE Elsafe - DESK PRODIGY CABLE BASKET 950MM BLACK",
                "2 x Black Cable Grommet in black 80mm diameter",
            ],
        )

    def test_parse_documents_imperial_job32_keeps_tall_blank_and_tap_clean(self) -> None:
        documents = [
            {
                "file_name": "37813 Imperial.pdf",
                "role": "spec",
                "pages": [
                    {
                        "page_no": 1,
                        "text": (
                            "Tall Pantry Doors - 7206 Danes Bow Handle Matt \n"
                            "Black 320mm - SO-7206-320-MB\n"
                            "No handles on Uppers - PTO Where reqHANDLES\n"
                            " Thermolaminated Vinyl Style 1 - Vienna - \n"
                            "Classic White Matt\n"
                            "Thermolaminated Vinyl Style 1 - Vienna - \n"
                            "Classic White Matt\n"
                            "Ceiling height:2430mm Cabinetry Height:2150mm\n"
                            "Shadowline:\n"
                            "KICKBOARDS As Doors\n"
                            "BASE CABINETRY COLOUR\n"
                            "UPPER CABINETRY COLOUR\n"
                            "BENCHTOP\n"
                            "Bulkhead:None\n"
                            "IMAGE\n"
                            "N/A\n"
                            "Hinges & Drawer Runners: Tiles ( Use same footprint as existing kicks)Floor Type & Kick refacing required:Softclose \n"
                            "Caesarstone\n"
                            "KITCHEN JOINERY SELECTION SHEET\n"
                            "92 Haldham Crescent, Regents Park\n"
                            "PRIVATE - Leah Mitchell\n"
                            "29/10/2025\n"
                            "Address:\n"
                            "Client:\n"
                            "Date:\n"
                            "20mm Stone \n"
                            "5131 Calacattra Nuvo - PR\n"
                            "Waterfall End\n"
                            "Polytec\n"
                            "Polytec\n"
                            "Titus Tekform\n"
                            "2163 Voda Profile Handle Matt Black\n"
                            "200mm - SO-2163-200-MB\n"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 3,
                        "text": (
                            "Taphole location: In Sink\n"
                            "SINKWARE (KITCHEN)\n"
                            "Veronar, matrix sink,double bowl, double drain, \n"
                            "stainless steel, Part Number:S220.SS.FG, Sink \n"
                            "Mounting -Topmount \n"
                            "Out of stock, available to back order \n"
                            "Furnware\n"
                            "TAPWARE (KITCHEN) Mixer Tap Clients own Taphole location: In Sink\n"
                            "TAPWARE (KITCHEN) Water Filter Tap Clients own\n"
                            "SINKWARE & TAPWARE\n"
                        ),
                        "needs_ocr": False,
                    },
                ],
            }
        ]
        snapshot = enrich_snapshot_rooms(parse_documents("37813", "Imperial", "spec", documents), documents)
        kitchen = snapshot["rooms"][0]
        self.assertEqual(snapshot["site_address"], "92 Haldham Crescent, Regents Park")
        self.assertIn("20mm Caesarstone", kitchen["bench_tops_wall_run"])
        self.assertIn("5131 Calacattra Nuvo - PR", kitchen["bench_tops_wall_run"])
        self.assertEqual(kitchen["door_colours_tall"], "")
        self.assertEqual(kitchen["bulkheads"], ["None"])
        self.assertEqual(kitchen["tap_info"], "Mixer Tap Clients own | Water Filter Tap Clients own")
        self.assertEqual(kitchen["accessories"], [])
        self.assertEqual(kitchen["drawers_soft_close"], "Soft Close")
        self.assertEqual(kitchen["hinges_soft_close"], "Soft Close")
        self.assertEqual(
            kitchen["handles"],
            [
                "7206 Danes Bow Handle Matt Black 320mm - SO-7206-320-MB",
                "No handles on Uppers - PTO Where req",
                "2163 Voda Profile Handle Matt Black 200mm - SO-2163-200-MB",
            ],
        )
        self.assertNotIn("IMAGE N/A", " ".join(kitchen["bulkheads"]))

    def test_parse_documents_imperial_job35_recovers_address_handle_fragments_and_sink_tap_from_pdf_text(self) -> None:
        documents = [
            {
                "file_name": "38211 Imperial.pdf",
                "role": "spec",
                "pages": [
                    {
                        "page_no": 1,
                        "raw_text": (
                            "3064 Square Handle Brushed Nickel\n"
                            "160mm - SO-3064-160-BN\n"
                            "KITCHEN JOINERY SELECTION SHEET\n"
                            "9 Greenland Court, Springfield\n"
                            "PRIVATE - Yasantha Warawita\n"
                            "12/03/2026\n"
                            "Ceiling height:\n"
                            "2415mm Cabinetry Height:2300mm\n"
                            "Shadowline:\n"
                            "KICKBOARDS\n"
                            "Classic White Matt\n"
                            "As Doors\n"
                            "BASE CABINETRY COLOUR\n"
                            "UPPER CABINETRY COLOUR\n"
                            "BENCHTOP\n"
                            "Bulkhead:Colourboard\n"
                            "n/a\n"
                            "Classic White Matt\n"
                            "Hinges & Drawer Runners: Tiles with Floating TBCFloor Type & Kick refacing required:Soft close\n"
                            "Ceasarstone\n"
                            "NOTESSUPPLIER\n"
                            "1 x Waterfall End\n"
                            "20mm Stone\n"
                            "AREA / ITEM SPECS / DESCRIPTION\n"
                            "6313 Turbine Grey\n"
                            "Polytec\n"
                            "Polytec\n"
                            "Titus Tekform\n"
                            "Polytec\n"
                            "Horizontal on Drawers and Vertical on\n"
                            "Doors\n"
                            "BIN\n"
                            "HANDLES\n"
                            "450mm Pull-Out Bin Hettich\n"
                            "AVAILABILITY IS SUBJECT TO SUPPLIER AT TIME OF INSTALL. BY SIGNING THIS I ACCEPT THE ABOVE SELECTIONS ARE CORRECT AND I AM HAPPY TO PROCEED.\n"
                        ),
                        "text": (
                            "3064 Square Handle Brushed Nickel\n"
                            "160mm - SO-3064-160-BN\n"
                            "KITCHEN JOINERY SELECTION SHEET\n"
                            "9 Greenland Court, Springfield\n"
                            "PRIVATE - Yasantha Warawita\n"
                            "12/03/2026\n"
                            "Ceiling height:\n"
                            "2415mm Cabinetry Height:2300mm\n"
                            "Shadowline:\n"
                            "KICKBOARDS\n"
                            "Classic White Matt\n"
                            "As Doors\n"
                            "BASE CABINETRY COLOUR\n"
                            "UPPER CABINETRY COLOUR\n"
                            "BENCHTOP\n"
                            "Bulkhead:Colourboard\n"
                            "n/a\n"
                            "Classic White Matt\n"
                            "Hinges & Drawer Runners: Tiles with Floating TBCFloor Type & Kick refacing required:Soft close\n"
                            "Ceasarstone\n"
                            "NOTESSUPPLIER\n"
                            "1 x Waterfall End\n"
                            "20mm Stone\n"
                            "AREA / ITEM SPECS / DESCRIPTION\n"
                            "6313 Turbine Grey\n"
                            "Polytec\n"
                            "Polytec\n"
                            "Titus Tekform\n"
                            "Polytec\n"
                            "Horizontal on Drawers and Vertical on\n"
                            "Doors\n"
                            "BIN\n"
                            "HANDLES\n"
                            "450mm Pull-Out Bin Hettich\n"
                            "AVAILABILITY IS SUBJECT TO SUPPLIER AT TIME OF INSTALL. BY SIGNING THIS I ACCEPT THE ABOVE SELECTIONS ARE CORRECT AND I AM HAPPY TO PROCEED.\n"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 3,
                        "raw_text": (
                            "Furnware Taphole Location: In Stone centered \n"
                            "behind sink\n"
                            "Veronar, Forge Undermount Sink, Double \n"
                            "Bowl, Satin Stainless Steel\n"
                            "Part Number:\n"
                            "SVF210SINK.SSS.FG\n"
                            " - UNDERMOUNT\n"
                            "TAPWARE (KITCHEN) Furnware\n"
                            "Veronar, otus, pull-out, goose neck mixer, \n"
                            "brushed nickel\n"
                            "Part Number:\n"
                            "PC1016SB.BRN\n"
                            "SIGNED DATE:\n"
                            "SPECS / DESCRIPTION IMAGE SUPPLIER NOTES\n"
                            "DESIGNER: SARAH ROSCOW CLIENT NAME: SIGNATURE:\n"
                            "Address:9 Greenland Court, Springfield\n"
                            "Client:PRIVATE - Yasantha Warawita\n"
                            "Date:12/03/2026\n"
                            "SINKWARE & TAPWARE\n"
                            "AREA / ITEM\n"
                            "SINKWARE (KITCHEN)\n"
                        ),
                        "text": (
                            "Furnware Taphole Location: In Stone centered \n"
                            "behind sink\n"
                            "Veronar, Forge Undermount Sink, Double \n"
                            "Bowl, Satin Stainless Steel\n"
                            "Part Number:\n"
                            "SVF210SINK.SSS.FG\n"
                            " - UNDERMOUNT\n"
                            "TAPWARE (KITCHEN) Furnware\n"
                            "Veronar, otus, pull-out, goose neck mixer, \n"
                            "brushed nickel\n"
                            "Part Number:\n"
                            "PC1016SB.BRN\n"
                            "SIGNED DATE:\n"
                            "SPECS / DESCRIPTION IMAGE SUPPLIER NOTES\n"
                            "DESIGNER: SARAH ROSCOW CLIENT NAME: SIGNATURE:\n"
                            "Address:9 Greenland Court, Springfield\n"
                            "Client:PRIVATE - Yasantha Warawita\n"
                            "Date:12/03/2026\n"
                            "SINKWARE & TAPWARE\n"
                            "AREA / ITEM\n"
                            "SINKWARE (KITCHEN)\n"
                        ),
                        "needs_ocr": False,
                    },
                ],
            }
        ]
        snapshot = enrich_snapshot_rooms(parse_documents("38211", "Imperial", "spec", documents), documents)
        kitchen = snapshot["rooms"][0]
        self.assertEqual(snapshot["site_address"], "9 Greenland Court, Springfield")
        self.assertEqual(kitchen["bench_tops_wall_run"], "20mm Caesarstone - 6313 Turbine Grey - 1 x Waterfall End")
        self.assertEqual(kitchen["door_colours_base"], "Polytec - Classic White Matt")
        self.assertEqual(kitchen["door_colours_overheads"], "Polytec - Classic White Matt")
        self.assertEqual(
            kitchen["handles"],
            ["Titus Tekform - 3064 Square Handle Brushed Nickel 160mm - SO-3064-160-BN - Horizontal on Drawers and Vertical on Doors"],
        )
        self.assertEqual(kitchen["accessories"], ["Hettich - 450mm Pull-Out Bin"])
        self.assertEqual(kitchen["drawers_soft_close"], "Soft Close")
        self.assertEqual(kitchen["hinges_soft_close"], "Soft Close")
        self.assertEqual(kitchen["flooring"], "Tiles with Floating TBC")
        self.assertEqual(
            kitchen["sink_info"],
            "Veronar, Forge Undermount Sink, Double Bowl, Satin Stainless Steel Part Number: SVF210SINK.SSS.FG UNDERMOUNT",
        )
        self.assertEqual(kitchen["tap_info"], "Veronar, otus, pull-out, goose neck mixer, brushed nickel Part Number: PC1016SB.BRN")

    def test_parse_documents_imperial_job34_row_boundaries_keep_sections_clean(self) -> None:
        documents = [
            {
                "file_name": "37558 Imperial.pdf",
                "role": "spec",
                "pages": [
                    {
                        "page_no": 1,
                        "text": (
                            "ACCESSORIES Veronar, Spice Tray Insert, To Suit 450mm\n"
                            "& 600MM Drawer, White -\n"
                            "VCT.450E.WH.FGx1 & VCT.600E.WH.FGx1\n"
                            "Furnware\n"
                            "450MM (SHORT) - 2 X 29LTR Hettich\n"
                            "Island Drawer GPO 1 - Side of Upper Bin \n"
                            "Draw 'By Builder'(see pic 1 shown)\n"
                            "Island Drawer GPO 2 - Rear panel of Utensil \n"
                            "Drawer - Hafele Trio 822.53.151 'By \n"
                            "Imperial' (See pic 2 shown)\n"
                            "KICKBOARDS AS DOORS Polytec\n"
                            "Pic 1 (GPO 1)             -            Pic 2 (GPO 2)\n"
                            "KITCHEN JOINERY SELECTION SHEET\n"
                            "Lot 532 Sandpiper Terrace, Worongary\n"
                            "Evoca\n"
                            "15.12.2025\n"
                            "Address:\n"
                            "Client:\n"
                            "Date:\n"
                            "SQUARE SET CEILING\n"
                            "Shadowline:15mm shadow to builders bulkheadBulkhead:by builder\n"
                            "Ceiling height:2740mm Cabinetry Height:2350mm\n"
                            "Hinges & Drawer Runners: tiledFloor Type & Kick refacing required:soft close\n"
                            "VERTICAL GRAIN\n"
                            "NOTESSUPPLIERAREA / ITEM SPECS / DESCRIPTION\n"
                            "BENCHTOP 40mm stone - Arissed By builder\n"
                            "IMAGE\n"
                            "ISLAND CABINETRY COLOUR (incl. BACK OF\n"
                            "ISLAND CURVE AND COLUMN) Notaio Walnut Woodmatt Polytec\n"
                            "BACK WALL & COFFEE NOOK INTERNAL \n"
                            "CABINETRY COLOUR\n"
                            "FLOATING SHELVES Notaio Walnut Woodmatt\n"
                            "Notaio Walnut Woodmatt\n"
                            "GPO'S Polytec VERTICAL GRAIN\n"
                            "VERTICAL GRAINPolytec\n"
                            "BIN\n"
                            "ALL COLOURS SHOWN ARE APPROXIMATE REPRESENTATIONS ONLY AND CANNOT BE RELIED ON COME INSTALLATION. IMPERIAL KITCHENS CAN NOT GUARENTEE STONE, LAMINATE, OR ANY BOARD COLOUR AVAILABILITY AT THE TIME OF THE CONSULTATION. PRODUCT AVAILABILITY IS \n"
                            "SUBJECT TO SUPPLIER AT TIME OF INSTALL. BY SIGNING THIS I ACCEPT THE ABOVE SELECTIONS ARE CORRECT AND I AM HAPPY TO PROCEED.\n"
                            "DESIGNER: CHLOE PARKER CLIENT NAME: SIGNATURE: SIGNED DATE:\n"
                            "LED's As per drawings\n"
                            "HANDLES\n"
                            "BASE- BEVEL EDGE FINGERPULL\n"
                            "UPPER - FINGERPULL\n"
                            "TALL - PTO\n"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 2,
                        "text": (
                            "BAR JOINERY SELECTION SHEET\n"
                            "LED's As per drawings\n"
                            "HANDLES\n"
                            "KICKBOARDS AS DOORS Polytec\n"
                            "BASE CABINETRY COLOUR Notaio Walnut Woodmatt Polytec VERTICAL GRAIN\n"
                            "FLOATING SHELVING COLOUR (WITH \n"
                            "INTERNAL STEEL SUPPORTS) Notaio Walnut Woodmatt Polytec VERTICAL GRAIN\n"
                            "BENCHTOP By builder40mm stone - Arissed\n"
                            "AREA / ITEM SPECS / DESCRIPTION IMAGE SUPPLIER NOTES\n"
                            "Bulkhead:N/A Shadowline:N/A\n"
                            "Hinges & Drawer Runners:soft close Floor Type & Kick refacing required:tiled\n"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 3,
                        "text": (
                            "PANTRY JOINERY SELECTION SHEET\n"
                            "Ceiling height:2740mm Cabinetry Height:2350mm SQUARE SET CEILING\n"
                            "Bulkhead:by builder\n"
                            "Shadowline:15mm shadow to builders bulkhead\n"
                            "Hinges & Drawer Runners:soft close Floor Type & Kick refacing required:Tiled\n"
                            "AREA / ITEM SPECS / DESCRIPTION IMAGE SUPPLIER NOTES\n"
                            "BASE CABINETRY COLOUR Notaio Walnut Woodmatt Polytec VERTICAL GRAIN\n"
                            "BENCHTOP BY BUILDER40MM STONE - Arissed\n"
                            "TALL CABINETRY COLOUR Notaio Walnut Woodmatt Polytec VERTICAL GRAIN\n"
                            "UPPER CABINETRY COLOUR Notaio Walnut Woodmatt Polytec VERTICAL GRAIN\n"
                            "HANDLES BASE- BEVEL EDGE FINGERPULL\n"
                            "UPPER - FINGERPULL\n"
                            "TALL - PTO\n"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 4,
                        "text": (
                            "LAUNDRY & MUD ROOM JOINERY SELECTION SHEET\n"
                            "Ceiling height:2740mm Cabinetry Height:2350mm SQUARE SET CEILING\n"
                            "Bulkhead:by builder\n"
                            "Shadowline:15mm shadow to builders bulkhead\n"
                            "Hinges & Drawer Runners:soft close Floor Type & Kick refacing required:Tiled\n"
                            "AREA / ITEM SPECS / DESCRIPTION IMAGE SUPPLIER NOTES\n"
                            "BASE CABINETRY COLOUR Notaio Walnut Woodmatt Polytec VERTICAL GRAIN\n"
                            "Mud room - 33mm Notaio Walnut laminate\n"
                            "BENCHTOP (SEAT)\n"
                            "BENCHTOP laundry - 20MM STONE BY BUILDER\n"
                            "HANGING RAIL Oval wardrobe tube, aluminium, 15mm x\n"
                            "30mm x 1.2m, gunmetal - Furnware\n"
                            "UPPER CABINETRY COLOUR Notaio Walnut Woodmatt Polytec VERTICAL GRAIN\n"
                            "HAMPER Tanova, Designer Laundry System, 1 X 65L\n"
                            "Metal Hamper, White\n"
                            "PART NO: LTDS45.165L.WH\n"
                            "HANDLES BASE- BEVEL EDGE FINGERPULL\n"
                            "UPPER - FINGERPULL\n"
                            "TALL - PTO\n"
                            "SUBJECT TO SUPPLIER AT TIME OF INSTALL. BY SIGNING THIS I ACCEPT THE ABOVE SELECTIONS ARE CORRECT AND I AM HAPPY TO PROCEED.\n"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 8,
                        "text": (
                            "Address:Lot 532 Sandpiper Terrace, Worongary\n"
                            "Client:Evoca\n"
                            "Date:15.12.2025\n"
                            "LIVING & OFFICE JOINERY SELECTION SHEET\n"
                            "Ceiling height:2740mm Cabinetry Height:AS PER DRAWINGS SQUARE SET CEILING\n"
                            "Bulkhead:NO Shadowline:NO\n"
                            "Hinges & Drawer Runners:soft close Floor Type & Kick refacing required:Tiled\n"
                            "AREA / ITEM SPECS / DESCRIPTION IMAGE SUPPLIER NOTES\n"
                            "CABINETRY COLOUR Notaio Walnut Woodmatt Polytec VERTICAL GRAIN\n"
                            "BENCHTOP Polytec VERTICAL GRAIN33mm laminate\n"
                            "Notaio Walnut Woodmatt - 10x10 edge\n"
                            "HANDLES\n"
                            "BASE- BEVEL EDGE FINGERPULL\n"
                            "UPPER - FINGERPULL\n"
                            "TALL - PTO\n"
                            "FLOATING SHELVES (WITH INTERNAL \n"
                            "STEEL SUPPORTS) Notaio Walnut Woodmatt Polytec VERTICAL GRAIN\n"
                            "KICKBOARDS AS DOORS OR WALL HUNG (REFER TO \n"
                            "DRAWINGS) Polytec\n"
                            "ALL COLOURS SHOWN ARE APPROXIMATE REPRESENTATIONS ONLY AND CANNOT BE RELIED ON COME INSTALLATION. IMPERIAL KITCHENS CAN NOT GUARENTEE STONE, LAMINATE, OR ANY BOARD COLOUR AVAILABILITY AT THE TIME OF THE CONSULTATION. PRODUCT AVAILABILITY IS \n"
                            "SUBJECT TO SUPPLIER AT TIME OF INSTALL. BY SIGNING THIS I ACCEPT THE ABOVE SELECTIONS ARE CORRECT AND I AM HAPPY TO PROCEED.\n"
                            "DESIGNER: CHLOE PARKER CLIENT NAME: SIGNATURE: SIGNED DATE:\n"
                            "otaio Walt Woomatt\n"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 10,
                        "text": (
                            "WALK IN ROBE JOINERY SELECTION SHEET\n"
                            "Ceiling height:N/A Cabinetry Height:as per drawings\n"
                            "Bulkhead:by builder\n"
                            "Shadowline:16mm shadowline\n"
                            "Hinges & Drawer Runners:soft close Floor Type & Kick refacing required:Hybrid\n"
                            "AREA / ITEM SPECS / DESCRIPTION IMAGE SUPPLIER NOTES\n"
                            "CABINETRY COLOUR PolytecStone Grey - Matt\n"
                            "RAIL Wardrobe rail for\n"
                            "JEWELLERY INSERT 806.39.404 Hafele\n"
                            "GLASS TOP Glass Infill above jewellery\n"
                            "LED LIGHTING\n"
                            "HANDLES Kethy Vertical on Doors and Horizontal on\n"
                            "drawers\n"
                            "Doors - S225-480-MBK - Matt Black \n"
                            "Anodised (MBK)\n"
                            "Drawers - S225-280-MBK - Matt Black \n"
                            "Anodised (MBK)\n"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 12,
                        "text": (
                            "Taphole location:\n"
                            "N / A - By others By others\n"
                            "Corner of Tub\n"
                            "Taphole location:\n"
                            "Taphole location: TUB (LAUNDRY) undermount - specs tbc\n"
                            "Taphole location:\n"
                            "Centre of Sink / Sink Pre-punched Hole \n"
                            "BASIN (UPPER BATHROOM)\n"
                            "BASIN (DS BATHROOM)\n"
                            "N / A - By others By others\n"
                            "BASIN (POWDER) above counter- specs tbc\n"
                            "above counter- specs tbc N / A - By others By others\n"
                            "above counter- specs tbc N / A - By others By others\n"
                            "Taphole location:\n"
                            "SINKWARE (KITCHEN)\n"
                            "SINKWARE (PANTRY)\n"
                            "undermount - specs tbc N / A - By others By others\n"
                            "undermount - specs tbc N / A - By others By others\n"
                            "Taphole location:\n"
                            "Ctr of sink\n"
                            "Taphole location:\n"
                            "DESIGNER: CHLOE PARKER CLIENT NAME: SIGNATURE: SIGNED DATE:\n"
                            "BASIN (WC) above counter- specs tbc N / A - By others By others\n"
                            "BASIN (MASTER ENSUITE) above counter- specs tbc N / A - By others By others\n"
                            "SINKWARE & TAPWARE\n"
                            "AREA / ITEM SPECS / DESCRIPTION IMAGE SUPPLIER NOTES\n"
                            "Address:Lot 532 Sandpiper Terrace, Worongary\n"
                            "Client:Evoca\n"
                            "Date:15.12.2025\n"
                            "Taphole location:\n"
                        ),
                        "needs_ocr": False,
                    },
                ],
            }
        ]
        snapshot = enrich_snapshot_rooms(parse_documents("37558.3", "Imperial", "spec", documents), documents)
        rooms = {row["room_key"]: row for row in snapshot["rooms"]}
        kitchen = rooms["kitchen"]
        self.assertEqual(kitchen["bench_tops_wall_run"], "40mm stone - Arissed - By Builder")
        self.assertEqual(kitchen["bench_tops_other"], "")
        self.assertEqual(kitchen["door_colours_base"], "Polytec - Notaio Walnut Woodmatt - VERTICAL GRAIN")
        self.assertEqual(kitchen["door_colours_island"], "Polytec - Notaio Walnut Woodmatt")
        self.assertEqual(kitchen["floating_shelf"], "Polytec - Notaio Walnut Woodmatt - VERTICAL GRAIN")
        self.assertEqual(kitchen["sink_info"], "undermount - specs tbc - Taphole location: Ctr of sink")
        self.assertEqual(kitchen["flooring"], "tiled")
        self.assertEqual(kitchen["toe_kick"], ["AS DOORS Polytec"])
        self.assertEqual(kitchen["drawers_soft_close"], "Soft Close")
        self.assertEqual(kitchen["hinges_soft_close"], "Soft Close")
        self.assertEqual(
            kitchen["accessories"],
            [
                "Veronar - Spice Tray Insert, To Suit 450mm & 600MM Drawer, White - VCT.450E.WH.FGx1 & VCT.600E.WH.FGx1",
                "Hettich - 450MM (SHORT) - 2 X 29LTR",
                "Island Drawer GPO 1 - Side of Upper Bin Draw 'By Builder'(see pic 1 shown)",
                "Island Drawer GPO 2 - Rear panel of Utensil Drawer - Hafele Trio 822.53.151 'By Imperial' (See pic 2 shown)",
            ],
        )

        bar = rooms["bar"]
        self.assertEqual(bar["bench_tops_other"], "40mm stone - Arissed - By Builder")
        self.assertIn("INTERNAL STEEL SUPPORTS", bar["floating_shelf"])
        self.assertNotIn("INTERNAL STEEL SUPPORTS", bar["bench_tops_other"])
        self.assertEqual(bar["door_colours_base"], "Polytec - Notaio Walnut Woodmatt VERTICAL GRAIN")
        self.assertEqual(bar["flooring"], "tiled")

        pantry = rooms["pantry"]
        self.assertEqual(pantry["bench_tops_other"], "40MM STONE - Arissed - By Builder")
        self.assertEqual(pantry["door_colours_base"], "Polytec - Notaio Walnut Woodmatt VERTICAL GRAIN")
        self.assertEqual(pantry["door_colours_overheads"], "Polytec - Notaio Walnut Woodmatt VERTICAL GRAIN")
        self.assertEqual(pantry["door_colours_tall"], "Polytec - Notaio Walnut Woodmatt VERTICAL GRAIN")
        self.assertEqual(pantry["sink_info"], "undermount - specs tbc - Taphole location: Ctr of sink")

        laundry = rooms["laundry_and_mud_room"]
        self.assertIn("33mm", laundry["bench_tops_other"])
        self.assertIn("Notaio Walnut laminate", laundry["bench_tops_other"])
        self.assertIn("20MM", laundry["bench_tops_other"])
        self.assertIn("STONE - By Builder", laundry["bench_tops_other"])
        self.assertEqual(laundry["door_colours_base"], "Polytec - Notaio Walnut Woodmatt VERTICAL GRAIN")
        self.assertEqual(laundry["flooring"], "tiled")
        self.assertEqual(laundry["accessories"], ["Tanova - Designer Laundry System, 1 X 65L Metal Hamper, White PART NO: LTDS45.165L.WH"])
        self.assertNotIn("SUBJECT TO SUPPLIER", " ".join(laundry["handles"]))

        living_office = rooms["living_and_office"]
        self.assertIn("33mm Polytec", living_office["bench_tops_other"])
        self.assertIn("10x10 edge", living_office["bench_tops_other"])
        self.assertIn("INTERNAL", living_office["floating_shelf"])
        self.assertIn("Notaio Walnut Woodmatt", living_office["floating_shelf"])
        self.assertEqual(living_office["flooring"], "tiled")

        wir = rooms["wir"]
        other_labels = {item["label"] for item in wir["other_items"]}
        self.assertIn("RAIL", other_labels)
        self.assertIn("JEWELLERY INSERT", other_labels)
        self.assertEqual(
            wir["handles"],
            [
                "Kethy Vertical on Doors and Horizontal on drawers",
                "Doors - S225-480-MBK - Matt Black Anodised (MBK)",
                "Drawers - S225-280-MBK - Matt Black Anodised (MBK)",
            ],
        )

    def test_parse_documents_imperial_job36_recovers_handles_flooring_and_room_boundaries_from_pdf_text(self) -> None:
        documents = [
            {
                "file_name": "38119 Imperial.pdf",
                "role": "spec",
                "pages": [
                    {
                        "page_no": 1,
                        "text": (
                            "Address:53 Thomas Street, Auchenflower\n"
                            "Client:PRIVATE - Stephen Mego\n"
                            "Date:19/02/2026\n"
                            "LAUNDRY JOINERY SELECTION SHEET\n"
                            "Ceiling height:2710mm Cabinetry Height:2400mm SQUARE SET\n"
                            "Bulkhead:20MM SHADOWLINE TO BUILDERS BULKHEAD Shadowline:20mm to existing\n"
                            "Hinges & Drawer Runners:Soft close Floor Type & Kick refacing required:Tiles\n"
                            "AREA / ITEM SPECS / DESCRIPTION IMAGE SUPPLIER NOTES\n"
                            "BASE CABINETRY COLOUR PolytecBoston Oak - Woodmatt\n"
                            "BENCHTOP Caesarstone\n"
                            "40mm Mitred Stone\n"
                            "506 Mirabel - PR\n"
                            "KICKBOARDS PolytecAs Doors\n"
                            "UPPER CABINETRY COLOUR Boston Oak- Woodmatt Polytec\n"
                            "HANDLES Furnware Horizontal on Drawers and Vertical on \n"
                            "Doors\n"
                            "Finger Pull on Uppers\n"
                            "Momo Barrington Eclipse Plain 96mm in \n"
                            "Matt Brass Part Number:BEPL96.MBR\n"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 2,
                        "text": (
                            "Address:53 Thomas Street, Auchenflower\n"
                            "Client:PRIVATE - Stephen Mego\n"
                            "Date:19/02/2026\n"
                            "ENSUITE JOINERY SELECTION SHEET\n"
                            "Ceiling height:2710mm Cabinetry Height:As Per Drawings SQUARE SET\n"
                            "Bulkhead:n/a Shadowline:n/a\n"
                            "Hinges & Drawer Runners:Softclose Floor Type & Kick refacing required:Tiled\n"
                            "AREA / ITEM SPECS / DESCRIPTION IMAGE SUPPLIER NOTES\n"
                            "BENCHTOP Caesarstone\n"
                            "80mm Mitred Stone\n"
                            "2141 Snow - PR\n"
                            "HANDLES Furnware HorizontalMomo Barrington Eclipse Plain 96mm in \n"
                            "Matt Brass Part Number:BEPL96.MBR\n"
                            "BASE CABINETRY COLOUR Florentine Walnut - Woodmatt Polytec \n"
                            "KICKBOARD Recessed Kick - MDF Imperial\n"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 3,
                        "text": (
                            "Address:53 Thomas Street, Auchenflower\n"
                            "Client:PRIVATE - Stephen Mego\n"
                            "Date:19/02/2026\n"
                            "POWDER JOINERY SELECTION SHEET\n"
                            "Ceiling height:2710mm Cabinetry Height:As Per Drawings SQUARE SET\n"
                            "Bulkhead:n/a Shadowline:n/a\n"
                            "Hinges & Drawer Runners:Soft close Floor Type & Kick refacing required:Tiled\n"
                            "AREA / ITEM SPECS / DESCRIPTION IMAGE SUPPLIER NOTES\n"
                            "BASE CABINETRY COLOUR Florentine Walnut - Woodmatt Polytec \n"
                            "BENCHTOP Caesarstone\n"
                            "40mm Mitred Stone\n"
                            "2141 Snow - PR\n"
                            "KICKBOARD Recessed Kick - MDF Imperial\n"
                            "HANDLES Momo Barrington Eclipse Plain 96mm in \n"
                            "Matt Brass Part Number:BEPL96.MBR Furnware Vertical\n"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 4,
                        "text": (
                            "Address:53 Thomas Street, Auchenflower\n"
                            "Client:PRIVATE - Stephen Mego\n"
                            "Date:19/02/2026\n"
                            " WIR JOINERY SELECTION SHEET\n"
                            "Ceiling height:2710mm Cabinetry Height:2300mm SQUARE SET\n"
                            "Bulkhead:Colourboard to match doors Shadowline:n/a\n"
                            "Hinges & Drawer Runners:Soft close Floor Type & Kick refacing required:Carpet\n"
                            "AREA / ITEM SPECS / DESCRIPTION IMAGE SUPPLIER NOTES\n"
                            "KICKBOARDS As Doors Polytec \n"
                            "BASE CABINETRY COLOUR Polytec Florentine Walnut - Woodmatt\n"
                            "HANGING RAIL\n"
                            "Oval wardrobe tube, aluminium, 15mm x \n"
                            "30mm x 3.6m, brushed brass\n"
                            "Part Number: 1400.36.BBR\n"
                            "Furnware\n"
                            "HANDLES Furnware Horizontal on Drawers and Vertical on \n"
                            "Doors\n"
                            "Momo Barrington Eclipse Plain 96mm in \n"
                            "Matt Brass Part Number:BEPL96.MBR\n"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 5,
                        "text": (
                            "Address:53 Thomas Street, Auchenflower\n"
                            "Client:PRIVATE - Stephen Mego\n"
                            "Date:19/02/2026\n"
                            "BED 1 ROBES JOINERY SELECTION SHEET\n"
                            "Ceiling height:2710mm Cabinetry Height:2400mm SQUARE SET\n"
                            "Bulkhead:16mm to meet existing frame Shadowline:n/a\n"
                            "Hinges & Drawer Runners:Soft close Floor Type & Kick refacing required:Carpet\n"
                            "AREA / ITEM SPECS / DESCRIPTION IMAGE SUPPLIER NOTES\n"
                            "Florentine Walnut - WoodmattBASE CABINETRY COLOUR Polytec \n"
                            "HANDLES Furnware Horizontal on Drawers and Vertical on \n"
                            "Doors\n"
                            "KICKBOARDS As Doors Polytec \n"
                            "Momo Barrington Eclipse Plain 96mm in \n"
                            "Matt Brass Part Number:BEPL96.MBR\n"
                            "HANGING RAIL\n"
                            "Oval wardrobe tube, aluminium, 15mm x \n"
                            "30mm x 3.6m, brushed brass\n"
                            "Part Number: 1400.36.BBR\n"
                            "Furnware\n"
                        ),
                        "needs_ocr": False,
                    },
                    {
                        "page_no": 7,
                        "text": (
                            "By OthersVenice 500 Semi-Inset Basin Solid Surface \n"
                            "White x 2 (Reece)\n"
                            "SIGNED DATE:\n"
                            "Fienza\n"
                            "SPECS / DESCRIPTION IMAGE SUPPLIER\n"
                            "Wall mounted Taps\n"
                            "NOTES\n"
                            "DESIGNER: SARAH ROSCOW CLIENT NAME: SIGNATURE:\n"
                            "Address:53 Thomas Street, Auchenflower\n"
                            "Client:PRIVATE - Stephen Mego\n"
                            "Date:19/02/2026\n"
                            "SINKWARE & TAPWARE\n"
                            "AREA / ITEM\n"
                            "SINKWARE (LAUNDRY)\n"
                            "Minka Solid Surface Wall Basin, 1 Tap \n"
                            "Hole - Product Code: CSB310-1SINKWARE POWDER)\n"
                            "Furnware Tapware location - In stone centered \n"
                            "behind sink\n"
                            "Veronar, Forge Undermount Sink, Double \n"
                            "Bowl, Satin Stainless Steel\n"
                            "Part Number: SVF210SINK.SSS.FG - \n"
                            "Undermounted\n"
                            "SINKWARE (ENSUITE)\n"
                        ),
                        "needs_ocr": False,
                    },
                ],
            }
        ]
        snapshot = enrich_snapshot_rooms(parse_documents("38119", "Imperial", "spec", documents), documents)
        rooms = {row["room_key"]: row for row in snapshot["rooms"]}
        laundry = rooms["laundry"]
        self.assertEqual(laundry["drawers_soft_close"], "Soft Close")
        self.assertEqual(laundry["hinges_soft_close"], "Soft Close")
        self.assertEqual(laundry["flooring"], "Tiled")
        self.assertEqual(
            laundry["handles"],
            [
                "Furnware - Momo Barrington Eclipse Plain 96mm in Matt Brass Part Number:BEPL96.MBR - Horizontal on Drawers and Vertical on Doors",
                "Finger Pull on Uppers",
            ],
        )
        self.assertEqual(laundry["sink_info"], "Minka Solid Surface Wall Basin, 1 Tap Hole - Product Code: CSB310-1")

        ensuite = rooms["ensuite"]
        self.assertEqual(ensuite["door_colours_base"], "Polytec - Florentine Walnut - Woodmatt")
        self.assertEqual(ensuite["toe_kick"], ["Recessed Kick - MDF Imperial"])
        self.assertNotIn("KICKBOARD", ensuite["bench_tops_other"])

        powder = rooms["powder"]
        self.assertEqual(powder["toe_kick"], ["Recessed Kick - MDF Imperial"])
        self.assertNotIn("KICKBOARD", powder["bench_tops_other"])

        wir = rooms["wir"]
        self.assertEqual(wir["door_colours_base"], "Polytec - Florentine Walnut - Woodmatt")
        self.assertTrue(any("Momo Barrington Eclipse Plain 96mm in Matt Brass Part Number:BEPL96.MBR" in entry for entry in wir["handles"]))

        bed_1_robes = rooms["bed_1_robes"]
        self.assertEqual(bed_1_robes["door_colours_base"], "Polytec - Florentine Walnut - Woodmatt")
        self.assertTrue(any("Momo Barrington Eclipse Plain 96mm in Matt Brass Part Number:BEPL96.MBR" in entry for entry in bed_1_robes["handles"]))

    def test_parse_documents_imperial_job38_layout_rows_drive_row_local_fields(self) -> None:
        documents = [
            {
                "file_name": "SIGNED FINAL COLOURS_FOXOVER 21 Shadowood st KENMORE 23 3 26.pdf",
                "role": "spec",
                "pages": [
                    {
                        "page_no": 1,
                        "raw_text": (
                            "Address:2510-076 - Private - 21 Shadowood Street, Kenmore Hills\n"
                            "Client:Eloise Cawcutt Foxover\n"
                            "Date:20.3.26\n"
                            "KITCHEN JOINERY SELECTION SHEET\n"
                            "Hinges & Drawer Runners:Soft Close Floor Type & Kick refacing required:NA\n"
                            "SPLASHBACK Tiles by client Installed By Imperial STD tile with white grout\n"
                            "BENCHTOP Caesarstone Fresh Concrete 20mm Pencil Round Edge\n"
                            "FEATURE TALL CABINETRY COLOUR + bar back Laminex Gumnut Natural Finish 2606\n"
                        ),
                        "text": "",
                        "needs_ocr": False,
                        "page_layout": {
                            "page_type": "joinery",
                            "section_label": "KITCHEN",
                            "room_label": "MICROWAVE LEAVE STANDARD SPACE BY CLIENT",
                            "room_blocks": [
                                {
                                    "room_label": "MICROWAVE LEAVE STANDARD SPACE BY CLIENT",
                                    "rows": [
                                        {"row_label": "SPLASHBACK", "row_kind": "material", "value_region_text": "Tiles by client", "supplier_region_text": "", "notes_region_text": "Installed By Imperial STD tile with white grout"},
                                        {"row_label": "BENCHTOP", "row_kind": "material", "value_region_text": "Fresh Concrete 20mm Pencil Round Edge", "supplier_region_text": "Caesarstone", "notes_region_text": ""},
                                        {"row_label": "FEATURE TALL CABINETRY COLOUR + bar back", "row_kind": "material", "value_region_text": "Gumnut Natural Finish 2606", "supplier_region_text": "Laminex", "notes_region_text": ""},
                                        {"row_label": "Hinges & Drawer Runners", "row_kind": "metadata", "value_region_text": "Soft Close", "supplier_region_text": "", "notes_region_text": ""},
                                        {"row_label": "Floor Type & Kick refacing required", "row_kind": "metadata", "value_region_text": "NA", "supplier_region_text": "", "notes_region_text": ""},
                                    ],
                                }
                            ],
                        },
                    },
                    {
                        "page_no": 2,
                        "raw_text": (
                            "GLASS INLAY DOORS TO OVERHEAD FEATURE CABINETRY\n"
                            "Reeded Glass Inlay Feature doors In Laminex Gumnut Natural Finish 2606 with Coloured internals and shelves in - Laminex Blackbutt Truescale internals\n"
                            "FEATURE TIMBER LOOK FLOATING SHELVES 51mm thick floating shelves Laminex Blackbutt Truescale Natural Finish 2618\n"
                            "BASE CABINETRY COLOUR Polytec Classic White Matt\n"
                            "KICKBOARDS MATCH ABOVE Polytec Classic White Matt or Laminex Gumnut Natural Finish 2606 or Laminex Blackbutt Truescale Natural Finish 2618\n"
                        ),
                        "text": "",
                        "needs_ocr": False,
                        "page_layout": {
                            "page_type": "joinery",
                            "section_label": "KITCHEN",
                            "room_label": "KITCHEN",
                            "room_blocks": [
                                {
                                    "room_label": "KITCHEN",
                                    "rows": [
                                        {"row_label": "GLASS INLAY DOORS TO OVERHEAD FEATURE CABINETRY", "row_kind": "material", "value_region_text": "Reeded Glass Inlay Feature doors In Laminex Gumnut Natural Finish 2606 with Coloured internals and shelves in", "supplier_region_text": "Laminex", "notes_region_text": "Laminex Blackbutt Truescale internals"},
                                        {"row_label": "FEATURE TIMBER LOOK FLOATING SHELVES", "row_kind": "material", "value_region_text": "51mm thick floating shelves Blackbutt Truescale Natural Finish 2618", "supplier_region_text": "Laminex", "notes_region_text": ""},
                                        {"row_label": "BASE CABINETRY COLOUR", "row_kind": "material", "value_region_text": "Classic White Matt", "supplier_region_text": "Polytec", "notes_region_text": "Includes single cabinet on bar back area"},
                                        {"row_label": "KICKBOARDS", "row_kind": "material", "value_region_text": "MATCH ABOVE", "supplier_region_text": "Polytec + Laminex", "notes_region_text": "Polytec Classic White Matt or Laminex Gumnut Natural Finish 2606 or Laminex Blackbutt Truescale Natural Finish 2618"},
                                    ],
                                }
                            ],
                        },
                    },
                    {
                        "page_no": 3,
                        "raw_text": (
                            "LIP PULL HANDLES - DRAWERS ABI INTERIORS Rappana Cabinetry pull extended 100mm brushed copper (10469) SUPPLIED BY CLIENT INSTALLED BY IMPERIAL Installed Horizontally\n"
                            "HANDLES - BASE CABS + OVERHEAD CABS ABI INTERIORS Elsa Cabinetry Knob- brushed copper (14494) for any of the other doors and the gas strut oh doors. SUPPLIED BY CLIENT INSTALLED BY IMPERIAL\n"
                            "FEATURE LIP PULL PANTRY HANDLES ABI INTERIORS 2 X Rappana Cabinetry Pull Extended 800mm - Brushed Copper SUPPLIED BY CLIENT INSTALLED BY IMPERIAL Installed Vertically to pantry doors only\n"
                        ),
                        "text": "",
                        "needs_ocr": False,
                        "page_layout": {
                            "page_type": "joinery",
                            "section_label": "KITCHEN",
                            "room_label": "KITCHEN",
                            "room_blocks": [
                                {
                                    "room_label": "KITCHEN",
                                    "rows": [
                                        {"row_label": "LIP PULL HANDLES - DRAWERS", "row_kind": "handle", "value_region_text": "Rappana Cabinetry pull extended 100mm brushed copper (10469)", "supplier_region_text": "ABI INTERIORS", "notes_region_text": "SUPPLIED BY CLIENT INSTALLED BY IMPERIAL Installed Horizontally"},
                                        {"row_label": "HANDLES - BASE CABS + OVERHEAD CABS", "row_kind": "handle", "value_region_text": "Elsa Cabinetry Knob- brushed copper (14494) for any of the other doors and the gas strut oh doors.", "supplier_region_text": "ABI INTERIORS", "notes_region_text": "SUPPLIED BY CLIENT INSTALLED BY IMPERIAL"},
                                        {"row_label": "FEATURE LIP PULL PANTRY HANDLES", "row_kind": "handle", "value_region_text": "2 X Rappana Cabinetry Pull Extended 800mm - Brushed Copper", "supplier_region_text": "ABI INTERIORS", "notes_region_text": "SUPPLIED BY CLIENT INSTALLED BY IMPERIAL Installed Vertically to pantry doors only"},
                                    ],
                                }
                            ],
                        },
                    },
                    {
                        "page_no": 4,
                        "raw_text": (
                            "NA\n"
                            "DINING BANQUETTE JOINERY SELECTION SHEET\n"
                            "BASE CABINETRY COLOUR Polytec Classic White Matt\n"
                            "BENCHTOP Polytec Classic White Matt Laminate Benchtop 33mm\n"
                            "HANDLES - BASE CABS Rappana Cabinetry Pull 50mm Brushed Copper SUPPLIED BY CLIENT Installed Horizontally\n"
                            "KICKBOARDS Polytec Classic White Matt\n"
                        ),
                        "text": "",
                        "needs_ocr": False,
                        "page_layout": {
                            "page_type": "joinery",
                            "section_label": "NA DINING BANQUETTE",
                            "room_label": "NA DINING BANQUETTE",
                            "room_blocks": [
                                {
                                    "room_label": "NA DINING BANQUETTE",
                                    "rows": [
                                        {"row_label": "BASE CABINETRY COLOUR", "row_kind": "material", "value_region_text": "Classic White Matt", "supplier_region_text": "Polytec", "notes_region_text": ""},
                                        {"row_label": "BENCHTOP", "row_kind": "material", "value_region_text": "Classic White Matt Laminate Benchtop 33mm", "supplier_region_text": "Polytec", "notes_region_text": ""},
                                        {"row_label": "KICKBOARDS", "row_kind": "material", "value_region_text": "Classic White Matt", "supplier_region_text": "Polytec", "notes_region_text": ""},
                                        {"row_label": "HANDLES - BASE CABS", "row_kind": "handle", "value_region_text": "Rappana Cabinetry Pull 50mm Brushed Copper", "supplier_region_text": "ABI INTERIORS", "notes_region_text": "SUPPLIED BY CLIENT Installed Horizontally"},
                                    ],
                                }
                            ],
                        },
                    },
                    {
                        "page_no": 7,
                        "raw_text": "SINKWARE (KITCHEN) ABEY Schock horizontal double bowl sink bronze N200BZ TOP MOUNT Taphole location: Tap Landing Centred to back",
                        "text": "",
                        "needs_ocr": False,
                        "page_layout": {
                            "page_type": "sinkware_tapware",
                            "section_label": "SINKWARE & TAPWARE",
                            "room_label": "KITCHEN",
                            "room_blocks": [
                                {
                                    "room_label": "KITCHEN",
                                    "rows": [
                                        {"row_label": "SINKWARE (KITCHEN)", "row_kind": "sink", "value_region_text": "ABEY Schock horizontal double bowl sink bronze N200BZ", "supplier_region_text": "", "notes_region_text": "TOP MOUNT Taphole location: Tap Landing Centred to back"},
                                    ],
                                }
                            ],
                        },
                    },
                    {
                        "page_no": 8,
                        "raw_text": "Tap Franke Eos Neo pull out tap copper TA9601CP TAPWARE (KITCHEN) BY CLIENT Eloise Cawcutt-Foxover",
                        "text": "",
                        "needs_ocr": False,
                        "page_layout": {
                            "page_type": "sinkware_tapware",
                            "section_label": "SINKWARE & TAPWARE",
                            "room_label": "KITCHEN",
                            "room_blocks": [
                                {
                                    "room_label": "KITCHEN",
                                    "rows": [
                                        {"row_label": "TAPWARE (KITCHEN)", "row_kind": "tap", "value_region_text": "Franke Eos Neo pull out tap copper TA9601CP", "supplier_region_text": "", "notes_region_text": "BY CLIENT"},
                                    ],
                                }
                            ],
                        },
                    },
                ],
            }
        ]
        snapshot = enrich_snapshot_rooms(parse_documents("38251", "Imperial", "spec", documents), documents)
        rooms = {row["room_key"]: row for row in snapshot["rooms"]}
        kitchen = rooms["kitchen"]
        self.assertEqual(snapshot["site_address"], "21 Shadowood Street, Kenmore Hills")
        self.assertEqual(kitchen["bench_tops_wall_run"], "20mm Caesarstone - Fresh Concrete - Pencil Round Edge")
        self.assertEqual(kitchen["door_colours_base"], "Polytec - Classic White Matt")
        self.assertIn("Reeded Glass Inlay", kitchen["door_colours_overheads"])
        self.assertEqual(kitchen["door_colours_tall"], "Laminex - Gumnut Natural Finish 2606")
        self.assertEqual(kitchen["door_colours_bar_back"], "Laminex - Gumnut Natural Finish 2606")
        self.assertIn("51mm", kitchen["floating_shelf"])
        self.assertIn("Blackbutt", kitchen["floating_shelf"])
        self.assertTrue(kitchen["toe_kick"])
        self.assertEqual(len(kitchen["handles"]), 3)
        self.assertEqual(kitchen["drawers_soft_close"], "Soft Close")
        self.assertEqual(kitchen["hinges_soft_close"], "Soft Close")
        self.assertEqual(kitchen["flooring"], "NA")
        self.assertNotIn("FEATURE TALL", kitchen["bench_tops_wall_run"])
        self.assertNotIn("Foxover", kitchen["tap_info"])
        self.assertIn("Franke Eos Neo", kitchen["tap_info"])
        dining = rooms["dining_banquette"]
        self.assertEqual(dining["original_room_label"], "DINING BANQUETTE")
        self.assertEqual(dining["door_colours_base"], "Polytec - Classic White Matt")
        self.assertIn("33mm Polytec - Classic White Matt", dining["bench_tops_other"])
        self.assertIn("Laminate Benchtop", dining["bench_tops_other"])
        self.assertEqual(dining["toe_kick"], ["Polytec - Classic White Matt"])
        self.assertEqual(
            dining["handles"],
            ["ABI Interiors - Rappana Cabinetry Pull 50mm Brushed Copper - SUPPLIED BY CLIENT Installed Horizontally"],
        )

    def test_job_detail_and_spec_list_titles_show_snapshot_site_address(self) -> None:
        builder_id = store.create_builder("Imperial", "imperial", "")
        job_id = store.create_job("37813.2", builder_id, "", "")
        store.upsert_snapshot(
            job_id,
            "raw_spec",
            {
                "job_no": "37813.2",
                "builder_name": "Imperial",
                "source_kind": "spec",
                "generated_at": "2026-03-28T07:06:09+00:00",
                "site_address": "92 Haldham Crescent, Regents Park",
                "analysis": {"mode": "heuristic_only", "parser_strategy": "global_conservative"},
                "rooms": [],
                "special_sections": [],
                "appliances": [],
                "others": {},
                "warnings": [],
                "source_documents": [],
            },
        )
        client = TestClient(app)
        self._login(client)
        detail_response = client.get(f"/jobs/{job_id}")
        self.assertEqual(detail_response.status_code, 200)
        self.assertIn("37813.2 - 92 Haldham Crescent, Regents Park", detail_response.text)
        spec_response = client.get(f"/jobs/{job_id}/spec-list")
        self.assertEqual(spec_response.status_code, 200)
        self.assertIn("Spec List for 37813.2 - 92 Haldham Crescent, Regents Park", spec_response.text)

    def test_imperial_office_benchtop_ignores_joinery_title_and_address_noise(self) -> None:
        documents = [
            {
                "file_name": "imperial-office-live.pdf",
                "role": "spec",
                "pages": [
                    {
                        "page_no": 7,
                        "text": (
                            "BENCHTOP\n"
                            "KICKBOARDS Polytec\n"
                            "Classic White Matt\n"
                            "Tasmanian Oak Matt \n"
                            "Laminate Benchtop \n"
                            "33mm square edge\n"
                            "BASE CABINETRY COLOUR Polytec\n"
                            "Classic White Matt\n"
                            "Hinges & Drawer Runners: NAFloor Type & Kick refacing required:SOFT CLOSE\n"
                            "NOTESSUPPLIERAREA / ITEM SPECS / DESCRIPTION IMAGE\n"
                            "Polytec\n"
                            "Polytec\n"
                            "Shadowline:NABulkhead:NA\n"
                            "Ceiling height:NA Cabinetry Height:760mm TO TOP OF BENCHTOP\n"
                            "OFFICE JOINERY SELECTION SHEET\n"
                            "16 Dovedale Cres ASHGROVE\n"
                            "Phill Deacon\n"
                            "12.9.25\n"
                            "Address:\n"
                            "Client:\n"
                            "Date:\n"
                            "Polytec\n"
                        ),
                        "needs_ocr": False,
                    }
                ],
            }
        ]
        snapshot = enrich_snapshot_rooms(parse_documents("37642", "Imperial", "spec", documents), documents)
        office = snapshot["rooms"][0]
        self.assertEqual(office["original_room_label"], "OFFICE")
        self.assertEqual(office["bench_tops_other"], "Tasmanian Oak Matt - Laminate Benchtop - 33mm square edge")
        self.assertNotIn("Dovedale Cres", office["bench_tops_other"])
        self.assertNotIn("JOINERY SELECTION SHEET", office["bench_tops_other"])

    def test_imperial_orientation_notes_do_not_become_tall_or_island_material(self) -> None:
        row = {
            "room_key": "kitchen",
            "original_room_label": "KITCHEN",
            "bench_tops": ["20mm Caesarstone - 5131 Calacattra Nuvo - PR Waterfall End"],
            "door_panel_colours": [],
            "door_colours_overheads": "Polytec - Thermolaminated Vinyl Style 1 - Vienna - Classic White Matt",
            "door_colours_base": "Polytec - Thermolaminated Vinyl Style 1 - Vienna - Classic White Matt",
            "door_colours_tall": "Vertical on Tall doors only",
            "door_colours_island": "Horizontal on all",
            "door_colours_bar_back": "",
            "has_explicit_overheads": True,
            "has_explicit_base": True,
            "has_explicit_tall": False,
            "has_explicit_island": False,
            "has_explicit_bar_back": False,
            "toe_kick": ["As Doors"],
            "bulkheads": ["None"],
            "handles": [],
            "floating_shelf": "",
            "led": "",
            "accessories": [],
            "other_items": [],
            "sink_info": "",
            "basin_info": "",
            "tap_info": "Mixer Tap Clients own | Water Filter Tap Clients own | Mixer Tap Clients own",
            "drawers_soft_close": "",
            "hinges_soft_close": "",
            "splashback": "",
            "flooring": "",
            "source_file": "imperial.pdf",
            "page_refs": "1",
            "evidence_snippet": "",
            "confidence": 0.7,
        }
        cleaned = parsing_module.apply_snapshot_cleaning_rules(
            {"rooms": [row], "appliances": [], "special_sections": [], "others": {}, "warnings": []}
        )["rooms"][0]
        self.assertEqual(cleaned["door_colours_tall"], "")
        self.assertEqual(cleaned["door_colours_island"], "")
        self.assertEqual(cleaned["tap_info"], "Mixer Tap Clients own | Water Filter Tap Clients own")

    def test_split_door_colour_groups_keeps_tall_open_shelves_as_tall(self) -> None:
        grouped = parsing_module._split_door_colour_groups(
            [
                "Door Colour 1 - Polytec Classic White Matt Finish Thermolaminate - Atlanta EM2 Profile - Base Cabinetry",
                "Door Colour 2 - Polytec Classic White Matt Finsih Melamine with Matching 1MM ABSE Edges - to Tall Open Shelves",
            ]
        )
        self.assertIn("Atlanta EM2 Profile", grouped["door_colours_base"])
        self.assertEqual(
            grouped["door_colours_tall"],
            "Polytec Classic White Matt Finsih Melamine with Matching 1MM ABSE Edges - to Tall Open Shelves",
        )

    def test_stable_hybrid_room_merge_keeps_base_accessories_and_rejects_orientation_only_ai_groups(self) -> None:
        merged = extraction_service._merge_single_room(
            {
                "room_key": "kitchen",
                "original_room_label": "KITCHEN",
                "accessories": [],
                "door_colours_tall": "",
                "door_colours_island": "",
            },
            {
                "room_key": "kitchen",
                "original_room_label": "KITCHEN",
                "accessories": ["OE ELSAFE DESK PRODIGY CABLE BASKET 950MM BLACK"],
                "door_colours_tall": "Polytec Vertical",
                "door_colours_island": "Polytec Horizontal on all",
            },
            stable_hybrid=True,
        )
        self.assertEqual(merged["accessories"], [])
        self.assertEqual(merged["door_colours_tall"], "")
        self.assertEqual(merged["door_colours_island"], "")

    def test_job_detail_page_hides_review_cards(self) -> None:
        builder_id = store.create_builder("Imperial", "imperial", "")
        job_id = store.create_job("37647", builder_id, "Imperial Test", "")
        client = TestClient(app)
        self._login(client)
        response = client.get(f"/jobs/{job_id}")
        self.assertEqual(response.status_code, 200)
        self.assertNotIn("<span class=\"eyebrow\">Review</span>", response.text)

    def test_shared_generic_overlay_filters_simonds_placeholder_fragments(self) -> None:
        section = {
            "original_section_label": "Laundry",
            "page_type": "joinery",
            "file_name": "simonds.pdf",
            "page_nos": [9],
            "text": "Laundry layout",
            "layout_rows": [
                {"row_label": "Manufacturer", "value_text": "Laminex", "supplier_text": "", "notes_text": "", "row_kind": "material"},
                {"row_label": "Finish", "value_text": "Natural", "supplier_text": "", "notes_text": "", "row_kind": "material"},
                {"row_label": "Profile", "value_text": "(N/A)", "supplier_text": "", "notes_text": "", "row_kind": "material"},
                {"row_label": "Colour", "value_text": "Chalk White", "supplier_text": "", "notes_text": "", "row_kind": "material"},
                {"row_label": "Kickboard", "value_text": "", "supplier_text": "", "notes_text": "", "row_kind": "material"},
                {"row_label": "Manufacturer", "value_text": "N/A", "supplier_text": "", "notes_text": "", "row_kind": "handle"},
                {"row_label": "Model", "value_text": "C137 Black 100m", "supplier_text": "", "notes_text": "", "row_kind": "handle"},
                {"row_label": "Fixing", "value_text": "Horizontal", "supplier_text": "", "notes_text": "", "row_kind": "handle"},
                {"row_label": "Cabinetry Handles", "value_text": "", "supplier_text": "", "notes_text": "", "row_kind": "handle"},
                {"row_label": "Manufacturer", "value_text": "(N/A)", "supplier_text": "", "notes_text": "", "row_kind": "handle"},
                {"row_label": "Model", "value_text": "up to 20mm Drop Down - No Handle", "supplier_text": "", "notes_text": "", "row_kind": "handle"},
                {"row_label": "Style", "value_text": "Soft Close", "supplier_text": "", "notes_text": "", "row_kind": "handle"},
                {"row_label": "Profile", "value_text": "Hanging Rail", "supplier_text": "", "notes_text": "", "row_kind": "handle"},
                {"row_label": "Overhead Cabinetry Handles", "value_text": "(N/A)", "supplier_text": "", "notes_text": "", "row_kind": "handle"},
                {"row_label": "Manufacturer", "value_text": "Everhard", "supplier_text": "", "notes_text": "", "row_kind": "sink"},
                {"row_label": "Range", "value_text": "Milano", "supplier_text": "", "notes_text": "", "row_kind": "sink"},
                {"row_label": "Style", "value_text": "Vegie Mixer", "supplier_text": "", "notes_text": "", "row_kind": "sink"},
                {"row_label": "Finish", "value_text": "Chrome", "supplier_text": "", "notes_text": "", "row_kind": "sink"},
                {"row_label": "Laundry Trough", "value_text": "", "supplier_text": "", "notes_text": "", "row_kind": "sink"},
                {"row_label": "Manufacturer", "value_text": "Alder", "supplier_text": "", "notes_text": "", "row_kind": "tap"},
                {"row_label": "Laundry Tapware", "value_text": "+ ALDER SACHI", "supplier_text": "", "notes_text": "", "row_kind": "tap"},
                {"row_label": "Manufacturer", "value_text": "(N/A)", "supplier_text": "", "notes_text": "", "row_kind": "accessory"},
                {"row_label": "Style", "value_text": "Excellence Squareline 45L", "supplier_text": "", "notes_text": "", "row_kind": "accessory"},
                {"row_label": "Finish", "value_text": "Stainless Steel", "supplier_text": "", "notes_text": "", "row_kind": "accessory"},
                {"row_label": "Accessories", "value_text": "N/A", "supplier_text": "", "notes_text": "", "row_kind": "accessory"},
                {"row_label": "Robe Hook", "value_text": "IN MATT BLACK", "supplier_text": "", "notes_text": "", "row_kind": "accessory"},
            ],
        }
        overlay = extraction_service._extract_generic_layout_overlay(section)
        self.assertEqual(overlay["toe_kick"], ["Laminex - Chalk White"])
        self.assertEqual(overlay["handles"], ["C137 Black 100m - Horizontal", "up to 20mm Drop Down - No Handle"])
        self.assertIn("Everhard", overlay["sink_info"])
        self.assertIn("Milano", overlay["sink_info"])
        self.assertIn("Alder", overlay["tap_info"])
        self.assertEqual(overlay["tap_info"], "Alder")
        self.assertEqual(overlay["accessories"], ["Excellence Squareline 45L - Stainless Steel"])

    def test_polish_generic_layout_room_preserves_explicit_overheads_for_non_kitchen(self) -> None:
        row = {
            "room_key": "laundry",
            "original_room_label": "Laundry",
            "door_colours_base": "old base",
            "door_colours_overheads": "old overhead",
            "has_explicit_base": False,
            "has_explicit_overheads": False,
        }
        overlay = {
            "door_colours_base": "Laminex - Chalk White - S/Edge",
            "door_colours_overheads": "Laminex - Blackened Legno - S/Edge",
            "has_explicit_base": True,
            "has_explicit_overheads": True,
        }
        polished = extraction_service._polish_generic_layout_room(row, overlay)
        cleaned = parsing_module.apply_snapshot_cleaning_rules({"rooms": [polished], "warnings": [], "others": {}, "appliances": []})
        laundry = cleaned["rooms"][0]
        self.assertEqual(laundry["door_colours_base"], "Laminex - Chalk White - S/Edge")
        self.assertEqual(laundry["door_colours_overheads"], "Laminex - Blackened Legno - S/Edge")

    def test_extract_generic_layout_overlay_skips_wet_area_accessory_noise_for_non_wet_rooms(self) -> None:
        section = {
            "original_section_label": "Kitchen",
            "file_name": "simonds.pdf",
            "page_nos": [15],
            "text": "Kitchen",
            "page_type": "joinery",
            "layout_rows": [
                {"row_label": "Robe Hook", "value_text": "IN MATT BLACK", "supplier_text": "", "notes_text": "", "row_kind": "material"},
            ],
        }
        overlay = extraction_service._extract_generic_layout_overlay(section)
        self.assertEqual(overlay["accessories"], [])

    def test_prepare_simonds_layout_text_inserts_exact_continuation_heading_and_strips_internal_paint_noise(self) -> None:
        raw_text = (
            "Selection Level 1 "
            "Study Manufacturer Laminex Profile S/Edge Range Natural Colour Blackended Legno Benchtop NOOK 1 AND NOOK 2 "
            "Internal Paint Selctions Internal Paint (N/A) Manufacturer Haymes Finish Newlife Ceiling "
            "Internal Fittings Selections Flooring Manufacturer Carpet Call "
            "Manufacturer Caesarstone - Mineral Profile 40mm Arris Range Caesarstone Standard M1 Colour Organic White Wall Run Benchtop "
            "Manufacturer Laminex Finish Natural Profile S/Edge Colour Blackened Legno Wall Run Base Cabinet Panels "
            "Manufacturer Franke Range Sirius Style SID 110-34 - Undermount Finish Carbon Black Kitchen Sink "
            "Bulters/WIP Manufacturer Caesarstone - Mineral Profile 40mm Arris Range Caesarstone Standard M1 Colour Organic White Benchtop"
        )
        prepared = extraction_service._prepare_simonds_layout_text(raw_text)
        self.assertIn("\nKitchen\n", prepared)
        self.assertNotIn("Internal Paint Selctions", prepared)
        self.assertNotIn("Internal Fittings Selections", prepared)
        self.assertIn("Bulters/WIP", prepared)

    def test_build_layout_from_pdf_tables_assigns_leading_rows_to_first_fallback_room_before_explicit_second_room(self) -> None:
        lines = [
            "Kitchen",
            "Benchtop",
            "Kitchen Sink",
            "Bulters/WIP",
            "Benchtop",
        ]
        page = {
            "table_rows": [
                [
                    ["Manufacturer", "Caesarstone - Mineral", "", ""],
                    ["Profile", "40mm Arris", "", ""],
                    ["Range", "Caesarstone Standard M1", "", ""],
                    ["Colour", "Organic White", "", ""],
                    ["Benchtop", "", "", ""],
                    ["Kitchen Sink", "Franke Sirius", "", ""],
                    ["Bulters/WIP", "", "", ""],
                    ["Manufacturer", "Caesarstone - Mineral", "", ""],
                    ["Profile", "40mm Arris", "", ""],
                    ["Range", "Caesarstone Standard M1", "", ""],
                    ["Colour", "Organic White", "", ""],
                    ["Benchtop", "", "", ""],
                ]
            ]
        }
        blocks = extraction_service._build_layout_from_pdf_tables("Simonds", "joinery", lines, page)
        labels = [block["room_label"] for block in blocks]
        self.assertEqual(labels, ["Kitchen", "Butlers/WIP"])
        self.assertTrue(any(row["row_label"] == "Kitchen Sink" for row in blocks[0]["rows"]))
        self.assertGreater(len(blocks[1]["rows"]), 0)

    def test_build_generic_layout_blocks_reassigns_prefix_properties_to_upcoming_anchor(self) -> None:
        rows = [
            {"row_label": "Cabinetry Handles", "value_region_text": "", "supplier_region_text": "", "notes_region_text": "", "row_kind": "handle"},
            {"row_label": "Manufacturer", "value_region_text": "N/A", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
            {"row_label": "Model", "value_region_text": "up to 20mm Drop Down - No Handle", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
            {"row_label": "Manufacturer", "value_region_text": "Franke", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
            {"row_label": "Range", "value_region_text": "Sirius", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
            {"row_label": "Kitchen Sink", "value_region_text": "", "supplier_region_text": "", "notes_region_text": "", "row_kind": "sink"},
        ]
        blocks = extraction_service._build_generic_layout_blocks(rows, page_type="joinery")
        self.assertEqual([(block["anchor_kind"], block["anchor_label"]) for block in blocks], [("handles", "Cabinetry Handles"), ("sink", "Kitchen Sink")])
        handle_rows = [row["row_label"] for row in blocks[0]["rows"]]
        sink_rows = [row["row_label"] for row in blocks[1]["rows"]]
        self.assertNotIn("Range", handle_rows)
        self.assertIn("Manufacturer", sink_rows)
        self.assertIn("Range", sink_rows)

    def test_build_generic_layout_blocks_keeps_material_properties_with_current_material_anchor(self) -> None:
        rows = [
            {"row_label": "Benchtops", "value_region_text": "", "supplier_region_text": "", "notes_region_text": "No shelf to cupboard underneath sink", "row_kind": "material"},
            {"row_label": "Manufacturer", "value_region_text": "Quantum Quartz", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
            {"row_label": "Colour", "value_region_text": "Champagne", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
            {"row_label": "Edge Profile", "value_region_text": "20mm Arissed", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
            {"row_label": "Underbench", "value_region_text": "including Island", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
            {"row_label": "Manufacturer", "value_region_text": "Polytec", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
        ]
        blocks = extraction_service._build_generic_layout_blocks(rows, page_type="joinery")
        self.assertEqual([(block["anchor_kind"], block["anchor_label"]) for block in blocks], [("bench", "Benchtops"), ("base", "Underbench")])
        bench_rows = [row["row_label"] for row in blocks[0]["rows"]]
        base_rows = [row["row_label"] for row in blocks[1]["rows"]]
        self.assertIn("Colour", bench_rows)
        self.assertIn("Edge Profile", bench_rows)
        self.assertNotIn("Colour", base_rows)

    def test_extract_generic_layout_overlay_rebuilds_docling_like_evoca_grouped_rows(self) -> None:
        section = {
            "original_section_label": "Kitchen",
            "section_key": "kitchen",
            "layout_rows": [
                {
                    "row_label": "15 CABINETS All Cabinets include Soft Close Hinges &Runners, lined internally with White Melamine &includes ABS edging. Benchtops over maximum length",
                    "value_region_text": "15 CABINETS All Cabinets include Soft Close Hinges &Runners, lined internally with White Melamine &includes ABS edging. Benchtops over maximum length",
                    "supplier_region_text": "",
                    "notes_region_text": "",
                    "row_kind": "material",
                },
                {
                    "row_label": "***One stone colour included, additional $350 charge for each additional stone colour used***",
                    "value_region_text": "***One stone colour included, additional $350 charge for each additional stone colour used***",
                    "supplier_region_text": "",
                    "notes_region_text": "",
                    "row_kind": "material",
                },
                {"row_label": "-", "value_region_text": "Benchtops", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
                {"row_label": "Manufacturer", "value_region_text": "Manufacturer", "supplier_region_text": "Quantum Quartz", "notes_region_text": "", "row_kind": "material"},
                {"row_label": "Colour", "value_region_text": "Colour", "supplier_region_text": "Champagne", "notes_region_text": "", "row_kind": "material"},
                {"row_label": "Edge Profile", "value_region_text": "Edge Profile", "supplier_region_text": "20mm Arissed", "notes_region_text": "", "row_kind": "material"},
                {"row_label": "Underbench", "value_region_text": "including Island Underbench including Island", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
                {"row_label": "Manufacturer", "value_region_text": "Manufacturer", "supplier_region_text": "Polytec", "notes_region_text": "", "row_kind": "material"},
                {"row_label": "Colour", "value_region_text": "&Finish Colour &Finish", "supplier_region_text": "Belgian Oak Matt", "notes_region_text": "", "row_kind": "material"},
                {"row_label": "Kickboard", "value_region_text": "Kickboard", "supplier_region_text": "Laminate", "notes_region_text": "", "row_kind": "material"},
                {"row_label": "Handles", "value_region_text": "Handles", "supplier_region_text": "4062-128-TG", "notes_region_text": "", "row_kind": "material"},
                {"row_label": "Door Handle", "value_region_text": "Door Handle", "supplier_region_text": "", "notes_region_text": "Vertical", "row_kind": "handle"},
                {"row_label": "Drawer Handle", "value_region_text": "Drawer Handle", "supplier_region_text": "", "notes_region_text": "Horizontal", "row_kind": "handle"},
                {"row_label": "Pantry Door Handle", "value_region_text": "** Pantry Door Handle**", "supplier_region_text": "#N/A", "notes_region_text": "", "row_kind": "handle"},
                {"row_label": "Overhead Cupboards", "value_region_text": "Overhead Cupboards", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
                {"row_label": "Manufacturer", "value_region_text": "Manufacturer", "supplier_region_text": "Polytec", "notes_region_text": "", "row_kind": "material"},
                {"row_label": "Colour", "value_region_text": "&Finish Colour &Finish", "supplier_region_text": "Belgian Oak Matt", "notes_region_text": "", "row_kind": "material"},
                {"row_label": "Handles", "value_region_text": "Handles", "supplier_region_text": "Finger Grip", "notes_region_text": "", "row_kind": "material"},
                {"row_label": "-", "value_region_text": "Pantry Doors", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
                {"row_label": "Manufacturer", "value_region_text": "Manufacturer", "supplier_region_text": "Polytec", "notes_region_text": "", "row_kind": "material"},
                {"row_label": "Colour", "value_region_text": "&Finish Colour &Finish", "supplier_region_text": "Belgian Oak Matt", "notes_region_text": "", "row_kind": "material"},
                {"row_label": "Kickboard", "value_region_text": "Kickboard", "supplier_region_text": "Laminate", "notes_region_text": "", "row_kind": "material"},
            ],
        }
        overlay = extraction_service._extract_generic_layout_overlay(section)
        self.assertEqual(overlay["bench_tops_wall_run"], "20mm Quantum Quartz - Champagne - Arissed")
        self.assertEqual(overlay["door_colours_base"], "Polytec - Belgian Oak Matt")
        self.assertEqual(overlay["door_colours_overheads"], "Polytec - Belgian Oak Matt")
        self.assertEqual(overlay["door_colours_tall"], "Polytec - Belgian Oak Matt")
        self.assertEqual(overlay["toe_kick"], ["Polytec - Belgian Oak Matt - Laminate"])
        self.assertEqual(overlay["handles"], ["4062-128-TG - Vertical - Horizontal", "Finger Grip"])

    def test_format_generic_material_from_parts_excludes_builder_surcharge_note(self) -> None:
        formatted = extraction_service._format_generic_material_from_parts(
            {
                "manufacturer": ["Quantum Quartz"],
                "colour": ["Champagne"],
                "profile": ["20mm Arissed"],
                "note": ["One stone colour included, additional $350 charge for each additional stone colour used"],
            }
        )
        self.assertEqual(formatted, "20mm Quantum Quartz - Champagne - Arissed")

    def test_extract_generic_layout_overlay_ignores_unrelated_study_paint_rows(self) -> None:
        section = {
            "original_section_label": "Study",
            "section_key": "study",
            "page_type": "joinery",
            "layout_rows": [
                {"row_label": "Benchtop", "value_region_text": "", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
                {"row_label": "Manufacturer", "value_region_text": "Laminex", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
                {"row_label": "Profile", "value_region_text": "S/Edge", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
                {"row_label": "Colour", "value_region_text": "Blackended Legno", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
                {"row_label": "Internal Paint", "value_region_text": "Haymes", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
                {"row_label": "Skirtings/Architraves", "value_region_text": "Greyology 1", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
            ],
        }
        overlay = extraction_service._extract_generic_layout_overlay(section)
        self.assertEqual(overlay["bench_tops_other"], "Laminex - Blackended Legno - S/Edge")

    def test_build_generic_layout_blocks_keeps_sink_model_with_current_sink_anchor(self) -> None:
        rows = [
            {"row_label": "Sink", "value_region_text": "", "supplier_region_text": "", "notes_region_text": "", "row_kind": "sink"},
            {"row_label": "Model", "value_region_text": "Burazzo 450mm Gun Metal Single Bowl Sink", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
            {"row_label": "Type", "value_region_text": "Top Mount", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
            {"row_label": "Sink Mixer", "value_region_text": "", "supplier_region_text": "", "notes_region_text": "", "row_kind": "tap"},
            {"row_label": "Type", "value_region_text": "Zara Gun Metal Pull-Out", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
            {"row_label": "Location", "value_region_text": "Centre of Sink", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
        ]
        blocks = extraction_service._build_generic_layout_blocks(rows, page_type="sinkware_tapware")
        self.assertEqual([(block["anchor_kind"], block["anchor_label"]) for block in blocks], [("sink", "Sink"), ("tap", "Sink Mixer")])
        sink_rows = [row["row_label"] for row in blocks[0]["rows"]]
        tap_rows = [row["row_label"] for row in blocks[1]["rows"]]
        self.assertIn("Model", sink_rows)
        self.assertIn("Type", sink_rows)
        self.assertIn("Location", tap_rows)

    def test_build_generic_layout_blocks_redirects_repeated_fixture_properties_to_future_anchor(self) -> None:
        rows = [
            {"row_label": "Kitchen Sink", "value_region_text": "", "supplier_region_text": "", "notes_region_text": "", "row_kind": "sink"},
            {"row_label": "Manufacturer", "value_region_text": "Franke", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
            {"row_label": "Range", "value_region_text": "Sirius", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
            {"row_label": "Style", "value_region_text": "SID 110-34 - Undermount", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
            {"row_label": "Finish", "value_region_text": "Carbon Black", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
            {"row_label": "Manufacturer", "value_region_text": "Alder", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
            {"row_label": "Range", "value_region_text": "Maxx", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
            {"row_label": "Model", "value_region_text": "Rectangle Sink Mixer", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
            {"row_label": "Finish", "value_region_text": "Matt Black", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
            {"row_label": "Kitchen Tapware", "value_region_text": "+ ALDER SACHI", "supplier_region_text": "", "notes_region_text": "", "row_kind": "tap"},
        ]
        blocks = extraction_service._build_generic_layout_blocks(rows, page_type="joinery")
        self.assertEqual([(block["anchor_kind"], block["anchor_label"]) for block in blocks], [("sink", "Kitchen Sink"), ("tap", "Kitchen Tapware")])
        sink_rows = [row["row_label"] for row in blocks[0]["rows"]]
        tap_rows = [row["row_label"] for row in blocks[1]["rows"]]
        self.assertEqual(sink_rows.count("Manufacturer"), 1)
        self.assertEqual(tap_rows.count("Manufacturer"), 1)
        self.assertIn("Model", tap_rows)

    def test_build_generic_layout_blocks_redirects_initial_tap_prefix_properties_to_future_anchor(self) -> None:
        rows = [
            {"row_label": "Kitchen Sink", "value_region_text": "", "supplier_region_text": "", "notes_region_text": "", "row_kind": "sink"},
            {"row_label": "Manufacturer", "value_region_text": "Franke", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
            {"row_label": "Range", "value_region_text": "Sirius", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
            {"row_label": "Model", "value_region_text": "SID 110-34", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
            {"row_label": "Manufacturer", "value_region_text": "Alder", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
            {"row_label": "Model", "value_region_text": "Rectangle Sink Mixer", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
            {"row_label": "Kitchen Tapware", "value_region_text": "", "supplier_region_text": "", "notes_region_text": "", "row_kind": "tap"},
        ]
        blocks = extraction_service._build_generic_layout_blocks(rows, page_type="joinery")
        self.assertEqual([(block["anchor_kind"], block["anchor_label"]) for block in blocks], [("sink", "Kitchen Sink"), ("tap", "Kitchen Tapware")])
        sink_rows = [row["row_label"] for row in blocks[0]["rows"]]
        tap_rows = [row["row_label"] for row in blocks[1]["rows"]]
        self.assertEqual(sink_rows.count("Manufacturer"), 1)
        self.assertEqual(tap_rows.count("Manufacturer"), 1)
        self.assertIn("Model", tap_rows)

    def test_normalize_layout_rows_splits_embedded_waterfall_end_anchor(self) -> None:
        rows = extraction_service._normalize_layout_rows(
            [
                {
                    "row_label": "Colour",
                    "value_region_text": "Organic White Waterfall End Panels",
                    "supplier_region_text": "",
                    "notes_region_text": "",
                    "row_kind": "material",
                },
                {
                    "row_label": "Manufacturer",
                    "value_region_text": "Laminex",
                    "supplier_region_text": "",
                    "notes_region_text": "",
                    "row_kind": "material",
                },
            ]
        )
        self.assertEqual(rows[0]["row_label"], "Colour")
        self.assertEqual(rows[0]["value_region_text"], "Organic White")
        self.assertEqual(rows[1]["row_label"], "Waterfall End Panels")

    def test_invalid_room_heading_candidate_rejects_quantity_towel_hook_line(self) -> None:
        self.assertTrue(
            extraction_service._looks_like_invalid_room_heading_candidate(
                "2No Bath towel hooks + 1No Hand towel hook"
            )
        )

    def test_shared_generic_polish_clears_placeholder_values_when_layout_block_is_present(self) -> None:
        row = {
            "room_key": "kitchen",
            "original_room_label": "Kitchen",
            "handles": ["Door Handle Drawer Handle **"],
            "sink_info": "Model Type #N/A",
            "tap_info": "Type Location Centre of Sink",
            "basin_info": "Not Applicable Model Type",
            "accessories": ["WC**"],
        }
        overlay = {
            "has_handles_block": True,
            "handles": [],
            "has_sink_block": True,
            "sink_info": "",
            "has_tap_block": True,
            "tap_info": "",
            "has_basin_block": True,
            "basin_info": "",
            "has_accessories_block": True,
            "accessories": [],
        }
        polished = extraction_service._polish_generic_layout_room(row, overlay)
        self.assertEqual(polished["handles"], [])
        self.assertEqual(polished["sink_info"], "")
        self.assertEqual(polished["tap_info"], "")
        self.assertEqual(polished["basin_info"], "")
        self.assertEqual(polished["accessories"], [])

    def test_shared_generic_polish_replaces_noisy_material_fields_when_explicit_layout_blocks_exist(self) -> None:
        row = {
            "room_key": "kitchen",
            "original_room_label": "Kitchen",
            "bench_tops": ["Manufacturer Laminex Range Natural Colour Blackened Legno Wall Run Benchtop"],
            "bench_tops_wall_run": "Manufacturer Laminex Range Natural Colour Blackened Legno Wall Run Benchtop",
            "bench_tops_island": "",
            "bench_tops_other": "",
            "door_colours_base": "Manufacturer Laminex Profile S/Edge Colour Blackened Legno Wall Run Base Cabinet Panels",
            "door_colours_overheads": "Manufacturer Laminex Profile S/Edge Colour Blackened Legno Overhead Cupboards",
            "door_colours_tall": "Manufacturer Laminex Profile S/Edge Colour Blackened Legno Pantry Doors",
            "floating_shelf": "Manufacturer Laminex Profile S/Edge Colour Blackened Legno Floating Shelf",
        }
        overlay = {
            "has_bench_block": True,
            "bench_tops_wall_run": "40mm Caesarstone - Organic White - Arris",
            "has_explicit_base": True,
            "door_colours_base": "Laminex - Blackened Legno",
            "has_explicit_overheads": True,
            "door_colours_overheads": "Laminex - Blackened Legno",
            "has_explicit_tall": True,
            "door_colours_tall": "Laminex - Blackened Legno",
            "has_floating_shelf_block": True,
            "floating_shelf": "Laminex - Blackened Legno - Vertical Grain",
        }
        polished = extraction_service._polish_generic_layout_room(row, overlay)
        self.assertEqual(polished["bench_tops_wall_run"], "40mm Caesarstone - Organic White - Arris")
        self.assertEqual(polished["door_colours_base"], "Laminex - Blackened Legno")
        self.assertEqual(polished["door_colours_overheads"], "Laminex - Blackened Legno")
        self.assertEqual(polished["door_colours_tall"], "Laminex - Blackened Legno")
        self.assertEqual(polished["floating_shelf"], "Laminex - Blackened Legno - Vertical Grain")

    def test_shared_generic_polish_clears_existing_materials_when_explicit_not_applicable_blocks_exist(self) -> None:
        row = {
            "room_key": "butlers",
            "original_room_label": "Butlers",
            "bench_tops": ["15 CABINETS | 20mm Quantum Quartz - Champagne - Arissed"],
            "bench_tops_wall_run": "15 CABINETS | 20mm Quantum Quartz - Champagne - Arissed",
            "bench_tops_island": "",
            "bench_tops_other": "",
            "door_colours_base": "Polytec - Belgian Oak Matt",
            "door_colours_overheads": "Polytec - Belgian Oak Matt",
            "door_colours_tall": "",
            "handles": ["4062-128-TG - Vertical"],
        }
        overlay = {
            "has_bench_block": True,
            "bench_tops_wall_run": "",
            "bench_tops_island": "",
            "bench_tops_other": "",
            "has_explicit_base": True,
            "door_colours_base": "",
            "has_explicit_overheads": True,
            "door_colours_overheads": "",
            "has_handles_block": True,
            "handles": [],
        }
        polished = extraction_service._polish_generic_layout_room(row, overlay)
        self.assertEqual(polished["bench_tops"], [])
        self.assertEqual(polished["bench_tops_wall_run"], "")
        self.assertEqual(polished["door_colours_base"], "")
        self.assertEqual(polished["door_colours_overheads"], "")
        self.assertEqual(polished["handles"], [])

    def test_shared_generic_polish_preserves_more_complete_existing_handles_over_truncated_overlay(self) -> None:
        row = {
            "room_key": "kitchen",
            "original_room_label": "Kitchen",
            "handles": [
                "Handless Lip Pull N/A",
                "(To All Lower Doors & Drawers excluding Pantry) C137 Caloundra Lip Pull Matt Black 100mm (Laid Horizontal to All)",
                "C137 Caloundra Lip Pull Matt Black 200mm (Laid Vertical)",
            ],
        }
        overlay = {
            "has_handles_block": True,
            "handles": [
                "Handless Lip Pull",
                "(To All Lower Doors & Drawers",
            ],
        }
        polished = extraction_service._polish_generic_layout_room(row, overlay)
        self.assertEqual(
            polished["handles"],
            [
                "Handless Lip Pull N/A",
                "(To All Lower Doors & Drawers excluding Pantry) C137 Caloundra Lip Pull Matt Black 100mm (Laid Horizontal to All)",
                "C137 Caloundra Lip Pull Matt Black 200mm (Laid Vertical)",
            ],
        )

    def test_shared_generic_polish_clears_noisy_fixture_fields_when_layout_blocks_exist(self) -> None:
        row = {
            "room_key": "kitchen",
            "original_room_label": "Kitchen",
            "sink_info": "Manufacturer Franke Range Sirius Model SID 110-34 Finish Carbon Black",
            "tap_info": "Manufacturer Alder Range Maxx Model Rectangle Sink Mixer Finish Matt Black Client Name: Test",
            "basin_info": "Manufacturer Caroma Range Liano II Model 400mm Round Above Counter Basin",
        }
        overlay = {
            "has_sink_block": True,
            "sink_info": "Franke - Sirius - SID 110-34 - Carbon Black",
            "has_tap_block": True,
            "tap_info": "Alder - Maxx - Rectangle Sink Mixer - Matt Black",
            "has_basin_block": True,
            "basin_info": "Caroma - Liano II - 400mm Round Above Counter Basin",
        }
        polished = extraction_service._polish_generic_layout_room(row, overlay)
        self.assertEqual(polished["sink_info"], "Franke - Sirius - SID 110-34 - Carbon Black")
        self.assertEqual(polished["tap_info"], "Alder - Maxx - Rectangle Sink Mixer - Matt Black")
        self.assertEqual(polished["basin_info"], "Caroma - Liano II - 400mm Round Above Counter Basin")

    def test_polish_generic_layout_room_clears_noisy_flooring_without_explicit_floor_block(self) -> None:
        row = {
            "room_key": "study",
            "original_room_label": "Study",
            "flooring": "Carpet Manufacturer Carpet Call Range Category 4 - Grand Escape Colour Kerella #78 Underlay Luxury Plus Underlay Timber Manufactuer Carpet Call",
        }
        polished = extraction_service._polish_generic_layout_room(row, {})
        self.assertEqual(polished["flooring"], "")

    def test_imperial_layout_row_fixture_entry_removes_client_name_tail(self) -> None:
        self.assertEqual(
            parsing_module._imperial_layout_row_fixture_entry(
                {
                    "value_text": "Franke Eos Neo pull out tap copper TA9601CP Eloise Cawcutt-Foxover 26-03-2026",
                    "supplier_text": "",
                    "notes_text": "BY CLIENT",
                },
                "tap",
            ),
            "Franke Eos Neo pull out tap copper TA9601CP",
        )

    def test_imperial_layout_row_fixture_entry_strips_leading_tap_label(self) -> None:
        self.assertEqual(
            parsing_module._imperial_layout_row_fixture_entry(
                {
                    "value_text": "Tap Franke Eos Neo pull out tap copper TA9601CP",
                    "supplier_text": "",
                    "notes_text": "",
                },
                "tap",
            ),
            "Franke Eos Neo pull out tap copper TA9601CP",
        )

    def test_imperial_layout_row_fixture_entry_removes_sink_tail_metadata(self) -> None:
        self.assertEqual(
            parsing_module._imperial_layout_row_fixture_entry(
                {
                    "value_text": "ABEY Schock horizontal double bowl sink bronze N200BZ TOP MOUNT Taphole location: Tap Landing Centred to back",
                    "supplier_text": "",
                    "notes_text": "Sink Pre-punched Hole Centre of Sink",
                },
                "sink",
            ),
            "ABEY Schock horizontal double bowl sink bronze N200BZ TOP MOUNT",
        )

    def test_imperial_layout_row_material_text_strips_heading_and_meta_noise(self) -> None:
        self.assertEqual(
            parsing_module._imperial_layout_row_material_text(
                {
                    "value_text": "## KITCHEN JOINERY SELECTION SHEET Polytec Classic White Matt",
                    "supplier_text": "",
                    "notes_text": "Cabinetry Height: 2400",
                }
            ),
            "Polytec - Classic White Matt",
        )

    def test_clean_imperial_layout_fragment_strips_private_name_date_prefix(self) -> None:
        self.assertEqual(
            parsing_module._clean_imperial_layout_fragment(
                "9 Greenland Court, Springfield PRIVATE - Yasantha Warawita 12/03/2026 Caesarstone"
            ),
            "Caesarstone",
        )

    def test_imperial_room_row_dedupes_floating_shelf_fragments(self) -> None:
        row = parsing_module.RoomRow(
            room_key="kitchen",
            original_room_label="KITCHEN",
            floating_shelf="Laminex - 51mmthick floating shelves Blackbutt Truescale Natural Finish 2618 | 51mm Laminex - thick floating shelves Blackbutt Truescale Natural Finish 2618",
        )
        row.floating_shelf = parsing_module._dedupe_delimited_fragments(
            parsing_module._collapse_repeated_token_sequence(row.floating_shelf)
        )
        self.assertEqual(row.floating_shelf, "Laminex - 51mmthick floating shelves Blackbutt Truescale Natural Finish 2618")

    def test_infer_page_type_detects_generic_sinkware_tables_without_explicit_header(self) -> None:
        text = (
            "Vanity Basin Tapware Matt Black\n"
            "Toilet Roll Holder\n"
            "Toilet Suite\n"
            "Shower Base\n"
            "Shower Frame\n"
            "Shower Mixer\n"
            "Floor Waste\n"
            "Wet Area Location\n"
        )
        page_type = extraction_service._infer_page_type_from_text("Simonds", "spec", text)
        self.assertEqual(page_type, "sinkware_tapware")

    def test_infer_page_type_prefers_joinery_when_joinery_signals_dominate_without_explicit_sinkware_heading(self) -> None:
        text = (
            "Kitchen\n"
            "Wall Run Benchtop\n"
            "Base Cabinet Panels\n"
            "Kickboard\n"
            "Cabinetry Handles\n"
            "Kitchen Sink\n"
            "Kitchen Tapware\n"
        )
        page_type = extraction_service._infer_page_type_from_text("Simonds", "spec", text)
        self.assertEqual(page_type, "joinery")

    def test_imperial_clean_flooring_value_preserves_extra_context(self) -> None:
        self.assertEqual(
            parsing_module._imperial_clean_flooring_value(["Tiles with Floating TBC"]),
            "Tiles with Floating TBC",
        )

    def test_heuristic_sink_tap_room_blocks_recovers_wet_area_location_room(self) -> None:
        lines = [
            "Additional Wet Area",
            "Location Ensuite 3",
            "Vanity Basin",
            "Manufacturer Caroma",
            "Range Liano II",
            "Model 400mm Round Above Counter Basin",
            "Vanity Basin Tapware",
            "Manufacturer Alder",
        ]
        blocks = extraction_service._heuristic_sink_tap_room_blocks(lines)
        self.assertEqual(len(blocks), 1)
        self.assertEqual(blocks[0]["room_label"], "Ensuite 3")

    def test_invalid_room_heading_candidate_rejects_field_titles(self) -> None:
        self.assertTrue(extraction_service._looks_like_invalid_room_heading_candidate("Floating Vanity"))
        self.assertTrue(extraction_service._looks_like_invalid_room_heading_candidate("Robe Sliding Type Frame Colour"))
        self.assertFalse(extraction_service._looks_like_invalid_room_heading_candidate("Dining Banquette"))

    def test_build_layout_from_pdf_tables_parses_evoca_sink_table(self) -> None:
        lines = [
            "20 PLUMBING FIXTURES & TAPWARE",
            "Kitchen",
            "Sink",
            "Model",
            "Type",
            "Accessories",
        ]
        page = {
            "table_rows": [
                [
                    ["20 PLUMBING FIXTURES & TAPWARE", "", "", ""],
                    ["", "Kitchen", "", ""],
                    ["-", "Sink\nModel\nType\nAccessories", "", ""],
                    ["", "", "Burazzo 450mm Gun Metal Single Bowl Sink (BU454525S-GM) ($370)", ""],
                    ["", "", "#N/A\nNot Applicable", ""],
                    ["-", "Sink Mixer\nType\nLocation", "", ""],
                    ["", "", "Zara Gun Metal Pull-Out (ZA120-GM)", ""],
                    ["", "", "Centre of Sink", ""],
                ]
            ]
        }
        blocks = extraction_service._build_layout_from_pdf_tables("Evoca", "sinkware_tapware", lines, page)
        self.assertEqual(len(blocks), 1)
        self.assertEqual(blocks[0]["room_label"], "Kitchen")
        labels = [row["row_label"] for row in blocks[0]["rows"]]
        self.assertIn("Model", labels)
        self.assertIn("Sink Mixer", labels)
        model_row = next(row for row in blocks[0]["rows"] if row["row_label"] == "Model")
        self.assertIn("Burazzo 450mm Gun Metal Single Bowl Sink", model_row["value_region_text"])

    def test_table_row_explicit_room_label_accepts_bathroom_and_rejects_placeholder(self) -> None:
        self.assertEqual(
            extraction_service._table_row_explicit_room_label(["", "Bathroom", None, None]),
            "Bathroom",
        )
        self.assertEqual(
            extraction_service._table_row_explicit_room_label(["", "#N/A\nNot Applicable", None, None]),
            "",
        )

    def test_build_generic_layout_blocks_supports_sink_prefix_rows(self) -> None:
        rows = [
            {"row_label": "Manufacturer", "value_text": "Caroma", "row_kind": "material"},
            {"row_label": "Range", "value_text": "Liano II", "row_kind": "material"},
            {"row_label": "Model", "value_text": "400mm Round Above Counter Basin", "row_kind": "material"},
            {"row_label": "Finish", "value_text": "White", "row_kind": "material"},
            {"row_label": "Vanity Basin", "value_text": "", "row_kind": "basin"},
            {"row_label": "Manufacturer", "value_text": "Alder", "row_kind": "material"},
            {"row_label": "Range", "value_text": "Samm", "row_kind": "material"},
            {"row_label": "Model", "value_text": "Wall Basin Mixer", "row_kind": "material"},
            {"row_label": "Finish", "value_text": "Matt Black", "row_kind": "material"},
            {"row_label": "Vanity Basin Tapware", "value_text": "", "row_kind": "tap"},
        ]
        blocks = extraction_service._build_generic_layout_blocks(rows, page_type="sinkware_tapware")
        self.assertEqual(len(blocks), 2)
        self.assertEqual(blocks[0]["anchor_label"], "Vanity Basin")
        self.assertEqual(blocks[1]["anchor_label"], "Vanity Basin Tapware")
        basin_values = [row.get("value_text", "") for row in blocks[0]["rows"]]
        tap_values = [row.get("value_text", "") for row in blocks[1]["rows"]]
        self.assertIn("Caroma", basin_values)
        self.assertIn("400mm Round Above Counter Basin", basin_values)
        self.assertIn("Alder", tap_values)
        self.assertIn("Wall Basin Mixer", tap_values)

    def test_extract_generic_layout_overlay_maps_island_base_panels_to_island_colour(self) -> None:
        section = {
            "original_section_label": "Kitchen",
            "file_name": "simonds.pdf",
            "page_nos": [8],
            "text": "Kitchen",
            "page_type": "joinery",
            "layout_rows": [
                {"row_label": "Island/Penisula Base Cabinet Panels", "value_text": "", "supplier_text": "", "notes_text": "", "row_kind": "material"},
                {"row_label": "Manufacturer", "value_text": "Laminex", "supplier_text": "", "notes_text": "", "row_kind": "material"},
                {"row_label": "Colour", "value_text": "Chalk White", "supplier_text": "", "notes_text": "", "row_kind": "material"},
            ],
        }
        overlay = extraction_service._extract_generic_layout_overlay(section)
        self.assertEqual(overlay["door_colours_island"], "Laminex - Chalk White")
        self.assertEqual(overlay["door_colours_base"], "")

    def test_polish_generic_layout_room_preserves_island_and_bar_back_fields(self) -> None:
        row = {
            "room_key": "kitchen",
            "original_room_label": "Kitchen",
            "door_colours_base": "",
            "door_colours_overheads": "",
            "door_colours_tall": "",
            "door_colours_island": "",
            "door_colours_bar_back": "",
            "floating_shelf": "",
        }
        overlay = {
            "door_colours_island": "Laminex - Chalk White",
            "door_colours_bar_back": "Laminex - Gumnut Natural Finish 2606",
            "floating_shelf": "",
        }
        polished = extraction_service._polish_generic_layout_room(row, overlay)
        self.assertEqual(polished["door_colours_island"], "Laminex - Chalk White")
        self.assertEqual(polished["door_colours_bar_back"], "Laminex - Gumnut Natural Finish 2606")

    def test_generic_fixture_formatter_keeps_real_tap_with_centre_of_sink_note(self) -> None:
        formatted = extraction_service._format_generic_fixture_from_parts(
            {"type": ["Zara Gun Metal Pull-Out (ZA120-GM)"], "location": ["Centre of Sink"]},
            kind="tap",
            anchor_label="Sink Mixer",
        )
        self.assertEqual(formatted, "Zara Gun Metal Pull-Out (ZA120-GM) - Centre of Sink")

    def test_generic_fixture_formatter_drops_accessory_tail_from_tap(self) -> None:
        formatted = extraction_service._format_generic_fixture_from_parts(
            {
                "manufacturer": ["Alder"],
                "range": ["Maxx"],
                "model": ["Rectangle Sink Mixer"],
                "finish": ["Matt Black"],
                "note": ["ALDER SACHI ROBE HOOK IN MATT BLACK"],
            },
            kind="tap",
            anchor_label="Kitchen Tapware",
        )
        self.assertEqual(formatted, "Alder - Maxx - Rectangle Sink Mixer - Matt Black")

    def test_sanitize_generic_fixture_field_removes_washing_machine_taps_noise_from_tap(self) -> None:
        cleaned = extraction_service._sanitize_generic_fixture_field(
            "Spin Gun Metal Gooseneck (SP120-GM) - Washing Machine Taps Chrome Washing Machine Stops, Located inside of cupboard - Corner of Tub",
            kind="tap",
        )
        self.assertEqual(cleaned, "Spin Gun Metal Gooseneck (SP120-GM)")

    def test_generic_overlay_reads_value_region_text_for_evoca_sink_and_tap(self) -> None:
        section = {
            "original_section_label": "Kitchen",
            "page_type": "sinkware_tapware",
            "file_name": "evoca.pdf",
            "page_nos": [12],
            "text": "Kitchen sink/tap page",
            "layout_rows": [
                {"row_label": "Sink", "value_region_text": "", "supplier_region_text": "", "notes_region_text": "", "row_kind": "sink"},
                {"row_label": "Model", "value_region_text": "Burazzo 450mm Gun Metal Single Bowl Sink (BU454525S-GM) ($370)", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
                {"row_label": "Sink Mixer", "value_region_text": "", "supplier_region_text": "", "notes_region_text": "", "row_kind": "tap"},
                {"row_label": "Type", "value_region_text": "Zara Gun Metal Pull-Out (ZA120-GM)", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
                {"row_label": "Location", "value_region_text": "Centre of Sink", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
            ],
        }
        overlay = extraction_service._extract_generic_layout_overlay(section)
        self.assertIn("Burazzo 450mm Gun Metal Single Bowl Sink", overlay["sink_info"])
        self.assertEqual(overlay["tap_info"], "Zara Gun Metal Pull-Out (ZA120-GM) - Centre of Sink")

    def test_extract_generic_inline_property_pairs_splits_joined_property_text(self) -> None:
        pairs = extraction_service._extract_generic_inline_property_pairs(
            "Manufacturer Laminex Range Formwrap Profile Coastal Colour Chalk White"
        )
        self.assertEqual(
            pairs,
            [
                ("manufacturer", "Laminex"),
                ("range", "Formwrap"),
                ("profile", "Coastal"),
                ("colour", "Chalk White"),
            ],
        )

    def test_generic_overlay_handles_joined_property_text_without_label_noise(self) -> None:
        section = {
            "original_section_label": "Kitchen",
            "page_type": "joinery",
            "file_name": "simonds.pdf",
            "page_nos": [8],
            "text": "Kitchen",
            "layout_rows": [
                {
                    "row_label": "Base Cabinet Panels",
                    "value_text": "Manufacturer Laminex Range Formwrap Profile Coastal Colour Chalk White",
                    "row_kind": "material",
                },
                {
                    "row_label": "Cabinetry Handles",
                    "value_text": "Manufacturer Hettich Model Salemi 9113368 Finish Brushed Nickel",
                    "row_kind": "handle",
                },
            ],
        }
        overlay = extraction_service._extract_generic_layout_overlay(section)
        self.assertEqual(overlay["door_colours_base"], "Laminex - Chalk White - Coastal")
        self.assertEqual(overlay["handles"], ["Hettich - Salemi 9113368 - Brushed Nickel"])

    def test_table_group_label_rows_shifts_not_applicable_from_accessories_anchor_to_first_child(self) -> None:
        rows = extraction_service._table_group_label_rows(
            "Accessories\nRobe Hook\nHand Towel Rail\nTowel Rail",
            ["Not Applicable\nGuest Towel Rail\nDouble Towel Rail\nToilet Roll Holder"],
            "sinkware_tapware",
        )
        self.assertEqual(rows[0]["row_label"], "Accessories")
        self.assertEqual(rows[0]["value_region_text"], "")
        self.assertEqual(rows[1]["row_label"], "Robe Hook")
        self.assertEqual(rows[1]["value_region_text"], "Not Applicable")
        self.assertEqual(rows[2]["value_region_text"], "Guest Towel Rail")

    def test_table_group_label_rows_skip_leading_generic_anchor_before_property_values(self) -> None:
        rows = extraction_service._table_group_label_rows(
            "Benchtops\nManufacturer\nColour\nEdge Profile",
            ["Quantum Quartz\nChampagne\n20mm Arissed"],
            "joinery",
        )
        self.assertEqual(rows[0]["row_label"], "Benchtops")
        self.assertEqual(rows[0]["value_region_text"], "")
        self.assertEqual(rows[1]["row_label"], "Manufacturer")
        self.assertEqual(rows[1]["value_region_text"], "Quantum Quartz")
        self.assertEqual(rows[2]["row_label"], "Colour")
        self.assertEqual(rows[2]["value_region_text"], "Champagne")
        self.assertEqual(rows[3]["row_label"], "Edge Profile")
        self.assertEqual(rows[3]["value_region_text"], "20mm Arissed")

    def test_table_group_label_rows_skip_leading_underbench_anchor_before_property_values(self) -> None:
        rows = extraction_service._table_group_label_rows(
            "Underbench including Island\nManufacturer\nColour & Finish\nKickboard\nHandles",
            ["Polytec\nBelgian Oak Matt\nLaminate\n4062-128-TG"],
            "joinery",
        )
        self.assertEqual(rows[0]["row_label"], "Underbench including Island")
        self.assertEqual(rows[0]["value_region_text"], "")
        self.assertEqual(rows[1]["value_region_text"], "Polytec")
        self.assertEqual(rows[2]["value_region_text"], "Belgian Oak Matt")
        self.assertEqual(rows[3]["value_region_text"], "Laminate")
        self.assertEqual(rows[4]["value_region_text"], "4062-128-TG")

    def test_table_group_label_rows_keeps_drawers_value_before_handles_on_evoca_style_rows(self) -> None:
        rows = extraction_service._table_group_label_rows(
            "Underbench\nManufacturer\nColour & Finish\nDrawers\nHandles\nDoor Handle\nDrawer Handle**",
            ["Polytec\nBelgian Oak Matt\nNot Included\n4062-128-TG\nVertical\nHorizontal"],
            "joinery",
        )
        self.assertEqual(rows[0]["row_label"], "Underbench")
        self.assertEqual(rows[0]["value_region_text"], "")
        self.assertEqual(rows[1]["value_region_text"], "Polytec")
        self.assertEqual(rows[2]["value_region_text"], "Belgian Oak Matt")
        self.assertEqual(rows[3]["row_label"], "Drawers")
        self.assertEqual(rows[3]["value_region_text"], "Not Included")
        self.assertEqual(rows[4]["row_label"], "Handles")
        self.assertEqual(rows[4]["value_region_text"], "4062-128-TG")
        self.assertEqual(rows[5]["value_region_text"], "Vertical")
        self.assertEqual(rows[6]["value_region_text"], "Horizontal")

    def test_extract_generic_layout_overlay_groups_compound_prefixed_rows(self) -> None:
        section = {
            "original_section_label": "Laundry",
            "page_type": "joinery",
            "file_name": "evoca.pdf",
            "page_nos": [9],
            "text": "Laundry",
            "layout_rows": [
                {"row_label": "Benchtops Manufacturer", "value_region_text": "Quantum Quartz", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
                {"row_label": "Benchtops Colour", "value_region_text": "Champagne", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
                {"row_label": "Benchtops Edge Profile", "value_region_text": "20mm Arissed", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
                {"row_label": "Underbench Manufacturer", "value_region_text": "Polytec", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
                {"row_label": "Underbench Colour & Finish", "value_region_text": "Belgian Oak Matt", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
                {"row_label": "Underbench Handles", "value_region_text": "4062-128-TG", "supplier_region_text": "", "notes_region_text": "", "row_kind": "handle"},
                {"row_label": "Underbench Door Handle", "value_region_text": "Vertical", "supplier_region_text": "", "notes_region_text": "", "row_kind": "handle"},
                {"row_label": "Underbench Drawer Handle", "value_region_text": "Horizontal", "supplier_region_text": "", "notes_region_text": "", "row_kind": "handle"},
            ],
        }
        overlay = extraction_service._extract_generic_layout_overlay(section)
        self.assertEqual(overlay["bench_tops_wall_run"], "20mm Quantum Quartz - Champagne - Arissed")
        self.assertEqual(overlay["door_colours_base"], "Polytec - Belgian Oak Matt")
        self.assertEqual(overlay["handles"], ["4062-128-TG - Vertical - Horizontal"])

    def test_extract_generic_layout_overlay_builds_island_bench_from_as_above_properties(self) -> None:
        section = {
            "original_section_label": "Kitchen",
            "page_type": "joinery",
            "file_name": "evoca.pdf",
            "page_nos": [8],
            "text": "Kitchen",
            "layout_rows": [
                {"row_label": "Benchtops", "value_region_text": "", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
                {"row_label": "Manufacturer", "value_region_text": "Quantum Quartz", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
                {"row_label": "Colour", "value_region_text": "Champagne", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
                {"row_label": "Island Colour", "value_region_text": "As Above", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
                {"row_label": "Edge Profile", "value_region_text": "20mm Arissed", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
                {"row_label": "Island Edge Profile", "value_region_text": "40mm Arissed", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
            ],
        }
        overlay = extraction_service._extract_generic_layout_overlay(section)
        self.assertEqual(overlay["bench_tops_wall_run"], "20mm Quantum Quartz - Champagne - Arissed")
        self.assertEqual(overlay["bench_tops_island"], "40mm Quantum Quartz - Champagne - Arissed")

    def test_sanitize_generic_fixture_field_dedupes_short_subset_duplicate(self) -> None:
        value = (
            "Alder - Samm - Wall Basin/Bath Mixer Set Backplate - 220mm - Matt Black"
            " | Alder - Samm - Mixer - Matt Black"
        )
        self.assertEqual(
            extraction_service._sanitize_generic_fixture_field(value, kind="tap"),
            "Alder - Samm - Wall Basin/Bath Mixer Set Backplate - 220mm - Matt Black",
        )

    def test_crosscheck_clarendon_snapshot_with_raw_restores_missing_cabinetry_fields(self) -> None:
        layout_snapshot = {
            "site_address": "Lot 8 (#25) Lake Serenity Boulevard Helensvale",
            "rooms": [
                {
                    "room_key": "kitchen",
                    "original_room_label": "KITCHEN",
                    "sink_info": "Abey Abey Large Single Bowl Undermount CUA720 Undermount Sink",
                    "tap_info": "Caroma Liano Ii Flag Pull Out Sink 96381BN56A Brushed Nickel",
                }
            ],
        }
        raw_snapshot = {
            "site_address": "Lot 8 (#25) Lake Serenity Boulevard Helensvale",
            "rooms": [
                {
                    "room_key": "kitchen",
                    "original_room_label": "KITCHEN",
                    "bench_tops_wall_run": "Quantum Zero White Swirl - 20MM Pencil Round Edge",
                    "bench_tops_island": "Quantum Zero White Swirl - 40MM Mitred Apron Edge",
                    "door_colours_base": "Polytec Aston White Smooth Finish Thermolaminate - Hampton EM9 Profile",
                    "toe_kick": ["Matching Melamine finish"],
                    "bulkheads": ["Bulkhead shadowline as matching Melamine finish"],
                    "handles": [
                        "Hettich - Belluno 9995772 200MM Long Brushed Stainless Steel Look",
                        "Hettich - Salemi 9113368 30MM Brushed Stainless Steel Look",
                    ],
                }
            ],
        }
        merged = extraction_service._crosscheck_clarendon_snapshot_with_raw(layout_snapshot, raw_snapshot)
        room = merged["rooms"][0]
        self.assertEqual(room["bench_tops_wall_run"], "Quantum Zero White Swirl - 20MM Pencil Round Edge")
        self.assertEqual(room["door_colours_base"], "Polytec Aston White Smooth Finish Thermolaminate - Hampton EM9 Profile")
        self.assertEqual(room["toe_kick"], ["Matching Melamine finish"])
        self.assertEqual(len(room["handles"]), 2)

    def test_extract_clarendon_fixture_overlays_recovers_laundry_supplier_and_detail(self) -> None:
        text = (
            "LAUNDRY SUPPLIER DESCRIPTION DESIGN COMMENTS "
            "Drop in Tub: EVERHARD "
            "Tap Style: METHVEN "
            "DESCRIPTION DETAILS "
            "PRISM LARGE SINGLE BOWL UNDERMOUNT PPR10LGB "
            "GASTON PULL DOWN MIXER BLACK / CHROME (02-1055) "
            "15MM CP QUARTER TURN WASHING MACHINE COCK (60822)"
        )
        overlays = extraction_service._extract_clarendon_fixture_overlays(text)
        laundry = overlays["laundry"]
        self.assertIn("Everhard", laundry["sink_info"])
        self.assertIn("Prism Large Single Bowl Undermount PPR10LGB", laundry["sink_info"])
        self.assertIn("gaston pull down mixer black / chrome (02-1055)", laundry["tap_info"].lower())
        self.assertIn("15mm cp quarter turn washing machine cock (60822)", laundry["tap_info"].lower())

    def test_extract_clarendon_fixture_overlays_recovers_arlo_laundry_sink_mixer_without_parentheses(self) -> None:
        text = (
            "LAUNDRY SUPPLIER DESCRIPTION DESIGN COMMENTS "
            "Drop in Tub: EVERHARD "
            "Tap Style: PHOENIX "
            "EVERHARD INDUSTRIES CLASSIC 45L UTILITY SINK (71245) "
            "ARLO SINK MIXER GOOSENECK 200MM_ 151-7310-40-1 BN "
            "15MM CP QUARTER TURN WASHING MACHINE COCK (60822)"
        )
        overlays = extraction_service._extract_clarendon_fixture_overlays(text)
        laundry = overlays["laundry"]
        self.assertIn("Everhard", laundry["sink_info"])
        self.assertIn("arlo sink mixer gooseneck 200mm 151-7310-40-1 bn", laundry["tap_info"].lower())
        self.assertIn("15mm cp quarter turn washing machine cock (60822)", laundry["tap_info"].lower())

    def test_select_clarendon_room_overlay_maps_butlers_pantry_to_walk_in_pantry(self) -> None:
        overlay = extraction_service._select_clarendon_room_overlay(
            {"room_key": "walk_in_pantry", "original_room_label": "WALK IN PANTRY"},
            {
                "butlers_pantry": {
                    **extraction_service._blank_clarendon_overlay(),
                    "door_colours_base": "Polytec Aston White Smooth Finish Thermolaminate - Hampton EM9 Profile",
                    "handles": ["Hettich - Belluno 9995772 200MM Long Brushed Stainless Steel Look"],
                }
            },
        )
        self.assertEqual(overlay["door_colours_base"], "Polytec Aston White Smooth Finish Thermolaminate - Hampton EM9 Profile")
        self.assertEqual(overlay["handles"], ["Hettich - Belluno 9995772 200MM Long Brushed Stainless Steel Look"])

    def test_select_clarendon_room_overlay_applies_vanities_overlay_to_ensuite(self) -> None:
        overlay = extraction_service._select_clarendon_room_overlay(
            {"room_key": "ensuite", "original_room_label": "Ensuite"},
            {
                "vanities": {
                    **extraction_service._blank_clarendon_overlay(),
                    "door_colours_base": "Polytec Aston White Smooth Finish Thermolaminate - Hampton EM9 Profile",
                    "handles": ["Hettich - Salemi 9113368 30MM Brushed Stainless Steel Look"],
                    "basin_info": "Eden Bench Mount Gloss White",
                    "tap_info": "Spin Gun Metal Tall Basin Mixer",
                }
            },
        )
        self.assertEqual(overlay["door_colours_base"], "Polytec Aston White Smooth Finish Thermolaminate - Hampton EM9 Profile")
        self.assertEqual(overlay["handles"], ["Hettich - Salemi 9113368 30MM Brushed Stainless Steel Look"])
        self.assertEqual(overlay["basin_info"], "Eden Bench Mount Gloss White")
        self.assertEqual(overlay["tap_info"], "Spin Gun Metal Tall Basin Mixer")

    def test_clarendon_schedule_room_key_maps_rumpus_room_explicitly(self) -> None:
        self.assertEqual(
            extraction_service._clarendon_schedule_room_key("RUMPUS ROOM COLOUR SCHEDULE"),
            "rumpus_room",
        )

    def test_clarendon_clean_benchtop_text_drops_shadowline_note(self) -> None:
        self.assertEqual(
            extraction_service._clarendon_clean_benchtop_text("Shadowline : Matching Melamine Finish *"),
            "",
        )

    def test_extract_generic_layout_overlay_uses_additional_section_location_as_room_label(self) -> None:
        section = {
            "original_section_label": "Additional Kitchen/Butlers/Kitchenette",
            "page_type": "joinery",
            "file_name": "simonds.pdf",
            "page_nos": [15],
            "text": "Additional Kitchen/Butlers/Kitchenette",
            "layout_rows": [
                {"row_label": "Location", "value_region_text": "Rumpus", "supplier_region_text": "", "notes_region_text": "", "row_kind": "other"},
                {"row_label": "Benchtop", "value_region_text": "", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
                {"row_label": "Manufacturer", "value_region_text": "Caesarstone - Mineral", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
                {"row_label": "Colour", "value_region_text": "Raven", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
            ],
        }
        overlay = extraction_service._extract_generic_layout_overlay(section)
        self.assertEqual(overlay["original_room_label"], "Rumpus")

    def test_extract_generic_layout_overlay_uses_document_location_for_additional_section(self) -> None:
        section = {
            "original_section_label": "Additional Bath/Ensuite/Powder",
            "page_type": "joinery",
            "file_name": "simonds.pdf",
            "page_nos": [13],
            "text": "Additional Bath/Ensuite/Powder",
            "layout_rows": [
                {"row_label": "Benchtop", "value_region_text": "", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
                {"row_label": "Manufacturer", "value_region_text": "Caesarstone - Mineral", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
            ],
        }
        documents = [
            {
                "file_name": "simonds.pdf",
                "pages": [
                    {
                        "page_no": 13,
                        "raw_text": (
                            "Additional Bath/Ensuite/Powder Additional Bath/Ensuite/Powder"
                            "LocationGuest Ensuite 2LocationManufacturerCaesarstone - Mineral"
                        ),
                    }
                ],
            }
        ]
        overlay = extraction_service._extract_generic_layout_overlay(section, documents=documents)
        self.assertEqual(overlay["original_room_label"], "Guest Ensuite 2")

    def test_extract_generic_layout_overlay_uses_document_location_for_additional_wet_area_section(self) -> None:
        section = {
            "original_section_label": "Additional Wet Area",
            "page_type": "joinery",
            "file_name": "simonds.pdf",
            "page_nos": [12],
            "text": "Additional Wet Area",
            "layout_rows": [
                {"row_label": "Benchtop", "value_region_text": "", "supplier_region_text": "", "notes_region_text": "", "row_kind": "material"},
            ],
        }
        documents = [
            {
                "file_name": "simonds.pdf",
                "pages": [
                    {
                        "page_no": 12,
                        "raw_text": (
                            "Additional Wet AreaLocationEnsuite 3Wet Area Location"
                            "ManufacturerCaesarstone - Mineral"
                        ),
                    }
                ],
            }
        ]
        overlay = extraction_service._extract_generic_layout_overlay(section, documents=documents)
        self.assertEqual(overlay["original_room_label"], "Ensuite 3")

    def test_resolve_generic_room_label_from_documents_handles_vanity_proxy_room(self) -> None:
        row = {
            "original_room_label": "Vanity",
            "source_file": "simonds.pdf",
            "page_refs": "12",
        }
        documents = [
            {
                "file_name": "simonds.pdf",
                "pages": [
                    {
                        "page_no": 12,
                        "raw_text": "Additional Wet AreaLocationEnsuite 3Wet Area LocationManufacturerCaesarstone - Mineral",
                    }
                ],
            }
        ]
        self.assertEqual(
            extraction_service._resolve_generic_room_label_from_documents(row, documents),
            "Ensuite 3",
        )

    def test_sanitize_generic_material_field_trims_bench_prefix_from_tall_colour(self) -> None:
        self.assertEqual(
            extraction_service._sanitize_generic_material_field(
                "40mm Caesarstone - Mineral - Organic White - Arris - Laminex Natural Blackened Legno",
                field_name="door_colours_tall",
                room_key="kitchen",
            ),
            "Laminex Natural Blackened Legno",
        )

    def test_sanitize_generic_material_field_drops_not_applicable_prefix_before_real_material(self) -> None:
        self.assertEqual(
            extraction_service._sanitize_generic_material_field(
                "Not Applicable Polytec - Belgian Oak Matt",
                field_name="door_colours_overheads",
                room_key="kitchen",
            ),
            "Polytec - Belgian Oak Matt",
        )

    def test_sanitize_generic_material_field_preserves_as_above_for_colour_fields(self) -> None:
        self.assertEqual(
            extraction_service._sanitize_generic_material_field(
                "As Above",
                field_name="door_colours_island",
                room_key="kitchen",
            ),
            "As Above",
        )

    def test_official_lookup_skips_appliance_placeholder_models(self) -> None:
        self.assertFalse(
            appliance_official._should_lookup(
                {
                    "appliance_type": "Cooktop",
                    "make": "",
                    "model_no": "as above BY CLIENT",
                }
            )
        )
        self.assertFalse(
            appliance_official._should_lookup(
                {
                    "appliance_type": "Microwave",
                    "make": "",
                    "model_no": "LEAVE STANDARD SPACE BY CLIENT",
                }
            )
        )

    def test_sanitize_generic_handle_entries_drops_not_included_noise_and_has_placeholder(self) -> None:
        self.assertEqual(
            extraction_service._sanitize_generic_handle_entries(
                [
                    "4062-128-TG - Vertical - Horizontal - shaving cabinets Not Included",
                    "has",
                ]
            ),
            ["4062-128-TG - Vertical - Horizontal"],
        )

    def test_sanitize_generic_fixture_field_drops_label_only_tap_fragments(self) -> None:
        self.assertEqual(
            extraction_service._sanitize_generic_fixture_field("Type Location Centre of", kind="tap"),
            "",
        )

    def test_prefer_imperial_raw_list_uses_split_raw_handles_over_compound_layout_entry(self) -> None:
        selected = extraction_service._prefer_imperial_raw_list(
            "handles",
            [
                "BASE - Kethy | UPPER - S225-480-MBK | TALL - S225-280-MBK | PTO designer",
            ],
            [
                "BASE - Kethy",
                "UPPER - S225-480-MBK",
                "TALL - S225-280-MBK",
            ],
        )
        self.assertEqual(
            selected,
            [
                "BASE - Kethy",
                "UPPER - S225-480-MBK",
                "TALL - S225-280-MBK",
            ],
        )

    def test_prefer_imperial_raw_list_uses_raw_toe_kick_when_layout_is_contaminated(self) -> None:
        selected = extraction_service._prefer_imperial_raw_list(
            "toe_kick",
            ["AS DOORS Polytec | MIRRORED SHAVING CABINET | External panels only"],
            ["AS DOORS Polytec"],
        )
        self.assertEqual(selected, ["AS DOORS Polytec"])

    def test_clean_generic_fragment_strips_additional_section_and_feature_batten_noise(self) -> None:
        self.assertEqual(
            extraction_service._clean_generic_fragment(
                "Additional Kitchen/Butlers/Kitchenette Dwarf Wall Capping finish to Rumpus Feature Battens"
            ),
            "",
        )

    def test_clean_generic_fragment_strips_duplicate_location_pattern(self) -> None:
        self.assertEqual(
            extraction_service._clean_generic_fragment("Location Centre of Sink Location"),
            "",
        )

    def test_polish_generic_layout_room_clears_duplicate_other_bench_for_non_kitchen_room(self) -> None:
        row = {
            "room_key": "study",
            "original_room_label": "Study",
            "bench_tops_wall_run": "40mm stone - Arissed - By Builder",
            "bench_tops_island": "33mm laminate - Square Edge",
            "bench_tops_other": "40mm stone - Arissed - By Builder | 33mm laminate - Square Edge",
            "bench_tops": [],
            "door_colours_base": "",
            "door_colours_overheads": "",
            "door_colours_tall": "",
            "door_colours_island": "",
            "door_colours_bar_back": "",
            "floating_shelf": "",
            "toe_kick": [],
            "bulkheads": [],
            "handles": [],
            "accessories": [],
            "other_items": [],
            "sink_info": "",
            "tap_info": "",
            "basin_info": "",
            "flooring": "",
        }
        overlay = {}
        polished = extraction_service._polish_generic_layout_room(row, overlay)
        self.assertEqual(polished["bench_tops_other"], "")

    def test_sanitize_generic_fixture_field_strips_accessory_tail_from_tap(self) -> None:
        self.assertEqual(
            extraction_service._sanitize_generic_fixture_field(
                "Alder - Maxx - Rectangle Sink Mixer - Matt Black + ALDER SACHI ROBE HOOK IN MATT BLACK",
                kind="tap",
            ),
            "Alder - Maxx - Rectangle Sink Mixer - Matt Black",
        )

    def test_sanitize_generic_fixture_field_drops_non_tap_secondary_fragment(self) -> None:
        self.assertEqual(
            extraction_service._sanitize_generic_fixture_field(
                "Alder - Samm - Wall Basin/Bath Mixer Set Backplate - 220mm - Matt Black | Alder - Fresco (WELS 6 stars)",
                kind="tap",
            ),
            "Alder - Samm - Wall Basin/Bath Mixer Set Backplate - 220mm - Matt Black",
        )

    def test_clean_fixture_text_strips_leading_tap_prefix(self) -> None:
        self.assertEqual(
            parsing_module._clean_fixture_text("Tap Franke Eos Neo pull out tap copper TA9601CP"),
            "Franke Eos Neo pull out tap copper TA9601CP",
        )

    def test_sanitize_generic_handle_entries_drops_electrical_section_pollution(self) -> None:
        self.assertEqual(
            extraction_service._sanitize_generic_handle_entries(
                [
                    "Switch Plates / GPO's Clipsal Iconic - White - TV Antenna Included",
                    "4062-128-TG - Vertical - Horizontal",
                ]
            ),
            ["4062-128-TG - Vertical - Horizontal"],
        )

    def test_polish_generic_layout_room_preserves_island_bench_as_above(self) -> None:
        room = {
            "original_room_label": "Kitchen",
            "room_key": "kitchen",
            "bench_tops_wall_run": "",
            "bench_tops_island": "",
            "bench_tops_other": "",
            "bench_tops": [],
            "toe_kick": [],
            "handles": [],
            "accessories": [],
        }
        overlay = {
            "bench_tops_wall_run": "20mm Quantum Quartz - Champagne - Arissed",
            "bench_tops_island": "As Above",
            "bench_tops_other": "",
            "has_bench_block": True,
        }
        polished = extraction_service._polish_generic_layout_room(room, overlay)
        self.assertEqual(polished["bench_tops_wall_run"], "20mm Quantum Quartz - Champagne - Arissed")
        self.assertEqual(polished["bench_tops_island"], "As Above")

    def test_format_generic_flooring_from_parts_preserves_original_wording(self) -> None:
        self.assertEqual(
            extraction_service._format_generic_flooring_from_parts(
                {
                    "note": ["Tiles with Floating TBC"],
                    "_ordered_fragments": ["Tiles with Floating TBC"],
                }
            ),
            "Tiles with Floating TBC",
        )

    def test_format_generic_material_from_parts_preserves_ordered_original_fragments(self) -> None:
        self.assertEqual(
            extraction_service._format_generic_material_from_parts(
                {
                    "manufacturer": ["Polytec"],
                    "colour": ["Aston White Smooth Finish Thermolaminate"],
                    "profile": ["Hampton EM9 Profile"],
                    "_ordered_fragments": [
                        "Polytec",
                        "Aston White Smooth Finish Thermolaminate",
                        "Hampton EM9 Profile",
                    ],
                }
            ),
            "Polytec - Aston White Smooth Finish Thermolaminate - Hampton EM9 Profile",
        )

    def test_sanitize_generic_material_field_preserves_not_applicable_when_it_is_the_value(self) -> None:
        self.assertEqual(
            extraction_service._sanitize_generic_material_field(
                "Not Applicable",
                field_name="flooring",
                room_key="laundry",
            ),
            "Not Applicable",
        )

    def test_strip_generic_anchor_tail_removes_trailing_cabinet_label_noise(self) -> None:
        self.assertEqual(
            extraction_service._strip_generic_anchor_tail("Chalk White Base Cabinet Panels"),
            "Chalk White",
        )
        self.assertEqual(
            extraction_service._strip_generic_anchor_tail("Blackened LegnoOverheads"),
            "Blackened Legno",
        )

    def test_sink_tap_blocks_split_shower_rail_and_screen_into_separate_anchors(self) -> None:
        rows = [
            {"row_label": "Basin Mixer", "value_region_text": "", "row_kind": "tap"},
            {"row_label": "Type", "value_region_text": "Spin Gun Metal In-wall Mixer (SP141-GM)", "row_kind": "material"},
            {"row_label": "Shower Rail / Rose", "value_region_text": "Omega Integrated Gun Metal Shower System", "row_kind": "material"},
            {"row_label": "Shower Screen", "value_region_text": "Semi-frameless with Clear Toughened Glass", "row_kind": "material"},
            {"row_label": "Shower Screen Colour", "value_region_text": "Gunmetal", "row_kind": "material"},
        ]
        blocks = extraction_service._build_generic_layout_blocks(rows, page_type="sinkware_tapware")
        labels = [block["anchor_label"] for block in blocks]
        self.assertIn("Basin Mixer", labels)
        self.assertIn("Shower Rail / Rose", labels)
        self.assertIn("Shower Screen", labels)

    def test_generic_sinkware_overlay_keeps_basin_mixer_separate_from_shower_rows(self) -> None:
        section = {
            "original_section_label": "Ensuite",
            "page_type": "sinkware_tapware",
            "file_name": "wet-area.pdf",
            "page_nos": [13],
            "text": "Ensuite wet area",
            "layout_rows": [
                {"row_label": "Basin", "value_region_text": "", "row_kind": "basin"},
                {"row_label": "Model", "value_region_text": "Eden Bench Mount Gloss White (FL135-W)", "row_kind": "material"},
                {"row_label": "Type", "value_region_text": "Overmount", "row_kind": "material"},
                {"row_label": "Basin Mixer", "value_region_text": "", "row_kind": "basin"},
                {"row_label": "Type", "value_region_text": "Spin Gun Metal Tall Basin Mixer (SP110-GM)", "row_kind": "material"},
                {"row_label": "Location", "value_region_text": "Centre of Basin", "row_kind": "material"},
                {"row_label": "Shower Rail / Rose", "value_region_text": "Omega Integrated Gun Metal Shower System", "row_kind": "material"},
                {"row_label": "Shower Screen", "value_region_text": "Semi-frameless with Clear Toughened Glass", "row_kind": "material"},
            ],
        }
        overlay = extraction_service._extract_generic_layout_overlay(section)
        self.assertIn("Spin Gun Metal Tall Basin Mixer", overlay["tap_info"])
        self.assertNotIn("Omega Integrated Gun Metal Shower System", overlay["tap_info"])
        other_labels = {item.get("label") for item in overlay["other_items"]}
        self.assertIn("Shower Rail / Rose", other_labels)
        self.assertIn("Shower Screen", other_labels)

    def test_generic_sinkware_overlay_keeps_cabinetry_rows_from_same_room(self) -> None:
        section = {
            "original_section_label": "Master Ensuite",
            "page_type": "sinkware_tapware",
            "file_name": "simonds-wet-area.pdf",
            "page_nos": [10],
            "text": "Master Ensuite joinery + wet area",
            "layout_rows": [
                {"row_label": "Benchtop", "value_region_text": "", "row_kind": "material"},
                {"row_label": "Manufacturer", "value_region_text": "Caesarstone - Mineral", "row_kind": "material"},
                {"row_label": "Profile", "value_region_text": "20mm Arris", "row_kind": "material"},
                {"row_label": "Range", "value_region_text": "Caesarstone Standard M1", "row_kind": "material"},
                {"row_label": "Colour", "value_region_text": "Organic White", "row_kind": "material"},
                {"row_label": "Base Cabinet Panels", "value_region_text": "", "row_kind": "material"},
                {"row_label": "Manufacturer", "value_region_text": "Laminex", "row_kind": "material"},
                {"row_label": "Finish", "value_region_text": "Natural", "row_kind": "material"},
                {"row_label": "Profile", "value_region_text": "S/Edge", "row_kind": "material"},
                {"row_label": "Colour", "value_region_text": "Blackened Legno", "row_kind": "material"},
                {"row_label": "Kickboard", "value_region_text": "", "row_kind": "material"},
                {"row_label": "Manufacturer", "value_region_text": "Laminex", "row_kind": "material"},
                {"row_label": "Finish", "value_region_text": "Natural", "row_kind": "material"},
                {"row_label": "Colour", "value_region_text": "Blackened Legno", "row_kind": "material"},
                {"row_label": "Cabinetry Handles", "value_region_text": "", "row_kind": "handle"},
                {"row_label": "Model", "value_region_text": "L Shaped Finger Pull", "row_kind": "material"},
            ],
        }
        overlay = extraction_service._extract_generic_layout_overlay(section)
        self.assertEqual(overlay["bench_tops_wall_run"], "20mm Caesarstone - Mineral - Organic White - Arris")
        self.assertEqual(overlay["door_colours_base"], "Laminex - Blackened Legno - S/Edge")
        self.assertEqual(overlay["toe_kick"], ["Laminex - Blackened Legno"])
        self.assertEqual(overlay["handles"], ["L Shaped Finger Pull"])

    def test_generic_accessory_formatter_drops_bare_anchor_without_values(self) -> None:
        formatted = extraction_service._format_generic_accessory_from_parts({}, anchor_label="Robe Hook")
        self.assertEqual(formatted, "")

    def test_placeholder_fixture_detection_catches_shower_label_only_string(self) -> None:
        self.assertTrue(
            extraction_service._looks_like_placeholder_fixture_text(
                "Shower Rail / Rose Shower Screen Shower Screen Colour"
            )
        )

    def _mark_raw_spec_qa_passed(self, job_id: int) -> None:
        verification = store.get_job_snapshot_verification(job_id, "raw_spec")
        self.assertIsNotNone(verification)
        checklist = []
        for item in verification["checklist"]:
            updated = dict(item)
            updated["status"] = "pass"
            updated["pdf_page_ref"] = updated.get("source_page_refs", "") or "1"
            checklist.append(updated)
        saved = store.save_snapshot_verification(int(verification["snapshot_id"]), checklist, checked_by="admin", notes="Automated test pass")
        self.assertIsNotNone(saved)
        self.assertEqual(saved["status"], "passed")

    def _qa_form_payload(self, verification: dict[str, object], csrf: str, item_status: str = "pass") -> dict[str, str]:
        payload: dict[str, str] = {"csrf_token": csrf, "item_count": str(len(verification.get("checklist", []))), "notes": "QA test"}
        for index, item in enumerate(verification.get("checklist", [])):
            row = dict(item)
            payload[f"section_type_{index}"] = str(row.get("section_type", ""))
            payload[f"entity_label_{index}"] = str(row.get("entity_label", ""))
            payload[f"field_name_{index}"] = str(row.get("field_name", ""))
            payload[f"extracted_value_{index}"] = str(row.get("extracted_value", ""))
            payload[f"source_page_refs_{index}"] = str(row.get("source_page_refs", ""))
            payload[f"pdf_page_ref_{index}"] = str(row.get("source_page_refs", "") or "1")
            payload[f"status_{index}"] = item_status
            payload[f"qa_note_{index}"] = ""
        return payload

    def _login(self, client: TestClient) -> str:
        login_page = client.get("/login")
        csrf = login_page.text.split('name="csrf_token" value="', 1)[1].split('"', 1)[0]
        response = client.post(
            "/login",
            data={"username": "admin", "password": "admin", "csrf_token": csrf},
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 303)
        return csrf


if __name__ == "__main__":
    unittest.main()
