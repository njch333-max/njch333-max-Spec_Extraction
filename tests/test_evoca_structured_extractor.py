from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from App.services import evoca_structured_extractor


def _page(page_no: int, rows: list[list[object]]) -> dict[str, object]:
    return {
        "page_no": page_no,
        "raw_text": " ".join(str(cell) for row in rows for cell in row if cell),
        "text": "",
        "table_rows": [rows],
    }


def _raw_line(top: float, words: list[tuple[str, float]]) -> dict[str, object]:
    return {
        "top": top,
        "bottom": top + 9,
        "center": top + 4.5,
        "text": " ".join(word for word, _ in words),
        "words": [
            {
                "text": word,
                "x0": x0,
                "x1": x0 + max(len(word), 1) * 5,
                "top": top,
                "bottom": top + 9,
                "center": top + 4.5,
            }
            for word, x0 in words
        ],
    }


def test_evoca_structured_extracts_section_room_group_rows() -> None:
    structured = evoca_structured_extractor.extract_evoca_pages(
        [
            _page(
                1,
                [
                    ["15 CABINETS", None, None, ""],
                    ["", "Kitchen", None, None],
                    ["-", "Benchtops\nManufacturer\nColour", "Quantum Quartz\nChampagne", None],
                ],
            )
        ],
        source_pdf="evoca.pdf",
    )

    section = structured["sections"][0]
    assert section["section_title"] == "15 CABINETS"
    room = section["rooms"][0]
    assert room["room_label"] == "Kitchen"
    group = room["groups"][0]
    assert group["group_label"] == "Benchtops"
    assert [(row["label"], row["value"]) for row in group["rows"]] == [
        ("Manufacturer", "Quantum Quartz"),
        ("Colour", "Champagne"),
    ]


def test_evoca_structured_splits_inline_child_label_values() -> None:
    structured = evoca_structured_extractor.extract_evoca_pages(
        [
            _page(
                12,
                [
                    ["20 PLUMBING FIXTURES & TAPWARE", None, None, ""],
                    ["", "Kitchen", None, None],
                    [
                        "-",
                        "Sink\n"
                        "Model Burazzo 750mm Stainless Steel Double Bowl Sink (BU754522D) ($185)\n"
                        "Type Undermount\n"
                        "Accessories Not Applicable",
                        "",
                        None,
                    ],
                    ["", "", "Burazzo 750mm Stainless Steel Double Bowl Sink (BU754522D) ($185)", None],
                ],
            )
        ],
        source_pdf="evoca.pdf",
    )

    group = structured["sections"][0]["rooms"][0]["groups"][0]
    assert [(row["label"], row["value"]) for row in group["rows"]] == [
        ("Model", "Burazzo 750mm Stainless Steel Double Bowl Sink (BU754522D) ($185)"),
        ("Type", "Undermount"),
        ("Accessories", "Not Applicable"),
    ]


def test_evoca_structured_preserves_not_applicable_rows() -> None:
    structured = evoca_structured_extractor.extract_evoca_pages(
        [
            _page(
                1,
                [
                    ["15 CABINETS", None, None, ""],
                    ["", "Butlers", None, None],
                    ["-", "Underbench\nManufacturer\nColour & Finish\nHandles", "Not Applicable", None],
                ],
            )
        ],
        source_pdf="evoca.pdf",
    )

    rows = structured["sections"][0]["rooms"][0]["groups"][0]["rows"]
    assert any(row["value"] == "Not Applicable" for row in rows)


def test_evoca_structured_promotes_extended_terminal_group_value() -> None:
    structured = evoca_structured_extractor.extract_evoca_pages(
        [
            _page(
                1,
                [
                    ["15 CABINETS", None, None, ""],
                    ["", "Kitchen", None, None],
                    [
                        "-",
                        "Benchtops\nManufacturer\nColour\nIsland Colour\nEdge Profile",
                        "Not Applicable - by owner after handover",
                        None,
                    ],
                ],
            )
        ],
        source_pdf="evoca.pdf",
    )

    rows = structured["sections"][0]["rooms"][0]["groups"][0]["rows"]
    assert [(row["label"], row["value"], bool(row.get("is_group_anchor"))) for row in rows] == [
        ("Benchtops", "Not Applicable - by owner after handover", True),
        ("Manufacturer", "", False),
        ("Colour", "", False),
        ("Island Colour", "", False),
        ("Edge Profile", "", False),
    ]


def test_evoca_structured_terminal_group_value_blocks_shifted_child_values() -> None:
    structured = evoca_structured_extractor.extract_evoca_pages(
        [
            _page(
                9,
                [
                    ["15 CABINETS", None, None, ""],
                    ["", "Bathroom", None, None],
                    ["-", "Benchtops\nManufacturer\nColour\nEdge Profile", "Not Applicable", None],
                    [None, None, "Polytec", None],
                    [None, None, "Taupe", None],
                ],
            )
        ],
        source_pdf="evoca.pdf",
    )

    rows = structured["sections"][0]["rooms"][0]["groups"][0]["rows"]
    assert [(row["label"], row["value"], bool(row.get("is_group_anchor"))) for row in rows] == [
        ("Benchtops", "Not Applicable", True),
        ("Manufacturer", "", False),
        ("Colour", "", False),
        ("Edge Profile", "", False),
    ]


def test_evoca_structured_promotes_client_supply_group_terminal_value() -> None:
    structured = evoca_structured_extractor.extract_evoca_pages(
        [
            _page(
                1,
                [
                    ["25 CARPET", None, None, ""],
                    ["-", "Carpets\nType\nColour\nUnderlay", "Client to supply & install after handover", None],
                ],
            )
        ],
        source_pdf="evoca.pdf",
    )

    rows = structured["sections"][0]["groups"][0]["rows"]
    assert [(row["label"], row["value"], bool(row.get("is_group_anchor"))) for row in rows] == [
        ("Carpets", "Client to supply & install after handover", True),
        ("Type", "", False),
        ("Colour", "", False),
        ("Underlay", "", False),
    ]


def test_evoca_structured_merges_multiline_anchor_value_without_child_labels() -> None:
    structured = evoca_structured_extractor.extract_evoca_pages(
        [
            _page(
                12,
                [
                    ["17 APPLIANCES, ACCESSORIES & HOT WATER UNIT", None, None, ""],
                    [
                        "-",
                        "Hot Water Unit",
                        (
                            "Ariston Primos 280 Litre\n"
                            "Heat Pump (Electric)\n"
                            "With tempering valve"
                        ),
                        None,
                    ],
                ],
            )
        ],
        source_pdf="evoca.pdf",
    )

    rows = structured["sections"][0]["groups"][0]["rows"]
    assert [(row["label"], row["value"], bool(row.get("is_group_anchor"))) for row in rows] == [
        (
            "Hot Water Unit",
            "Ariston Primos 280 Litre Heat Pump (Electric) With tempering valve",
            True,
        )
    ]
    assert all(row["label"] != "Unassigned Source Text" for row in rows)


def test_evoca_structured_skips_non_required_output_sections_without_leakage() -> None:
    structured = evoca_structured_extractor.extract_evoca_pages(
        [
            _page(
                1,
                [
                    ["16 ELECTRICAL / ALARM SYSTEM / CCTV / SOLAR PV SYSTEM", None, None, ""],
                    ["-", "Alarm System", "Included", None],
                    ["17 APPLIANCES, ACCESSORIES & HOT WATER UNIT", None, None, ""],
                    ["-", "Hot Water Unit", "Ariston", None],
                    ["18 AIR-CONDITIONING", None, None, ""],
                    ["-", "Ducted Reverse Cycle", "Included", None],
                ],
            ),
            _page(
                2,
                [
                    ["19 PLUMBING & GAS", None, None, ""],
                    ["-", "Gas Type", "Natural Gas", None],
                    ["20 PLUMBING FIXTURES & TAPWARE", None, None, ""],
                    ["", "Kitchen", None, None],
                    ["-", "Sink\nModel", "Burazzo", None],
                    ["21 MIRRORS", None, None, ""],
                    ["", "Bathroom", None, None],
                    ["-", "Mirrors\nType", "Polished edge", None],
                ],
            ),
            _page(
                3,
                [
                    ["22 WINDOW FURNISHINGS", None, None, ""],
                    ["-", "Vertical / Roller Blinds\nType", "Roller", None],
                    ["23 TILING / HARD FLOORING", None, None, ""],
                    ["-", "Main Floor Tile\nType", "Ceramic", None],
                ],
            ),
        ],
        source_pdf="evoca.pdf",
    )

    assert [section["section_code"] for section in structured["sections"]] == ["17", "20", "23"]
    output_json = json.dumps(structured["sections"])
    for excluded_text in (
        "Alarm System",
        "Ducted Reverse Cycle",
        "Gas Type",
        "Mirrors",
        "Vertical / Roller Blinds",
    ):
        assert excluded_text not in output_json
    skipped = [title for page in structured["pages"] for title in page["sections_skipped"]]
    assert skipped == [
        "16 ELECTRICAL / ALARM SYSTEM / CCTV / SOLAR PV SYSTEM",
        "18 AIR-CONDITIONING",
        "19 PLUMBING & GAS",
        "21 MIRRORS",
        "22 WINDOW FURNISHINGS",
    ]


def test_evoca_structured_carries_section_across_pages() -> None:
    structured = evoca_structured_extractor.extract_evoca_pages(
        [
            _page(
                1,
                [
                    ["15 CABINETS", None, None, ""],
                    ["", "Kitchen", None, None],
                    ["-", "Benchtops\nManufacturer", "Quantum Quartz", None],
                ],
            ),
            _page(
                2,
                [
                    ["", "Laundry", None, None],
                    ["-", "Benchtops\nManufacturer", "Quantum Quartz", None],
                ],
            ),
        ],
        source_pdf="evoca.pdf",
    )

    assert len(structured["sections"]) == 1
    section = structured["sections"][0]
    assert section["page_start"] == 1
    assert section["page_end"] == 2
    assert [room["room_label"] for room in section["rooms"]] == ["Kitchen", "Laundry"]


def test_evoca_structured_detects_section_level_appliances() -> None:
    structured = evoca_structured_extractor.extract_evoca_pages(
        [
            _page(
                1,
                [
                    ["17 APPLIANCES, ACCESSORIES & HOT WATER UNIT", None, None, ""],
                    ["-", "Appliances\nFreestanding Cooker\nHot Plate", "Not Applicable", None],
                    [None, None, "Fisher & Paykel 900mm Induction Cooktop", None],
                ],
            )
        ],
        source_pdf="evoca.pdf",
    )

    section = structured["sections"][0]
    assert section["section_title"] == "17 APPLIANCES, ACCESSORIES & HOT WATER UNIT"
    assert section["rooms"] == []
    group = section["groups"][0]
    assert group["group_label"] == "Appliances"
    assert [row["label"] for row in group["rows"]] == ["Freestanding Cooker", "Hot Plate"]


def test_evoca_structured_workbook_exports_expected_columns(tmp_path: Path) -> None:
    structured = evoca_structured_extractor.extract_evoca_pages(
        [
            _page(
                1,
                [
                    ["20 PLUMBING FIXTURES & TAPWARE", None, None, ""],
                    ["", "Kitchen", None, None],
                    ["-", "Sink\nModel\nType", "", None],
                    [None, None, "Burazzo Sink", None],
                    [None, None, "Top Mount", None],
                ],
            )
        ],
        source_pdf="evoca.pdf",
    )
    output = tmp_path / "evoca.xlsx"
    evoca_structured_extractor.write_structured_workbook(structured, output)

    workbook = load_workbook(output)
    assert workbook.sheetnames[0] == "_summary"
    assert "20_PLUMBING" in workbook.sheetnames
    sheet = workbook["20_PLUMBING"]
    headers = [cell.value for cell in sheet[1]]
    assert headers == ["Page", "Order", "Room", "Group", "Label", "Value", "Anchor", "Source Text"]
    values = [row[5].value for row in sheet.iter_rows(min_row=2)]
    assert "Burazzo Sink" in values


def test_evoca_structured_promotes_group_anchor_value_before_child_rows() -> None:
    structured = evoca_structured_extractor.extract_evoca_pages(
        [
            _page(
                1,
                [
                    ["15 CABINETS", None, None, ""],
                    ["", "Kitchen", None, None],
                    [
                        "",
                        "Overhead Cupboards\nManufacturer\nColour & Finish\nHandles",
                        "Push to Open Above Oven\nPolytec",
                        None,
                    ],
                    [None, None, "Rojo Walnut Woodmatt", None],
                    [None, None, "Finger Grip", None],
                ],
            )
        ],
        source_pdf="evoca.pdf",
    )

    evoca_structured_extractor._rescue_missing_values(
        structured,
        {
            1: {
                "overhead cupboards": ["Push to Open Above Oven"],
                "manufacturer": ["Polytec"],
                "colour & finish": ["Rojo Walnut Woodmatt"],
                "handles": ["Finger Grip"],
            }
        },
    )

    group = structured["sections"][0]["rooms"][0]["groups"][0]
    assert [(row["label"], row["value"], bool(row.get("is_group_anchor"))) for row in group["rows"]] == [
        ("Overhead Cupboards", "Push to Open Above Oven", True),
        ("Manufacturer", "Polytec", False),
        ("Colour & Finish", "Rojo Walnut Woodmatt", False),
        ("Handles", "Finger Grip", False),
    ]
    assert structured["diagnostics"]["anchor_value_groups"] == 1
    assert structured["diagnostics"]["anchor_value_child_realignments"] == 3


def test_evoca_structured_keeps_room_note_before_next_group() -> None:
    structured = evoca_structured_extractor.extract_evoca_pages(
        [
            _page(
                1,
                [
                    ["15 CABINETS", None, None, ""],
                    ["", "Kitchen", None, None],
                    ["", "No shelf to cupboard underneath sink", None, None],
                    [
                        "-",
                        "Benchtops\nManufacturer\nColour\nIsland Colour\nEdge Profile",
                        "Quantum Quartz\nStatuario Zero\nAs Above\n20mm Arissed",
                        None,
                    ],
                ],
            )
        ],
        source_pdf="evoca.pdf",
    )

    room = structured["sections"][0]["rooms"][0]
    assert [(note["label"], note["value"]) for note in room["notes"]] == [
        ("Note", "No shelf to cupboard underneath sink")
    ]
    assert [group["group_label"] for group in room["groups"]] == ["Benchtops"]
    rows = room["groups"][0]["rows"]
    assert [(row["label"], row["value"]) for row in rows[:2]] == [
        ("Manufacturer", "Quantum Quartz"),
        ("Colour", "Statuario Zero"),
    ]


def test_evoca_structured_rescue_is_bounded_to_matching_group_block() -> None:
    structured = evoca_structured_extractor.extract_evoca_pages(
        [
            _page(
                1,
                [
                    ["20 PLUMBING FIXTURES & TAPWARE", None, None, ""],
                    ["", "Bathroom", None, None],
                    ["-", "Basin\nModel\nType", "", None],
                    [None, None, "Eden Bench Mount Gloss White (FL135-W)", None],
                    [None, None, "Overmount", None],
                    ["-", "Basin Mixer\nType\nLocation", "", None],
                    ["", "Ensuite", None, None],
                    ["-", "Basin\nModel\nType", "", None],
                    [None, None, "Eden Bench Mount Gloss White (FL135-W)", None],
                    [None, None, "Overmount", None],
                ],
            )
        ],
        source_pdf="evoca.pdf",
    )
    entry_key = evoca_structured_extractor.LOOKUP_ENTRIES_KEY
    evoca_structured_extractor._rescue_missing_values(
        structured,
        {
            1: {
                entry_key: [
                    {"key": "bathroom", "label": "Bathroom", "value": ""},
                    {"key": "basin", "label": "Basin", "value": ""},
                    {"key": "model", "label": "Model", "value": "Eden Bench Mount Gloss White (FL135-W)"},
                    {"key": "type", "label": "Type", "value": "Overmount"},
                    {"key": "basin mixer", "label": "Basin Mixer", "value": ""},
                    {"key": "type", "label": "Type", "value": ""},
                    {"key": "location", "label": "Location", "value": "Centre of Basin"},
                    {"key": "ensuite", "label": "Ensuite", "value": ""},
                    {"key": "basin", "label": "Basin", "value": ""},
                    {"key": "model", "label": "Model", "value": "Eden Bench Mount Gloss White (FL135-W)"},
                    {"key": "type", "label": "Type", "value": "Overmount"},
                ],
                "model": ["Eden Bench Mount Gloss White (FL135-W)", "Eden Bench Mount Gloss White (FL135-W)"],
                "type": ["Overmount", "Overmount"],
                "location": ["Centre of Basin"],
            }
        },
    )

    bathroom = structured["sections"][0]["rooms"][0]
    basin_mixer = bathroom["groups"][1]
    assert basin_mixer["group_label"] == "Basin Mixer"
    assert [(row["label"], row["value"]) for row in basin_mixer["rows"]] == [
        ("Type", ""),
        ("Location", "Centre of Basin"),
    ]


def test_evoca_structured_raw_text_fallback_fills_group_bounded_missing_pair() -> None:
    structured = evoca_structured_extractor.extract_evoca_pages(
        [
            _page(
                1,
                [
                    ["15 CABINETS", None, None, ""],
                    ["", "Kitchen", None, None],
                    ["", "No shelf to cupboard underneath sink", None, None],
                    [
                        "-",
                        "Benchtops\nManufacturer\nColour\nIsland Colour\nEdge Profile",
                        "",
                        None,
                    ],
                    [None, None, "As Above", None],
                    [None, None, "20mm Arissed", None],
                    ["-", "Underbench\nManufacturer\nColour & Finish", "Polytec\nAston White Matte", None],
                ],
            )
        ],
        source_pdf="evoca.pdf",
    )
    entry_key = evoca_structured_extractor.LOOKUP_ENTRIES_KEY
    line_key = evoca_structured_extractor.LOOKUP_LINES_KEY
    evoca_structured_extractor._rescue_missing_values(
        structured,
        {
            1: {
                entry_key: [
                    {"key": "kitchen", "label": "Kitchen", "value": ""},
                    {"key": "benchtops", "label": "Benchtops", "value": ""},
                    {"key": "manufacturer", "label": "Manufacturer", "value": "Quantum Quartz"},
                    {"key": "colour", "label": "Colour", "value": ""},
                    {"key": "island colour", "label": "Island Colour", "value": "As Above"},
                    {"key": "edge profile", "label": "Edge Profile", "value": "20mm Arissed"},
                    {"key": "underbench", "label": "Underbench", "value": ""},
                ],
                line_key: [
                    _raw_line(100, [("Kitchen", 40)]),
                    _raw_line(120, [("-", 30), ("Benchtops", 45)]),
                    _raw_line(135, [("Manufacturer", 70), ("Quantum", 200), ("Quartz", 238)]),
                    _raw_line(150, [("Colour", 70), ("Statuario", 200), ("Zero", 238)]),
                    _raw_line(165, [("Island", 70), ("Colour", 100), ("As", 200), ("Above", 215)]),
                    _raw_line(180, [("Edge", 70), ("Profile", 95), ("20mm", 200), ("Arissed", 228)]),
                    _raw_line(210, [("-", 30), ("Underbench", 45)]),
                ],
                "manufacturer": ["Quantum Quartz"],
                "island colour": ["As Above"],
                "edge profile": ["20mm Arissed"],
            }
        },
    )

    group = structured["sections"][0]["rooms"][0]["groups"][0]
    assert [(row["label"], row["value"], row["source_method"]) for row in group["rows"]] == [
        ("Manufacturer", "Quantum Quartz", "pdfplumber_text_rescue"),
        ("Colour", "Statuario Zero", "pdfplumber_raw_text_fallback"),
        ("Island Colour", "As Above", "pdfplumber_text_rescue"),
        ("Edge Profile", "20mm Arissed", "pdfplumber_text_rescue"),
    ]
    assert structured["diagnostics"]["raw_text_fallback_groups"] == 1
    assert structured["diagnostics"]["raw_text_fallback_pairs_filled"] == 1


def test_evoca_structured_raw_text_cursor_advances_past_terminal_group() -> None:
    structured = evoca_structured_extractor.extract_evoca_pages(
        [
            _page(
                1,
                [
                    ["15 CABINETS", None, None, ""],
                    ["", "Ensuite", None, None],
                    ["-", "Benchtops\nManufacturer\nColour\nEdge Profile", "", None],
                    ["", "Ensuite 2", None, None],
                    ["-", "Benchtops\nManufacturer\nColour\nEdge Profile", "Not Applicable", None],
                    ["", "Powder", None, None],
                    ["-", "Benchtops\nManufacturer\nColour\nEdge Profile", "", None],
                ],
            )
        ],
        source_pdf="evoca.pdf",
    )

    line_key = evoca_structured_extractor.LOOKUP_LINES_KEY
    evoca_structured_extractor._rescue_missing_values(
        structured,
        {
            1: {
                line_key: [
                    _raw_line(100, [("Ensuite", 40)]),
                    _raw_line(120, [("-", 30), ("Benchtops", 45)]),
                    _raw_line(135, [("Manufacturer", 70), ("Quantum", 200), ("Quartz", 238)]),
                    _raw_line(150, [("Colour", 70), ("Verona", 200), ("Gold", 235)]),
                    _raw_line(165, [("Edge", 70), ("Profile", 95), ("20mm", 200), ("Arissed", 228)]),
                    _raw_line(200, [("Ensuite", 40), ("2", 78)]),
                    _raw_line(220, [("-", 30), ("Benchtops", 45), ("Not", 200), ("Applicable", 220)]),
                    _raw_line(235, [("Manufacturer", 70)]),
                    _raw_line(250, [("Colour", 70)]),
                    _raw_line(265, [("Edge", 70), ("Profile", 95)]),
                    _raw_line(300, [("Powder", 40)]),
                    _raw_line(320, [("-", 30), ("Benchtops", 45)]),
                    _raw_line(335, [("Manufacturer", 70), ("Quantum", 200), ("Quartz", 238)]),
                    _raw_line(350, [("Colour", 70), ("Polar", 200)]),
                    _raw_line(365, [("Edge", 70), ("Profile", 95), ("20mm", 200), ("Arissed", 228)]),
                ],
            }
        },
    )

    powder = structured["sections"][0]["rooms"][2]
    assert powder["room_label"] == "Powder"
    assert [(row["label"], row["value"], row["source_method"]) for row in powder["groups"][0]["rows"]] == [
        ("Manufacturer", "Quantum Quartz", "pdfplumber_raw_text_fallback"),
        ("Colour", "Polar", "pdfplumber_raw_text_fallback"),
        ("Edge Profile", "20mm Arissed", "pdfplumber_raw_text_fallback"),
    ]


def test_evoca_structured_pairs_drawer_label_only_rows_with_wrapped_values() -> None:
    structured = evoca_structured_extractor.extract_evoca_pages(
        [
            _page(
                1,
                [
                    ["15 CABINETS", None, None, ""],
                    ["", "Kitchen", None, None],
                    [
                        "-",
                        "Drawers",
                        (
                            "1x Set of 4 Drawers with Cutlery Tray to 1st Drawer, 5 x banks of 3 drawers\n"
                            "2 x pot drawers below wall oven/microwave towers\n"
                            "2 x 29L bins"
                        ),
                        None,
                    ],
                    ["", "Standard", None, None],
                    ["", "Pot", None, None],
                    ["", "Bin", None, None],
                ],
            )
        ],
        source_pdf="evoca.pdf",
    )

    rows = structured["sections"][0]["rooms"][0]["groups"][0]["rows"]
    assert [(row["label"], row["value"]) for row in rows] == [
        ("Standard", "1x Set of 4 Drawers with Cutlery Tray to 1st Drawer, 5 x banks of 3 drawers"),
        ("Pot", "2 x pot drawers below wall oven/microwave towers"),
        ("Bin", "2 x 29L bins"),
    ]
    assert all(row["label"] != "Unassigned Source Text" for row in rows)


def test_evoca_structured_splits_unanchored_empty_value_group_header_after_non_terminal_group() -> None:
    structured = evoca_structured_extractor.extract_evoca_pages(
        [
            _page(
                13,
                [
                    ["20 PLUMBING FIXTURES & TAPWARE", None, None, ""],
                    ["", "Bathroom", None, None],
                    [
                        "-",
                        "Bath Mixer / Spout\nModel\nBath Spout Model",
                        "Spin Brushed Nickel In-wall Mixer (SP141-BN)\nOmega Brushed Nickel Swivel 220mm Bath Spout (OMG220-BN)",
                        None,
                    ],
                    ["", "Shower\nMixer\nShower Rail / Rose\nShower Screen\nShower Screen Colour", "", None],
                    [None, None, "Spin Brushed Nickel In-wall Mixer (SP141-BN)", None],
                    [
                        None,
                        None,
                        "Spin Brushed Nickel Shower Rail with Eden Hand Shower Head (R166-BN & EDEN-BN)",
                        None,
                    ],
                    [None, None, "Semi-frameless with Clear Toughened Glass", None],
                    [None, None, "Brushed Nickel", None],
                    ["-", "Accessories & Toilet Suite\nToilet Roll Holder", "Not Applicable", None],
                ],
            )
        ],
        source_pdf="evoca.pdf",
    )

    groups = structured["sections"][0]["rooms"][0]["groups"]
    assert [group["group_label"] for group in groups] == ["Bath Mixer / Spout", "Shower", "Accessories & Toilet Suite"]
    assert [(row["label"], row["value"]) for row in groups[0]["rows"]] == [
        ("Model", "Spin Brushed Nickel In-wall Mixer (SP141-BN)"),
        ("Bath Spout Model", "Omega Brushed Nickel Swivel 220mm Bath Spout (OMG220-BN)"),
    ]
    assert [(row["label"], row["value"]) for row in groups[1]["rows"]] == [
        ("Mixer", "Spin Brushed Nickel In-wall Mixer (SP141-BN)"),
        (
            "Shower Rail / Rose",
            "Spin Brushed Nickel Shower Rail with Eden Hand Shower Head (R166-BN & EDEN-BN)",
        ),
        ("Shower Screen", "Semi-frameless with Clear Toughened Glass"),
        ("Shower Screen Colour", "Brushed Nickel"),
    ]
    assert all(row["label"] != "Unassigned Source Text" for group in groups for row in group["rows"])


def test_evoca_structured_anchor_synthesis_repairs_same_page_blank_dash_underbench_group() -> None:
    structured = evoca_structured_extractor.extract_evoca_pages(
        [
            _page(
                8,
                [
                    ["15 CABINETS", None, None, ""],
                    ["", "Kitchen", None, None],
                    [
                        "-",
                        "Benchtops\nManufacturer\nColour\nEdge Profile",
                        "Quantum Quartz\nLuna White\n20mm Arissed",
                        None,
                    ],
                    ["-", None, "Polytec", None],
                    [None, None, "Blosson White", None],
                    ["-", "Overhead Cupboards\nManufacturer", "Polytec", None],
                ],
            )
        ],
        source_pdf="evoca.pdf",
    )
    evoca_structured_extractor._repair_missing_anchor_groups(
        structured,
        {
            8: {
                evoca_structured_extractor.LOOKUP_LINES_KEY: [
                    _raw_line(100, [("15", 35), ("CABINETS", 55)]),
                    _raw_line(120, [("Kitchen", 45)]),
                    _raw_line(140, [("-", 35), ("Benchtops", 50)]),
                    _raw_line(160, [("Manufacturer", 70), ("Quantum", 200), ("Quartz", 238)]),
                    _raw_line(180, [("Colour", 70), ("Luna", 200), ("White", 230)]),
                    _raw_line(200, [("Edge", 70), ("Profile", 95), ("20mm", 200), ("Arissed", 230)]),
                    _raw_line(220, [("-", 35), ("Underbench", 50), ("including", 112), ("Island", 160)]),
                    _raw_line(240, [("Manufacturer", 70), ("Polytec", 200)]),
                    _raw_line(260, [("Colour", 70), ("&", 104), ("Finish", 116), ("Blosson", 200), ("White", 238)]),
                    _raw_line(280, [("-", 35), ("Overhead", 50), ("Cupboards", 95)]),
                ]
            }
        },
    )

    room = structured["sections"][0]["rooms"][0]
    assert [group["group_label"] for group in room["groups"]] == [
        "Benchtops",
        "Underbench including Island",
        "Overhead Cupboards",
    ]
    assert [(row["label"], row["value"], row["source_method"]) for row in room["groups"][1]["rows"]] == [
        ("Manufacturer", "Polytec", "pdfplumber_raw_text_anchor_synthesis"),
        ("Colour & Finish", "Blosson White", "pdfplumber_raw_text_anchor_synthesis"),
    ]
    assert all(
        row["label"] != "Unassigned Source Text" for row in room["groups"][0]["rows"]
    )
    assert structured["diagnostics"]["raw_text_anchor_synthesized_same_page_groups"] == 1
    assert structured["diagnostics"]["raw_text_anchor_synthesized_same_page_pairs_filled"] == 2
    assert structured["diagnostics"]["raw_text_cross_page_groups"] == 0


def test_evoca_structured_anchor_synthesis_repairs_same_page_blank_dash_accessories_group() -> None:
    structured = evoca_structured_extractor.extract_evoca_pages(
        [
            _page(
                13,
                [
                    ["20 PLUMBING FIXTURES & TAPWARE", None, None, ""],
                    ["", "Ensuite", None, None],
                    ["-", "Shower\nMixer", "Spin Chrome In-wall Mixer", None],
                    ["-", None, "Spin Chrome Guest Towel Rail (SP53-CH)", None],
                    [None, None, "Spin Chrome Toilet Roll Holder (SP51-CH)", None],
                ],
            )
        ],
        source_pdf="evoca.pdf",
    )
    evoca_structured_extractor._repair_missing_anchor_groups(
        structured,
        {
            13: {
                evoca_structured_extractor.LOOKUP_LINES_KEY: [
                    _raw_line(100, [("20", 35), ("PLUMBING", 55), ("FIXTURES", 120), ("&", 180), ("TAPWARE", 195)]),
                    _raw_line(120, [("Ensuite", 45)]),
                    _raw_line(140, [("-", 35), ("Shower", 50)]),
                    _raw_line(160, [("Mixer", 70), ("Spin", 200), ("Chrome", 230)]),
                    _raw_line(180, [("-", 35), ("Accessories", 50), ("&", 115), ("Toilet", 130), ("Suite", 160)]),
                    _raw_line(200, [("Hand", 70), ("Towel", 100), ("Rail", 130), ("Spin", 200), ("Chrome", 230), ("Guest", 268), ("Towel", 302), ("Rail", 334), ("(SP53-CH)", 360)]),
                    _raw_line(220, [("Toilet", 70), ("Roll", 108), ("Holder", 134), ("Spin", 200), ("Chrome", 230), ("Toilet", 268), ("Roll", 302), ("Holder", 328), ("(SP51-CH)", 365)]),
                ]
            }
        },
    )

    room = structured["sections"][0]["rooms"][0]
    assert [group["group_label"] for group in room["groups"]] == ["Shower", "Accessories & Toilet Suite"]
    assert [(row["label"], row["value"], row["source_method"]) for row in room["groups"][1]["rows"]] == [
        ("Hand Towel Rail", "Spin Chrome Guest Towel Rail (SP53-CH)", "pdfplumber_raw_text_anchor_synthesis"),
        ("Toilet Roll Holder", "Spin Chrome Toilet Roll Holder (SP51-CH)", "pdfplumber_raw_text_anchor_synthesis"),
    ]
    assert all(row["label"] != "Unassigned Source Text" for row in room["groups"][0]["rows"])


def test_evoca_structured_raw_text_fallback_bounds_repeated_drawer_groups() -> None:
    structured = evoca_structured_extractor.extract_evoca_pages(
        [
            _page(
                1,
                [
                    ["15 CABINETS", None, None, ""],
                    ["", "Butlers", None, None],
                    ["-", "Drawers\nStandard\nPot\nBin", "", None],
                    ["", "Laundry", None, None],
                    ["", "Drawers\nStandard", "1 x drawer below laminated laundry tower", None],
                ],
            )
        ],
        source_pdf="evoca.pdf",
    )

    line_key = evoca_structured_extractor.LOOKUP_LINES_KEY
    evoca_structured_extractor._rescue_missing_values(
        structured,
        {
            1: {
                line_key: [
                    _raw_line(100, [("Butlers", 40)]),
                    _raw_line(120, [("-", 30), ("Drawers", 45)]),
                    _raw_line(135, [("Standard", 70), ("1x", 200), ("Set", 218), ("of", 238), ("4", 252)]),
                    _raw_line(150, [("Pot", 70), ("5", 200), ("x", 212), ("Banks", 224), ("of", 255), ("3", 270), ("Drawers", 282)]),
                    _raw_line(165, [("Bin", 70), ("2", 200), ("x", 212), ("29L", 224), ("Bins", 246)]),
                    _raw_line(200, [("Laundry", 40)]),
                    _raw_line(220, [("Drawers", 45)]),
                    _raw_line(235, [("Standard", 70), ("1", 200), ("x", 212), ("drawer", 224), ("below", 258)]),
                ],
            }
        },
    )

    butlers = structured["sections"][0]["rooms"][0]
    laundry = structured["sections"][0]["rooms"][1]
    assert [(row["label"], row["value"], row["source_method"]) for row in butlers["groups"][0]["rows"]] == [
        ("Standard", "1x Set of 4", "pdfplumber_raw_text_fallback"),
        ("Pot", "5 x Banks of 3 Drawers", "pdfplumber_raw_text_fallback"),
        ("Bin", "2 x 29L Bins", "pdfplumber_raw_text_fallback"),
    ]
    assert [(row["label"], row["value"]) for row in laundry["groups"][0]["rows"]] == [
        ("Standard", "1 x drawer below laminated laundry tower")
    ]


def test_evoca_structured_raw_text_fallback_does_not_cross_group_boundary() -> None:
    structured = evoca_structured_extractor.extract_evoca_pages(
        [
            _page(
                1,
                [
                    ["20 PLUMBING FIXTURES & TAPWARE", None, None, ""],
                    ["", "Bathroom", None, None],
                    ["-", "Basin\nModel\nType", "", None],
                    [None, None, "Eden Bench Mount Gloss White (FL135-W)", None],
                    [None, None, "Overmount", None],
                    ["-", "Basin Mixer\nType\nLocation", "", None],
                ],
            )
        ],
        source_pdf="evoca.pdf",
    )
    entry_key = evoca_structured_extractor.LOOKUP_ENTRIES_KEY
    line_key = evoca_structured_extractor.LOOKUP_LINES_KEY
    evoca_structured_extractor._rescue_missing_values(
        structured,
        {
            1: {
                entry_key: [
                    {"key": "bathroom", "label": "Bathroom", "value": ""},
                    {"key": "basin", "label": "Basin", "value": ""},
                    {"key": "model", "label": "Model", "value": "Eden Bench Mount Gloss White (FL135-W)"},
                    {"key": "type", "label": "Type", "value": "Overmount"},
                    {"key": "basin mixer", "label": "Basin Mixer", "value": ""},
                    {"key": "type", "label": "Type", "value": ""},
                    {"key": "location", "label": "Location", "value": "Centre of Basin"},
                ],
                line_key: [
                    _raw_line(100, [("Bathroom", 40)]),
                    _raw_line(120, [("-", 30), ("Basin", 45)]),
                    _raw_line(135, [("Model", 70), ("Eden", 200), ("Bench", 230), ("Mount", 265)]),
                    _raw_line(150, [("Type", 70), ("Overmount", 200)]),
                    _raw_line(180, [("-", 30), ("Basin", 45), ("Mixer", 75)]),
                    _raw_line(195, [("Type", 70)]),
                    _raw_line(210, [("Location", 70), ("Centre", 200), ("of", 235), ("Basin", 250)]),
                ],
                "model": ["Eden Bench Mount Gloss White (FL135-W)"],
                "type": ["Overmount"],
                "location": ["Centre of Basin"],
            }
        },
    )

    bathroom = structured["sections"][0]["rooms"][0]
    basin = bathroom["groups"][0]
    basin_mixer = bathroom["groups"][1]
    assert [(row["label"], row["value"]) for row in basin["rows"]] == [
        ("Model", "Eden Bench Mount Gloss White (FL135-W)"),
        ("Type", "Overmount"),
    ]
    assert [(row["label"], row["value"]) for row in basin_mixer["rows"]] == [
        ("Type", ""),
        ("Location", "Centre of Basin"),
    ]


def test_evoca_structured_raw_text_fallback_uses_label_owned_next_line_value() -> None:
    structured = evoca_structured_extractor.extract_evoca_pages(
        [
            _page(
                1,
                [
                    ["20 PLUMBING FIXTURES & TAPWARE", None, None, ""],
                    ["", "Bathroom", None, None],
                    ["-", "Shower\nMixer\nShower Rail / Rose\nShower Screen", "", None],
                    [None, None, "Semi-frameless with Clear Toughened Glass", None],
                ],
            )
        ],
        source_pdf="evoca.pdf",
    )
    entry_key = evoca_structured_extractor.LOOKUP_ENTRIES_KEY
    line_key = evoca_structured_extractor.LOOKUP_LINES_KEY
    evoca_structured_extractor._rescue_missing_values(
        structured,
        {
            1: {
                entry_key: [
                    {"key": "bathroom", "label": "Bathroom", "value": ""},
                    {"key": "shower", "label": "Shower", "value": ""},
                    {"key": "mixer", "label": "Mixer", "value": ""},
                    {"key": "shower rail / rose", "label": "Shower Rail / Rose", "value": ""},
                    {"key": "shower screen", "label": "Shower Screen", "value": "Semi-frameless with Clear Toughened Glass"},
                ],
                line_key: [
                    _raw_line(100, [("Bathroom", 40)]),
                    _raw_line(120, [("-", 30), ("Shower", 45)]),
                    _raw_line(135, [("Mixer", 70), ("Adler", 200), ("Soho", 230), ("54380", 260)]),
                    _raw_line(150, [("Shower", 70), ("Rail", 105), ("/", 128), ("Rose", 138)]),
                    _raw_line(165, [("Alder", 200), ("Dual", 230), ("Shower", 255), ("Round", 290), ("Rail", 320)]),
                    _raw_line(180, [("Shower", 70), ("Screen", 105), ("Semi-frameless", 200)]),
                ],
                "shower screen": ["Semi-frameless with Clear Toughened Glass"],
            }
        },
    )

    group = structured["sections"][0]["rooms"][0]["groups"][0]
    assert [(row["label"], row["value"], row["source_method"]) for row in group["rows"]] == [
        ("Mixer", "Adler Soho 54380", "pdfplumber_raw_text_fallback"),
        ("Shower Rail / Rose", "Alder Dual Shower Round Rail", "pdfplumber_raw_text_fallback"),
        ("Shower Screen", "Semi-frameless with Clear Toughened Glass", "pdfplumber_text_rescue"),
    ]


def test_evoca_structured_raw_text_fallback_keeps_label_words_inside_value_column() -> None:
    structured = evoca_structured_extractor.extract_evoca_pages(
        [
            _page(
                1,
                [
                    ["20 PLUMBING FIXTURES & TAPWARE", None, None, ""],
                    ["", "Ensuite", None, None],
                    ["-", "Accessories\nToilet Suite\nFloor Waste", "", None],
                ],
            )
        ],
        source_pdf="evoca.pdf",
    )

    line_key = evoca_structured_extractor.LOOKUP_LINES_KEY
    evoca_structured_extractor._rescue_missing_values(
        structured,
        {
            1: {
                line_key: [
                    _raw_line(100, [("Ensuite", 40)]),
                    _raw_line(120, [("-", 30), ("Accessories", 45)]),
                    _raw_line(
                        135,
                        [
                            ("Toilet", 70),
                            ("Suite", 105),
                            ("Lana", 200),
                            ("Rimless", 226),
                            ("Back", 266),
                            ("to", 291),
                            ("Wall", 306),
                            ("Toilet", 333),
                            ("Suite", 368),
                            ("Gloss", 400),
                            ("White", 430),
                            ("(6002-R-W)", 462),
                        ],
                    ),
                    _raw_line(150, [("Floor", 70), ("Waste", 100), ("Tile", 200), ("Insert", 225)]),
                ],
            }
        },
    )

    group = structured["sections"][0]["rooms"][0]["groups"][0]
    assert [(row["label"], row["value"]) for row in group["rows"]] == [
        ("Toilet Suite", "Lana Rimless Back to Wall Toilet Suite Gloss White (6002-R-W)"),
        ("Floor Waste", "Tile Insert"),
    ]


def test_evoca_structured_raw_text_fallback_rejects_footer_continuation() -> None:
    structured = evoca_structured_extractor.extract_evoca_pages(
        [
            _page(
                1,
                [
                    ["15 CABINETS", None, None, ""],
                    ["", "Bathroom", None, None],
                    ["-", "Underbench\nHandles\nDrawer Handle", "", None],
                ],
            )
        ],
        source_pdf="evoca.pdf",
    )
    entry_key = evoca_structured_extractor.LOOKUP_ENTRIES_KEY
    line_key = evoca_structured_extractor.LOOKUP_LINES_KEY
    evoca_structured_extractor._rescue_missing_values(
        structured,
        {
            1: {
                entry_key: [
                    {"key": "bathroom", "label": "Bathroom", "value": ""},
                    {"key": "underbench", "label": "Underbench", "value": ""},
                    {"key": "handles", "label": "Handles", "value": "2163 Voda Profile Handle"},
                    {"key": "drawer handle", "label": "Drawer Handle", "value": ""},
                ],
                line_key: [
                    _raw_line(100, [("Bathroom", 40)]),
                    _raw_line(120, [("-", 30), ("Underbench", 45)]),
                    _raw_line(135, [("Handles", 70), ("2163", 200), ("Voda", 230), ("Profile", 260), ("Handle", 300)]),
                    _raw_line(150, [("Drawer", 70), ("Handle", 105)]),
                    _raw_line(700, [("Page", 200), ("9", 230), ("of", 242), ("83", 255), ("Client", 300), ("Initials:______________________", 335)]),
                ],
                "handles": ["2163 Voda Profile Handle"],
            }
        },
    )

    rows = structured["sections"][0]["rooms"][0]["groups"][0]["rows"]
    assert [(row["label"], row["value"]) for row in rows] == [
        ("Handles", "2163 Voda Profile Handle"),
        ("Drawer Handle", ""),
    ]


def test_evoca_structured_appends_extent_continuation_without_business_label() -> None:
    structured = evoca_structured_extractor.extract_evoca_pages(
        [
            _page(
                15,
                [
                    ["23 TILING / HARD FLOORING", None, None],
                    ["-", "Main Floor Tile\nType\nColour\nFabric\nExtent", "Roller"],
                    [None, None, "Chalk"],
                    [
                        None,
                        None,
                        "Essentials\nTo all Aluminium Sliding Doors & Clear Glazed Windows Excluding Wet Area's & Kitchen / Butlers\nSplashback window**",
                    ],
                ],
            )
        ],
        source_pdf="evoca.pdf",
    )

    rows = structured["sections"][0]["groups"][0]["rows"]
    assert [(row["label"], row["value"]) for row in rows] == [
        ("Type", "Roller"),
        ("Colour", "Chalk"),
        ("Fabric", "Essentials"),
        (
            "Extent",
            "To all Aluminium Sliding Doors & Clear Glazed Windows Excluding Wet Area's & Kitchen / Butlers Splashback window**",
        ),
    ]
    assert all(row["label"] != "Continuation" for row in rows)


def test_evoca_structured_merges_wrapped_shower_rail_before_screen_colour() -> None:
    structured = evoca_structured_extractor.extract_evoca_pages(
        [
            _page(
                13,
                [
                    ["20 PLUMBING FIXTURES & TAPWARE", None, None, ""],
                    ["", "Ensuite", None, None],
                    ["-", "Shower\nMixer\nShower Rail / Rose\nShower Screen\nShower Screen Colour", "", None],
                    [None, None, "Spin Gun Metal In-wall Mixer (SP141-GM)", None],
                    [
                        None,
                        None,
                        "Omega Integrated Gun Metal Shower System with Eden Hand Shower Head & 250mm Round\nMonsoon Shower (OMG02-GM & EDEN-GM & MS250R-GM)",
                        None,
                    ],
                    [None, None, "Semi-frameless with Clear Toughened Glass", None],
                    [None, None, "Gunmetal", None],
                ],
            )
        ],
        source_pdf="evoca.pdf",
    )

    rows = structured["sections"][0]["rooms"][0]["groups"][0]["rows"]
    assert [(row["label"], row["value"]) for row in rows] == [
        ("Mixer", "Spin Gun Metal In-wall Mixer (SP141-GM)"),
        (
            "Shower Rail / Rose",
            "Omega Integrated Gun Metal Shower System with Eden Hand Shower Head & 250mm Round Monsoon Shower (OMG02-GM & EDEN-GM & MS250R-GM)",
        ),
        ("Shower Screen", "Semi-frameless with Clear Toughened Glass"),
        ("Shower Screen Colour", "Gunmetal"),
    ]
    assert all(row["label"] != "Continuation" for row in rows)


def test_evoca_structured_keeps_unowned_extra_value_as_diagnostic() -> None:
    structured = evoca_structured_extractor.extract_evoca_pages(
        [
            _page(
                13,
                [
                    ["20 PLUMBING FIXTURES & TAPWARE", None, None, ""],
                    ["", "Powder", None, None],
                    [
                        "-",
                        "Accessories & Toilet Suite\nHand Towel Rail\nToilet Roll Holder\nToilet Suite",
                        "Spin Gun Metal robe hook (SP54-GM)",
                        None,
                    ],
                    [None, None, "Spin Gun Metal Toilet Roll Holder (SP51-GM)", None],
                    [None, None, "Lana Rimless Back to Wall Toilet Suite Gloss White (6002-R-W)", None],
                    ["", "WC**", None, None],
                ],
            )
        ],
        source_pdf="evoca.pdf",
    )

    rows = structured["sections"][0]["rooms"][0]["groups"][0]["rows"]
    assert [(row["label"], row["value"]) for row in rows] == [
        ("Hand Towel Rail", "Spin Gun Metal robe hook (SP54-GM)"),
        ("Toilet Roll Holder", "Spin Gun Metal Toilet Roll Holder (SP51-GM)"),
        ("Toilet Suite", "Lana Rimless Back to Wall Toilet Suite Gloss White (6002-R-W)"),
        ("Unassigned Source Text", "WC"),
    ]
    assert rows[-1]["is_diagnostic"] is True
    assert all(row["label"] != "Continuation" for row in rows)


def test_evoca_structured_does_not_append_lost_mixer_anchor_to_basin_type() -> None:
    structured = evoca_structured_extractor.extract_evoca_pages(
        [
            _page(
                13,
                [
                    ["20 PLUMBING FIXTURES & TAPWARE", None, None, ""],
                    ["", "Ensuite 2", None, None],
                    ["-", "Basin\nModel\nType", "", None],
                    [None, None, "Byron Bench Mount Gloss White (FL4149-W) with Overflow", None],
                    [None, None, "Overmount", None],
                    [None, None, "Spin Gun Metal Tall Basin Mixer (SP110-GM)", None],
                ],
            )
        ],
        source_pdf="evoca.pdf",
    )

    rows = structured["sections"][0]["rooms"][0]["groups"][0]["rows"]
    assert [(row["label"], row["value"]) for row in rows] == [
        ("Model", "Byron Bench Mount Gloss White (FL4149-W) with Overflow"),
        ("Type", "Overmount"),
        ("Unassigned Source Text", "Spin Gun Metal Tall Basin Mixer (SP110-GM)"),
    ]
    assert rows[-1]["is_diagnostic"] is True


def test_evoca_structured_repairs_cross_page_missing_benchtops_group() -> None:
    structured = evoca_structured_extractor.extract_evoca_pages(
        [
            _page(
                11,
                [
                    ["15 CABINETS", None, None, ""],
                    ["", "Powder", None, None],
                ],
            ),
            _page(
                12,
                [
                    [None, None, "Polar", None],
                    [None, None, "20mm Arissed", None],
                    ["-", "Underbench\nManufacturer", "Polytec", None],
                ],
            ),
        ],
        source_pdf="evoca.pdf",
    )
    evoca_structured_extractor._repair_cross_page_missing_groups(
        structured,
        {
            11: {
                evoca_structured_extractor.LOOKUP_LINES_KEY: [
                    _raw_line(100, [("15", 35), ("CABINETS", 55)]),
                    _raw_line(120, [("Powder", 45)]),
                    _raw_line(140, [("-", 35), ("Benchtops", 50)]),
                    _raw_line(160, [("Manufacturer", 70), ("Quantum", 200), ("Quartz", 238)]),
                    _raw_line(700, [("Page", 200), ("11", 230), ("of", 245), ("91", 260), ("Client", 300)]),
                ]
            },
            12: {
                evoca_structured_extractor.LOOKUP_LINES_KEY: [
                    _raw_line(100, [("Colour", 70), ("Polar", 200)]),
                    _raw_line(120, [("Edge", 70), ("Profile", 95), ("20mm", 200), ("Arissed", 230)]),
                    _raw_line(140, [("-", 35), ("Underbench", 50)]),
                    _raw_line(160, [("Manufacturer", 70), ("Polytec", 200)]),
                ]
            },
        },
    )

    room = structured["sections"][0]["rooms"][0]
    assert room["page_end"] == 12
    assert [group["group_label"] for group in room["groups"]] == ["Benchtops", "Underbench"]
    assert [(row["label"], row["value"], row["source_method"]) for row in room["groups"][0]["rows"]] == [
        ("Manufacturer", "Quantum Quartz", "pdfplumber_raw_text_anchor_synthesis"),
        ("Colour", "Polar", "pdfplumber_raw_text_anchor_synthesis"),
        ("Edge Profile", "20mm Arissed", "pdfplumber_raw_text_anchor_synthesis"),
    ]
    assert room["notes"] == []
    assert structured["diagnostics"]["raw_text_cross_page_groups"] == 1
    assert structured["diagnostics"]["raw_text_cross_page_pairs_filled"] == 3


def test_evoca_structured_repairs_cross_page_missing_basin_mixer_group() -> None:
    structured = evoca_structured_extractor.extract_evoca_pages(
        [
            _page(
                13,
                [
                    ["20 PLUMBING FIXTURES & TAPWARE", None, None, ""],
                    ["", "Ensuite 2", None, None],
                    ["-", "Basin\nModel\nType", "", None],
                    [None, None, "Byron Bench Mount Gloss White (FL4149-W) with Overflow", None],
                    [None, None, "Overmount", None],
                    [None, None, "Spin Gun Metal Tall Basin Mixer (SP110-GM)", None],
                ],
            ),
            _page(14, [["-", "Bath", "Not Applicable", None]]),
        ],
        source_pdf="evoca.pdf",
    )
    evoca_structured_extractor._repair_cross_page_missing_groups(
        structured,
        {
            13: {
                evoca_structured_extractor.LOOKUP_LINES_KEY: [
                    _raw_line(100, [("20", 35), ("PLUMBING", 55), ("FIXTURES", 120), ("&", 180), ("TAPWARE", 195)]),
                    _raw_line(120, [("Ensuite", 45), ("2", 95)]),
                    _raw_line(140, [("-", 35), ("Basin", 50)]),
                    _raw_line(160, [("Model", 70), ("Byron", 200)]),
                    _raw_line(180, [("Type", 70), ("Overmount", 200)]),
                    _raw_line(200, [("-", 35), ("Basin", 50), ("Mixer", 85)]),
                    _raw_line(220, [("Type", 70), ("Spin", 200), ("Gun", 230), ("Metal", 255), ("Tall", 285), ("Basin", 310), ("Mixer", 345), ("(SP110-GM)", 380)]),
                    _raw_line(700, [("Page", 200), ("13", 230), ("of", 245), ("91", 260), ("Client", 300)]),
                ]
            },
            14: {
                evoca_structured_extractor.LOOKUP_LINES_KEY: [
                    _raw_line(100, [("Location", 70), ("Centre", 200), ("of", 235), ("Basin", 250)]),
                    _raw_line(120, [("-", 35), ("Bath", 50), ("Not", 200), ("Applicable", 225)]),
                ]
            },
        },
    )

    room = structured["sections"][0]["rooms"][0]
    assert [group["group_label"] for group in room["groups"]] == ["Basin", "Basin Mixer", "Bath"]
    assert [(row["label"], row["value"], row["source_method"]) for row in room["groups"][1]["rows"]] == [
        ("Type", "Spin Gun Metal Tall Basin Mixer (SP110-GM)", "pdfplumber_raw_text_anchor_synthesis"),
        ("Location", "Centre of Basin", "pdfplumber_raw_text_anchor_synthesis"),
    ]
    assert [(row["label"], row["value"]) for row in room["groups"][0]["rows"]] == [
        ("Model", "Byron Bench Mount Gloss White (FL4149-W) with Overflow"),
        ("Type", "Overmount"),
    ]


def test_evoca_structured_cross_page_repair_removes_owned_group_notes() -> None:
    structured = evoca_structured_extractor.extract_evoca_pages(
        [
            _page(
                15,
                [
                    ["20 PLUMBING FIXTURES & TAPWARE", None, None, ""],
                    ["", "Ensuite 5", None, None],
                    ["-", "Basin\nModel\nType", "", None],
                    [None, None, "Byron Bench Mount Gloss White (FL4149-W) with Overflow", None],
                    [None, None, "Overmount", None],
                    ["", "", "Spin Gun Metal Tall Basin Mixer (SP110-GM)", None],
                ],
            ),
            _page(16, [["", "", "Centre of Basin", None], ["-", "Shower\nMixer", "", None]]),
        ],
        source_pdf="evoca.pdf",
    )
    evoca_structured_extractor._repair_cross_page_missing_groups(
        structured,
        {
            15: {
                evoca_structured_extractor.LOOKUP_LINES_KEY: [
                    _raw_line(100, [("20", 35), ("PLUMBING", 55), ("FIXTURES", 120), ("&", 180), ("TAPWARE", 195)]),
                    _raw_line(120, [("Ensuite", 45), ("5", 95)]),
                    _raw_line(140, [("-", 35), ("Basin", 50)]),
                    _raw_line(160, [("Model", 70), ("Byron", 200)]),
                    _raw_line(180, [("Type", 70), ("Overmount", 200)]),
                    _raw_line(200, [("-", 35), ("Basin", 50), ("Mixer", 85)]),
                    _raw_line(220, [("Type", 70), ("Spin", 200), ("Gun", 230), ("Metal", 255), ("Tall", 285), ("Basin", 310), ("Mixer", 345), ("(SP110-GM)", 380)]),
                ]
            },
            16: {
                evoca_structured_extractor.LOOKUP_LINES_KEY: [
                    _raw_line(100, [("Location", 70), ("Centre", 200), ("of", 235), ("Basin", 250)]),
                    _raw_line(120, [("-", 35), ("Shower", 50)]),
                ]
            },
        },
    )

    room = structured["sections"][0]["rooms"][0]
    basin_rows = room["groups"][0]["rows"]
    assert [(row["label"], row["value"]) for row in basin_rows] == [
        ("Model", "Byron Bench Mount Gloss White (FL4149-W) with Overflow"),
        ("Type", "Overmount"),
    ]
    assert [(row["label"], row["value"]) for row in room["groups"][1]["rows"]] == [
        ("Type", "Spin Gun Metal Tall Basin Mixer (SP110-GM)"),
        ("Location", "Centre of Basin"),
    ]
